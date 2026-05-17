"""
Build XLSX modello finanziario arricchito v2.0.
Estende il modello esistente HALE-Financial-Model-M3.xlsx con:
- Sheet aggiuntivi: Monte Carlo proxy (10000 sims), Sensitivity analysis estesa, NPV scenarios
- Stile coerente brand Firmamento
- Logo + intestazione fiscale
"""

import os
import shutil
import random
import math
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              NamedStyle, Color)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.trendline import Trendline

BASE = "/home/user/HALE/studio-di-fattibilita"
ASSETS = "/home/user/HALE/cad"
OUT_DIR = os.path.join(BASE, "deliverables-FINAL")
LOGO = os.path.join(ASSETS, "LogoFirmamento Technologies.png")
SRC_MODEL = os.path.join(BASE, "allegati", "financial-model", "HALE-Financial-Model-M3.xlsx")
OUT_XLSX = os.path.join(OUT_DIR, "HALE-Financial-Model-ENRICHED-v2.0.xlsx")

# ============== STILE BRAND ==============
NAVY = "0a1a3d"
GOLD = "f0c95c"
GREY_LIGHT = "f4f4f6"
GREY_TEXT = "3a3a3a"
WHITE = "ffffff"
RED = "c03333"
GREEN = "2e7d32"

navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
grey_fill = PatternFill(start_color=GREY_LIGHT, end_color=GREY_LIGHT, fill_type="solid")
red_fill = PatternFill(start_color="ffe0e0", end_color="ffe0e0", fill_type="solid")
green_fill = PatternFill(start_color="e0f5e0", end_color="e0f5e0", fill_type="solid")

bold_white = Font(name="Calibri", size=11, bold=True, color=WHITE)
bold_navy = Font(name="Calibri", size=11, bold=True, color=NAVY)
normal = Font(name="Calibri", size=10, color="1a1a1a")
small_italic = Font(name="Calibri", size=9, italic=True, color=GREY_TEXT)

thin_border = Border(
    left=Side(style="thin", color="cfcfd4"),
    right=Side(style="thin", color="cfcfd4"),
    top=Side(style="thin", color="cfcfd4"),
    bottom=Side(style="thin", color="cfcfd4"),
)
header_border = Border(
    left=Side(style="thin", color=NAVY),
    right=Side(style="thin", color=NAVY),
    top=Side(style="thin", color=NAVY),
    bottom=Side(style="medium", color=GOLD),
)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center", wrap_text=True)

# ============== COPY E ESTENDI MODELLO ==============
def setup_workbook():
    shutil.copy(SRC_MODEL, OUT_XLSX)
    wb = load_workbook(OUT_XLSX)
    return wb

def add_cover_sheet(wb):
    """Sheet 0 cover con logo + dati fiscali."""
    if "Cover" in wb.sheetnames:
        ws = wb["Cover"]
        wb.remove(ws)
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for c in "BCDEFGHIJ":
        ws.column_dimensions[c].width = 14

    # Logo
    try:
        img = XLImage(LOGO)
        img.width = 280
        img.height = 130
        ws.add_image(img, "B2")
    except Exception:
        pass

    # Titolo
    ws.merge_cells("B11:J12")
    ws["B11"] = "Studio di Fattibilita HALE/VTOL"
    ws["B11"].font = Font(name="Calibri", size=22, bold=True, color=NAVY)
    ws["B11"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 30
    ws.row_dimensions[12].height = 18

    ws.merge_cells("B13:J14")
    ws["B13"] = "Modello Finanziario Esteso v2.0 + Sensitivity + Monte Carlo proxy"
    ws["B13"].font = Font(name="Calibri", size=14, italic=True, color=GREY_TEXT)
    ws["B13"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 22
    ws.row_dimensions[14].height = 12

    # Dati fiscali tabella
    ws.merge_cells("B16:J16")
    ws["B16"] = "Identificativi fiscali del proponente"
    ws["B16"].font = bold_white
    ws["B16"].fill = navy_fill
    ws["B16"].alignment = center
    ws.row_dimensions[16].height = 22

    fiscal = [
        ("Denominazione sociale", "FIRMAMENTO TECHNOLOGIES SOCIETA' COOPERATIVA"),
        ("Indirizzo", "Via Brigata Liguria 105 R, 16121 Genova (GE)"),
        ("Partita IVA", "IT03038500991"),
        ("Codice Fiscale", "03038500991"),
        ("REA", "528629"),
        ("PEC", "firmamentotechnologies@pec.it"),
        ("Bando di riferimento", "Coopfond Cooding Prototypes (Legacoop)"),
        ("Versione documento", "Bozza M+11, Maggio 2026"),
        ("Conformita", "D.Lgs. 36/2023 art. 41 + Allegato I.7"),
        ("Metodologia", "NASA SE Handbook Rev 2 + GAO Cost Estimating Guide 2020"),
    ]
    row = 17
    for label, val in fiscal:
        ws.merge_cells(f"B{row}:D{row}")
        ws.merge_cells(f"E{row}:J{row}")
        ws[f"B{row}"] = label
        ws[f"B{row}"].font = bold_navy
        ws[f"B{row}"].fill = grey_fill
        ws[f"B{row}"].alignment = left
        ws[f"B{row}"].border = thin_border
        ws[f"E{row}"] = val
        ws[f"E{row}"].font = normal
        ws[f"E{row}"].alignment = left
        ws[f"E{row}"].border = thin_border
        ws.row_dimensions[row].height = 18
        row += 1

    # Indice sheet
    row += 1
    ws.merge_cells(f"B{row}:J{row}")
    ws[f"B{row}"] = "Contenuto del workbook"
    ws[f"B{row}"].font = bold_white
    ws[f"B{row}"].fill = navy_fill
    ws[f"B{row}"].alignment = center
    ws.row_dimensions[row].height = 22

    row += 1
    sheets_index = [
        ("Cover", "Identificativi + indice"),
        ("Assumptions", "Parametri di input (CapEx, OpEx, Revenue, WACC)"),
        ("CapEx_Y1", "Quadro Economico Y1 Percorso 6A (art. 41 D.Lgs. 36/2023)"),
        ("OpEx_RECONCILED", "OpEx run-rate Y2 €1.18M (baseline + regulatory team)"),
        ("Revenue_RECALIBRATED", "Revenue Y1-Y5 post audit Cluster D"),
        ("CashFlow_10y", "Cash Flow proiezione Y1-Y10 scenario base"),
        ("NPV_IRR", "Calcolo NPV / IRR / Payback / ROI"),
        ("Sensitivity", "Sensitivity analysis su 5 driver primari"),
        ("Scenarios", "Worst / Base / Best scenario comparison"),
        ("MonteCarlo_proxy", "Monte Carlo proxy 10.000 simulazioni (NPV distribution)"),
        ("Funding_Mix", "Mix finanziamento Y1 raccomandato"),
        ("Phase_B_6B", "Phase B 6B HALE R&D budget €5.5-13.5M"),
    ]
    for sname, desc in sheets_index:
        ws.merge_cells(f"B{row}:D{row}")
        ws.merge_cells(f"E{row}:J{row}")
        ws[f"B{row}"] = sname
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        ws[f"B{row}"].fill = grey_fill
        ws[f"B{row}"].alignment = left
        ws[f"B{row}"].border = thin_border
        ws[f"E{row}"] = desc
        ws[f"E{row}"].font = normal
        ws[f"E{row}"].alignment = left
        ws[f"E{row}"].border = thin_border
        ws.row_dimensions[row].height = 16
        row += 1

    # Footer
    row += 2
    ws.merge_cells(f"B{row}:J{row}")
    ws[f"B{row}"] = "Modello finanziario per istruttoria Coopfond/Regione Liguria. Confidence aggregato: MEDIUM-LOW (richiede validazione esterna RINA/DNV per investment-grade VC)."
    ws[f"B{row}"].font = small_italic
    ws[f"B{row}"].alignment = center
    ws.row_dimensions[row].height = 20

def add_sensitivity_sheet(wb):
    if "Sensitivity" in wb.sheetnames:
        wb.remove(wb["Sensitivity"])
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 18

    # Title
    ws.merge_cells("B2:G2")
    ws["B2"] = "Sensitivity Analysis, impatto su NPV 10y di variazione +-20% sui driver primari"
    ws["B2"].font = bold_white
    ws["B2"].fill = navy_fill
    ws["B2"].alignment = center
    ws.row_dimensions[2].height = 28

    headers = ["Driver", "Baseline", "Scenario -20%", "Scenario +20%", "Δ NPV -20%", "Δ NPV +20%"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=4, column=2 + i, value=h)
        cell.font = bold_white
        cell.fill = navy_fill
        cell.alignment = center
        cell.border = header_border
    ws.row_dimensions[4].height = 22

    sensitivity_data = [
        ("Revenue Y1 (€k, RECALIBRATED)", 260, 208, 312, -1.2, +1.2),
        ("Pricing PA (€k/anno baseline)", 75, 60, 90, -0.9, +0.9),
        ("CapEx Y1 (€k)", 1400, 1120, 1680, +0.5, -0.5),
        ("OpEx Y2 run-rate (€k/anno)", 1180, 944, 1416, +1.5, -1.5),
        ("Cooperative engagement (n. su 10)", 8, 6, 10, -0.4, +0.3),
        ("Mix funding committed (%)", 60, 40, 80, -0.7, +0.4),
        ("WACC blended (%)", 12, 14.4, 9.6, -0.6, +0.7),
        ("Break-even year", 5.5, 6.6, 4.4, -0.8, +0.6),
        ("ARR Y3 scale-up (€M)", 2.5, 2.0, 3.0, -1.0, +1.0),
        ("Capital intensity Y10 small fleet (€B)", 1.0, 0.8, 1.5, +0.3, -0.4),
    ]
    row = 5
    for label, baseline, m20, p20, dnpv_m, dnpv_p in sensitivity_data:
        ws.cell(row=row, column=2, value=label).font = bold_navy
        ws.cell(row=row, column=3, value=baseline).font = normal
        ws.cell(row=row, column=4, value=m20).font = normal
        ws.cell(row=row, column=5, value=p20).font = normal
        ws.cell(row=row, column=6, value=dnpv_m).font = normal
        ws.cell(row=row, column=7, value=dnpv_p).font = normal
        for col in range(2, 8):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = center if col > 2 else left
            if row % 2 == 1:
                ws.cell(row=row, column=col).fill = grey_fill
        ws.row_dimensions[row].height = 18
        row += 1

    # Note
    row += 1
    ws.merge_cells(f"B{row}:G{row}")
    ws[f"B{row}"] = ("Note: Δ NPV espresso in €M, NPV scenario base = +€3.5M (Cap. 8 §8.6.1). "
                    "Driver Revenue Y1 e OpEx Y2 sono i piu sensibili. Sensitivity OpEx riflette inclusione regulatory team mandatory.")
    ws[f"B{row}"].font = small_italic
    ws[f"B{row}"].alignment = left
    ws.row_dimensions[row].height = 36

def add_scenarios_sheet(wb):
    if "Scenarios" in wb.sheetnames:
        wb.remove(wb["Scenarios"])
    ws = wb.create_sheet("Scenarios")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for c in "BCDEF":
        ws.column_dimensions[c].width = 22

    ws.merge_cells("B2:F2")
    ws["B2"] = "Scenarios Worst / Base / Best, modello finanziario MVP Y1-Y5 RECALIBRATED post Cluster D"
    ws["B2"].font = bold_white
    ws["B2"].fill = navy_fill
    ws["B2"].alignment = center
    ws.row_dimensions[2].height = 28

    headers = ["Metrica", "Worst (P10)", "Base (P50) RECALIBRATED", "Best (P90)", "Note"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=4, column=2 + i, value=h)
        cell.font = bold_white
        cell.fill = navy_fill
        cell.alignment = center
        cell.border = header_border
    ws.row_dimensions[4].height = 30

    rows = [
        ("Revenue Y1 (€k)", "130", "260", "380", "Post recal Cluster D"),
        ("Revenue Y3 (€M)", "0.8", "2.0", "3.5", "Scale-up SNAI"),
        ("Revenue Y5 (€M)", "2.0", "5.0", "12.0", "Multi-regione + HAPS subscale"),
        ("CapEx Y1 (€k)", "1960", "1400", "970", "Overrun aerospace + contingency"),
        ("OpEx Y2 baseline (€k)", "480", "370", "280", "Tecnico-only"),
        ("OpEx Y2 RECONCILED (+3 FTE reg.) (€k)", "1300", "1180", "1050", "Cap. 5 §5.17 mandatory"),
        ("Break-even cumulato", "Y7+", "Y5-Y6", "Y4", "Post recal"),
        ("NPV 10y (WACC 12%) (€M)", "-2.0", "+3.5", "+22.5", "Sensitivity ampia"),
        ("IRR 10y (%)", "<5", "12-18", "28-40", "Post recal"),
        ("Payback (anni)", "8-10", "6", "3.5", "Post recal"),
        ("P scenario (probabilita)", "20%", "55-60%", "20-25%", "Stima qualitativa"),
    ]
    row = 5
    for r in rows:
        for ci, val in enumerate(r):
            cell = ws.cell(row=row, column=2 + ci, value=val)
            cell.border = thin_border
            if ci == 0:
                cell.font = bold_navy
                cell.alignment = left
            else:
                cell.font = normal
                cell.alignment = center
            if row % 2 == 1:
                cell.fill = grey_fill
            # Color worst/best
            if ci == 1 and any(c.isdigit() for c in str(val)):
                if any(neg in str(val) for neg in ["-", "Y7", "<5", "1960", "1300"]):
                    cell.fill = red_fill
            elif ci == 3 and any(c.isdigit() for c in str(val)):
                if "Y4" in str(val) or "+22.5" in str(val) or "28-40" in str(val):
                    cell.fill = green_fill
        ws.row_dimensions[row].height = 20
        row += 1

    # Verdetto box
    row += 1
    ws.merge_cells(f"B{row}:F{row}")
    ws[f"B{row}"] = ("Verdetto consolidato: scenario base (P 55-60%) prevede HOLD CON PIANO REGOLATORIO RAFFORZATO + GO CONDIZIONATO "
                    "(P 5-15%) al verificarsi simultaneo delle 5 hard conditions C1-C5. Le 3 FTE regulatory team (CISO + DPO + Head Regulatory) "
                    "post Cap. 5 §5.17 sono fixed cost obbligatorio nello scenario operativo.")
    ws[f"B{row}"].font = small_italic
    ws[f"B{row}"].fill = PatternFill(start_color="fdf8e8", end_color="fdf8e8", fill_type="solid")
    ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws[f"B{row}"].border = thin_border
    ws.row_dimensions[row].height = 50

def add_montecarlo_sheet(wb):
    """Monte Carlo proxy: 10.000 sim NPV distribuzione."""
    if "MonteCarlo_proxy" in wb.sheetnames:
        wb.remove(wb["MonteCarlo_proxy"])
    ws = wb.create_sheet("MonteCarlo_proxy")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for c in "BCDEFGHIJ":
        ws.column_dimensions[c].width = 14

    ws.merge_cells("B2:J2")
    ws["B2"] = "Monte Carlo proxy, 10.000 simulazioni NPV 10y (WACC 12%)"
    ws["B2"].font = bold_white
    ws["B2"].fill = navy_fill
    ws["B2"].alignment = center
    ws.row_dimensions[2].height = 28

    # Genera 10000 sims (proxy, deterministico seed)
    random.seed(42)
    sims = []
    for i in range(10000):
        # Driver Revenue Y1 normale: mean 260, sd 50 (range 110-410)
        rev_y1 = max(80, random.gauss(260, 60))
        # CapEx Y1 normale: mean 1400, sd 250 (range 700-2200)
        capex = max(700, random.gauss(1400, 250))
        # OpEx Y2 normale: mean 1180, sd 100 (range 880-1480)
        opex = max(800, random.gauss(1180, 110))
        # WACC normale: mean 12, sd 2
        wacc = max(7, random.gauss(12, 2))
        # Scale-up factor Y3: log-normal
        scale_y3 = max(0.5, random.lognormvariate(0.7, 0.4))  # mean ~2.5x revenue Y1
        # ARR Y5: scale_y3 * 2-3x
        arr_y5 = rev_y1 * scale_y3 * random.uniform(1.8, 3.2)

        # NPV proxy semplificato (€M)
        revenue_stream = sum([
            rev_y1 / 1000,
            rev_y1 / 1000 * 1.8,
            rev_y1 / 1000 * scale_y3,
            arr_y5 / 1000 * 0.7,
            arr_y5 / 1000,
            arr_y5 / 1000 * 1.4,
            arr_y5 / 1000 * 1.8,
            arr_y5 / 1000 * 2.2,
            arr_y5 / 1000 * 2.6,
            arr_y5 / 1000 * 3.0,
        ]) / ((1 + wacc / 100) ** 5)
        cost_stream = sum([
            capex / 1000 + opex / 1000,
            opex / 1000 * 1.1,
            opex / 1000 * 1.3,
            opex / 1000 * 1.5,
            opex / 1000 * 1.7,
            opex / 1000 * 1.9,
            opex / 1000 * 2.0,
            opex / 1000 * 2.1,
            opex / 1000 * 2.2,
            opex / 1000 * 2.3,
        ]) / ((1 + wacc / 100) ** 5)
        npv = revenue_stream - cost_stream
        sims.append(npv)

    # Statistiche
    sims_sorted = sorted(sims)
    n = len(sims)
    mean = sum(sims) / n
    median = sims_sorted[n // 2]
    p10 = sims_sorted[int(n * 0.10)]
    p25 = sims_sorted[int(n * 0.25)]
    p75 = sims_sorted[int(n * 0.75)]
    p90 = sims_sorted[int(n * 0.90)]
    sd = math.sqrt(sum((x - mean) ** 2 for x in sims) / n)
    p_negative = sum(1 for x in sims if x < 0) / n * 100
    p_above_10 = sum(1 for x in sims if x > 10) / n * 100

    # Tabella statistiche
    headers = ["Statistica", "Valore (€M)"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=4, column=2 + i, value=h)
        cell.font = bold_white
        cell.fill = navy_fill
        cell.alignment = center
        cell.border = header_border
    ws.row_dimensions[4].height = 22

    stats = [
        ("Mean", f"{mean:+.2f}"),
        ("Median", f"{median:+.2f}"),
        ("Std deviation", f"{sd:.2f}"),
        ("P10 (worst decile)", f"{p10:+.2f}"),
        ("P25", f"{p25:+.2f}"),
        ("P75", f"{p75:+.2f}"),
        ("P90 (best decile)", f"{p90:+.2f}"),
        ("Min", f"{sims_sorted[0]:+.2f}"),
        ("Max", f"{sims_sorted[-1]:+.2f}"),
        ("% sim NPV < 0", f"{p_negative:.1f}%"),
        ("% sim NPV > +€10M", f"{p_above_10:.1f}%"),
    ]
    row = 5
    for label, val in stats:
        ws.cell(row=row, column=2, value=label).font = bold_navy
        ws.cell(row=row, column=2).alignment = left
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3, value=val).font = normal
        ws.cell(row=row, column=3).alignment = center
        ws.cell(row=row, column=3).border = thin_border
        if row % 2 == 1:
            ws.cell(row=row, column=2).fill = grey_fill
            ws.cell(row=row, column=3).fill = grey_fill
        ws.row_dimensions[row].height = 18
        row += 1

    # Istogramma NPV (bin 1M)
    row += 2
    ws.merge_cells(f"B{row}:J{row}")
    ws[f"B{row}"] = "Distribuzione NPV (istogramma bins €1M)"
    ws[f"B{row}"].font = bold_white
    ws[f"B{row}"].fill = navy_fill
    ws[f"B{row}"].alignment = center
    ws.row_dimensions[row].height = 22

    row += 1
    ws.cell(row=row, column=2, value="Bin (€M)").font = bold_white
    ws.cell(row=row, column=2).fill = navy_fill
    ws.cell(row=row, column=2).alignment = center
    ws.cell(row=row, column=2).border = header_border
    ws.cell(row=row, column=3, value="Count").font = bold_white
    ws.cell(row=row, column=3).fill = navy_fill
    ws.cell(row=row, column=3).alignment = center
    ws.cell(row=row, column=3).border = header_border
    ws.cell(row=row, column=4, value="%").font = bold_white
    ws.cell(row=row, column=4).fill = navy_fill
    ws.cell(row=row, column=4).alignment = center
    ws.cell(row=row, column=4).border = header_border
    ws.row_dimensions[row].height = 22

    # Bin: da -10 a +30 a step 2
    bins = {}
    for npv in sims:
        b = int(npv // 2) * 2
        if b < -10:
            b = -10
        if b > 28:
            b = 28
        bins[b] = bins.get(b, 0) + 1

    chart_start = row + 1
    bin_keys = sorted(bins.keys())
    for b in bin_keys:
        row += 1
        cnt = bins[b]
        pct = cnt / n * 100
        ws.cell(row=row, column=2, value=f"{b} to {b+2}").font = normal
        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3, value=cnt).font = normal
        ws.cell(row=row, column=3).alignment = center
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4, value=f"{pct:.1f}%").font = normal
        ws.cell(row=row, column=4).alignment = center
        ws.cell(row=row, column=4).border = thin_border
        if row % 2 == 1:
            for c in [2, 3, 4]:
                ws.cell(row=row, column=c).fill = grey_fill
        ws.row_dimensions[row].height = 16

    # Chart bar
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = "NPV 10y distribuzione, Monte Carlo proxy (10.000 sim)"
    chart.y_axis.title = "Numero simulazioni"
    chart.x_axis.title = "NPV bin (€M)"
    data = Reference(ws, min_col=3, min_row=chart_start, max_row=row, max_col=3)
    cats = Reference(ws, min_col=2, min_row=chart_start, max_row=row, max_col=2)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20
    ws.add_chart(chart, "F4")

    # Caveat
    row += 2
    ws.merge_cells(f"B{row}:J{row}")
    ws[f"B{row}"] = ("CAVEAT: Monte Carlo proxy basato su assunzioni gaussiane (Revenue, CapEx, OpEx) + log-normale (Scale Y3). "
                    "Modello SEMPLIFICATO per validazione preliminare. NON sostituisce validazione esterna RINA/DNV con DCF dettagliato + correlation matrix risk drivers. "
                    "Confidence aggregato: LOW (Regola 7 epistemic-rigor, base-rate aerospace cost overrun 30-150% GAO 2020).")
    ws[f"B{row}"].font = small_italic
    ws[f"B{row}"].fill = PatternFill(start_color="fdf8e8", end_color="fdf8e8", fill_type="solid")
    ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws[f"B{row}"].border = thin_border
    ws.row_dimensions[row].height = 60

def add_funding_mix_sheet(wb):
    if "Funding_Mix" in wb.sheetnames:
        wb.remove(wb["Funding_Mix"])
    ws = wb.create_sheet("Funding_Mix")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 22

    ws.merge_cells("B2:G2")
    ws["B2"] = "Mix finanziamento raccomandato Y1 Percorso 6A (€0.75-1.75M target)"
    ws["B2"].font = bold_white
    ws["B2"].fill = navy_fill
    ws["B2"].alignment = center
    ws.row_dimensions[2].height = 28

    headers = ["Fonte", "Target €k", "%", "Status M+3", "Owner action", "Deadline"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=4, column=2 + i, value=h)
        cell.font = bold_white
        cell.fill = navy_fill
        cell.alignment = center
        cell.border = header_border
    ws.row_dimensions[4].height = 26

    rows = [
        ("Coopfond Cooding Prototypes 2026", "50", "5%", "DR-002 pending", "financial-cfo + snai", "M+3"),
        ("Coopfond Cooding-Invest", "150-300", "15-20%", "Q2 2026", "business-model + Coopfond", "M+6"),
        ("Regione Liguria FESR 2021-2027", "300-500", "25-40%", "OQ-010 LoI", "snai + Regione", "M+9"),
        ("PNRR Aerospazio / IS4Aerospace", "0-300", "0-20%", "Partnership Polito", "aerospace-SE + MIMIT", "M+9"),
        ("Equity privato (founder + seed)", "200-500", "15-35%", "Round seed Q1 2026", "CEO + CFO", "M+6"),
        ("R&D tax credit (L. 160/2019)", "50-150", "5-15%", "Cumulabile post-spesa", "CFO", "Y1 ex-post"),
        ("TOTALE Y1", "750-1750", "100%", "Mix da consolidare", "CFO", "M+10 (60% committed)"),
    ]
    row = 5
    for r in rows:
        is_total = (r[0] == "TOTALE Y1")
        for ci, val in enumerate(r):
            cell = ws.cell(row=row, column=2 + ci, value=val)
            cell.border = thin_border
            if is_total:
                cell.font = bold_navy
                cell.fill = gold_fill
                cell.alignment = center if ci > 0 else left
            else:
                cell.font = bold_navy if ci == 0 else normal
                cell.alignment = left if ci in [0, 3, 4] else center
                if row % 2 == 1:
                    cell.fill = grey_fill
        ws.row_dimensions[row].height = 22
        row += 1

def main():
    wb = setup_workbook()
    add_cover_sheet(wb)
    add_sensitivity_sheet(wb)
    add_scenarios_sheet(wb)
    add_montecarlo_sheet(wb)
    add_funding_mix_sheet(wb)
    wb.save(OUT_XLSX)
    size_mb = os.path.getsize(OUT_XLSX) / 1024 / 1024
    n_sheets = len(wb.sheetnames)
    print(f"XLSX generato: {OUT_XLSX} ({size_mb:.2f} MB, {n_sheets} sheets)")
    print(f"Sheets: {wb.sheetnames}")

if __name__ == "__main__":
    main()
