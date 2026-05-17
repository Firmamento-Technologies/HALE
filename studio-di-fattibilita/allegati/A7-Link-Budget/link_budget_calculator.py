#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==============================================================================
A.7 -- Link Budget Calculator  (HAPS / VTOL UAS)
Firmamento Technologies -- Studio di Fattibilita' HALE / VTOL
==============================================================================

Scope
-----
Calcolo dei link budget RF per i quattro link principali del progetto duale
6A (VTOL pilota Pentema) + 6B (HALE stratosferico):

  1. Percorso 6A -- C2 link (Command & Control)
       a) 2.4 GHz ISM
       b) 5.8 GHz ISM
       c) Iridium L-band (SATCOM secondary)
  2. Percorso 6A -- Payload data downlink
       a) 5 GHz UHF licensed (accordo AGCOM, 16QAM)
       b) 5 GHz UHF licensed (64QAM)
  3. Percorso 6B -- Service link HAPS -> UE
       a) S-band 2.1 GHz  (3GPP NR-NTN n255/n256)
       b) 700 MHz (5G NR rural)
  4. Percorso 6B -- Feeder link HAPS -> Gateway
       a) Ka-band 31 GHz  (HAPS ITU dedicated)
       b) Ka-band 28 GHz  (HAPS gateway alternative)

Reference standards
-------------------
- ITU-R P.618-14 (rain fade slant path)              -- fonti/R-REC-P.618-14.md
- ITU-R P.676-13 (gaseous absorption, oxygen + H2O)  -- sintetico
- ITU-R P.840-9  (cloud and fog attenuation)         -- sintetico
- ITU-R P.837-7  (rainfall rate climatic zones)
- ITU-R P.838-3  (specific rain attenuation k, alpha)
- ITU-R P.839-4  (rain height)
- 3GPP TR 38.811 v15  (NR-NTN channel models, HAPS scenario)
- 3GPP TR 38.821 v16  (Solutions for NR-NTN)
- 3GPP TR 36.763      (IoT-NTN)
- Universidad de Vigo paper -- fonti/Link_budget_uvigo.md

Output
------
- LINK-BUDGET-v1.0.xlsx                (multi-sheet workbook)
- link_budget_C2_6A.png                 (C/N0 vs distance, banda)
- link_budget_service_link_6B.png       (throughput vs SNR per modulation)
- rain_fade_ITU.png                     (rain fade vs availability ITU-R P.618-14)
- coverage_vs_gain.png                  (coverage area vs HAPS antenna gain)

Methodology
-----------
1. Free-space path loss (Friis):  L_fs = 20 log10(4 pi d / lambda)
2. Atmospheric oxygen + water vapour: ITU-R P.676 simplified zenith model
   scaled by 1/sin(elev) for non-zenith paths
3. Rain fade: ITU-R P.618-14 procedure (Section 2.2.1.1) with
   coefficients k, alpha from ITU-R P.838-3
4. Cloud: ITU-R P.840 simplified (K_l * L) factor, ~0.3 dB for sub-10 GHz
5. Scintillation: ITU-R P.618 § 2.4.1 (Karasawa) light approximation
6. Polarization + body loss: empirical
7. C/N0 = EIRP - L_total + G/T + 228.6
8. SNR  = C/N0 - 10 log10(BW)
9. Throughput: practical (Modulation+coding) and Shannon ceiling
10. Margin = SNR_available - SNR_required_for_modulation

Epistemic discipline
--------------------
- All assumptions stated with confidence level (high/med/low)
- Falsifying observations listed at end of each scenario
- Sensitivity sweep (distance, frequency, rain rate)
- Sources cited inline

Author : Telecom-NTN Payload Expert (synthetic)  --  Claude Code
Date   : 2026-05-17
==============================================================================
"""

from __future__ import annotations

import os
import math
from dataclasses import dataclass, asdict, field
from typing import Dict, List, Tuple, Optional

import numpy as np
import pandas as pd
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# =============================================================================
#  0. OUTPUT PATHS
# =============================================================================

OUT_DIR  = "/home/user/HALE/studio-di-fattibilita/allegati/A7-Link-Budget"
os.makedirs(OUT_DIR, exist_ok=True)

XLSX_PATH        = os.path.join(OUT_DIR, "LINK-BUDGET-v1.0.xlsx")
PNG_C2_6A        = os.path.join(OUT_DIR, "link_budget_C2_6A.png")
PNG_SERVICE_6B   = os.path.join(OUT_DIR, "link_budget_service_link_6B.png")
PNG_RAIN         = os.path.join(OUT_DIR, "rain_fade_ITU.png")
PNG_COVERAGE     = os.path.join(OUT_DIR, "coverage_vs_gain.png")


# =============================================================================
#  1. PHYSICAL CONSTANTS
# =============================================================================

C_LIGHT      = 2.998e8          # speed of light [m/s]
BOLTZ_K_DBW  = -228.6           # 10*log10(k)  [dBW/K/Hz]
T0_KELVIN    = 290.0            # reference noise temperature [K]


# =============================================================================
#  2. ITU-R P.838-3  --  Specific rain attenuation coefficients k, alpha
#  Vertical polarization, elevation 0 deg, horizontal pol given for reference.
#  Values fitted from ITU-R P.838-3 Table 1 (selected freqs).
# =============================================================================

# (frequency_GHz, k_H, alpha_H, k_V, alpha_V)
ITU_P838 = {
    1.0  : (0.0000259, 0.9691, 0.0000308, 0.8592),
    2.0  : (0.0000847, 1.0664, 0.0000998, 0.9490),
    4.0  : (0.0001071, 1.6009, 0.0002461, 1.2476),
    6.0  : (0.0007056, 1.5900, 0.0004878, 1.5728),
    7.0  : (0.001425,  1.4745, 0.001344,  1.3797),
    8.0  : (0.002455,  1.4810, 0.002115,  1.3905),
    10.0 : (0.01217,   1.2571, 0.01129,   1.2156),
    12.0 : (0.02386,   1.1825, 0.02455,   1.1216),
    15.0 : (0.04481,   1.1233, 0.05008,   1.0440),
    20.0 : (0.09164,   1.0568, 0.09611,   0.9847),
    25.0 : (0.1571,    0.9991, 0.1533,    0.9491),
    28.0 : (0.2051,    0.9679, 0.1963,    0.9277),
    30.0 : (0.2403,    0.9485, 0.2291,    0.9129),
    31.0 : (0.2581,    0.9395, 0.2455,    0.9061),
    35.0 : (0.3374,    0.9047, 0.3224,    0.8761),
    40.0 : (0.4431,    0.8673, 0.4274,    0.8421),
    47.0 : (0.5926,    0.8270, 0.5728,    0.8068),
}


def itu_p838_coeffs(freq_ghz: float, polarization: str = "circular") -> Tuple[float, float]:
    """Return (k, alpha) interpolated linearly from ITU-R P.838-3 table.
       For circular polarization use the average of H and V (ITU-R P.530 method).
    """
    freqs = sorted(ITU_P838.keys())
    if freq_ghz <= freqs[0]:
        kH, aH, kV, aV = ITU_P838[freqs[0]]
    elif freq_ghz >= freqs[-1]:
        kH, aH, kV, aV = ITU_P838[freqs[-1]]
    else:
        # piecewise log-log interpolation
        for i in range(len(freqs)-1):
            f1, f2 = freqs[i], freqs[i+1]
            if f1 <= freq_ghz <= f2:
                t = (math.log(freq_ghz) - math.log(f1)) / (math.log(f2) - math.log(f1))
                k1H, a1H, k1V, a1V = ITU_P838[f1]
                k2H, a2H, k2V, a2V = ITU_P838[f2]
                kH = math.exp(math.log(k1H) + t*(math.log(k2H) - math.log(k1H)))
                aH = a1H + t*(a2H - a1H)
                kV = math.exp(math.log(k1V) + t*(math.log(k2V) - math.log(k1V)))
                aV = a1V + t*(a2V - a1V)
                break
    if polarization == "horizontal":
        return kH, aH
    if polarization == "vertical":
        return kV, aV
    # circular = average  (ITU-R P.530-17 simplified)
    return 0.5*(kH + kV), 0.5*(aH + aV)


# =============================================================================
#  3. ITU-R P.618-14  --  Rain attenuation procedure
#  (Section 2.2.1.1 -- simplified, elevation > 5 deg)
#  Adapted for Italia North (Liguria 44 N, 9 E) -- zone K (R_0.01 = 32 mm/h).
# =============================================================================

# ITU-R P.837 climatic zones R_0.01 [mm/h]
ITU_RAIN_ZONES = {
    "A" : 8,    "B" : 12,   "C" : 15,   "D" : 19,   "E" : 22,
    "F" : 28,   "G" : 30,   "H" : 32,   "J" : 35,   "K" : 42,
    "L" : 60,   "M" : 63,   "N" : 95,   "P" : 145,  "Q" : 115,
}

# For Italia Nord (Liguria 44N) the ITU-R P.837 climatic map gives zone K.
# Reference: ITU-R Recommendation P.837-7 Annex 2 Figure 1.
# R_0.01 = 42 mm/h  (vs ~28-32 in older charts; conservative 32 used in some
# Italian aerospace literature -- we use 32 as median and 42 as worst-case).
R0_LIGURIA_NOMINAL  = 32.0     # mm/h  (median Italian climatic data)
R0_LIGURIA_WORST    = 42.0     # mm/h  (ITU-R P.837 zone K extreme)

# Rain height for 44 N latitude (ITU-R P.839 simplified)
def rain_height_km(latitude_deg: float) -> float:
    """Rain height h_R [km] above mean sea level, ITU-R P.839-4 simplified."""
    abs_lat = abs(latitude_deg)
    if abs_lat <= 23.0:
        return 5.0
    if abs_lat >= 71.0:
        return 0.0
    # linear between 23 and 71 deg (P.839 simplification)
    return 5.0 - 0.075 * (abs_lat - 23.0)


def rain_attenuation_db(
    freq_ghz       : float,
    elev_deg       : float,
    R0_01          : float,
    h_station_km   : float,
    latitude_deg   : float,
    polarization   : str   = "circular",
    availability   : float = 99.5,
) -> float:
    """ITU-R P.618-14 procedure for slant-path rain attenuation [dB].
       Full implementation of section 2.2.1.1 (steps 1-10).

    Parameters
    ----------
    freq_ghz       : carrier frequency [GHz]
    elev_deg       : elevation angle from earth station to space station [deg]
    R0_01          : rainfall rate exceeded 0.01% of average year [mm/h]
    h_station_km   : earth station altitude AMSL [km]
    latitude_deg   : earth station latitude [deg]
    polarization   : 'horizontal' | 'vertical' | 'circular'
    availability   : % of time link is available (e.g. 99.5 -> 0.5% outage)

    Returns
    -------
    A : rain attenuation at requested availability [dB]
    """
    # ---- guard rails ----
    if R0_01 <= 0.0 or elev_deg <= 0.0:
        return 0.0
    if elev_deg < 5.0:
        elev_deg = 5.0  # ITU-R P.618 ground rule, switch to far-field formula

    elev_rad = math.radians(elev_deg)
    hR  = rain_height_km(latitude_deg)
    if hR <= h_station_km:
        return 0.0

    Re = 8500.0   # effective Earth radius [km]

    # ---- Step 2 : slant-path length L_s [km] ----
    if elev_deg >= 5.0:
        Ls = (hR - h_station_km) / math.sin(elev_rad)
    else:
        # far-field formula (eq. 2)
        num = 2 * (hR - h_station_km)
        den = math.sqrt(math.sin(elev_rad)**2 + 2*(hR - h_station_km)/Re) + math.sin(elev_rad)
        Ls = num / den

    # ---- Step 3 : horizontal projection L_G ----
    LG = Ls * math.cos(elev_rad)

    f  = freq_ghz

    # ---- Step 4 : R_0.01 already given ----

    # ---- Step 5 : specific rain attenuation gamma_R [dB/km] ----
    k, alpha = itu_p838_coeffs(f, polarization)
    gammaR = k * (R0_01 ** alpha)

    # ---- Step 6 : horizontal reduction factor r_0.01 (eq. 5) ----
    r001 = 1.0 / (1.0 + 0.78 * math.sqrt(LG * gammaR / f)
                  - 0.38 * (1.0 - math.exp(-2.0 * LG)))

    # ---- Step 7 : vertical adjustment factor v_0.01 (eq. between 5 and 6) ----
    zeta_rad = math.atan((hR - h_station_km) / (LG * r001))
    zeta_deg = math.degrees(zeta_rad)
    if zeta_deg > elev_deg:
        LR = (LG * r001) / math.cos(elev_rad)
    else:
        LR = (hR - h_station_km) / math.sin(elev_rad)
    abs_lat = abs(latitude_deg)
    if abs_lat < 36:
        chi = 36 - abs_lat
    else:
        chi = 0.0
    chi_rad = math.radians(chi)
    v001 = 1.0 / (1.0 + math.sqrt(math.sin(elev_rad))
                  * (31.0 * (1.0 - math.exp(-elev_deg/(1.0 + chi)))
                     * math.sqrt(LR * gammaR) / (f * f) - 0.45))

    # ---- Step 8 : effective path length L_E ----
    LE = LR * v001

    # ---- Step 9 : attenuation @ 0.01% (eq. 7) ----
    A001 = gammaR * LE

    # ---- Step 10 : scale to requested availability (eq. 8) ----
    p_outage = 100.0 - availability       # [%]
    if p_outage <= 0.001:
        p_outage = 0.001
    if abs(p_outage - 0.01) < 1e-9:
        return A001

    # beta computation per text just above eq. 8
    if p_outage >= 1.0 or abs_lat >= 36:
        beta = 0.0
    elif p_outage < 1.0 and abs_lat < 36 and elev_deg >= 25.0:
        beta = -0.005 * (abs_lat - 36)
    else:
        beta = -0.005 * (abs_lat - 36) + 1.8 - 4.25 * math.sin(elev_rad)

    # eq. 8 exponent (note: (1 - p) factor, p in fraction not %)
    p_frac = p_outage / 100.0
    exp_ = -(0.655 + 0.033 * math.log(p_outage)
             - 0.045 * math.log(max(A001, 1e-3))
             - beta * (1.0 - p_frac) * math.sin(elev_rad))
    Ap = A001 * (p_outage / 0.01) ** exp_
    return max(0.0, Ap)


# =============================================================================
#  4. ITU-R P.676  --  Gaseous absorption (oxygen + water vapour)
#  Simplified zenith model for elevation > 5 deg
# =============================================================================

def gaseous_attenuation_db(
    freq_ghz   : float,
    elev_deg   : float,
    h_station_km : float = 0.5,
) -> float:
    """ITU-R P.676-13 simplified zenith gaseous attenuation, scaled to
       slant path by 1/sin(elev). Valid for f < 350 GHz, elev > 5 deg.
       Values represent mid-latitude summer atmosphere, 7.5 g/m3 water vapour.
    """
    f = freq_ghz
    if elev_deg < 5.0:
        elev_deg = 5.0
    elev_rad = math.radians(elev_deg)

    # Zenith oxygen attenuation [dB], simplified from ITU-R P.676 Annex 2
    if f < 6:
        gamma_o = 0.0067 + 0.0001 * f**2
    elif f < 22:
        gamma_o = 0.0067 + 0.0001 * f**2 + 0.0002*(f-6)**1.5
    elif f < 60:
        # 22-60 GHz: includes 60 GHz oxygen complex
        gamma_o = 0.012 + 0.003*(f-22) + 0.005*(f-22)**1.4
    else:
        gamma_o = 0.30 + 0.001*(f-60)**2

    # Zenith water vapour attenuation [dB] -- 7.5 g/m3
    if f < 10:
        gamma_w = 0.001 + 0.0005 * f
    elif f < 22:
        gamma_w = 0.005 + 0.001*(f-10)**1.5
    elif f < 30:
        # 22.235 GHz water peak
        gamma_w = 0.08 + 0.5 * math.exp(-((f - 22.235)/2.0)**2)
    elif f < 50:
        gamma_w = 0.05 + 0.0015 * f
    else:
        gamma_w = 0.10 + 0.002 * f

    # account for station altitude reducing column thickness (rough scaling)
    altitude_scale = math.exp(-h_station_km / 8.0)  # 8 km scale height for H2O
    gamma_total_zenith = gamma_o + gamma_w * altitude_scale

    # slant path
    return gamma_total_zenith / math.sin(elev_rad)


def cloud_attenuation_db(freq_ghz: float, elev_deg: float, L_cloud_kg_m2: float = 1.0) -> float:
    """ITU-R P.840 cloud attenuation (specific attenuation coefficient K_l).
       L_cloud_kg_m2 = integrated cloud liquid water content [kg/m^2]
       Mediterranean median ~ 0.5-1.0 kg/m^2.
    """
    if elev_deg < 5.0:
        elev_deg = 5.0
    f = freq_ghz
    # Simplified Liebe model -- K_l [(dB/km)/(g/m^3)]
    Kl = 0.000244 * f * (f + 18.0) / 100.0  # crude but monotonic
    A_cloud = Kl * L_cloud_kg_m2 / math.sin(math.radians(elev_deg))
    return A_cloud


def free_space_loss_db(freq_hz: float, distance_m: float) -> float:
    """Friis free-space path loss [dB]."""
    if distance_m <= 0:
        return 0.0
    lam = C_LIGHT / freq_hz
    return 20.0 * math.log10(4.0 * math.pi * distance_m / lam)


# =============================================================================
#  5. MODULATION / CODING -- SNR REQUIREMENTS  (3GPP + DVB-S2X reference)
# =============================================================================

MODULATIONS = {
    # name          : (spectral_eff bps/Hz, SNR_req dB at PER 1e-3, source)
    "BPSK 1/2"      : (0.45,  -2.5, "DVB-S2X"),
    "QPSK 1/2"      : (1.00,   1.0, "DVB-S2X"),
    "QPSK 3/4"      : (1.50,   4.0, "DVB-S2X / 3GPP"),
    "QPSK 5/6"      : (1.65,   5.5, "3GPP NR-NTN"),
    "16QAM 1/2"     : (2.00,   8.0, "3GPP NR / DVB-S2X"),
    "16QAM 3/4"     : (3.00,  10.5, "3GPP NR-NTN"),
    "16QAM 5/6"     : (3.32,  11.5, "DVB-S2X"),
    "64QAM 3/4"     : (4.50,  16.5, "3GPP NR / LTE"),
    "64QAM 5/6"     : (5.00,  18.5, "3GPP NR / LTE"),
    "256QAM 3/4"    : (6.00,  22.5, "3GPP NR Cat.18+"),
    "256QAM 5/6"    : (6.65,  24.0, "DVB-S2X"),
}


def shannon_bps(snr_db: float, bw_hz: float) -> float:
    """Shannon capacity [bps]."""
    snr_lin = 10.0 ** (snr_db/10.0)
    return bw_hz * math.log2(1.0 + snr_lin)


def practical_bps(modulation: str, bw_hz: float) -> float:
    """Practical throughput at a given modulation [bps]."""
    eff, _, _ = MODULATIONS[modulation]
    return eff * bw_hz


def best_modulation_for_snr(snr_available_db: float, margin_db: float = 3.0) -> Tuple[str, float, float]:
    """Pick highest spectral efficiency whose SNR_req+margin <= SNR_available.
       Returns (mod_name, spectral_eff, SNR_req_with_margin)."""
    best = None
    for name, (eff, snr_req, _) in MODULATIONS.items():
        if snr_req + margin_db <= snr_available_db:
            if best is None or eff > best[1]:
                best = (name, eff, snr_req + margin_db)
    return best or ("BPSK 1/2", 0.45, -2.5 + margin_db)


# =============================================================================
#  6. LINK BUDGET DATA CLASS
# =============================================================================

@dataclass
class LinkBudget:
    """Single-config link budget container."""

    # identity
    name              : str    = ""
    direction         : str    = ""        # 'uplink' | 'downlink' | 'feeder' | 'C2'
    band_label        : str    = ""
    description       : str    = ""

    # geometry
    freq_hz           : float  = 2.4e9
    distance_km       : float  = 25.0
    elev_deg          : float  = 30.0
    h_station_km      : float  = 0.5      # earth station altitude
    h_haps_km         : float  = 20.0
    latitude_deg      : float  = 44.0
    polarization      : str    = "circular"

    # transmitter
    p_tx_w            : float  = 5.0
    g_tx_dbi          : float  = 6.0
    l_tx_db           : float  = 1.0      # feedline + connectors

    # receiver
    g_rx_dbi          : float  = 0.0
    l_rx_db           : float  = 1.0
    t_sys_k           : float  = 290.0
    nf_db             : float  = 5.0      # receiver noise figure

    # channel
    bw_hz             : float  = 1e6
    availability_pct  : float  = 99.5     # target % availability
    r_rain_mm_h       : float  = R0_LIGURIA_NOMINAL

    # other losses (manual override)
    l_pol_db          : float  = 0.5
    l_scint_db        : float  = 0.3
    l_body_db         : float  = 0.0      # body loss for handheld UE
    l_pointing_db     : float  = 0.5      # mispointing for HAPS antenna
    l_fade_margin_target_db : float = 6.0

    # required modulation (for SNR_required selection)
    modulation_target : str    = "QPSK 1/2"
    coding_overhead   : float  = 0.0      # extra dB needed (FEC margin)

    # ---- derived / computed (filled by .compute()) ----
    eirp_dbw          : float = 0.0
    l_fs_db           : float = 0.0
    l_atm_db          : float = 0.0
    l_rain_db         : float = 0.0
    l_cloud_db        : float = 0.0
    l_other_db        : float = 0.0
    l_total_db        : float = 0.0
    g_t_db_per_k      : float = 0.0
    c_n0_db_hz        : float = 0.0
    snr_db            : float = 0.0
    snr_required_db   : float = 0.0
    margin_db         : float = 0.0
    shannon_bps       : float = 0.0
    practical_bps     : float = 0.0
    verdict           : str   = "n/a"
    notes             : List[str] = field(default_factory=list)

    # ------------------------------------------------------------------
    def compute(self) -> None:
        """Compute full link budget and verdict."""
        # ---- EIRP ----
        p_tx_dbw = 10.0 * math.log10(self.p_tx_w) - 0.0   # W to dBW = 10log10(W)
        # Note: dBW = dBm - 30, here p_tx_w is in Watts so dBW = 10log10(W)
        self.eirp_dbw = p_tx_dbw + self.g_tx_dbi - self.l_tx_db

        # ---- Free-space path loss ----
        d_m = self.distance_km * 1000.0
        self.l_fs_db = free_space_loss_db(self.freq_hz, d_m)

        # ---- Atmospheric (gaseous) ----
        f_ghz = self.freq_hz / 1e9
        self.l_atm_db = gaseous_attenuation_db(f_ghz, self.elev_deg, self.h_station_km)

        # ---- Cloud ----
        self.l_cloud_db = cloud_attenuation_db(f_ghz, self.elev_deg)

        # ---- Rain (ITU-R P.618-14) ----
        self.l_rain_db = rain_attenuation_db(
            freq_ghz     = f_ghz,
            elev_deg     = self.elev_deg,
            R0_01        = self.r_rain_mm_h,
            h_station_km = self.h_station_km,
            latitude_deg = self.latitude_deg,
            polarization = self.polarization,
            availability = self.availability_pct,
        )

        # ---- Other losses ----
        self.l_other_db = (
            self.l_pol_db
            + self.l_scint_db
            + self.l_body_db
            + self.l_pointing_db
        )

        # ---- Total loss ----
        self.l_total_db = (
            self.l_fs_db + self.l_atm_db + self.l_rain_db + self.l_cloud_db
            + self.l_other_db
        )

        # ---- G/T receiver ----
        t_sys_effective = self.t_sys_k * 10.0 ** (self.nf_db/10.0)
        self.g_t_db_per_k = (self.g_rx_dbi - self.l_rx_db) - 10.0 * math.log10(t_sys_effective)

        # ---- C/N0 ----
        self.c_n0_db_hz = (
            self.eirp_dbw - self.l_total_db + self.g_t_db_per_k - BOLTZ_K_DBW
        )

        # ---- SNR ----
        bw_dbhz = 10.0 * math.log10(self.bw_hz)
        self.snr_db = self.c_n0_db_hz - bw_dbhz

        # ---- SNR required ----
        eff, snr_req, _src = MODULATIONS[self.modulation_target]
        self.snr_required_db = snr_req + self.coding_overhead

        # ---- Margin ----
        self.margin_db = self.snr_db - self.snr_required_db

        # ---- Throughput ----
        self.shannon_bps = shannon_bps(self.snr_db, self.bw_hz)
        self.practical_bps = practical_bps(self.modulation_target, self.bw_hz)

        # ---- Verdict ----
        if self.margin_db >= self.l_fade_margin_target_db:
            self.verdict = "OK"
        elif self.margin_db >= 0:
            self.verdict = "MARGINAL"
        else:
            self.verdict = "FAIL"

    # ------------------------------------------------------------------
    def as_dict_summary(self) -> Dict[str, str]:
        """One-line dict for tabulation."""
        return {
            "Link"          : self.name,
            "Banda"         : self.band_label,
            "Freq [GHz]"    : f"{self.freq_hz/1e9:.3f}",
            "Dist [km]"     : f"{self.distance_km:.1f}",
            "Elev [deg]"    : f"{self.elev_deg:.1f}",
            "EIRP [dBW]"    : f"{self.eirp_dbw:.2f}",
            "L_fs [dB]"     : f"{self.l_fs_db:.2f}",
            "L_atm [dB]"    : f"{self.l_atm_db:.2f}",
            "L_rain [dB]"   : f"{self.l_rain_db:.2f}",
            "L_other [dB]"  : f"{self.l_other_db:.2f}",
            "L_total [dB]"  : f"{self.l_total_db:.2f}",
            "G/T [dB/K]"    : f"{self.g_t_db_per_k:.2f}",
            "C/N0 [dB-Hz]"  : f"{self.c_n0_db_hz:.2f}",
            "BW [MHz]"      : f"{self.bw_hz/1e6:.2f}",
            "SNR [dB]"      : f"{self.snr_db:.2f}",
            "SNR_req [dB]"  : f"{self.snr_required_db:.2f}",
            "Margin [dB]"   : f"{self.margin_db:.2f}",
            "Modulation"    : self.modulation_target,
            "Throughput [Mbps]"      : f"{self.practical_bps/1e6:.2f}",
            "Shannon [Mbps]"         : f"{self.shannon_bps/1e6:.2f}",
            "Availability [%]"       : f"{self.availability_pct:.2f}",
            "Verdict"       : self.verdict,
        }


# =============================================================================
#  7. SCENARIO BUILDERS
# =============================================================================

def build_c2_6A_scenarios() -> List[LinkBudget]:
    """Percorso 6A -- C2 link (Command & Control), 5-50 km LOS."""
    scenarios: List[LinkBudget] = []

    # ---- 2.4 GHz ISM @ typical 20 km LOS ----
    s = LinkBudget(
        name              = "6A-C2-2.4ISM-20km",
        direction         = "C2",
        band_label        = "2.4 GHz ISM",
        description       = "C2 RF link 2.4 GHz ISM banda, range LOS 20 km Pentema-area",
        freq_hz           = 2.4e9,
        distance_km       = 20.0,
        elev_deg          = 6.0,             # quasi-orizzontale, AGL ~1500 m
        h_station_km      = 1.1,             # Pentema 1100 m
        h_haps_km         = 1.5,             # VTOL al ceiling AGL
        polarization      = "circular",
        p_tx_w            = 1.0,             # 30 dBm radio C2 typical
        g_tx_dbi          = 8.0,             # ground antenna directional
        l_tx_db           = 1.5,
        g_rx_dbi          = 3.0,             # UAV onboard antenna
        l_rx_db           = 1.0,
        nf_db             = 5.0,
        bw_hz             = 1e6,             # 1 MHz channel (C2 is low BW)
        availability_pct  = 99.9,            # C2 critical, 99.9
        r_rain_mm_h       = R0_LIGURIA_NOMINAL,
        l_pol_db          = 0.5,
        l_scint_db        = 0.2,
        l_pointing_db     = 1.0,
        l_fade_margin_target_db = 12.0,      # SyR/COMMS-007 fade margin C2
        modulation_target = "QPSK 1/2",
    )
    s.notes.append("ISM 2.4 GHz, no AGCOM individual licensing needed; ETSI EN 300 328 compliance.")
    s.notes.append("Falsifying observation: se in test sito Pentema il RSSI tipico < -85 dBm => margine perso, switch SATCOM.")
    scenarios.append(s)

    # ---- 2.4 GHz @ 50 km worst-case ----
    s2 = LinkBudget(
        name              = "6A-C2-2.4ISM-50km",
        direction         = "C2",
        band_label        = "2.4 GHz ISM",
        description       = "C2 RF link 2.4 GHz ISM, worst-case range 50 km",
        freq_hz           = 2.4e9,
        distance_km       = 50.0,
        elev_deg          = 3.0,
        h_station_km      = 1.1,
        h_haps_km         = 2.0,
        polarization      = "circular",
        p_tx_w            = 2.0,
        g_tx_dbi          = 12.0,
        l_tx_db           = 1.5,
        g_rx_dbi          = 5.0,
        l_rx_db           = 1.0,
        nf_db             = 5.0,
        bw_hz             = 1e6,
        availability_pct  = 99.9,
        r_rain_mm_h       = R0_LIGURIA_NOMINAL,
        l_pol_db          = 0.5,
        l_scint_db        = 0.3,
        l_pointing_db     = 1.5,
        l_fade_margin_target_db = 12.0,
        modulation_target = "QPSK 1/2",
    )
    s2.notes.append("Range 50 km e' bordo operativo VTOL; valutare antenna tracking ground.")
    scenarios.append(s2)

    # ---- 5.8 GHz ISM ----
    s3 = LinkBudget(
        name              = "6A-C2-5.8ISM-20km",
        direction         = "C2",
        band_label        = "5.8 GHz ISM",
        description       = "C2 link 5.8 GHz ISM (banda alternativa, throughput superiore)",
        freq_hz           = 5.8e9,
        distance_km       = 20.0,
        elev_deg          = 6.0,
        h_station_km      = 1.1,
        h_haps_km         = 1.5,
        polarization      = "circular",
        p_tx_w            = 1.0,
        g_tx_dbi          = 10.0,
        l_tx_db           = 1.5,
        g_rx_dbi          = 5.0,
        l_rx_db           = 1.0,
        nf_db             = 4.5,
        bw_hz             = 5e6,
        availability_pct  = 99.9,
        r_rain_mm_h       = R0_LIGURIA_NOMINAL,
        l_pol_db          = 0.5,
        l_scint_db        = 0.2,
        l_pointing_db     = 1.0,
        l_fade_margin_target_db = 12.0,
        modulation_target = "QPSK 1/2",
    )
    s3.notes.append("5.8 GHz path loss e' ~7.5 dB superiore a 2.4 GHz; compensato da maggior gain antenna.")
    scenarios.append(s3)

    # ---- Iridium L-band SATCOM secondary ----
    s4 = LinkBudget(
        name              = "6A-C2-IridiumL-SATCOM",
        direction         = "C2",
        band_label        = "Iridium L-band 1.6 GHz",
        description       = "SATCOM Iridium L-band secondary C2 link (shadow-zone fallback)",
        freq_hz           = 1.6e9,
        distance_km       = 781.0,           # tipico slant LEO Iridium @50 deg elev
        elev_deg          = 30.0,            # mean elevation, lat 44 N
        h_station_km      = 1.1,
        h_haps_km         = 781.0,
        polarization      = "circular",
        p_tx_w            = 3.0,             # Iridium Certus modem typical
        g_tx_dbi          = 3.0,             # patch antenna
        l_tx_db           = 2.0,
        g_rx_dbi          = 23.0,            # Iridium satellite phased array (effective per spot)
        l_rx_db           = 2.0,
        nf_db             = 4.0,
        bw_hz             = 41.7e3,          # Iridium NEXT user channel ~41.7 kHz
        availability_pct  = 99.0,            # Iridium spec
        r_rain_mm_h       = R0_LIGURIA_NOMINAL,
        l_pol_db          = 0.5,
        l_scint_db        = 0.5,
        l_pointing_db     = 0.5,
        l_fade_margin_target_db = 6.0,
        modulation_target = "QPSK 1/2",
    )
    s4.notes.append("Iridium Certus throughput up to ~700 kbps; sufficiente per telemetria + comandi.")
    s4.notes.append("Latency RTT ~ 250-300 ms (LEO), compatibile con EASA SC-Light-UAS 250 ms one-way.")
    scenarios.append(s4)

    return scenarios


def build_payload_downlink_6A_scenarios() -> List[LinkBudget]:
    """Percorso 6A -- Payload data downlink (EO/IR/Telecom)."""
    scenarios: List[LinkBudget] = []

    # ---- 5 GHz UHF licensed (5030-5091 MHz, AGCOM banda aeronautica UAS) 16QAM ----
    s = LinkBudget(
        name              = "6A-Payload-5GHz-16QAM-20km",
        direction         = "downlink",
        band_label        = "5 GHz UHF licensed (AGCOM)",
        description       = "Payload data downlink 5 GHz UHF licensed, 16QAM modulation",
        freq_hz           = 5.06e9,          # mid-band 5030-5091 MHz aeronautical
        distance_km       = 20.0,
        elev_deg          = 6.0,
        h_station_km      = 1.1,
        h_haps_km         = 1.5,
        polarization      = "circular",
        p_tx_w            = 2.0,             # UAV downlink radio
        g_tx_dbi          = 3.0,
        l_tx_db           = 1.5,
        g_rx_dbi          = 18.0,            # ground antenna directional tracking
        l_rx_db           = 1.5,
        nf_db             = 3.5,
        bw_hz             = 20e6,            # 20 MHz channel
        availability_pct  = 99.5,
        r_rain_mm_h       = R0_LIGURIA_NOMINAL,
        l_pol_db          = 0.5,
        l_scint_db        = 0.2,
        l_pointing_db     = 1.0,
        l_fade_margin_target_db = 6.0,
        modulation_target = "16QAM 3/4",
    )
    s.notes.append("Banda 5030-5091 MHz allocata ITU per UAS C2; uso payload secondario possibile in coordinated test.")
    s.notes.append("Compliance AGCOM: PNRF nota AGCOM 18/14/CONS, va negoziata licenza individuale per esercizio.")
    scenarios.append(s)

    # ---- 5 GHz UHF 64QAM (best modulation) ----
    s2 = LinkBudget(
        name              = "6A-Payload-5GHz-64QAM-20km",
        direction         = "downlink",
        band_label        = "5 GHz UHF licensed (AGCOM)",
        description       = "Payload downlink 5 GHz UHF, 64QAM (high throughput EO/IR)",
        freq_hz           = 5.06e9,
        distance_km       = 20.0,
        elev_deg          = 6.0,
        h_station_km      = 1.1,
        h_haps_km         = 1.5,
        polarization      = "circular",
        p_tx_w            = 5.0,             # higher Tx power
        g_tx_dbi          = 5.0,
        l_tx_db           = 1.5,
        g_rx_dbi          = 22.0,            # ground antenna 1.5 m parabolic
        l_rx_db           = 1.5,
        nf_db             = 3.5,
        bw_hz             = 20e6,
        availability_pct  = 99.5,
        r_rain_mm_h       = R0_LIGURIA_NOMINAL,
        l_pol_db          = 0.5,
        l_scint_db        = 0.2,
        l_pointing_db     = 1.0,
        l_fade_margin_target_db = 6.0,
        modulation_target = "64QAM 3/4",
    )
    s2.notes.append("64QAM richiede SNR > 16.5 dB; verificare antenna pointing + tracking automatico.")
    scenarios.append(s2)

    # ---- 50 km worst-case 16QAM ----
    s3 = LinkBudget(
        name              = "6A-Payload-5GHz-16QAM-50km",
        direction         = "downlink",
        band_label        = "5 GHz UHF licensed (AGCOM)",
        description       = "Payload downlink @ 50 km worst-case range",
        freq_hz           = 5.06e9,
        distance_km       = 50.0,
        elev_deg          = 3.0,
        h_station_km      = 1.1,
        h_haps_km         = 2.0,
        polarization      = "circular",
        p_tx_w            = 5.0,
        g_tx_dbi          = 5.0,
        l_tx_db           = 1.5,
        g_rx_dbi          = 28.0,            # large parabolic ground 3 m
        l_rx_db           = 1.5,
        nf_db             = 3.5,
        bw_hz             = 10e6,
        availability_pct  = 99.5,
        r_rain_mm_h       = R0_LIGURIA_NOMINAL,
        l_pol_db          = 0.5,
        l_scint_db        = 0.3,
        l_pointing_db     = 1.5,
        l_fade_margin_target_db = 6.0,
        modulation_target = "16QAM 1/2",
    )
    s3.notes.append("Range 50 km e' al limite; ridurre BW da 20 a 10 MHz per recover SNR.")
    scenarios.append(s3)

    return scenarios


def build_service_link_6B_scenarios() -> List[LinkBudget]:
    """Percorso 6B -- HAPS service link (HAPS -> UE)."""
    scenarios: List[LinkBudget] = []

    # ---- S-band 2.1 GHz NR-NTN n255, downlink (nadir 25 km) ----
    s = LinkBudget(
        name              = "6B-Service-S2.1GHz-25km-NR-NTN",
        direction         = "downlink",
        band_label        = "S-band 2.1 GHz (3GPP NR-NTN n255)",
        description       = "HAPS->UE service link S-band, NR-NTN n255, slant 25 km nadir",
        freq_hz           = 2.1e9,
        distance_km       = 25.0,            # nadir + 10 deg offset
        elev_deg          = 53.0,            # corrispondente
        h_station_km      = 0.1,             # UE a quota mare
        h_haps_km         = 20.0,
        polarization      = "circular",
        p_tx_w            = 25.0,            # 14 dBW HAPS payload per beam
        g_tx_dbi          = 24.0,            # AESA beamforming gain
        l_tx_db           = 1.0,
        g_rx_dbi          = 0.0,             # omni UE handheld
        l_rx_db           = 1.0,
        nf_db             = 7.0,             # UE LNA
        t_sys_k           = 290.0,
        bw_hz             = 20e6,
        availability_pct  = 99.5,
        r_rain_mm_h       = R0_LIGURIA_NOMINAL,
        l_pol_db          = 0.5,
        l_scint_db        = 0.3,
        l_body_db         = 3.0,             # UE body loss handheld
        l_pointing_db     = 0.5,
        l_fade_margin_target_db = 6.0,
        modulation_target = "16QAM 3/4",
    )
    s.notes.append("3GPP TR 38.821 Sec.6.1.3.1 (HAPS-A): suburban LOS conditions assumed.")
    s.notes.append("UE coverage area: ~25-50 km diametro per beam (footprint nadir ~25 km @20 km alt).")
    scenarios.append(s)

    # ---- S-band 2.1 GHz @ slant 100 km (basso elevation 12 deg) ----
    s2 = LinkBudget(
        name              = "6B-Service-S2.1GHz-100km-NR-NTN",
        direction         = "downlink",
        band_label        = "S-band 2.1 GHz (3GPP NR-NTN n255)",
        description       = "HAPS->UE service link S-band, low elevation (slant 100 km)",
        freq_hz           = 2.1e9,
        distance_km       = 100.0,
        elev_deg          = 12.0,
        h_station_km      = 0.1,
        h_haps_km         = 20.0,
        polarization      = "circular",
        p_tx_w            = 25.0,
        g_tx_dbi          = 24.0,
        l_tx_db           = 1.0,
        g_rx_dbi          = 0.0,
        l_rx_db           = 1.0,
        nf_db             = 7.0,
        t_sys_k           = 290.0,
        bw_hz             = 10e6,            # ridotto per recovering SNR
        availability_pct  = 99.5,
        r_rain_mm_h       = R0_LIGURIA_NOMINAL,
        l_pol_db          = 0.5,
        l_scint_db        = 0.5,
        l_body_db         = 3.0,
        l_pointing_db     = 1.0,
        l_fade_margin_target_db = 6.0,
        modulation_target = "QPSK 3/4",
    )
    s2.notes.append("Slant 100 km e' bordo di cella; modulation derating 16QAM->QPSK accettato.")
    scenarios.append(s2)

    # ---- 700 MHz 5G NR rural (downlink) ----
    s3 = LinkBudget(
        name              = "6B-Service-700MHz-25km-5G-NR",
        direction         = "downlink",
        band_label        = "700 MHz (5G NR n28 rural)",
        description       = "HAPS->UE service link 700 MHz 5G NR rural, slant 25 km",
        freq_hz           = 0.706e9,
        distance_km       = 25.0,
        elev_deg          = 53.0,
        h_station_km      = 0.1,
        h_haps_km         = 20.0,
        polarization      = "circular",
        p_tx_w            = 50.0,            # higher power lower band
        g_tx_dbi          = 18.0,            # antenna phys size limit
        l_tx_db           = 1.0,
        g_rx_dbi          = 0.0,
        l_rx_db           = 1.0,
        nf_db             = 7.0,
        t_sys_k           = 290.0,
        bw_hz             = 10e6,
        availability_pct  = 99.7,
        r_rain_mm_h       = R0_LIGURIA_NOMINAL,
        l_pol_db          = 0.5,
        l_scint_db        = 0.3,
        l_body_db         = 3.0,
        l_pointing_db     = 0.5,
        l_fade_margin_target_db = 6.0,
        modulation_target = "16QAM 1/2",
    )
    s3.notes.append("700 MHz path loss minore (better penetration); ma antenna AESA fisica grande.")
    s3.notes.append("Banda 700 MHz allocata TIM/VOD/W3 in Italia -- richiede accordo operator MOCN/RAN-sharing.")
    scenarios.append(s3)

    # ---- Service link UPLINK (UE -> HAPS) -- bottleneck UE power ----
    s4 = LinkBudget(
        name              = "6B-Service-S2.1GHz-UPLINK-25km",
        direction         = "uplink",
        band_label        = "S-band 2.1 GHz (3GPP NR-NTN n256 UL)",
        description       = "UE->HAPS service link uplink, UE 200 mW, bottleneck scenario",
        freq_hz           = 2.1e9,
        distance_km       = 25.0,
        elev_deg          = 53.0,
        h_station_km      = 0.1,
        h_haps_km         = 20.0,
        polarization      = "circular",
        p_tx_w            = 0.2,             # 23 dBm UE typical
        g_tx_dbi          = 0.0,             # omni
        l_tx_db           = 1.0,
        g_rx_dbi          = 30.0,            # HAPS antenna bigger
        l_rx_db           = 1.0,
        nf_db             = 2.0,             # HAPS LNA cooled
        t_sys_k           = 200.0,           # sky-cold + LNA
        bw_hz             = 1.4e6,           # PRACH allocation NR-NTN
        availability_pct  = 99.5,
        r_rain_mm_h       = R0_LIGURIA_NOMINAL,
        l_pol_db          = 0.5,
        l_scint_db        = 0.3,
        l_body_db         = 3.0,
        l_pointing_db     = 0.5,
        l_fade_margin_target_db = 3.0,
        modulation_target = "QPSK 1/2",
    )
    s4.notes.append("Uplink UE-limited: P_tx UE = 23 dBm cap; budget piu' stretto del downlink.")
    s4.notes.append("Mitigazione: PRACH ripetizioni (NR-NTN), riduzione BW UE side, HARQ aggregation.")
    scenarios.append(s4)

    return scenarios


def build_feeder_link_6B_scenarios() -> List[LinkBudget]:
    """Percorso 6B -- Feeder link HAPS -> Gateway (Ka-band)."""
    scenarios: List[LinkBudget] = []

    # ---- Ka 31 GHz HAPS-dedicated, nominal weather ----
    s = LinkBudget(
        name              = "6B-Feeder-Ka31GHz-nominal",
        direction         = "feeder",
        band_label        = "Ka-band 31 GHz (HAPS dedicated ITU)",
        description       = "Feeder HAPS->Gateway, Ka 31 GHz, 99.5% availability zone K",
        freq_hz           = 31.0e9,
        distance_km       = 25.0,
        elev_deg          = 53.0,
        h_station_km      = 0.5,             # gateway ground at 500 m
        h_haps_km         = 20.0,
        polarization      = "vertical",
        p_tx_w            = 5.0,             # HAPS feeder TX
        g_tx_dbi          = 40.0,            # parabolic 0.5 m onboard
        l_tx_db           = 1.5,
        g_rx_dbi          = 50.0,            # Gateway 2 m parabolic
        l_rx_db           = 1.5,
        nf_db             = 2.5,             # Gateway cooled LNA
        t_sys_k           = 150.0,           # cooled LNA + sky temp
        bw_hz             = 250e6,           # 250 MHz channel
        availability_pct  = 99.5,
        r_rain_mm_h       = R0_LIGURIA_NOMINAL,
        l_pol_db          = 0.3,
        l_scint_db        = 1.0,             # higher at Ka
        l_pointing_db     = 1.5,
        l_fade_margin_target_db = 5.0,
        modulation_target = "16QAM 3/4",
    )
    s.notes.append("ITU-R RR Articolo 1.66A: 31-31.3 GHz allocata HAPS gateway downlink.")
    s.notes.append("Rain fade Liguria zona K @ 31 GHz 99.5%: 5-10 dB tipico; 99.9% = 15-22 dB.")
    scenarios.append(s)

    # ---- Ka 31 GHz, 99.9% availability worst-weather ----
    s2 = LinkBudget(
        name              = "6B-Feeder-Ka31GHz-worst",
        direction         = "feeder",
        band_label        = "Ka-band 31 GHz (HAPS dedicated ITU)",
        description       = "Feeder Ka 31 GHz, 99.9% availability (storm scenario)",
        freq_hz           = 31.0e9,
        distance_km       = 25.0,
        elev_deg          = 53.0,
        h_station_km      = 0.5,
        h_haps_km         = 20.0,
        polarization      = "vertical",
        p_tx_w            = 5.0,
        g_tx_dbi          = 40.0,
        l_tx_db           = 1.5,
        g_rx_dbi          = 50.0,
        l_rx_db           = 1.5,
        nf_db             = 2.5,
        t_sys_k           = 200.0,           # higher T noise in rain
        bw_hz             = 250e6,
        availability_pct  = 99.9,
        r_rain_mm_h       = R0_LIGURIA_WORST,
        l_pol_db          = 0.5,
        l_scint_db        = 1.5,
        l_pointing_db     = 1.5,
        l_fade_margin_target_db = 3.0,
        modulation_target = "QPSK 1/2",       # degraded modulation under storm
    )
    s2.notes.append("Site diversity (gateway secondary site 10+ km) raccomandato per recover.")
    s2.notes.append("Falsifying obs: se margin < 0 dB anche a QPSK 1/2 => site diversity OBBLIGATORIA.")
    scenarios.append(s2)

    # ---- Ka 28 GHz gateway alternative ----
    s3 = LinkBudget(
        name              = "6B-Feeder-Ka28GHz-nominal",
        direction         = "feeder",
        band_label        = "Ka-band 28 GHz (HAPS gateway alt)",
        description       = "Feeder HAPS->Gateway alternative 28 GHz, 99.5% availability",
        freq_hz           = 27.95e9,
        distance_km       = 25.0,
        elev_deg          = 53.0,
        h_station_km      = 0.5,
        h_haps_km         = 20.0,
        polarization      = "vertical",
        p_tx_w            = 5.0,
        g_tx_dbi          = 40.0,
        l_tx_db           = 1.5,
        g_rx_dbi          = 50.0,
        l_rx_db           = 1.5,
        nf_db             = 2.5,
        t_sys_k           = 150.0,
        bw_hz             = 200e6,
        availability_pct  = 99.5,
        r_rain_mm_h       = R0_LIGURIA_NOMINAL,
        l_pol_db          = 0.3,
        l_scint_db        = 0.8,
        l_pointing_db     = 1.5,
        l_fade_margin_target_db = 5.0,
        modulation_target = "16QAM 3/4",
    )
    s3.notes.append("ITU-R RR 27.9-28.2 GHz HAPS gateway uplink (banda alternativa).")
    scenarios.append(s3)

    return scenarios


# =============================================================================
#  8. SENSITIVITY ANALYSIS
# =============================================================================

def sensitivity_freq_distance(
    base_lb       : LinkBudget,
    freq_list_ghz : List[float],
    dist_list_km  : List[float],
) -> pd.DataFrame:
    """Sweep frequency x distance; return DataFrame of C/N0 [dB-Hz]."""
    grid = np.zeros((len(freq_list_ghz), len(dist_list_km)))
    for i, f_ghz in enumerate(freq_list_ghz):
        for j, dkm in enumerate(dist_list_km):
            lb = LinkBudget(**asdict(base_lb))
            lb.freq_hz = f_ghz * 1e9
            lb.distance_km = dkm
            lb.compute()
            grid[i, j] = lb.c_n0_db_hz
    df = pd.DataFrame(grid,
                      index=[f"{f:.2f} GHz" for f in freq_list_ghz],
                      columns=[f"{d:.0f} km" for d in dist_list_km])
    df.index.name = "Frequency"
    df.columns.name = "Distance"
    return df


def sensitivity_rain_fade(
    freq_ghz_list   : List[float],
    avail_list_pct  : List[float],
    elev_deg        : float = 53.0,
    R0_01           : float = R0_LIGURIA_NOMINAL,
    h_station_km    : float = 0.5,
    latitude_deg    : float = 44.0,
) -> pd.DataFrame:
    """Sensitivity matrix: rain fade [dB] vs frequency vs availability."""
    grid = np.zeros((len(freq_ghz_list), len(avail_list_pct)))
    for i, f in enumerate(freq_ghz_list):
        for j, ap in enumerate(avail_list_pct):
            grid[i, j] = rain_attenuation_db(
                f, elev_deg, R0_01, h_station_km, latitude_deg,
                polarization="vertical", availability=ap,
            )
    df = pd.DataFrame(grid,
                      index=[f"{f:.1f} GHz" for f in freq_ghz_list],
                      columns=[f"{a:.3f}%" for a in avail_list_pct])
    df.index.name = "Frequency"
    df.columns.name = "Availability"
    return df


def coverage_radius_from_gain(
    haps_alt_km   : float,
    g_tx_dbi      : float,
    eirp_other_dbw_const : float = 14.0,
    freq_hz       : float = 2.1e9,
    bw_hz         : float = 20e6,
    snr_required_db : float = 11.0,
    ue_gt_db_per_k  : float = -25.6,
    l_other_fixed_db : float = 10.0,
    l_rain_db     : float = 0.0,
) -> float:
    """Estimate max cell radius [km] where SNR_required is met.
       Uses simple closed-form: solve C/N0 = SNR_req + 10log10(BW).
       Other losses lumped (incl. rain fade for the band)."""
    eirp_dbw = eirp_other_dbw_const + g_tx_dbi
    bw_dbhz = 10.0 * math.log10(bw_hz)
    c_n0_req = snr_required_db + bw_dbhz
    # rearrange: L_total = EIRP + G/T + 228.6 - C/N0_req
    l_total_allow = eirp_dbw + ue_gt_db_per_k - BOLTZ_K_DBW - c_n0_req
    l_fs_allow = l_total_allow - l_other_fixed_db - l_rain_db
    if l_fs_allow <= 0:
        return 0.0
    # solve L_fs = 20 log10(4 pi d f / c)
    lam = C_LIGHT / freq_hz
    d_m = lam * 10.0 ** (l_fs_allow / 20.0) / (4 * math.pi)
    slant_km = d_m / 1000.0
    # convert slant to ground radius using HAPS altitude
    if slant_km <= haps_alt_km:
        return 0.0
    radius_km = math.sqrt(slant_km**2 - haps_alt_km**2)
    return radius_km


# =============================================================================
#  9. PLOTS
# =============================================================================

def plot_c2_6A(scenarios: List[LinkBudget]) -> None:
    """C/N0 vs distance for C2 link variants."""
    dist_grid = np.linspace(5, 60, 30)
    fig, ax = plt.subplots(figsize=(10, 6))

    for sc in scenarios:
        c_n0_list = []
        for d in dist_grid:
            lb = LinkBudget(**asdict(sc))
            lb.distance_km = float(d)
            lb.compute()
            c_n0_list.append(lb.c_n0_db_hz)
        ax.plot(dist_grid, c_n0_list, marker="o", linewidth=1.2,
                label=f"{sc.band_label} ({sc.p_tx_w:.1f} W, G_tx={sc.g_tx_dbi:.0f} dBi)")

    ax.axhline(60, ls="--", color="red", alpha=0.5,
               label="C/N0 minimo C2 (BW 1 MHz, QPSK 1/2 + 12 dB margin)")
    ax.set_xlabel("Distance / slant range  [km]")
    ax.set_ylabel("C/N0  [dB-Hz]")
    ax.set_title("Percorso 6A -- C2 Link: C/N0 vs Distance per banda")
    ax.legend(loc="lower left", fontsize=8)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(PNG_C2_6A, dpi=140)
    plt.close(fig)


def plot_service_link_6B(scenarios: List[LinkBudget]) -> None:
    """Throughput vs SNR per modulation -- 6B service link."""
    snr_grid = np.linspace(-5, 35, 80)
    fig, ax = plt.subplots(figsize=(10, 6))

    # Shannon curve, BW = 20 MHz reference
    bw_ref = 20e6
    shannon_curve = [shannon_bps(s, bw_ref)/1e6 for s in snr_grid]
    ax.plot(snr_grid, shannon_curve, color="black", linestyle="--",
            label="Shannon ceiling (BW 20 MHz)")

    for mod_name, (eff, snr_req, src) in MODULATIONS.items():
        if "BPSK" in mod_name or "256" in mod_name:
            continue
        rate_mbps = eff * bw_ref / 1e6
        ax.scatter([snr_req], [rate_mbps], s=80, marker="o", label=f"{mod_name}")
        ax.text(snr_req + 0.3, rate_mbps + 1, mod_name, fontsize=7)

    # mark the operating points of the scenarios
    for sc in scenarios:
        rate_mbps = sc.practical_bps / 1e6
        ax.scatter([sc.snr_db], [rate_mbps], marker="*", s=200, color="red", zorder=5,
                   label=f"OP-{sc.name}" if sc == scenarios[0] else None)
        ax.annotate(sc.name.split("-")[-1],
                    xy=(sc.snr_db, rate_mbps),
                    xytext=(sc.snr_db + 1, rate_mbps + 4),
                    fontsize=6.5, color="darkred")

    ax.set_xlabel("SNR available  [dB]")
    ax.set_ylabel("Throughput  [Mbps]  (BW 20 MHz)")
    ax.set_title("Percorso 6B -- Service Link: Throughput vs SNR per Modulation")
    ax.legend(loc="upper left", fontsize=7)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(PNG_SERVICE_6B, dpi=140)
    plt.close(fig)


def plot_rain_fade_ITU() -> None:
    """Rain fade vs availability (ITU-R P.618-14) for multiple bands."""
    avail_grid = np.array([99.0, 99.3, 99.5, 99.7, 99.8, 99.9, 99.95, 99.99])
    fig, ax = plt.subplots(figsize=(10, 6))

    freqs = [2.1, 5.06, 14.0, 28.0, 31.0, 47.0]
    elev_deg = 53.0
    h_station = 0.5
    lat = 44.0

    for f in freqs:
        fades_nominal = [rain_attenuation_db(f, elev_deg, R0_LIGURIA_NOMINAL,
                                             h_station, lat, "vertical", a)
                         for a in avail_grid]
        ax.semilogy(avail_grid, fades_nominal, marker="o", linewidth=1.5,
                    label=f"{f:.2f} GHz (R0.01 = 32 mm/h)")

    # worst-case zone K Italia
    for f in [28.0, 31.0]:
        fades_worst = [rain_attenuation_db(f, elev_deg, R0_LIGURIA_WORST,
                                           h_station, lat, "vertical", a)
                       for a in avail_grid]
        ax.semilogy(avail_grid, fades_worst, marker="x", linewidth=1.0, linestyle="--",
                    label=f"{f:.2f} GHz worst (R0.01 = 42 mm/h)")

    ax.set_xlabel("Availability  [%]")
    ax.set_ylabel("Rain fade  [dB]  (slant path, Liguria 44 N)")
    ax.set_title("ITU-R P.618-14 -- Rain fade vs availability, zona K Italia")
    ax.legend(loc="upper left", fontsize=8)
    ax.grid(True, which="both", alpha=0.3)
    fig.tight_layout()
    fig.savefig(PNG_RAIN, dpi=140)
    plt.close(fig)


def plot_coverage_vs_gain() -> None:
    """Cell radius vs HAPS antenna gain (analytic coverage)."""
    g_grid = np.linspace(5, 50, 60)
    fig, ax = plt.subplots(figsize=(10, 6))

    # (label, freq, BW, rain_fade_db, ue_gt, color)
    configs = [
        ("S-band 2.1 GHz (UE handheld G/T=-25.6 dB/K)", 2.1e9, 20e6, 0.0, -25.6, "tab:blue"),
        ("700 MHz 5G NR (UE G/T=-25.6 dB/K)",           0.706e9, 10e6, 0.0, -25.6, "tab:orange"),
        ("Ka-band 31 GHz (Gateway G/T=+24 dB/K)",       31e9, 250e6, 4.0, 24.0, "tab:green"),
    ]

    for label, f, bw, rain, gt, color in configs:
        radii = [coverage_radius_from_gain(20.0, g, freq_hz=f, bw_hz=bw,
                                           snr_required_db=11.0,
                                           ue_gt_db_per_k=gt,
                                           l_rain_db=rain)
                 for g in g_grid]
        ax.plot(g_grid, radii, linewidth=1.8, label=label, color=color)

    ax.axhline(50, ls="--", color="red", alpha=0.6,
               label="Target copertura cella 50 km (HAPS service area)")
    ax.axhline(25, ls=":",  color="purple", alpha=0.6,
               label="Target spot beam 25 km (cella nadir)")

    ax.set_xlabel("HAPS antenna gain (per beam)  [dBi]")
    ax.set_ylabel("Cell ground radius  [km]")
    ax.set_title("Copertura cella vs HAPS antenna gain (analitica)\n"
                 "HAPS 20 km, BW vario, P_tx 25 W, SNR_req 11 dB")
    ax.legend(loc="upper left", fontsize=8)
    ax.grid(True, alpha=0.3)
    ax.set_ylim(0, 250)
    fig.tight_layout()
    fig.savefig(PNG_COVERAGE, dpi=140)
    plt.close(fig)


# =============================================================================
# 10. EXCEL WORKBOOK BUILDER
# =============================================================================

HEADER_FILL  = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
SUBHDR_FILL  = PatternFill("solid", fgColor="BDD7EE")
SUBHDR_FONT  = Font(bold=True, size=10)
WARN_FILL    = PatternFill("solid", fgColor="FFC7CE")
OK_FILL      = PatternFill("solid", fgColor="C6EFCE")
MARG_FILL    = PatternFill("solid", fgColor="FFEB9C")
THIN_BORDER  = Border(left=Side(style="thin", color="888888"),
                      right=Side(style="thin", color="888888"),
                      top=Side(style="thin", color="888888"),
                      bottom=Side(style="thin", color="888888"))


def style_header(ws, row: int, max_col: int) -> None:
    for c in range(1, max_col+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_subheader(ws, row: int, max_col: int) -> None:
    for c in range(1, max_col+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SUBHDR_FILL
        cell.font = SUBHDR_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def autosize(ws, max_col: int, base: int = 14) -> None:
    for c in range(1, max_col+1):
        col_letter = get_column_letter(c)
        max_len = base
        for cell in ws[col_letter]:
            v = cell.value
            if v is None:
                continue
            length = len(str(v))
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 2, 38)


def add_lb_sheet(wb: Workbook, title: str, scenarios: List[LinkBudget]) -> None:
    """Write a multi-block sheet with one full link-budget block per scenario."""
    ws = wb.create_sheet(title)
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True, color="1F4E78")
    ws["A2"] = ("Source: ITU-R P.618-14, P.676, P.840 + 3GPP TR 38.811/38.821 -- "
                "computed by link_budget_calculator.py")
    ws["A2"].font = Font(italic=True, size=9)

    row = 4
    for sc in scenarios:
        # ---- title bar ----
        ws.cell(row=row, column=1, value=sc.name).font = Font(bold=True, size=12, color="1F4E78")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        ws.cell(row=row, column=1, value=sc.description).font = Font(italic=True, size=9)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 2

        # ---- 4-col block: Section | Parameter | Value | Unit ----
        ws.cell(row=row, column=1, value="Section")
        ws.cell(row=row, column=2, value="Parameter")
        ws.cell(row=row, column=3, value="Value")
        ws.cell(row=row, column=4, value="Unit / Note")
        style_header(ws, row, 4)
        row += 1

        rows = [
            ("Geometry", "Frequency",         f"{sc.freq_hz/1e9:.4f}",  "GHz"),
            ("Geometry", "Distance / slant",  f"{sc.distance_km:.2f}", "km"),
            ("Geometry", "Elevation angle",   f"{sc.elev_deg:.2f}",    "deg"),
            ("Geometry", "Earth-station alt", f"{sc.h_station_km:.2f}", "km AMSL"),
            ("Geometry", "HAPS / UAV alt",    f"{sc.h_haps_km:.2f}",   "km AMSL"),
            ("Geometry", "Polarization",      sc.polarization,         "-"),

            ("Transmitter", "P_tx",                 f"{sc.p_tx_w:.3f}", "W"),
            ("Transmitter", "P_tx",                 f"{10*math.log10(sc.p_tx_w):.2f}", "dBW"),
            ("Transmitter", "G_tx",                 f"{sc.g_tx_dbi:.2f}", "dBi"),
            ("Transmitter", "L_tx feedline",        f"{sc.l_tx_db:.2f}", "dB"),
            ("Transmitter", "EIRP",                 f"{sc.eirp_dbw:.2f}", "dBW   <<"),

            ("Propagation", "Free-space PL",        f"{sc.l_fs_db:.2f}", "dB"),
            ("Propagation", "Gaseous (ITU-R P.676)", f"{sc.l_atm_db:.2f}", "dB"),
            ("Propagation", "Cloud (ITU-R P.840)",  f"{sc.l_cloud_db:.2f}", "dB"),
            ("Propagation", f"Rain fade ({sc.availability_pct:.2f}%, ITU-R P.618-14)",
                                                    f"{sc.l_rain_db:.2f}", "dB"),
            ("Propagation", "Polarization loss",    f"{sc.l_pol_db:.2f}", "dB"),
            ("Propagation", "Scintillation",        f"{sc.l_scint_db:.2f}", "dB"),
            ("Propagation", "Body loss",            f"{sc.l_body_db:.2f}", "dB"),
            ("Propagation", "Pointing loss",        f"{sc.l_pointing_db:.2f}", "dB"),
            ("Propagation", "TOTAL L_path",         f"{sc.l_total_db:.2f}", "dB   <<"),

            ("Receiver", "G_rx",                    f"{sc.g_rx_dbi:.2f}", "dBi"),
            ("Receiver", "L_rx",                    f"{sc.l_rx_db:.2f}", "dB"),
            ("Receiver", "T_sys",                   f"{sc.t_sys_k:.0f}", "K"),
            ("Receiver", "NF",                      f"{sc.nf_db:.2f}", "dB"),
            ("Receiver", "G/T",                     f"{sc.g_t_db_per_k:.2f}", "dB/K   <<"),

            ("Budget", "C/N0",                      f"{sc.c_n0_db_hz:.2f}", "dB-Hz"),
            ("Budget", "BW",                        f"{sc.bw_hz/1e6:.2f}", "MHz"),
            ("Budget", "SNR available",             f"{sc.snr_db:.2f}", "dB"),
            ("Budget", "Modulation",                sc.modulation_target, "-"),
            ("Budget", "SNR required",              f"{sc.snr_required_db:.2f}", "dB"),
            ("Budget", "Link margin",               f"{sc.margin_db:.2f}", "dB   <<"),
            ("Budget", "Margin target",             f"{sc.l_fade_margin_target_db:.2f}", "dB"),
            ("Budget", "Throughput (practical)",    f"{sc.practical_bps/1e6:.2f}", "Mbps"),
            ("Budget", "Throughput (Shannon)",      f"{sc.shannon_bps/1e6:.2f}", "Mbps"),
            ("Budget", "Verdict",                   sc.verdict, "OK/MARG/FAIL"),
        ]
        for r in rows:
            for cc, val in enumerate(r, start=1):
                ws.cell(row=row, column=cc, value=val).border = THIN_BORDER
            # colour verdict row
            if r[1] == "Verdict":
                fill = OK_FILL if sc.verdict == "OK" else (MARG_FILL if sc.verdict == "MARGINAL" else WARN_FILL)
                for cc in range(1, 5):
                    ws.cell(row=row, column=cc).fill = fill
            row += 1

        # ---- notes ----
        if sc.notes:
            ws.cell(row=row, column=1, value="Notes / falsifying observations:").font = Font(bold=True, size=9)
            row += 1
            for n in sc.notes:
                ws.cell(row=row, column=1, value=f"- {n}").font = Font(size=9, italic=True)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                row += 1
        row += 2  # blank rows between scenarios

    autosize(ws, 4, base=24)


def add_cover_sheet(wb: Workbook, all_scenarios: Dict[str, List[LinkBudget]]) -> None:
    ws = wb.create_sheet("Cover", 0)
    ws["A1"] = "A.7 -- LINK BUDGET MODEL v1.0"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
    ws["A2"] = "Firmamento Technologies -- Studio di Fattibilita' HALE / VTOL"
    ws["A2"].font = Font(size=12, italic=True)
    ws["A3"] = "Generato: 2026-05-17 -- Volume 2 Allegato A.7"
    ws["A3"].font = Font(size=10, italic=True, color="888888")

    row = 5
    ws.cell(row=row, column=1, value="Scope").font = Font(bold=True, size=12)
    row += 1
    scope_lines = [
        "Modello di calcolo link budget RF per i quattro link principali HALE/VTOL:",
        "  1. Percorso 6A -- C2 link (2.4 / 5.8 GHz ISM + Iridium SATCOM)",
        "  2. Percorso 6A -- Payload data downlink (5 GHz UHF licensed AGCOM)",
        "  3. Percorso 6B -- Service link HAPS->UE (S-band, 700 MHz 5G NR)",
        "  4. Percorso 6B -- Feeder link HAPS->Gateway (Ka-band 28/31 GHz)",
        "",
        "Conformita': ITU-R P.618-14, P.676-13, P.840-9 + 3GPP TR 38.811/38.821",
    ]
    for line in scope_lines:
        ws.cell(row=row, column=1, value=line)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Conventions").font = Font(bold=True, size=12)
    row += 1
    convs = [
        ("Units", "dBW per power, dBi per antenna gain, dB-Hz per C/N0, dB/K per G/T."),
        ("Polarization", "circular = average of horizontal and vertical (per ITU-R P.530)."),
        ("Rain fade", "ITU-R P.618-14, zona K Italia (R0.01 = 32 mm/h nominale, 42 worst-case)."),
        ("Availability", "99.5% nominal; C2 link 99.9%; feeder Ka 99.9% worst-weather."),
        ("Latitude", "44.0 N (Pentema/Liguria); rain height 5.0 km AMSL."),
        ("Modulation table", "From 3GPP TR 38.821 + DVB-S2X reference."),
    ]
    for k, v in convs:
        ws.cell(row=row, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=v).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Sheets in this workbook").font = Font(bold=True, size=12)
    row += 1
    sheets = [
        ("Cover",                  "Documento di copertura, scope, units"),
        ("LB_C2_6A",               "Link budget C2 Percorso 6A (4 configurazioni)"),
        ("LB_PayloadDownlink_6A",  "Link budget payload downlink 6A (3 configurazioni)"),
        ("LB_ServiceLink_6B",      "Link budget HAPS service link 6B (4 configurazioni)"),
        ("LB_FeederLink_6B",       "Link budget HAPS feeder link Ka (3 configurazioni)"),
        ("Summary",                "Tabella sintetica tutti i link"),
        ("Sensitivity_Freq_Dist",  "Sensitivity C/N0 vs frequency vs distance"),
        ("Sensitivity_Rain",       "Sensitivity rain fade vs frequency vs availability"),
        ("Coverage_Map",           "Coverage analitica raggio cella vs antenna gain"),
        ("Compliance_AGCOM",       "Allocation spettro AGCOM/ITU per i link analizzati"),
        ("Modulation_Table",       "Tabella modulation/coding (SNR required, spectral eff)"),
    ]
    for n, d in sheets:
        ws.cell(row=row, column=1, value=n).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=d)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Verdict (preliminary)").font = Font(bold=True, size=12, color="C00000")
    row += 1
    total = sum(len(v) for v in all_scenarios.values())
    n_ok = sum(1 for v in all_scenarios.values() for sc in v if sc.verdict == "OK")
    n_marg = sum(1 for v in all_scenarios.values() for sc in v if sc.verdict == "MARGINAL")
    n_fail = sum(1 for v in all_scenarios.values() for sc in v if sc.verdict == "FAIL")
    ws.cell(row=row, column=1, value=f"Scenari totali analizzati: {total}")
    row += 1
    ws.cell(row=row, column=1, value=f"  OK (margine >= target): {n_ok}").fill = OK_FILL
    row += 1
    ws.cell(row=row, column=1, value=f"  MARGINAL (margine 0 - target): {n_marg}").fill = MARG_FILL
    row += 1
    ws.cell(row=row, column=1, value=f"  FAIL (margine < 0): {n_fail}").fill = WARN_FILL
    row += 1

    autosize(ws, 6, base=20)


def add_summary_sheet(wb: Workbook, all_scenarios: Dict[str, List[LinkBudget]]) -> None:
    ws = wb.create_sheet("Summary")
    flat: List[Dict] = []
    for sc_list in all_scenarios.values():
        for sc in sc_list:
            flat.append(sc.as_dict_summary())
    if not flat:
        return
    df = pd.DataFrame(flat)
    # header
    for c, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=c, value=col)
    style_header(ws, 1, len(df.columns))
    # body
    for r, rec in enumerate(flat, start=2):
        for c, col in enumerate(df.columns, start=1):
            val = rec[col]
            ws.cell(row=r, column=c, value=val).border = THIN_BORDER
            if col == "Verdict":
                if val == "OK":
                    ws.cell(row=r, column=c).fill = OK_FILL
                elif val == "MARGINAL":
                    ws.cell(row=r, column=c).fill = MARG_FILL
                else:
                    ws.cell(row=r, column=c).fill = WARN_FILL
    autosize(ws, len(df.columns), base=12)


def add_sens_freq_dist_sheet(wb: Workbook) -> None:
    """Sensitivity C/N0 vs frequency vs distance, baseline service link 6B."""
    base = LinkBudget(
        name              = "BASE-6B-S-band",
        band_label        = "S-band 2.1 GHz",
        freq_hz           = 2.1e9,
        distance_km       = 25.0,
        elev_deg          = 53.0,
        h_station_km      = 0.1,
        h_haps_km         = 20.0,
        polarization      = "circular",
        p_tx_w            = 25.0,
        g_tx_dbi          = 24.0,
        l_tx_db           = 1.0,
        g_rx_dbi          = 0.0,
        l_rx_db           = 1.0,
        nf_db             = 7.0,
        bw_hz             = 20e6,
        availability_pct  = 99.5,
        r_rain_mm_h       = R0_LIGURIA_NOMINAL,
        l_pol_db          = 0.5, l_scint_db = 0.3, l_body_db = 3.0, l_pointing_db = 0.5,
        modulation_target = "16QAM 3/4",
    )
    freq_list = [0.7, 2.1, 5.0, 14.0, 28.0, 31.0, 47.0]
    dist_list = [10, 25, 50, 75, 100, 150]
    df = sensitivity_freq_distance(base, freq_list, dist_list)
    ws = wb.create_sheet("Sensitivity_Freq_Dist")
    ws["A1"] = "Sensitivity: C/N0 [dB-Hz]  vs  Frequency x Distance"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A2"] = ("Base config: HAPS 20 km, EIRP 37 dBW, BW 20 MHz, UE G/T -25.6 dB/K, "
                "rain fade zona K Italia, 99.5% availability")
    ws["A2"].font = Font(italic=True, size=9)
    # write headers
    ws.cell(row=4, column=1, value="Frequency \\ Distance")
    for j, c in enumerate(df.columns, start=2):
        ws.cell(row=4, column=j, value=c)
    style_header(ws, 4, len(df.columns)+1)
    for i, idx in enumerate(df.index, start=5):
        ws.cell(row=i, column=1, value=idx).font = Font(bold=True)
        for j, c in enumerate(df.columns, start=2):
            v = df.loc[idx, c]
            ws.cell(row=i, column=j, value=round(float(v), 2)).border = THIN_BORDER
    autosize(ws, len(df.columns)+1, base=14)


def add_sens_rain_sheet(wb: Workbook) -> None:
    freq_list = [2.1, 5.0, 14.0, 28.0, 31.0, 47.0]
    avail_list = [99.0, 99.3, 99.5, 99.7, 99.9, 99.95, 99.99]
    df = sensitivity_rain_fade(freq_list, avail_list, elev_deg=53.0)
    ws = wb.create_sheet("Sensitivity_Rain")
    ws["A1"] = "Sensitivity: Rain fade [dB] vs Frequency x Availability (zona K Italia)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A2"] = ("ITU-R P.618-14, lat 44 N, h_station 0.5 km, elev 53 deg, R0.01 = 32 mm/h")
    ws["A2"].font = Font(italic=True, size=9)
    ws.cell(row=4, column=1, value="Frequency \\ Avail")
    for j, c in enumerate(df.columns, start=2):
        ws.cell(row=4, column=j, value=c)
    style_header(ws, 4, len(df.columns)+1)
    for i, idx in enumerate(df.index, start=5):
        ws.cell(row=i, column=1, value=idx).font = Font(bold=True)
        for j, c in enumerate(df.columns, start=2):
            v = df.loc[idx, c]
            ws.cell(row=i, column=j, value=round(float(v), 2)).border = THIN_BORDER
            if float(v) > 15.0:
                ws.cell(row=i, column=j).fill = WARN_FILL
            elif float(v) > 5.0:
                ws.cell(row=i, column=j).fill = MARG_FILL
            else:
                ws.cell(row=i, column=j).fill = OK_FILL
    autosize(ws, len(df.columns)+1, base=12)


def add_coverage_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Coverage_Map")
    ws["A1"] = "Coverage analitica -- Raggio cella vs HAPS Antenna Gain"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A2"] = ("Formula: closed-form solve da L_fs(d) tale che SNR_required = 11 dB. "
                "HAPS alt 20 km, P_tx 25 W. S-band BW 20 MHz, 700 MHz BW 10 MHz, Ka BW 250 MHz.")
    ws["A2"].font = Font(italic=True, size=9)
    ws["A3"] = ("S-band e 700 MHz: receiver = UE handheld (G/T=-25.6 dB/K). "
                "Ka 31 GHz: receiver = Gateway (G/T=+24 dB/K, rain fade 4 dB).")
    ws["A3"].font = Font(italic=True, size=9)
    ws["A5"] = "Antenna gain [dBi]"
    ws["B5"] = "Cell radius S-band 2.1 GHz [km]"
    ws["C5"] = "Cell radius 700 MHz [km]"
    ws["D5"] = "Cell radius Ka 31 GHz (Gateway) [km]"
    style_header(ws, 5, 4)
    g_list = list(range(5, 51, 2))
    for i, g in enumerate(g_list, start=6):
        r_s = coverage_radius_from_gain(20.0, g, freq_hz=2.1e9, bw_hz=20e6, ue_gt_db_per_k=-25.6)
        r_700 = coverage_radius_from_gain(20.0, g, freq_hz=0.706e9, bw_hz=10e6, ue_gt_db_per_k=-25.6)
        r_ka = coverage_radius_from_gain(20.0, g, freq_hz=31e9, bw_hz=250e6,
                                         ue_gt_db_per_k=24.0, l_rain_db=4.0)
        ws.cell(row=i, column=1, value=g)
        ws.cell(row=i, column=2, value=round(r_s, 1))
        ws.cell(row=i, column=3, value=round(r_700, 1))
        ws.cell(row=i, column=4, value=round(r_ka, 1))
        for cc in range(1, 5):
            ws.cell(row=i, column=cc).border = THIN_BORDER
    autosize(ws, 4, base=22)


def add_compliance_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Compliance_AGCOM")
    ws["A1"] = "Compliance Spettro AGCOM / ITU per i link analizzati"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A2"] = ("Riferimenti: PNRF (D.M. MIMIT), ITU-R RR Art.1.66A HAPS, "
                "AGCOM Delibera 93/26/CONS, 3GPP TR 38.811/38.821")
    ws["A2"].font = Font(italic=True, size=9)

    header = ["Link", "Banda", "Range [MHz]", "Allocation primaria",
              "Status Italia", "Licenza richiesta", "Note"]
    for c, h in enumerate(header, start=1):
        ws.cell(row=4, column=c, value=h)
    style_header(ws, 4, len(header))

    rows = [
        ("6A C2 (2.4 ISM)",     "2.4 GHz ISM",       "2400-2483.5",
         "Mobile + Amateur (license-exempt EN 300 328)",
         "Aperta",            "No (EN 300 328 compliance)",
         "Limite EIRP 100 mW (20 dBm); spread spectrum richiesto."),

        ("6A C2 (5.8 ISM)",     "5.8 GHz ISM",       "5725-5875",
         "Mobile + Amateur (license-exempt EN 300 440)",
         "Aperta",            "No (EN 300 440 compliance)",
         "Limite EIRP 25 mW; aumentabile con licenza individuale."),

        ("6A C2 (Iridium L)",   "L-band Iridium",    "1616-1626.5",
         "MSS (Mobile Satellite Service)",
         "Roaming Iridium",   "Subscription Iridium",
         "Iridium ha global license; Italia parte ITU regione 1."),

        ("6A Payload (5 GHz UHF)", "Banda aeronautica UAS", "5030-5091",
         "AM(R)S - Aeronautical Mobile Route",
         "Aperta (ITU-R P.689 + AGCOM 18/14/CONS)",
         "Si (licenza individuale + parere MIMIT)",
         "Banda allocata ITU per UAS C2 link; payload secondario tollerato."),

        ("6B Service S-band",   "S-band NTN 3GPP n255/n256", "1980-2010 / 2170-2200",
         "MSS + IMT (3GPP NR-NTN)",
         "Allocata operatori MSS (Inmarsat, Eutelsat) + 3GPP NTN",
         "Si (AGCOM, accordo con MSS operator o NTN ecosystem partner)",
         "Spettro condiviso con MSS satellite; coordinamento ITU richiesto."),

        ("6B Service 700 MHz",  "5G NR n28 rural",   "703-733 / 758-788",
         "IMT (5G NR FDD)",
         "Allocata TIM, Vodafone, WindTre (asta 2018)",
         "Si (accordo MOCN/RAN-sharing con operator)",
         "Nessuna banda HAPS-dedicata sotto 1 GHz; richiede sub-leasing."),

        ("6B Feeder Ka 28 GHz", "Ka HAPS gateway uplink", "27900-28200",
         "HAPS Earth-to-space (ITU-R RR Art.1.66A)",
         "Riservata HAPS post-WRC-19; AGCOM disciplina in attesa",
         "Si (licenza individuale + coordinamento internazionale ITU)",
         "Condivisione con FSS; site coordination obbligatoria."),

        ("6B Feeder Ka 31 GHz", "Ka HAPS gateway downlink", "31000-31300",
         "HAPS space-to-Earth (ITU-R RR Art.1.66A)",
         "Riservata HAPS post-WRC-19",
         "Si (licenza individuale)",
         "Banda gemella a 28 GHz; uplink/downlink pairing standard ITU."),
    ]
    for i, r in enumerate(rows, start=5):
        for c, v in enumerate(r, start=1):
            ws.cell(row=i, column=c, value=v).border = THIN_BORDER
            ws.cell(row=i, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws, 7, base=18)


def add_modulation_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Modulation_Table")
    ws["A1"] = "Modulation / Coding -- SNR required for PER 1e-3"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A2"] = ("Reference: 3GPP TR 38.821 Table 6.1.3.3-1 + DVB-S2X Annex M3 "
                "(AWGN channel, soft-decision FEC)")
    ws["A2"].font = Font(italic=True, size=9)
    header = ["Modulation", "Spectral eff [bps/Hz]", "SNR required [dB]", "Source"]
    for c, h in enumerate(header, start=1):
        ws.cell(row=4, column=c, value=h)
    style_header(ws, 4, len(header))
    for i, (name, (eff, snr, src)) in enumerate(MODULATIONS.items(), start=5):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=eff)
        ws.cell(row=i, column=3, value=snr)
        ws.cell(row=i, column=4, value=src)
        for cc in range(1, 5):
            ws.cell(row=i, column=cc).border = THIN_BORDER
    autosize(ws, 4, base=18)


# =============================================================================
# 11. MAIN
# =============================================================================

def main() -> None:
    print("[1/6] Building scenarios ...")
    scen_c2_6A         = build_c2_6A_scenarios()
    scen_payload_6A    = build_payload_downlink_6A_scenarios()
    scen_service_6B    = build_service_link_6B_scenarios()
    scen_feeder_6B     = build_feeder_link_6B_scenarios()

    print("[2/6] Computing link budgets ...")
    all_scen = {
        "LB_C2_6A":             scen_c2_6A,
        "LB_PayloadDownlink_6A": scen_payload_6A,
        "LB_ServiceLink_6B":    scen_service_6B,
        "LB_FeederLink_6B":     scen_feeder_6B,
    }
    for sc_list in all_scen.values():
        for sc in sc_list:
            sc.compute()

    print("[3/6] Generating plots ...")
    plot_c2_6A(scen_c2_6A)
    plot_service_link_6B(scen_service_6B)
    plot_rain_fade_ITU()
    plot_coverage_vs_gain()

    print("[4/6] Building Excel workbook ...")
    wb = Workbook()
    wb.remove(wb.active)
    add_cover_sheet(wb, all_scen)
    add_lb_sheet(wb, "LB_C2_6A",              scen_c2_6A)
    add_lb_sheet(wb, "LB_PayloadDownlink_6A", scen_payload_6A)
    add_lb_sheet(wb, "LB_ServiceLink_6B",     scen_service_6B)
    add_lb_sheet(wb, "LB_FeederLink_6B",      scen_feeder_6B)
    add_summary_sheet(wb, all_scen)
    add_sens_freq_dist_sheet(wb)
    add_sens_rain_sheet(wb)
    add_coverage_sheet(wb)
    add_compliance_sheet(wb)
    add_modulation_sheet(wb)
    wb.save(XLSX_PATH)
    print(f"   -> {XLSX_PATH}")

    print("[5/6] Console summary:")
    print("-" * 90)
    print(f"{'Link':<40}  {'Margin [dB]':>12}  {'Verdict':>10}")
    print("-" * 90)
    for sc_list in all_scen.values():
        for sc in sc_list:
            print(f"{sc.name:<40}  {sc.margin_db:>12.2f}  {sc.verdict:>10}")
    print("-" * 90)

    print("[6/6] Done.")
    print()
    print("Output files:")
    print(f"  {XLSX_PATH}")
    print(f"  {PNG_C2_6A}")
    print(f"  {PNG_SERVICE_6B}")
    print(f"  {PNG_RAIN}")
    print(f"  {PNG_COVERAGE}")


if __name__ == "__main__":
    main()
