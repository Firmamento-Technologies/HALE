#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_risk_register.py
======================

Costruisce l'Allegato A.2 — Risk Register v1.0 dello Studio di Fattibilita'
HALE/VTOL di Firmamento Technologies.

Metodologia: NASA NPR 8000.4 (Continuous Risk Management) + FMECA
(MIL-STD-1629A) + FTA (ARP4761) + ISO 31000.

Output:
- RISK-REGISTER-v1.0.xlsx   (multi-sheet, 22 fogli)
- RISK-REGISTER-v1.0-full.csv (tutti i rischi consolidati)
- A2-RISK-REGISTER-REPORT.md  (report narrativo)

Autore: senior risk manager + safety engineer aerospace
Versione: v1.0  (M+3, 2026-05-17)
"""

from __future__ import annotations

import csv
import os
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ---------------------------------------------------------------------------
# Configurazione globale
# ---------------------------------------------------------------------------
OUT_DIR = Path("/home/user/HALE/studio-di-fattibilita/allegati/A2-Risk-Register")
OUT_DIR.mkdir(parents=True, exist_ok=True)

XLSX_PATH = OUT_DIR / "RISK-REGISTER-v1.0.xlsx"
CSV_PATH = OUT_DIR / "RISK-REGISTER-v1.0-full.csv"
MD_PATH = OUT_DIR / "A2-RISK-REGISTER-REPORT.md"

TODAY = "2026-05-17"
VERSION = "v1.0"

# ---------------------------------------------------------------------------
# Helpers di stile per openpyxl
# ---------------------------------------------------------------------------
THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_HEADER = PatternFill("solid", fgColor="1F3864")
FILL_SUB = PatternFill("solid", fgColor="D9E1F2")
FILL_RED = PatternFill("solid", fgColor="C00000")
FILL_ORANGE = PatternFill("solid", fgColor="ED7D31")
FILL_YELLOW = PatternFill("solid", fgColor="FFC000")
FILL_GREEN = PatternFill("solid", fgColor="70AD47")
FILL_GREY = PatternFill("solid", fgColor="BFBFBF")

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_SUB = Font(name="Calibri", size=10, bold=True, color="1F3864")
FONT_BODY = Font(name="Calibri", size=10)
FONT_BODY_BOLD = Font(name="Calibri", size=10, bold=True)
FONT_WHITE_BOLD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def color_for_score(score: int) -> PatternFill:
    """Color coding NASA NPR 8000.4 P x I matrix."""
    if score >= 15:
        return FILL_RED
    if score >= 8:
        return FILL_YELLOW
    if score >= 4:
        return FILL_GREEN
    return PatternFill("solid", fgColor="E2EFDA")


def font_for_score(score: int) -> Font:
    if score >= 15:
        return FONT_WHITE_BOLD
    return FONT_BODY


def status_color(status: str) -> str:
    return {
        "Open": "C00000",
        "Open-Critical": "C00000",
        "Open-High": "ED7D31",
        "Mitigated": "70AD47",
        "Monitor": "FFC000",
        "Closed": "A6A6A6",
        "Showstopper": "C00000",
    }.get(status, "000000")


def write_header(ws, headers, row=1, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = CENTER
        cell.border = BORDER
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def write_row(ws, row_idx, values, score_col=None, fill_score=True):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=c, value=v)
        cell.alignment = WRAP
        cell.font = FONT_BODY
        cell.border = BORDER
    if score_col and fill_score:
        score_val = values[score_col - 1]
        if isinstance(score_val, int):
            ws.cell(row=row_idx, column=score_col).fill = color_for_score(score_val)
            ws.cell(row=row_idx, column=score_col).font = font_for_score(score_val)
            ws.cell(row=row_idx, column=score_col).alignment = CENTER

# ---------------------------------------------------------------------------
# DATASET RISCHI — schema standardizzato
# ---------------------------------------------------------------------------
# Campi: id, categoria, descrizione, trigger, P, I, status, owner,
#        response, mitigation, residual_P, residual_I, phase, confidence,
#        ewi (early warning indicator), falsifying_obs (None se non top-10)
#
# Confidence: high / medium-high / medium / medium-low / low
# Status: Open-Critical / Open-High / Mitigated / Monitor / Closed / Showstopper
# Response: Avoid / Mitigate / Transfer / Accept / Mitigate+Transfer
#
# Tutti gli ID coerenti con Cap. 5/6/10. Score = P x I; Residual Score = Pr x Ir
# ---------------------------------------------------------------------------

RISKS: list[dict] = [

    # ----- TECNICI (RSK-TEC) -------------------------------------------------
    {
        "id": "RSK-TEC-001", "cat": "Tecnico",
        "desc": "Energy balance HALE inverno 44N - deficit -50% confermato da simulazione (vs +0-15% stima a mano)",
        "trigger": "Simulazione 365gg M+3 mostra -50.1% margin solstizio dic; perennial flight NON fattibile baseline 2026-28",
        "P": 5, "I": 5, "status": "Showstopper",
        "owner": "propulsion-energy-engineer",
        "response": "Mitigate",
        "mitigation": "Plan A obbligato E5 'Seasonal-only mar-ott' (~7 mesi). Plan B Y6+: migrazione SS Li 450 Wh/kg o PEM+LH2. Plan C: ridimensionamento R&D-only fino tech 2030+",
        "rP": 5, "rI": 4, "phase": "Y3-Y5 (Phase B 6B)",
        "confidence": "high",
        "ewi": "Sim. allegato A.7 + monthly clear-sky variability + LiS pack TRL update trimestrale",
        "fo": "Se al gate G5 (M+24) sim. con dati operativi reali conferma deficit >30% giorni anche scenario E5, Percorso 6B terminato come operativo perennial",
        "top": True,
    },
    {
        "id": "RSK-TEC-002", "cat": "Tecnico",
        "desc": "Aeroelasticita ala high-AR (AR>=25) - flutter, divergenza, instabilita non lineare",
        "trigger": "Analisi aeroelastica preliminare M+12 mostra flutter speed < 1.3x Vdive o divergenza < 1.5x Vc",
        "P": 3, "I": 5, "status": "Showstopper",
        "owner": "aero-structures-engineer",
        "response": "Mitigate",
        "mitigation": "Aeroelastic analysis non-lineare (NASTRAN+ZAERO o MSC.Nastran SOL145) + GVT (Ground Vibration Test) + flight test subscale + winglet/passive damping design. Tilted spar caps + balance mass.",
        "rP": 2, "rI": 4, "phase": "Y3-Y4 (Phase B 6B)",
        "confidence": "medium-high",
        "ewi": "Output FEA aeroelastico subscale + risultati GVT + base rate Helios/PHASA flutter events",
        "fo": "Se subscale flight test M+18-24 mostra divergence o flutter sotto envelope, ridisegno radicale ala richiesto - costo +1-2M EUR + delay 6-12 mesi",
        "top": True,
    },
    {
        "id": "RSK-TEC-003", "cat": "Tecnico/Regolatorio",
        "desc": "Type Certification HALE timeline > 5 anni - no precedente HALE solare civile EU con TC emesso",
        "trigger": "EASA non apre RMT HAPS o Special Condition path entro 2028",
        "P": 4, "I": 4, "status": "Showstopper",
        "owner": "aviation-regulatory + sovereign-strategist",
        "response": "Mitigate+Accept",
        "mitigation": "Parallel approach: ops 6A genera revenue + esperienza mentre TC HALE matura. Engagement EASA Innovation Network + consortium CIRA/TAS per Special Condition collettiva",
        "rP": 4, "rI": 3, "phase": "Y4-Y8",
        "confidence": "medium-high",
        "ewi": "EASA RMT HAPS calendar + advisory bodies pubblicazioni + AALTO/Skydweller TC progress",
        "fo": "Se a Y5 (M+60) EASA non ha aperto RMT HAPS, Percorso 6B operativo commerciale rinviato a Y8+, scenario No-Go pieno se anche window IRIS2 chiusa",
        "top": True,
    },
    {
        "id": "RSK-TEC-004", "cat": "Tecnico",
        "desc": "Integrazione payload modulare 6A - incompatibilita elettrica/SW vs ICD payload bay JOUAV CW-30E",
        "trigger": "Test bench pre-deploy M+7 rivela mismatch power budget o data bus payload-FCS",
        "P": 3, "I": 3, "status": "Open-High",
        "owner": "systems-engineer",
        "response": "Mitigate",
        "mitigation": "ICD rigoroso e validato pre-quotation; test bed integrazione hardware-in-the-loop entro M+6; vendor compatibility matrix",
        "rP": 2, "rI": 3, "phase": "Y1 (M+6 → M+9)",
        "confidence": "medium",
        "ewi": "Esiti test bench HIL trimestrale + vendor response time tickets ICD",
        "fo": None,
        "top": True,
    },
    {
        "id": "RSK-TEC-005", "cat": "Tecnico",
        "desc": "Cybersecurity link C2 - jamming, spoofing GNSS, command injection",
        "trigger": "Eventi jamming GNSS regione (Mar Ligure ports) o report incidenti UAS via",
        "P": 2, "I": 4, "status": "Open-High",
        "owner": "avionics-gnc-engineer + CISO (new)",
        "response": "Mitigate",
        "mitigation": "Frequency hopping (FHSS) + crypto authentication AES-256 + multi-GNSS L1/L5 + dead-reckoning IMU fallback + Lost-Link procedure DAL-C",
        "rP": 1, "rI": 4, "phase": "Y1+ continuous",
        "confidence": "medium-high",
        "ewi": "Report GPS interference EASA bulletin + NOTAM jamming Mediterraneo",
        "fo": None,
        "top": True,
    },
    {
        "id": "RSK-TEC-006", "cat": "Tecnico/Privacy",
        "desc": "Privacy by design fail (DPIA bocciata Garante per sorveglianza territoriale)",
        "trigger": "Garante Privacy emette provvedimento negativo o richiesta integrazione DPIA",
        "P": 2, "I": 4, "status": "Open-High",
        "owner": "data-privacy-counsel + DPO (new)",
        "response": "Mitigate",
        "mitigation": "Edge anonymization (blur on-board obbligatorio); geofence aree residenziali; DPIA preliminare entro M+6; consultazione preventiva Garante",
        "rP": 1, "rI": 3, "phase": "Y1 (M+0 → M+9)",
        "confidence": "medium-high",
        "ewi": "Comunicati Garante AI Act/UAS + precedenti DPIA aerei",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-TEC-007", "cat": "Tecnico/Regolatorio",
        "desc": "Lost-Link behaviour mismatch SORA OSO #9 - mancata RtB automatica",
        "trigger": "ENAC in pre-application richiede dimostrazione Lost-Link Profile + RtB testato",
        "P": 2, "I": 4, "status": "Open-High",
        "owner": "avionics-gnc-engineer",
        "response": "Mitigate",
        "mitigation": "Lost-Link Profile documentato + Return-to-Base testato in flight test pre-deploy + procedura GS notifica + safety case ARP4761",
        "rP": 1, "rI": 3, "phase": "Y1 (M+6)",
        "confidence": "medium-high",
        "ewi": "Test report RtB + ENAC pre-app feedback",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-TEC-008", "cat": "Tecnico",
        "desc": "Batterie thermal runaway - LiPo 6A o LiS 6B - propagation cell-to-cell",
        "trigger": "Incidente settore (Boeing 787, Tesla, drone delivery) o test interno cella defect",
        "P": 2, "I": 5, "status": "Open-High",
        "owner": "propulsion-energy-engineer + safety",
        "response": "Mitigate+Transfer",
        "mitigation": "Cell-level fuse + intumescent barriers + BMS multi-sensor + hangar dedicato ATEX + assicurazione casco con clausola batteria",
        "rP": 1, "rI": 4, "phase": "Y1+ continuous",
        "confidence": "high",
        "ewi": "Reports thermal events FAA/EASA + cell vendor recall",
        "fo": None,
        "top": True,
    },
    {
        "id": "RSK-TEC-009", "cat": "Tecnico",
        "desc": "Materiali compositi - delaminazione/fatica lino in strutture secondarie (cicli umidita)",
        "trigger": "Test panel POLIMI/POLITO mostra degradazione > soglia dopo 1000h ciclica umido",
        "P": 3, "I": 2, "status": "Open-Medium",
        "owner": "aero-structures-engineer",
        "response": "Mitigate",
        "mitigation": "Lino solo in strutture secondarie con coating idrorepellente; longherone primario CFRP; LCA + qualification path posticipato",
        "rP": 2, "rI": 2, "phase": "Y3-Y4",
        "confidence": "medium",
        "ewi": "Pubblicazioni Pinato/Biogear test cicli umidi + qualification panel internal",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-TEC-010", "cat": "Tecnico",
        "desc": "GNSS spoofing/jamming Liguria - prossimita conflict zones Mar Mediterraneo",
        "trigger": "Incidenti GPS reportati Mar Ligure o esercitazioni NATO area",
        "P": 3, "I": 3, "status": "Open-High",
        "owner": "avionics-gnc-engineer",
        "response": "Mitigate",
        "mitigation": "Multi-constellation (GPS+Galileo+GLONASS+BeiDou) + RAIM + IMU dead-reckoning + opzione Galileo PRS Phase B",
        "rP": 2, "rI": 3, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "Report EASA GPS interference + UK CAA bulletin spoofing",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-TEC-011", "cat": "Tecnico",
        "desc": "Solar cell degradation in stratosfera - UV + thermal cycling ridotta efficienza GaAs",
        "trigger": "Test cell exposure stratosferico (Sahara HALE program) mostra > 5%/year degradation",
        "P": 2, "I": 3, "status": "Monitor",
        "owner": "propulsion-energy-engineer",
        "response": "Mitigate",
        "mitigation": "Margine design 30% su energy budget; cell qualification per ambiente stratosferico; vendor Spectrolab/Azur Space datasheet flight heritage",
        "rP": 2, "rI": 2, "phase": "Y4+",
        "confidence": "medium",
        "ewi": "Vendor flight heritage data + EuroHAPS testbed reports",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-TEC-012", "cat": "Tecnico",
        "desc": "FCS DAL-C HALE custom - mancanza track record civile + costo qualification ~2-5M EUR",
        "trigger": "Vendor FCS DAL-C EU rifiuta development partnership o costo > budget",
        "P": 3, "I": 4, "status": "Open-High",
        "owner": "avionics-gnc-engineer",
        "response": "Mitigate",
        "mitigation": "Engagement early vendor EU (UAVOS, MicroPilot, Honeywell EU); partnership con CIRA su FCS Italian sovereign; budget riservato 2-3M EUR R&D",
        "rP": 2, "rI": 4, "phase": "Y3-Y4",
        "confidence": "medium",
        "ewi": "Quotation vendor FCS + RMT EASA on autonomy",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-TEC-013", "cat": "Tecnico",
        "desc": "Link budget HAPS service link rain fade - storm scenario 1% time supera margin",
        "trigger": "Simulazione ITU-R P.618-14 stormy day worst-case mostra margin < 6 dB",
        "P": 2, "I": 3, "status": "Monitor",
        "owner": "telecom-ntn-payload-expert",
        "response": "Mitigate",
        "mitigation": "Adaptive coding modulation (ACM) + site diversity gateway Ka-band + buffer queue + scheduled re-transmission",
        "rP": 1, "rI": 2, "phase": "Y4+",
        "confidence": "high",
        "ewi": "Live link metrics post test integrato",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-TEC-014", "cat": "Tecnico",
        "desc": "Single Event Upset (SEU) avionica stratosferica - aumentato flusso particelle 20 km",
        "trigger": "FCS reboot rate > 1/100h in flight test stratosferico",
        "P": 3, "I": 3, "status": "Open-Medium",
        "owner": "avionics-gnc-engineer",
        "response": "Mitigate",
        "mitigation": "FPGA rad-tolerant (Microchip RTPolarFire o Xilinx Versal AI Edge) + 2oo3 voting + watchdog timer + memory ECC",
        "rP": 2, "rI": 2, "phase": "Y3-Y5",
        "confidence": "medium",
        "ewi": "Flight test stratosferico subscale + heritage SEU dataset Zephyr",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-TEC-015", "cat": "Tecnico",
        "desc": "Riduzione TRL gap M+24 - integrato HALE subsystem critici < TRL 5",
        "trigger": "Gate G5 review M+24 mostra TRL integrato propulsione/avionica/payload < 5",
        "P": 3, "I": 5, "status": "Open-Critical",
        "owner": "systems-engineer + propulsion-energy-engineer",
        "response": "Mitigate",
        "mitigation": "Roadmap TRL puntuale per sottosistema + milestone trimestrali + partnership prime per acceleration TRL (DR-013 finding)",
        "rP": 3, "rI": 4, "phase": "Y3 (gate G5)",
        "confidence": "medium",
        "ewi": "TRL milestone tracker + subsystem demo report + partnership signed",
        "fo": "Se TRL gap > 2 a M+24, Phase C-D non finanziabile; ridimensionamento seasonal-only + R&D-only mode",
        "top": False,
    },
    {
        "id": "RSK-TEC-016", "cat": "Tecnico",
        "desc": "NTN payload winter unsustainable - margin -58.9% con P_payload 500 W (scenario E5)",
        "trigger": "Simulazione M+3 (DR-014) conferma deficit anche con payload pulse-mode",
        "P": 4, "I": 4, "status": "Open-Critical",
        "owner": "propulsion-energy-engineer + telecom-ntn-payload-expert",
        "response": "Mitigate",
        "mitigation": "NTN seasonal-only + payload pulse-mode duty-cycle < 30% + dedicated battery bank NTN",
        "rP": 3, "rI": 3, "phase": "Y4-Y5",
        "confidence": "medium-high",
        "ewi": "Sim. allegato A.7 update; product roadmap 5G NTN bypass requirements",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-TEC-017", "cat": "Tecnico",
        "desc": "Ground Station single point of failure - GS fissa Pentema unica fonte C2",
        "trigger": "Power outage o danno infrastruttura GS Pentema durante volo BVLOS",
        "P": 2, "I": 4, "status": "Open-High",
        "owner": "vtol-uas-specialist + ops",
        "response": "Mitigate",
        "mitigation": "GS mobile ridondante + UPS GS fissa + SATCOM fallback + procedura RtB Lost-Link",
        "rP": 1, "rI": 3, "phase": "Y1+",
        "confidence": "high",
        "ewi": "Live GS uptime + UPS test mensile",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-TEC-018", "cat": "Tecnico",
        "desc": "Calibrazione IR sensor persa - false positive antincendio o missed hotspot",
        "trigger": "Validation post-deploy mostra false alarm rate > 5% o missed detection > 2%",
        "P": 3, "I": 4, "status": "Open-High",
        "owner": "earth-observation-expert",
        "response": "Mitigate",
        "mitigation": "NUC (Non-Uniformity Correction) frequente + cross-check RGB + validation periodica vs ground truth (VVF, satellite)",
        "rP": 2, "rI": 3, "phase": "Y1+",
        "confidence": "medium-high",
        "ewi": "FAR/MDR weekly metrics + IR cal report mensile",
        "fo": None,
        "top": False,
    },

    # ----- REGOLATORI (RSK-REG) ---------------------------------------------
    {
        "id": "RSK-REG-001", "cat": "Regolatorio",
        "desc": "Mancanza framework HAPS EASA/ENAC - no Special Condition aperto HALE solare civile",
        "trigger": "EASA non apre RMT HAPS nel calendario 2026-2028; ENAC non rilascia AMC HAPS",
        "P": 5, "I": 4, "status": "Showstopper",
        "owner": "aviation-regulatory + sovereign-strategist",
        "response": "Mitigate",
        "mitigation": "Engagement EASA Innovation Network + consortium CIRA/TAS per Special Condition collettiva; partecipazione ASD-Eurospace HAPS WG; lobby DG MOVE/DG DEFIS",
        "rP": 4, "rI": 4, "phase": "Y3-Y6 (Phase B/C)",
        "confidence": "high",
        "ewi": "EASA RMT calendar (semestrale) + Special Condition published + ASD-Eurospace minutes",
        "fo": "Se al gate G5 (M+24) RMT HAPS non aperto e nessuna Special Condition in dialogo formale, Phase B 6B sospesa fino 2028+",
        "top": True,
    },
    {
        "id": "RSK-REG-002", "cat": "Regolatorio",
        "desc": "SORA SAIL III Pentema BVLOS - richiede percorso Certified Category EASA",
        "trigger": "ENAC in pre-application valuta SAIL > III per scenario Pentema (popolato + critical)",
        "P": 3, "I": 3, "status": "Open-High",
        "owner": "aviation-regulatory",
        "response": "Mitigate",
        "mitigation": "Pre-application meeting ENAC M+3-6 + M1 (operational mitigation) + M2 (technical mitigation, parachute + flight termination)",
        "rP": 2, "rI": 3, "phase": "Y1 (M+3-9)",
        "confidence": "medium-high",
        "ewi": "ENAC pre-app feedback + SAIL determination doc",
        "fo": None,
        "top": True,
    },
    {
        "id": "RSK-REG-003", "cat": "Regolatorio",
        "desc": "AGCOM spettro HAPS Italia - banda S non assegnata per HAPS commercial",
        "trigger": "AGCOM emette consultazione spettro senza include HAPS service link",
        "P": 3, "I": 3, "status": "Monitor",
        "owner": "telecom-ntn-payload-expert + aviation-regulatory",
        "response": "Mitigate",
        "mitigation": "Engagement AGCOM precoce (DR-005); domanda licenza sperimentale L1+L5; partecipazione consultazione ITU-R WP 5C",
        "rP": 2, "rI": 3, "phase": "Y3-Y4",
        "confidence": "medium",
        "ewi": "AGCOM delibera spettro UAS/HAPS + ITU-R WRC outcomes",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-004", "cat": "Regolatorio",
        "desc": "U-Space transizione 2027 - ENAC procedura USSP italiana non definita",
        "trigger": "ENAC delay implementazione Reg.UE 2021/664 oltre 2027",
        "P": 3, "I": 2, "status": "Monitor",
        "owner": "aviation-regulatory",
        "response": "Mitigate",
        "mitigation": "Engagement ENAC U-Space tavolo nazionale; partecipazione D-Flight pilot Pentema; SORA fallback continua",
        "rP": 2, "rI": 2, "phase": "Y2-Y3",
        "confidence": "medium",
        "ewi": "ENAC U-Space roadmap pubblicato + D-Flight expansion",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-005", "cat": "Regolatorio",
        "desc": "GDPR enforcement DPIA - DPA ricorre vs operazioni EO sorveglianza territoriale",
        "trigger": "Garante apre procedimento o richiede stop attivita",
        "P": 2, "I": 4, "status": "Open-High",
        "owner": "data-privacy-counsel + DPO",
        "response": "Mitigate",
        "mitigation": "DPIA pubblica + edge anonymization + consultazione preventiva Garante + geofence + retention policy stringente",
        "rP": 1, "rI": 3, "phase": "Y1+",
        "confidence": "medium-high",
        "ewi": "Comunicati Garante + provvedimenti UAS/EO settore",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-006", "cat": "Regolatorio",
        "desc": "Geographical zones ENAC - Pentema in zona vincolata (parco Antola, ZSC)",
        "trigger": "Mappa D-Flight aggiornata classifica area Pentema come restricted",
        "P": 2, "I": 3, "status": "Monitor",
        "owner": "aviation-regulatory",
        "response": "Mitigate",
        "mitigation": "Verifica D-Flight pre-mission; coordinamento Ente Parco Antola; NOTAM dedicato; finestre temporali",
        "rP": 1, "rI": 3, "phase": "Y1+",
        "confidence": "high",
        "ewi": "D-Flight zone update + Ente Parco notice",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-007", "cat": "Regolatorio",
        "desc": "EASA SORA 2.5 transizione - documentazione SORA pre-existente da aggiornare",
        "trigger": "ENAC richiede re-submission SORA secondo ED Decision 2025/018/R post Sep 2025",
        "P": 3, "I": 2, "status": "Open-Medium",
        "owner": "aviation-regulatory",
        "response": "Mitigate",
        "mitigation": "SORA documentation SORA 2.5 compliant fin da prima applicazione; consulenza ENAC pre-app",
        "rP": 2, "rI": 2, "phase": "Y1",
        "confidence": "medium-high",
        "ewi": "EASA AMC update + ENAC circolari implementative",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-008", "cat": "Regolatorio",
        "desc": "EASA Part-21 design organisation approval (DOA) - richiesto per HALE Phase C",
        "trigger": "Transition Phase B->C richiede DOA accreditata o partner con DOA",
        "P": 4, "I": 4, "status": "Open-High",
        "owner": "aviation-regulatory + systems-engineer",
        "response": "Mitigate",
        "mitigation": "Partnership con DOA esistente (Leonardo, Tekever, AALTO) per Phase C; preparazione DOA Firmamento Y5+",
        "rP": 3, "rI": 3, "phase": "Y5+",
        "confidence": "medium",
        "ewi": "DOA holders pubblicati EASA + partnership negotiation status",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-009", "cat": "Regolatorio",
        "desc": "Notified Body shortage - CE marking AI Act + Direttiva Macchine + radio",
        "trigger": "Mercato NB satura su AI Act 2026-27; tempo attesa > 12 mesi",
        "P": 3, "I": 3, "status": "Monitor",
        "owner": "systems-engineer + data-privacy",
        "response": "Mitigate",
        "mitigation": "Engagement early NB (TUV, IMQ, Bureau Veritas); preparazione pre-assessment documentazione",
        "rP": 2, "rI": 2, "phase": "Y2-Y3",
        "confidence": "medium",
        "ewi": "NB capacity report ACCREDIA + waiting time tracking",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-010", "cat": "Regolatorio",
        "desc": "ITAR/EAR US restrictions - componenti US-origin (chip Honeywell/Collins, celle Spectrolab)",
        "trigger": "US BIS pubblica restrictions su componenti HAPS-related EU",
        "P": 2, "I": 4, "status": "Open-High",
        "owner": "supply-chain + sovereign-strategist",
        "response": "Mitigate",
        "mitigation": "Mapping componenti US-origin + EU sovereign alternatives roadmap (Azur Space, ACC, STM) + stock buffer 12 mesi",
        "rP": 2, "rI": 3, "phase": "Y2+",
        "confidence": "medium",
        "ewi": "BIS Entity List update + ITAR notice settore aerospace",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-011", "cat": "Regolatorio",
        "desc": "VIA (Valutazione Impatto Ambientale) Pentema - hangar + GS in area protetta",
        "trigger": "ARPA Liguria richiede VIA per infrastruttura ground HALE",
        "P": 2, "I": 2, "status": "Monitor",
        "owner": "environmental + ops",
        "response": "Mitigate",
        "mitigation": "VIA preliminare M+6; coordinamento con Ente Parco Antola; design infrastruttura minimale + reversibile",
        "rP": 1, "rI": 2, "phase": "Y1",
        "confidence": "medium-high",
        "ewi": "ARPA notice + Ente Parco feedback",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-012", "cat": "Regolatorio",
        "desc": "Reg.UE 2019/947 evoluzione - revisione 2027 puo cambiare classificazione SORA",
        "trigger": "EASA NPA pubblicata con revisione SORA framework",
        "P": 3, "I": 2, "status": "Monitor",
        "owner": "aviation-regulatory",
        "response": "Mitigate",
        "mitigation": "Monitoring NPA EASA; partecipazione consultation; flessibilita design",
        "rP": 2, "rI": 2, "phase": "Y2-Y3",
        "confidence": "medium",
        "ewi": "EASA NPA calendar + ED Decision pubblicazioni",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-013", "cat": "Regolatorio",
        "desc": "Coordinamento ENAV - manca AIP procedure UAS BVLOS Pentema",
        "trigger": "ENAV non rilascia procedura AIP entro M+9",
        "P": 3, "I": 3, "status": "Open-High",
        "owner": "aviation-regulatory + vtol-uas-specialist",
        "response": "Mitigate",
        "mitigation": "Engagement ENAV Genova ACC; NOTAM dedicato per ogni missione; coordinamento con CAA militare Mar Ligure",
        "rP": 2, "rI": 3, "phase": "Y1",
        "confidence": "medium-high",
        "ewi": "ENAV response time + AIP Italy SUP",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-014", "cat": "Regolatorio",
        "desc": "Garante Privacy - precedenti negativi su EO/UAS in spazi pubblici",
        "trigger": "Provvedimento Garante simile (es. sorveglianza droni Comuni) rivela limiti",
        "P": 3, "I": 3, "status": "Open-Medium",
        "owner": "data-privacy-counsel",
        "response": "Mitigate",
        "mitigation": "Analisi precedenti Garante (DR-006); DPIA conforme; documentazione necessita/proporzionalita",
        "rP": 2, "rI": 2, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "Garante provvedimenti pubblicati + sentenze Cassazione",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-015", "cat": "Regolatorio",
        "desc": "Cross-border BVLOS - assenza EU regulation harmonizzata per cross-border ops HAPS",
        "trigger": "Operazioni cross-border (Liguria-Francia, Liguria-Svizzera) bloccate da MS divergenti",
        "P": 4, "I": 3, "status": "Monitor",
        "owner": "aviation-regulatory + sovereign-strategist",
        "response": "Accept+Mitigate",
        "mitigation": "Focus Y1-Y3 su operazioni domestiche IT; engagement EASA cross-border WG Y3+",
        "rP": 3, "rI": 2, "phase": "Y3+",
        "confidence": "medium",
        "ewi": "EASA cross-border framework calendar",
        "fo": None,
        "top": False,
    },
    # ---- 15 NUOVI showstopper §5.16 -----------------------------------------
    {
        "id": "RSK-REG-016", "cat": "Regolatorio",
        "desc": "AI Act (Reg.UE 2024/1689) - payload IR + biometric classification high-risk Annex III",
        "trigger": "Garante o AgID classifica sistemi IR onboard come 'alto rischio AI Act'",
        "P": 3, "I": 4, "status": "Open-High",
        "owner": "data-privacy-counsel + DPO",
        "response": "Mitigate",
        "mitigation": "Privacy by design hardware (blur on-board obbligatorio); esclusione esplicita riconoscimento biometrico; engagement Garante + AgID pre-clearance; conformity assessment Notified Body se richiesto",
        "rP": 2, "rI": 3, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "Garante delibere AI Act + AgID circolari + EU AI Office guidance",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-017", "cat": "Regolatorio",
        "desc": "EUSPA accreditation downstream services - mancato accesso fondi/GOVSATCOM",
        "trigger": "EUSPA emette guideline che escludono HAPS dai servizi spaziali downstream",
        "P": 2, "I": 3, "status": "Monitor",
        "owner": "sovereign-strategist",
        "response": "Mitigate",
        "mitigation": "Engagement EUSPA via CASSINI accelerator; workshop downstream services HAPS; mapping intersezione GOVSATCOM",
        "rP": 2, "rI": 2, "phase": "Y3+",
        "confidence": "medium",
        "ewi": "EUSPA workplan pubblicato + CASSINI HAPS call",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-018", "cat": "Regolatorio",
        "desc": "EUROCONTROL Network Manager - coordinamento ATM-ANS HAPS FL400+ EU airspace",
        "trigger": "EUROCONTROL non rilascia procedure operative HAPS perennial entro Y4-Y5",
        "P": 4, "I": 4, "status": "Open-Critical",
        "owner": "avionics-gnc-engineer + aviation-regulatory",
        "response": "Mitigate",
        "mitigation": "Engagement EUROCONTROL precocemente (Y2-Y3); partecipazione workshop UAM/HAPS Network Manager; contributo definizione procedure",
        "rP": 3, "rI": 3, "phase": "Y3+",
        "confidence": "medium",
        "ewi": "EUROCONTROL Network Manager workplan + HAPS procedure pubblicate",
        "fo": "Se EUROCONTROL declina procedure HAPS perennial entro Y5, operativita cross-border bloccata; ridimensionamento operazioni IT-only",
        "top": True,
    },
    {
        "id": "RSK-REG-019", "cat": "Regolatorio",
        "desc": "Part-IS EASA Reg.UE 2023/203 - ISMS obbligatorio da feb 2026, CISO assente",
        "trigger": "Audit ENAC Part-IS rileva non-conformita sostanziali ISMS",
        "P": 4, "I": 4, "status": "Open-Critical",
        "owner": "aviation-regulatory + CISO (new)",
        "response": "Mitigate",
        "mitigation": "Assunzione CISO entro M+6; ISMS implementazione entro M+9; certificazione ISO/IEC 27001 entro M+12; audit interno + pre-audit ENAC",
        "rP": 3, "rI": 3, "phase": "Y1 (M+0 → M+12) urgente",
        "confidence": "high",
        "ewi": "CISO hire date + ISMS gap analysis + ISO 27001 cert progress",
        "fo": "Se al M+9 ISMS non implementato, ENAC sospende operazioni commerciali continuative fino remediation",
        "top": True,
    },
    {
        "id": "RSK-REG-020", "cat": "Regolatorio",
        "desc": "Codice Penale art. 432-bis - responsabilita penale PIC + organizzazione D.Lgs.231",
        "trigger": "Incidente UAS BVLOS con conseguenze (near-miss vs aviazione manned o danni terzi)",
        "P": 2, "I": 5, "status": "Open-High",
        "owner": "aviation-regulatory + ops + legal",
        "response": "Mitigate+Transfer",
        "mitigation": "Assicurazione RC+casco BVLOS adeguata (>=5M EUR); D.Lgs.231 modello organizzativo; SOPs rigorosi; training PIC; safety reporting culture",
        "rP": 1, "rI": 4, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "Incident reports settore UAS BVLOS Italia + near-miss tracking interno",
        "fo": "Se incidente con conseguenza penale a Y1-Y2, intero modello operativo a rischio; Garante interviene parallelo",
        "top": True,
    },
    {
        "id": "RSK-REG-021", "cat": "Regolatorio",
        "desc": "AgID/PSN hosting dati PA - cloud non qualificato blocca contratti Regione/PC",
        "trigger": "Verifica AgID al M+9 rivela cloud Aruba/OVH non PSN-qualified per livello criticita",
        "P": 4, "I": 4, "status": "Open-Critical",
        "owner": "data-privacy + IT + DPO",
        "response": "Mitigate",
        "mitigation": "Migrazione provider PSN-qualified (TIM Enterprise, Polo PSN, CDP Cloud) o cloud qualificato AgID (Aruba qualified, Engineering, Reply, Almaviva); audit AgID compliance",
        "rP": 2, "rI": 3, "phase": "Y1+",
        "confidence": "high",
        "ewi": "AgID qualifica list update + audit compliance internal",
        "fo": "Se al M+9 dati Pentema non in cloud PSN-qualified, contratti PA pluriennali rifiutati",
        "top": True,
    },
    {
        "id": "RSK-REG-022", "cat": "Regolatorio",
        "desc": "ATEX (Reg.UE 2014/34) batterie LiS storage - hangar classificato zona 1",
        "trigger": "ASL/VVF ispezione hangar Pentema classifica zona ATEX senza protezioni adeguate",
        "P": 3, "I": 3, "status": "Open-Medium",
        "owner": "propulsion-engineer + safety + ops",
        "response": "Mitigate",
        "mitigation": "Hangar dedicato con ventilazione anti-esplosione (LEL detection + emergency vent); procedure ATEX; formazione operatori; assicurazione adeguata",
        "rP": 2, "rI": 2, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "Ispezioni ASL/VVF programma + thermal events battery settore",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-023", "cat": "Regolatorio",
        "desc": "RoHS (Dir.2011/65/UE) - componenti elettronici cinesi JOUAV potenzialmente non compliant",
        "trigger": "Audit RoHS BOM rivela componente critico non-RoHS",
        "P": 3, "I": 2, "status": "Monitor",
        "owner": "systems-engineer + supply-chain",
        "response": "Mitigate",
        "mitigation": "Verifica DoC ogni componente BoM; audit RoHS pre-deployment; sostituzione componenti non-compliant",
        "rP": 2, "rI": 2, "phase": "Y1",
        "confidence": "medium-high",
        "ewi": "BOM audit RoHS quarterly + vendor DoC update",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-024", "cat": "Regolatorio",
        "desc": "Codice Navigazione R.D. 327/1942 - diritti sorvolo proprieta private + aree militari",
        "trigger": "Contestazione legale da proprietari o sorvolo area militare non mappata",
        "P": 2, "I": 3, "status": "Monitor",
        "owner": "regulatory + community + ops",
        "response": "Mitigate",
        "mitigation": "Mappa aree militari/sensibili Liguria; consenso proprietari grandi appezzamenti; NOTAM coordinato; engagement comunita Pentema",
        "rP": 1, "rI": 2, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "Mappa NOTAM/restricted aggiornata + community feedback",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-025", "cat": "Regolatorio",
        "desc": "Affidamento PA art.50 D.Lgs.36/2023 - contratto Regione > 140k EUR richiede gara",
        "trigger": "Regione Liguria contratto pluriennale > 300k EUR bocciato fase amministrativa",
        "P": 4, "I": 4, "status": "Open-Critical",
        "owner": "snai-funding + legal + business-model",
        "response": "Mitigate",
        "mitigation": "Pre-engagement Regione + accordo quadro (art.59 D.Lgs.36); partnership Coopfond come veicolo non-gara; accordi di programma SNAI; gara con specificita tecniche Firmamento-friendly",
        "rP": 3, "rI": 3, "phase": "Y0+",
        "confidence": "high",
        "ewi": "Bando Regione pubblicato + legal review procedura",
        "fo": "Se contratto Regione bocciato per gara, rinvio M+6-12 + competitor risk (Leonardo, Telespazio)",
        "top": True,
    },
    {
        "id": "RSK-REG-026", "cat": "Regolatorio",
        "desc": "Insurance BVLOS - mancata copertura broker aviation o premio > 100k EUR/anno",
        "trigger": "Tender broker non riceve quotation accettabile per BVLOS Pentema",
        "P": 4, "I": 3, "status": "Open-High",
        "owner": "financial-cfo + ops",
        "response": "Transfer+Mitigate",
        "mitigation": "Tender con broker specializzati aviation (Marsh, Aon, Willis Italia); copertura RC+casco+cyber+privacy; piano captive insurance Y3+",
        "rP": 3, "rI": 2, "phase": "Y0+",
        "confidence": "medium",
        "ewi": "Market hardening aviation insurance + quotation timing",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-027", "cat": "Regolatorio",
        "desc": "NIS2 D.Lgs.138/2024 - registrazione ACN omessa, sanzioni fino 10M EUR / 2% fatturato",
        "trigger": "Firmamento classificata 'soggetto essenziale' senza registrazione entro 30gg",
        "P": 4, "I": 4, "status": "Open-Critical",
        "owner": "CISO (new) + legal",
        "response": "Mitigate",
        "mitigation": "Registrazione preventiva ACN entro M+1; ISMS Part-IS allineato; notifica incidenti 24h procedure",
        "rP": 2, "rI": 3, "phase": "Y0+ immediato",
        "confidence": "high",
        "ewi": "ACN classification notice + sanction publications",
        "fo": "Se incidente cyber senza registrazione, sanzione amministrativa + reputazione + esclusione bandi PA",
        "top": True,
    },
    {
        "id": "RSK-REG-028", "cat": "Regolatorio",
        "desc": "Galileo PRS - dual-use civile-difesa Phase B richiede CASD/Difesa authorization",
        "trigger": "Phase B 6B dichiara dual-use; CASD richiede clearance personale + tecnologia",
        "P": 2, "I": 3, "status": "Monitor",
        "owner": "sovereign-strategist + CASD",
        "response": "Mitigate",
        "mitigation": "Y1 usare GNSS standard L1/L5 + Galileo open service + anti-spoofing software-based; engagement CASD solo se dual-use Fase 3+",
        "rP": 1, "rI": 2, "phase": "Y3+",
        "confidence": "medium",
        "ewi": "Phase B scope clearance + CASD engagement plan",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-029", "cat": "Regolatorio",
        "desc": "Direttiva Macchine Reg.UE 2023/1230 - ground equipment GS+carrelli CE marking obbligatorio",
        "trigger": "Phase B 6B ground equipment custom non CE-marked",
        "P": 3, "I": 3, "status": "Monitor",
        "owner": "systems-engineer",
        "response": "Mitigate",
        "mitigation": "Design CE-compliant fin da concept; collaborazione con ente notificato; risk assessment + DoC",
        "rP": 2, "rI": 2, "phase": "Y3+",
        "confidence": "medium",
        "ewi": "Ground equipment design review + NB engagement",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REG-030", "cat": "Regolatorio",
        "desc": "ENAV procedure FL400+ - HAPS perennial sopra FL400/FL650 senza procedure dedicate",
        "trigger": "ENAV declina procedure HAPS perennial entro Y4",
        "P": 4, "I": 4, "status": "Open-Critical",
        "owner": "avionics + sovereign-strategist + aviation-regulatory",
        "response": "Mitigate",
        "mitigation": "Engagement ENAV precoce (Y2); contributo definizione procedure standard EUROCONTROL; testing spazio aereo segregato (Sardinia, GATB Apulia)",
        "rP": 3, "rI": 3, "phase": "Y3+",
        "confidence": "medium-high",
        "ewi": "ENAV operational instruction + AIP Italy SUP HAPS",
        "fo": "Se ENAV declina procedure entro Y5, operativita italiana 6B bloccata; ridimensionamento test bed estero",
        "top": True,
    },

    # ----- FINANZIARI (RSK-FIN) ----------------------------------------------
    {
        "id": "RSK-FIN-001", "cat": "Finanziario",
        "desc": "Mancanza commitment funding Phase B 6B - 5.5-13.5M EUR mix EDF+Horizon+PNRR+equity",
        "trigger": "Gate G5 M+24 mostra funding mix Phase B < 30%",
        "P": 4, "I": 5, "status": "Showstopper",
        "owner": "financial-cfo + sovereign-strategist",
        "response": "Mitigate",
        "mitigation": "Mix funding: EDF (DG DEFIS) + Horizon Europe + PNRR Aerospazio + Series B equity (CDP, EIB); fasi graduali; partnership prime per cost sharing",
        "rP": 3, "rI": 4, "phase": "Y2-Y3 (gate G5)",
        "confidence": "medium-high",
        "ewi": "Calendar bandi EDF/Horizon + Series B pipeline + LoI investitori",
        "fo": "Se al gate G5 funding < 30% committed, DEFER 6B a M+36 con re-review; se < 15%, Hold permanente fino 2030+",
        "top": True,
    },
    {
        "id": "RSK-FIN-002", "cat": "Finanziario",
        "desc": "Mancato grant Coopfond Cooding (50k EUR seed)",
        "trigger": "Coopfond non finanzia Firmamento entro M+6",
        "P": 2, "I": 2, "status": "Open-Medium",
        "owner": "financial-cfo + snai-funding",
        "response": "Accept",
        "mitigation": "Alternative funding (FESR Liguria, bandi tematici Horizon, equity seed); bridge financing founder",
        "rP": 2, "rI": 2, "phase": "Y0",
        "confidence": "medium-high",
        "ewi": "Coopfond decisione + alternative grant calendar",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-FIN-003", "cat": "Finanziario",
        "desc": "Sovracosti CapEx 6A - quotation vendor real-world supera baseline 700k-2M EUR +30-50%",
        "trigger": "Quotation JOUAV/Tekever + payload + GS supera budget +30%",
        "P": 3, "I": 3, "status": "Open-High",
        "owner": "financial-cfo + systems-engineer",
        "response": "Mitigate",
        "mitigation": "Quotation parallel vendor multipli; contingency 20% in budget; scope reduction MVP se necessario",
        "rP": 2, "rI": 3, "phase": "Y1",
        "confidence": "medium",
        "ewi": "Quotation report vendor mensile + market index UAS",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-FIN-004", "cat": "Finanziario",
        "desc": "OpEx Y1 underestimato - +450-800k EUR per 3 FTE regulatory (CISO+DPO+Head Reg.Aff)",
        "trigger": "Audit regulatory M+3 rivela 3 FTE addizionali non-budget",
        "P": 5, "I": 3, "status": "Open-Critical",
        "owner": "financial-cfo",
        "response": "Mitigate",
        "mitigation": "Aggiornamento Cap.8 OpEx Y1 con +450-800k EUR fixed cost; revisione mix funding equity Series A; cost sharing su CISO/DPO via partnership cooperative",
        "rP": 4, "rI": 2, "phase": "Y0-Y1",
        "confidence": "high",
        "ewi": "FTE hire pipeline + payroll forecast",
        "fo": None,
        "top": True,
    },
    {
        "id": "RSK-FIN-005", "cat": "Finanziario",
        "desc": "Slittamento grant FESR/PNRR - tempi PA italiani median 18-30 mesi",
        "trigger": "Grant FESR Liguria comunicato slittato > 12 mesi vs piano",
        "P": 4, "I": 3, "status": "Open-High",
        "owner": "financial-cfo + snai-funding",
        "response": "Mitigate",
        "mitigation": "Bridge financing (banca + founder); pipeline grant multiple parallele; flessibilita timeline progetto",
        "rP": 3, "rI": 2, "phase": "Y1-Y3",
        "confidence": "high",
        "ewi": "Grant decisional timeline tracking + Regione comunicazioni",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-FIN-006", "cat": "Finanziario",
        "desc": "WACC effettivo > 18% se grant mix < 30% - NPV diventa negativo (sensitivity Cap.8)",
        "trigger": "Grant mix committed < 30% al M+12",
        "P": 3, "I": 4, "status": "Open-High",
        "owner": "financial-cfo",
        "response": "Mitigate",
        "mitigation": "Pricing premium servizi + cost optimization OpEx; renegotiate equity terms; cost-shared infrastructure con cooperative",
        "rP": 2, "rI": 3, "phase": "Y1-Y3",
        "confidence": "medium-high",
        "ewi": "Grant commitment % monthly + NPV sensitivity update",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-FIN-007", "cat": "Finanziario",
        "desc": "Currency risk - JOUAV CN forniture USD-denominated + EUR weakness",
        "trigger": "EUR/USD < 0.95 sustained o tariffe import +10%",
        "P": 3, "I": 2, "status": "Monitor",
        "owner": "financial-cfo + supply-chain",
        "response": "Transfer",
        "mitigation": "Hedging EUR/USD forward per acquisti CapEx; quotation EUR-denominated vendor EU (Tekever, Quantum) come benchmark",
        "rP": 2, "rI": 2, "phase": "Y1",
        "confidence": "medium",
        "ewi": "EUR/USD daily + tariffe USA-CN update",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-FIN-008", "cat": "Finanziario",
        "desc": "Cash flow gap Y2-Y3 - revenue ramp lento vs OpEx fixed",
        "trigger": "Revenue Y2 < 500k EUR (50% target)",
        "P": 3, "I": 4, "status": "Open-High",
        "owner": "financial-cfo + business-model",
        "response": "Mitigate",
        "mitigation": "Bridge financing Series A; cost flex su FTE variable; contratti pluriennali Regione anticipated payment",
        "rP": 2, "rI": 3, "phase": "Y2-Y3",
        "confidence": "medium",
        "ewi": "Revenue actuals monthly + cash burn rate",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-FIN-009", "cat": "Finanziario",
        "desc": "Series A non chiude (Y2) - valutazione investitori < 8M EUR pre-money",
        "trigger": "Valutazione Series A < 5M EUR o tempi > 12 mesi",
        "P": 3, "I": 4, "status": "Open-High",
        "owner": "financial-cfo + CEO",
        "response": "Mitigate",
        "mitigation": "Investor relations pipeline (CDP Venture, EIC Fund, BPER); milestone-based bridge financing; alternative grant heavy mix",
        "rP": 2, "rI": 3, "phase": "Y2",
        "confidence": "medium",
        "ewi": "Investor meetings pipeline + comparable Series A aerospace IT",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-FIN-010", "cat": "Finanziario",
        "desc": "ROI insurance premium + safety provision - +50-150k EUR/anno OpEx ricorrente",
        "trigger": "Premio assicurativo BVLOS + cyber > 100k EUR/anno",
        "P": 4, "I": 2, "status": "Open-High",
        "owner": "financial-cfo + risk",
        "response": "Accept+Transfer",
        "mitigation": "Tender broker + adeguamento periodico; captive insurance scenario Y3+",
        "rP": 3, "rI": 2, "phase": "Y1+",
        "confidence": "high",
        "ewi": "Premio quote annuale + claims tracking",
        "fo": None,
        "top": False,
    },

    # ----- MERCATO (RSK-MKT) -------------------------------------------------
    {
        "id": "RSK-MKT-001", "cat": "Mercato",
        "desc": "Adozione lenta PA - cicli appalti pubblici 12-24 mesi vs piano 6-9 mesi",
        "trigger": "Contratti pluriennali Regione Liguria non firmati entro M+12",
        "P": 4, "I": 3, "status": "Open-High",
        "owner": "snai-funding + business-model",
        "response": "Mitigate",
        "mitigation": "Anchor customer Regione + LoI pre-formale; contratti pluriennali quadro; partnership cooperative come veicolo di servizi",
        "rP": 3, "rI": 3, "phase": "Y1-Y2",
        "confidence": "high",
        "ewi": "Bandi Regione publication + LoI tracking",
        "fo": None,
        "top": True,
    },
    {
        "id": "RSK-MKT-002", "cat": "Mercato",
        "desc": "Competitor Tier 1 AALTO-Leonardo JV - cattura 2-3 Regioni SNAI con pricing aggressivo",
        "trigger": "AALTO-Leonardo annuncia JV o pilota multi-regionale entro Y2",
        "P": 3, "I": 4, "status": "Open-High",
        "owner": "competitive-intelligence + business-model",
        "response": "Mitigate",
        "mitigation": "Differenziazione cooperativa + sovranita IT; speed to market 6A; partnership CIRA/POLITO; lock-in cooperative Legacoop",
        "rP": 3, "rI": 3, "phase": "Y2-Y3",
        "confidence": "medium",
        "ewi": "AALTO press releases + Leonardo strategy update",
        "fo": None,
        "top": True,
    },
    {
        "id": "RSK-MKT-003", "cat": "Mercato",
        "desc": "Sentinel/Copernicus competizione EO - servizi free open data minano pricing premium",
        "trigger": "Use case Pentema antincendio/SAR realizzabile con Sentinel a costo zero",
        "P": 3, "I": 3, "status": "Open-Medium",
        "owner": "earth-observation-expert + business-model",
        "response": "Mitigate",
        "mitigation": "Posizionamento high-resolution + low-latency vs Sentinel weekly + 10m; servizi value-added analytics + report customized",
        "rP": 2, "rI": 3, "phase": "Y1+",
        "confidence": "high",
        "ewi": "Sentinel-2C/3C launch + Copernicus services expansion",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-MKT-004", "cat": "Mercato",
        "desc": "Pricing pressure - PA italiana clienti price-sensitive vs servizi premium",
        "trigger": "Tender Regione cost ceiling < 50k EUR/anno per servizio EO/SAR",
        "P": 4, "I": 3, "status": "Open-High",
        "owner": "business-model + financial-cfo",
        "response": "Mitigate",
        "mitigation": "Bundle servizi multipli (EO+IR+NTN+monitoring) per economia scale; cost-shared infrastruttura cooperative; modello canone vs ore-volo",
        "rP": 3, "rI": 3, "phase": "Y1-Y3",
        "confidence": "medium-high",
        "ewi": "Tender history pricing + benchmark drone services market IT",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-MKT-005", "cat": "Mercato",
        "desc": "Single-customer concentration Liguria - alternanza politica regionale evapora anchor",
        "trigger": "Elezioni Regione Liguria 2025+ portano cambio amministrazione + scope revision",
        "P": 3, "I": 4, "status": "Open-High",
        "owner": "business-model + sovereign-strategist",
        "response": "Mitigate",
        "mitigation": "Espansione 2-3 regioni SNAI Y3 (Piemonte, Marche, Calabria, Basilicata); LoI pre-formalizzata multi-region; contratti pluriennali con clausole continuita amministrative",
        "rP": 2, "rI": 4, "phase": "Y2-Y3",
        "confidence": "medium",
        "ewi": "Elezioni regionali calendar + cambio assessori Liguria",
        "fo": None,
        "top": True,
    },
    {
        "id": "RSK-MKT-006", "cat": "Mercato",
        "desc": "Adozione cooperative lenta - 10 cooperative pilota non confermano partecipazione",
        "trigger": "M+6 < 6 cooperative confermano partecipazione formale",
        "P": 3, "I": 3, "status": "Open-Medium",
        "owner": "business-model + Legacoop",
        "response": "Mitigate",
        "mitigation": "Workshop cooperative Q+1; valore proposto chiaro + canone accessibile; pilot gratuito M+0-3; community engagement",
        "rP": 2, "rI": 3, "phase": "Y0-Y1",
        "confidence": "medium-high",
        "ewi": "Cooperative engagement metrics + LoI signed",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-MKT-007", "cat": "Mercato",
        "desc": "Servizi NTN 6B mercato - 5G NTN adozione lenta vs piano IRIS2/Starlink",
        "trigger": "5G NTN deployment commerciale < 10% device EU entro Y4",
        "P": 3, "I": 3, "status": "Monitor",
        "owner": "telecom-ntn-payload-expert + business-model",
        "response": "Mitigate",
        "mitigation": "Servizi NTN target settori dedicated (emergency, IoT rurale) vs mass market; pricing capacity wholesale",
        "rP": 3, "rI": 2, "phase": "Y4+",
        "confidence": "medium",
        "ewi": "5G NTN device shipments + IRIS2/Starlink Direct-to-Device launches",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-MKT-008", "cat": "Mercato",
        "desc": "Window IRIS2 chiude - posizionamento 'complementare' falsificato (vedi §5.16bis)",
        "trigger": "Y3 Commissione UE/SpaceRISE escludono layer stratosferici da architettura sovrana",
        "P": 2, "I": 5, "status": "Open-High",
        "owner": "sovereign-strategist + CEO",
        "response": "Mitigate",
        "mitigation": "Position paper 'Italian Stratospheric Sovereignty' Y1-Y2; engagement DG CNECT + DG DEFIS + SpaceRISE; alternative narrative 'EU sovereign multi-layer'",
        "rP": 2, "rI": 4, "phase": "Y2-Y3",
        "confidence": "medium",
        "ewi": "IRIS2 architecture publications + SpaceRISE workplan + UE press releases",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-MKT-009", "cat": "Mercato",
        "desc": "Mercato VTOL UAS satura - 50+ vendor commerciali, commoditization pricing",
        "trigger": "Prezzo medio servizio EO con drone < 5 EUR/ha entro Y3",
        "P": 3, "I": 2, "status": "Monitor",
        "owner": "business-model",
        "response": "Mitigate",
        "mitigation": "Differenziazione service-only + cooperative + sovranita; bundling value-added analytics; vertical integration use cases SNAI",
        "rP": 2, "rI": 2, "phase": "Y2+",
        "confidence": "medium-high",
        "ewi": "Mercato benchmark drone services IT/EU pricing",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-MKT-010", "cat": "Mercato",
        "desc": "Cambiamento normativa SNAI - riduzione fondi Aree Interne 2027+",
        "trigger": "PSNAI 2027+ riduce dotazione fondi aree interne",
        "P": 2, "I": 3, "status": "Monitor",
        "owner": "snai-funding",
        "response": "Mitigate",
        "mitigation": "Diversificazione cliente PA (PC, Difesa civile, agricoltura); fondi Horizon + INTERREG",
        "rP": 2, "rI": 2, "phase": "Y3+",
        "confidence": "medium",
        "ewi": "DPCM PSNAI publications + budget aree interne",
        "fo": None,
        "top": False,
    },

    # ----- OPERATIVI (RSK-OPS) -----------------------------------------------
    {
        "id": "RSK-OPS-001", "cat": "Operativo",
        "desc": "Operazioni invernali Appennino Ligure - neve, ghiaccio, basse temp riducono giorni op.",
        "trigger": "Inverno 2026-27 mostra < 60% giorni operativi vs piano",
        "P": 3, "I": 3, "status": "Open-High",
        "owner": "vtol-uas-specialist + ops",
        "response": "Mitigate",
        "mitigation": "Training pilota inverno; finestre operative meteo-dipendenti; de-icing batterie+celle; planning seasonal",
        "rP": 2, "rI": 3, "phase": "Y1+",
        "confidence": "high",
        "ewi": "Meteo Pentema mensile + giorni op. tracking",
        "fo": None,
        "top": True,
    },
    {
        "id": "RSK-OPS-002", "cat": "Operativo",
        "desc": "Incidente UAS BVLOS sul territorio - danno a persone/cose/aviazione manned",
        "trigger": "Near-miss vs aviazione manned o danno a terzi",
        "P": 2, "I": 5, "status": "Open-Critical",
        "owner": "aviation-regulatory + ops + safety",
        "response": "Mitigate+Transfer",
        "mitigation": "SORA M2 mitigation (parachute, flight termination); GRC mitigation (geofence, Lost-Link); assicurazione adeguata; SOPs rigorosi; training continuo",
        "rP": 1, "rI": 4, "phase": "Y1+",
        "confidence": "high",
        "ewi": "Incident reports settore IT/EU + near-miss interno + FAR/MDR drone",
        "fo": "Se incidente con vittime, intero modello operativo sospeso + responsabilita penale (RSK-REG-020)",
        "top": True,
    },
    {
        "id": "RSK-OPS-003", "cat": "Operativo",
        "desc": "Reclutamento pilota UAS specialist BVLOS - mercato compresso in Italia",
        "trigger": "Tempo hire pilota BVLOS > 6 mesi",
        "P": 4, "I": 3, "status": "Open-High",
        "owner": "HR + business-model",
        "response": "Mitigate",
        "mitigation": "Training in-house pilota junior; partnership con scuole UAS (es. Volandia, scuole Aero IT); contratti competitivi",
        "rP": 3, "rI": 2, "phase": "Y1",
        "confidence": "medium-high",
        "ewi": "Job market UAS pilots IT + scuole UAS graduates",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-OPS-004", "cat": "Operativo",
        "desc": "Disponibilita piattaforma - lead time JOUAV/Tekever 6-12 mesi standard",
        "trigger": "Order placement > M+6 ritarda first flight oltre M+12",
        "P": 3, "I": 3, "status": "Open-Medium",
        "owner": "vtol-uas-specialist + supply-chain",
        "response": "Mitigate",
        "mitigation": "Order placement early M+3; stock spare parts; rental option backup",
        "rP": 2, "rI": 2, "phase": "Y1",
        "confidence": "medium-high",
        "ewi": "Vendor lead time updates + order confirmations",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-OPS-005", "cat": "Operativo",
        "desc": "Manutenzione UAS - downtime non programmato > 20% target",
        "trigger": "Downtime cumulativo Y1 > 60 giorni",
        "P": 3, "I": 3, "status": "Open-Medium",
        "owner": "ops + maintenance",
        "response": "Mitigate",
        "mitigation": "Piano manutenzione preliminare; spare parts kit on-site; vendor support contract; secondo UAV backup Y2",
        "rP": 2, "rI": 3, "phase": "Y1+",
        "confidence": "medium-high",
        "ewi": "Downtime tracking weekly + MTBF metrics",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-OPS-006", "cat": "Operativo",
        "desc": "Coordinamento PC Liguria - tempi attivazione emergency > target",
        "trigger": "Test esercitazione PC mostra tempo attivazione > 4h",
        "P": 3, "I": 3, "status": "Monitor",
        "owner": "ops + community-engagement",
        "response": "Mitigate",
        "mitigation": "MoU PC Liguria + procedure attivazione 24/7; training congiunto PC+Firmamento; SLA contrattuale",
        "rP": 2, "rI": 2, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "Esercitazioni PC + tempo attivazione actuals",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-OPS-007", "cat": "Operativo",
        "desc": "Meteo estremo - alluvioni Liguria > 100mm/24h blocca operazioni",
        "trigger": "Frequenza eventi estremi 2026 > 2 vs media",
        "P": 4, "I": 2, "status": "Open-High",
        "owner": "ops + business-model",
        "response": "Accept+Mitigate",
        "mitigation": "Buffer giorni operativi piano; UAV ricovero indoor; servizi differiti accettati cliente PA",
        "rP": 3, "rI": 2, "phase": "Y1+",
        "confidence": "high",
        "ewi": "Meteo storico Liguria + ARPAL allerte",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-OPS-008", "cat": "Operativo",
        "desc": "Comunita Pentema - contestazioni rumore/privacy/sicurezza limitano accettazione",
        "trigger": "Reclami formali residenti > 5/anno o petizione",
        "P": 2, "I": 3, "status": "Open-Medium",
        "owner": "community-engagement + ops",
        "response": "Mitigate",
        "mitigation": "Workshop comunita pre-deploy M+6; canone partecipazione + benefit locali; trasparenza operazioni; finestre orarie",
        "rP": 1, "rI": 2, "phase": "Y1+",
        "confidence": "medium-high",
        "ewi": "Community feedback monitoring + reclami tracking",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-OPS-009", "cat": "Operativo",
        "desc": "Scalabilita ops Y2-Y3 - team ops sottodimensionato per 2-3 regioni",
        "trigger": "Y2 espansione 2 regioni richiede team 2x ma hire delay",
        "P": 3, "I": 3, "status": "Monitor",
        "owner": "HR + ops",
        "response": "Mitigate",
        "mitigation": "Pianificazione hire Y2 (3-4 FTE ops); training accelerato; tooling automazione operations management",
        "rP": 2, "rI": 2, "phase": "Y2-Y3",
        "confidence": "medium",
        "ewi": "Hire pipeline + tools deployment",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-OPS-010", "cat": "Operativo",
        "desc": "Data pipeline processing latency - target near-real-time vs realta batch",
        "trigger": "Test pipeline EO data delivery > 1h vs target 15 min",
        "P": 3, "I": 2, "status": "Monitor",
        "owner": "earth-observation-expert + IT",
        "response": "Mitigate",
        "mitigation": "Edge processing UAV + cloud parallel; ottimizzazione pipeline; SLA realistico con cliente",
        "rP": 2, "rI": 2, "phase": "Y1+",
        "confidence": "medium-high",
        "ewi": "Pipeline latency monitoring + SLA compliance",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-OPS-011", "cat": "Operativo",
        "desc": "Sicurezza personale - DUVRI ops montane + formazione D.Lgs.81/2008",
        "trigger": "Audit ASL rivela DUVRI/formazione carente",
        "P": 2, "I": 3, "status": "Open-Medium",
        "owner": "safety + HR",
        "response": "Mitigate",
        "mitigation": "DUVRI formalizzato M+3; formazione D.Lgs.81 obbligatoria; PPE adeguati; medical surveillance",
        "rP": 1, "rI": 2, "phase": "Y1+",
        "confidence": "high",
        "ewi": "Audit ASL programma + formazione completion",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-OPS-012", "cat": "Operativo",
        "desc": "Hangar Pentema disponibilita - struttura fisica adeguata non disponibile in loco",
        "trigger": "Survey M+3 non trova hangar suitable Pentema",
        "P": 2, "I": 3, "status": "Open-Medium",
        "owner": "ops + community",
        "response": "Mitigate",
        "mitigation": "Survey siti alternativi Torriglia/Casella; container temporaneo; partnership cooperative locali per spazi",
        "rP": 1, "rI": 2, "phase": "Y1",
        "confidence": "medium-high",
        "ewi": "Site survey results + lease negotiations",
        "fo": None,
        "top": False,
    },

    # ----- SUPPLY CHAIN (RSK-SUP) --------------------------------------------
    {
        "id": "RSK-SUP-001", "cat": "Supply Chain",
        "desc": "Lead time JOUAV (vendor CN) - escalation tariffaria USA-CN blocca import",
        "trigger": "USA/UE pubblicano restrizioni componenti CN aerospace",
        "P": 3, "I": 3, "status": "Open-High",
        "owner": "vtol-uas-specialist + supply-chain",
        "response": "Mitigate",
        "mitigation": "Plan B Tekever (PT) pronto entro M+9; stock spare parts 12 mesi; valutazione Quantum (DE) terzo backup",
        "rP": 2, "rI": 3, "phase": "Y1+",
        "confidence": "high",
        "ewi": "USA BIS notices + UE FDI screening + tariffe announcements",
        "fo": "Se sanzioni bloccano JOUAV entro M+6, attivazione Plan B Tekever; CapEx 6A +100-200k EUR + delay 3-6 mesi",
        "top": True,
    },
    {
        "id": "RSK-SUP-002", "cat": "Supply Chain",
        "desc": "Single source celle solari multi-junction GaAs - concentrazione Spectrolab (US)",
        "trigger": "Spectrolab export restriction o capacity allocation military priority",
        "P": 3, "I": 4, "status": "Open-High",
        "owner": "propulsion-energy-engineer + supply-chain",
        "response": "Mitigate",
        "mitigation": "Azur Space (DE) qualification path; engagement Enel 3SUN (IT); EU CRMA Strategic Project status",
        "rP": 2, "rI": 3, "phase": "Y3+",
        "confidence": "medium",
        "ewi": "Spectrolab capacity announcements + Azur Space EU production",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-SUP-003", "cat": "Supply Chain",
        "desc": "Batterie LiS - capacity allocation Northvolt/Italvolt incerta",
        "trigger": "Northvolt bankruptcy o Italvolt ritardo > 2 anni production",
        "P": 4, "I": 4, "status": "Open-High",
        "owner": "propulsion-energy-engineer + supply-chain",
        "response": "Mitigate",
        "mitigation": "Multi-vendor LiS pipeline (Oxis Energy heir, Lyten, NexTech); custom pack assembly con cell suppliers diversi",
        "rP": 3, "rI": 3, "phase": "Y3+",
        "confidence": "medium",
        "ewi": "Battery vendors financials + production timelines",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-SUP-004", "cat": "Supply Chain",
        "desc": "Chip GNC aerospace-grade - Honeywell/Collins concentration + ITAR risk",
        "trigger": "ITAR application su componenti aerospace-grade per HAPS EU",
        "P": 2, "I": 4, "status": "Open-High",
        "owner": "avionics-gnc-engineer + supply-chain",
        "response": "Mitigate",
        "mitigation": "STM (FR-IT) qualification path; engagement Leonardo Elettronica; dual sourcing US+EU",
        "rP": 2, "rI": 3, "phase": "Y3+",
        "confidence": "medium",
        "ewi": "BIS Entity List update + STM/Leonardo capacity",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-SUP-005", "cat": "Supply Chain",
        "desc": "FPGA radhard DAL-B - Microchip RTPolarFire single source EU acquisable",
        "trigger": "Microchip allocation military priority o delay > 12 mesi",
        "P": 2, "I": 3, "status": "Monitor",
        "owner": "avionics-gnc-engineer + supply-chain",
        "response": "Mitigate",
        "mitigation": "Xilinx Versal AI Edge alternative; stock buffer 24 mesi; design flessibile FPGA-vendor agnostic",
        "rP": 2, "rI": 2, "phase": "Y3+",
        "confidence": "medium",
        "ewi": "Microchip lead times + Xilinx alternative qualification",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-SUP-006", "cat": "Supply Chain",
        "desc": "Materiali compositi CFRP - Toray (JP) + Hexcel (FR-US) supplier concentration",
        "trigger": "Disruption shipping Asia-EU o ITAR Hexcel US-derived",
        "P": 2, "I": 3, "status": "Monitor",
        "owner": "aero-structures-engineer + supply-chain",
        "response": "Mitigate",
        "mitigation": "Solvay (BE) qualification + Italian carbon producers (Mapei, Saati) + stock 12 mesi",
        "rP": 2, "rI": 2, "phase": "Y3+",
        "confidence": "medium",
        "ewi": "Shipping rates + Solvay/Mapei capacity",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-SUP-007", "cat": "Supply Chain",
        "desc": "Lino fibra Italia - capacita produzione artigianale, lead time variabile",
        "trigger": "Produttore lino IT incapace fornire qualita aeronautica certificata",
        "P": 3, "I": 2, "status": "Monitor",
        "owner": "aero-structures-engineer + supply-chain",
        "response": "Mitigate",
        "mitigation": "Multi-vendor lino IT (Linificio Canapificio Nazionale, Biogear partner); test qualifica panel; backup vendor BE/NL",
        "rP": 2, "rI": 2, "phase": "Y3+",
        "confidence": "medium",
        "ewi": "Lino vendor capacity + qualifica test results",
        "fo": None,
        "top": False,
    },

    # ----- PRIVACY/LEGALE (RSK-PRV) ------------------------------------------
    {
        "id": "RSK-PRV-001", "cat": "Privacy/Legale",
        "desc": "DPIA blocca casi d'uso sorveglianza territoriale - Garante richiede riduzione scope",
        "trigger": "DPIA review M+6 da Garante richiede scope reduction",
        "P": 2, "I": 3, "status": "Open-Medium",
        "owner": "data-privacy-counsel + DPO",
        "response": "Mitigate",
        "mitigation": "Privacy by design fin da concept; edge anonymization; geofence aree residenziali; DPIA consultazione preventiva Garante",
        "rP": 1, "rI": 2, "phase": "Y1",
        "confidence": "medium-high",
        "ewi": "Garante DPIA feedback + provvedimenti settore",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-PRV-002", "cat": "Privacy/Legale",
        "desc": "Contenzioso privacy individuale - cittadino ricorre vs sorvolo proprieta privata",
        "trigger": "Ricorso TAR/Garante da cittadino Pentema",
        "P": 2, "I": 3, "status": "Open-Medium",
        "owner": "data-privacy-counsel + legal",
        "response": "Mitigate+Transfer",
        "mitigation": "Trasparenza operazioni; consenso preventivo proprietari grandi appezzamenti; assicurazione legal protection; mediation community",
        "rP": 1, "rI": 2, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "Reclami tracking + comunicazioni TAR",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-PRV-003", "cat": "Privacy/Legale",
        "desc": "GDPR breach data subject - leak dati EO + biometrici",
        "trigger": "Cyber incident con esfiltrazione dati sensibili",
        "P": 2, "I": 4, "status": "Open-High",
        "owner": "DPO + CISO",
        "response": "Mitigate+Transfer",
        "mitigation": "ISMS Part-IS + ISO 27001; crittografia at-rest + in-transit; incident response plan; assicurazione cyber",
        "rP": 1, "rI": 3, "phase": "Y1+",
        "confidence": "medium-high",
        "ewi": "Security incidents tracking + threat intel feeds",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-PRV-004", "cat": "Privacy/Legale",
        "desc": "Contenzioso cooperative - dispute interno governance/revenue sharing",
        "trigger": "Cooperative member ricorre vs Firmamento o tra loro",
        "P": 2, "I": 3, "status": "Monitor",
        "owner": "legal + business-model",
        "response": "Mitigate",
        "mitigation": "Accordo cooperativa formalizzato + statuto chiaro; arbitrato Legacoop; comunicazione trasparente",
        "rP": 1, "rI": 2, "phase": "Y1+",
        "confidence": "medium-high",
        "ewi": "Member feedback + assemblea decisions",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-PRV-005", "cat": "Privacy/Legale",
        "desc": "Liability contrattuale PA - SLA non rispettato in emergenza (PC, antincendio)",
        "trigger": "Servizio non disponibile durante emergenza con conseguenze",
        "P": 2, "I": 4, "status": "Open-High",
        "owner": "legal + ops",
        "response": "Mitigate+Transfer",
        "mitigation": "SLA realistici con buffer; assicurazione professional indemnity; ridondanza operativa; clausole force majeure",
        "rP": 1, "rI": 3, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "SLA actuals tracking + dispute history",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-PRV-006", "cat": "Privacy/Legale",
        "desc": "AI Act conformity assessment - Notified Body bottleneck (cf RSK-REG-016/009)",
        "trigger": "Conformity assessment AI Act > 12 mesi ritardo",
        "P": 3, "I": 3, "status": "Monitor",
        "owner": "data-privacy + systems-engineer",
        "response": "Mitigate",
        "mitigation": "Engagement early NB; pre-assessment documentazione; design AI Act-compliant fin da concept",
        "rP": 2, "rI": 2, "phase": "Y2-Y3",
        "confidence": "medium",
        "ewi": "NB capacity reports + assessment timelines",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-PRV-007", "cat": "Privacy/Legale",
        "desc": "IP infringement - patent claim su tecnologia HALE solare o NTN",
        "trigger": "Cease & desist letter da competitor (AALTO, Skydweller)",
        "P": 2, "I": 4, "status": "Monitor",
        "owner": "legal + IP counsel",
        "response": "Mitigate",
        "mitigation": "Freedom-to-Operate analysis pre-Phase B; patent portfolio Firmamento; differenziazione design vs competitor",
        "rP": 1, "rI": 3, "phase": "Y3+",
        "confidence": "medium",
        "ewi": "Patent landscape monitoring + competitor IP filings",
        "fo": None,
        "top": False,
    },

    # ----- CYBERSECURITY (RSK-SEC) -------------------------------------------
    {
        "id": "RSK-SEC-001", "cat": "Cybersecurity",
        "desc": "Cyber attack ground segment - data breach Pentema (NIS2 incident reportable)",
        "trigger": "Intrusion detected su GS Pentema o cloud Aruba/OVH",
        "P": 2, "I": 4, "status": "Open-High",
        "owner": "CISO (new) + IT",
        "response": "Mitigate",
        "mitigation": "Zero-trust architecture; segregazione rete; NIS2 readiness + ISMS Part-IS; SOC managed service; ISO 27001",
        "rP": 1, "rI": 3, "phase": "Y1+",
        "confidence": "medium-high",
        "ewi": "SIEM alerts + ACN threat intel + sector incidents",
        "fo": None,
        "top": True,
    },
    {
        "id": "RSK-SEC-002", "cat": "Cybersecurity",
        "desc": "Hijacking C2 link UAV BVLOS - takeover malicious operator",
        "trigger": "Anomalia command pattern detected su FCS",
        "P": 1, "I": 5, "status": "Open-High",
        "owner": "CISO + avionics-gnc-engineer",
        "response": "Mitigate",
        "mitigation": "Crypto authentication AES-256 + mutual TLS; rolling keys; flight termination automatica su anomaly; air-gap FCS critical",
        "rP": 1, "rI": 4, "phase": "Y1+",
        "confidence": "medium-high",
        "ewi": "FCS anomaly detection + sector threat reports",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-SEC-003", "cat": "Cybersecurity",
        "desc": "Ransomware ground segment - servizi PA bloccati + reputazione",
        "trigger": "Ransomware infection rilevato",
        "P": 2, "I": 4, "status": "Open-High",
        "owner": "CISO + IT",
        "response": "Mitigate",
        "mitigation": "Backup offsite 3-2-1; immutable backups; EDR su endpoint; ZTNA su accessi remoti; cyber insurance",
        "rP": 1, "rI": 3, "phase": "Y1+",
        "confidence": "medium-high",
        "ewi": "Ransomware sector trends + EDR alerts",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-SEC-004", "cat": "Cybersecurity",
        "desc": "Supply chain attack - componente HW/SW compromesso fornitore",
        "trigger": "CVE critico componente in BoM o vendor compromise public",
        "P": 3, "I": 3, "status": "Monitor",
        "owner": "CISO + supply-chain",
        "response": "Mitigate",
        "mitigation": "SBOM (Software Bill of Materials) maintained; vendor security assessment; firmware signing; CVE monitoring continuous",
        "rP": 2, "rI": 3, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "CVE feeds + vendor advisory notifications",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-SEC-005", "cat": "Cybersecurity",
        "desc": "Insider threat - dipendente o cooperative member compromette dati o operations",
        "trigger": "Anomaly access pattern o data exfiltration interno",
        "P": 2, "I": 3, "status": "Monitor",
        "owner": "CISO + HR",
        "response": "Mitigate",
        "mitigation": "RBAC + least privilege; DLP solution; background check personale critico; offboarding procedure",
        "rP": 1, "rI": 2, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "User behavior analytics + access pattern monitoring",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-SEC-006", "cat": "Cybersecurity",
        "desc": "DDoS service link HAPS - degradazione servizi NTN/EO downlink",
        "trigger": "DDoS pattern detected su gateway Ka-band o cloud endpoint",
        "P": 2, "I": 3, "status": "Monitor",
        "owner": "CISO + telecom-ntn",
        "response": "Mitigate",
        "mitigation": "DDoS protection upstream provider (Cloudflare, Akamai); rate limiting; geo-blocking",
        "rP": 1, "rI": 2, "phase": "Y3+",
        "confidence": "medium-high",
        "ewi": "Traffic anomaly detection",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-SEC-007", "cat": "Cybersecurity",
        "desc": "Compromissione DOA partner - Phase C dependency su DOA esterna",
        "trigger": "DOA partner compromised o licenza sospesa",
        "P": 2, "I": 4, "status": "Monitor",
        "owner": "CISO + sovereign-strategist",
        "response": "Mitigate",
        "mitigation": "Multi-partner DOA option; due diligence partner cybersecurity; contract clauses continuity",
        "rP": 1, "rI": 3, "phase": "Y5+",
        "confidence": "medium",
        "ewi": "Partner security posture review",
        "fo": None,
        "top": False,
    },

    # ----- RISORSE UMANE (RSK-HR) --------------------------------------------
    {
        "id": "RSK-HR-001", "cat": "Risorse Umane",
        "desc": "Difficolta reclutamento pilota UAS specialist BVLOS",
        "trigger": "Tempo hire pilota > 6 mesi",
        "P": 4, "I": 3, "status": "Open-High",
        "owner": "HR + business-model",
        "response": "Mitigate",
        "mitigation": "Training in-house pilota junior; partnership scuole UAS; contratti competitivi + benefits cooperative",
        "rP": 3, "rI": 2, "phase": "Y1",
        "confidence": "medium-high",
        "ewi": "Job market UAS pilots IT + LinkedIn data + scuole UAS",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-HR-002", "cat": "Risorse Umane",
        "desc": "Reclutamento CISO + DPO + Head Regulatory - 3 ruoli senior in mercato compresso",
        "trigger": "Tempo hire > 9 mesi (vs target 3-6 mesi)",
        "P": 4, "I": 4, "status": "Open-Critical",
        "owner": "HR + CEO",
        "response": "Mitigate",
        "mitigation": "Headhunter specializzati; contratti competitivi (180-220k EUR/anno per CISO senior); part-time fractional CISO/DPO M+0-6; partnership consulting Legal/Cyber",
        "rP": 3, "rI": 3, "phase": "Y0-Y1",
        "confidence": "medium",
        "ewi": "Hire pipeline weekly + headhunter pipeline status",
        "fo": "Se ruoli senior non riempiti M+9, NIS2/Part-IS compliance a rischio + cap.OpEx esplode con consulting fees",
        "top": True,
    },
    {
        "id": "RSK-HR-003", "cat": "Risorse Umane",
        "desc": "Turnover senior team - exit founder member o key engineer",
        "trigger": "Member team key fa exit non programmato",
        "P": 2, "I": 4, "status": "Open-High",
        "owner": "CEO + HR",
        "response": "Mitigate",
        "mitigation": "Founder agreement con vesting + restrictive covenants; equity incentive plan team; succession planning",
        "rP": 1, "rI": 3, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "Engagement surveys + 1-on-1 reviews + market salary benchmarks",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-HR-004", "cat": "Risorse Umane",
        "desc": "Formazione team Part-IS ISMS NIS2 - mancanza certificazioni interne",
        "trigger": "Audit Part-IS rivela formazione team carente",
        "P": 3, "I": 2, "status": "Open-Medium",
        "owner": "HR + CISO",
        "response": "Mitigate",
        "mitigation": "Training plan certificazioni (ISO 27001 LA, CISM, CISSP); budget formazione 10k EUR/FTE/anno; e-learning piattaforma",
        "rP": 2, "rI": 2, "phase": "Y1+",
        "confidence": "medium-high",
        "ewi": "Training completion rates + certifications obtained",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-HR-005", "cat": "Risorse Umane",
        "desc": "Diversita team - female engineering ratio < 20% riduce eligibility bandi EU",
        "trigger": "Bandi EU Gender Equality Plan richiesto e Firmamento non compliant",
        "P": 3, "I": 2, "status": "Monitor",
        "owner": "HR + CEO",
        "response": "Mitigate",
        "mitigation": "GEP formalizzato; hire diversificato; partnership universita STEM women; mentoring programma",
        "rP": 2, "rI": 2, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "Team composition tracking + GEP compliance",
        "fo": None,
        "top": False,
    },

    # ----- REPUTAZIONALE (RSK-REP) -------------------------------------------
    {
        "id": "RSK-REP-001", "cat": "Reputazionale",
        "desc": "Incidente UAS BVLOS pubblico - copertura mediatica nazionale negativa",
        "trigger": "Incidente con conseguenze + coverage stampa nazionale",
        "P": 2, "I": 5, "status": "Open-Critical",
        "owner": "CEO + comms",
        "response": "Mitigate+Transfer",
        "mitigation": "Crisis communication plan; PR specialist on retainer; safety record positive + trasparenza; assicurazione legal protection + comms",
        "rP": 1, "rI": 4, "phase": "Y1+",
        "confidence": "medium-high",
        "ewi": "Settore drone incidents IT/EU + sentiment monitoring",
        "fo": None,
        "top": True,
    },
    {
        "id": "RSK-REP-002", "cat": "Reputazionale",
        "desc": "Contestazioni comunita Pentema - copertura locale negativa + petizione",
        "trigger": "Petizione comunita formalmente presentata o coverage giornalismo locale negativo",
        "P": 2, "I": 3, "status": "Open-Medium",
        "owner": "community-engagement + CEO",
        "response": "Mitigate",
        "mitigation": "Engagement comunita continuo + benefit locali; mediazione Comune Torriglia + ENTE Parco; community advisory board",
        "rP": 1, "rI": 2, "phase": "Y1+",
        "confidence": "medium-high",
        "ewi": "Community sentiment tracking + media monitoring locale",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REP-003", "cat": "Reputazionale",
        "desc": "Controversie cooperative - conflitto interno reputation rete Legacoop",
        "trigger": "Cooperative member exit pubblico con motivazioni negative",
        "P": 2, "I": 3, "status": "Monitor",
        "owner": "business-model + Legacoop",
        "response": "Mitigate",
        "mitigation": "Trasparenza governance; mediazione Legacoop; comunicazione one-voice",
        "rP": 1, "rI": 2, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "Member feedback + Legacoop communications",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REP-004", "cat": "Reputazionale",
        "desc": "Greenwashing accusation - claim ESG/lino contestato come marketing non sostanziale",
        "trigger": "Articolo investigativo o NGO accusa greenwashing",
        "P": 3, "I": 3, "status": "Open-Medium",
        "owner": "CEO + comms + sustainability",
        "response": "Mitigate",
        "mitigation": "LCA (Life Cycle Assessment) entro M+12; claim ESG verificabili + report sostenibilita; certification (B-Corp aspirazionale)",
        "rP": 2, "rI": 2, "phase": "Y1+",
        "confidence": "medium",
        "ewi": "Media monitoring ESG/sustainability + NGO communications",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-REP-005", "cat": "Reputazionale",
        "desc": "Geopolitical narrative misalignment - linguaggio 'alternativa Starlink' attiva RSK-GEO-001",
        "trigger": "Comunicazione interna o esterna usa framing competitive vs Starlink",
        "P": 3, "I": 4, "status": "Open-High",
        "owner": "CEO + comms + sovereign-strategist",
        "response": "Mitigate",
        "mitigation": "Communication guidelines: 'complementare a IRIS2' SEMPRE; mai 'vs Starlink'; press training; review pre-publication",
        "rP": 2, "rI": 3, "phase": "Y1+",
        "confidence": "high",
        "ewi": "Communications audit monthly + media coverage analysis",
        "fo": None,
        "top": False,
    },

    # ----- GEOPOLITICI (RSK-GEO) — reference ai RESERVED (no detail) --------
    {
        "id": "RSK-GEO-001", "cat": "Geopolitico",
        "desc": "[RESERVED — vedi riferimenti/RESERVED-rischi-geopolitici.md §2] Frizione con USA / posizione narrativa stratosferica",
        "trigger": "[RESERVED — accesso ristretto]",
        "P": 3, "I": 4, "status": "Open-High",
        "owner": "sovereign-strategist + CEO",
        "response": "Mitigate",
        "mitigation": "[RESERVED] Linguaggio 'complementare IRIS2'; diversification supply; dialogo Atlantico NATO DIANA; avoid US dependency core",
        "rP": 2, "rI": 3, "phase": "Y4+",
        "confidence": "medium",
        "ewi": "[RESERVED — Early Warning Indicators trimestrale]",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-GEO-002", "cat": "Geopolitico",
        "desc": "[RESERVED] Classificazione strategica nazionale + obblighi governance",
        "trigger": "[RESERVED]",
        "P": 2, "I": 4, "status": "Monitor",
        "owner": "sovereign-strategist + legal + CEO",
        "response": "Mitigate",
        "mitigation": "[RESERVED] Engagement preventivo istituzioni; ownership stabile IT; notifica preventiva; compliance NIS2",
        "rP": 2, "rI": 3, "phase": "Y3+",
        "confidence": "medium",
        "ewi": "[RESERVED]",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-GEO-003", "cat": "Geopolitico",
        "desc": "[RESERVED] Dipendenza supply chain non-EU - escalation geopolitica blocca componenti",
        "trigger": "[RESERVED]",
        "P": 3, "I": 4, "status": "Open-High",
        "owner": "sovereign-strategist + supply-chain",
        "response": "Mitigate",
        "mitigation": "[RESERVED] Supply chain mapping; EU sovereign suppliers roadmap; stock buffer; EU CRMA Strategic Project",
        "rP": 2, "rI": 3, "phase": "Y2+",
        "confidence": "medium",
        "ewi": "[RESERVED]",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-GEO-004", "cat": "Geopolitico",
        "desc": "[RESERVED] Misalignment con programmi sovrani EU multi-orbita",
        "trigger": "[RESERVED]",
        "P": 3, "I": 4, "status": "Open-High",
        "owner": "sovereign-strategist + CEO",
        "response": "Mitigate",
        "mitigation": "[RESERVED] Position paper complementarita; engagement DG CNECT + DG DEFIS; joint workshop",
        "rP": 2, "rI": 3, "phase": "Y2-Y4",
        "confidence": "medium",
        "ewi": "[RESERVED]",
        "fo": None,
        "top": False,
    },
    {
        "id": "RSK-GEO-005", "cat": "Geopolitico",
        "desc": "[RESERVED] Acquisizione difensiva da incumbent IT settore aerospace",
        "trigger": "[RESERVED]",
        "P": 3, "I": 4, "status": "Open-High",
        "owner": "CEO + Board",
        "response": "Mitigate",
        "mitigation": "[RESERVED] Capital structure resistente; anchor investors non-ostili; speed to value; pluralismo cooperative",
        "rP": 2, "rI": 3, "phase": "Y3-Y4",
        "confidence": "medium",
        "ewi": "[RESERVED]",
        "fo": None,
        "top": False,
    },
]


def calc_score(r):
    r["score"] = r["P"] * r["I"]
    r["rscore"] = r["rP"] * r["rI"]
    return r


for r in RISKS:
    calc_score(r)

# ---------------------------------------------------------------------------
# FMECA / FTA / EWI datasets
# ---------------------------------------------------------------------------
FMECA_PAYLOAD = [
    ("Camera RGB Phase One iXM 100", "No image", "Sensor failure (electronic)", "Loss EO mission", "Mission abort", 3, 2, 2, "Ridondanza dual sensor + healthcheck onboard"),
    ("Camera RGB Phase One iXM 100", "Blur image", "Vibrazioni gimbal / IBIS fail", "Quality degradata", "Re-fly necessario", 2, 3, 3, "Gimbal damping + IBIS + post-processing deblur"),
    ("Camera RGB Phase One iXM 100", "Lens fog / cold", "Termica + condensa", "Quality degradata", "Re-fly", 2, 3, 2, "Lens heater + dry-air purge"),
    ("Gimbal Phase One IXM mount", "Stuck position", "Motor failure singolo asse", "Off-target framing", "Reduced area coverage", 3, 3, 2, "Service interval + ridondanza motor 2oo3"),
    ("Gimbal", "Vibration excess", "Bearings wear", "Image blur", "Quality degradata", 2, 3, 3, "Maintenance schedule + bearing monitor"),
    ("IR sensor WIRIS Pro LWIR", "Calibrazione persa", "Termica + tempo + drift", "False alarm / missed hotspot", "False positive antincendio", 4, 3, 4, "NUC frequente + crosscheck con RGB + ground truth"),
    ("IR sensor", "Sensor saturation", "Sole diretto / sorgente HOT", "Image saturated", "Detection persa", 3, 2, 3, "Auto exposure + filter + look-angle planning"),
    ("Storage on-board SSD aerospace", "Corruzione data", "Bit flip / temperatura / vibrazione", "Loss data missione", "Re-fly necessario", 3, 2, 3, "RAID-style ridondanza + ECC + on-line backup"),
    ("Storage", "Write failure full", "Capacity exceeded", "Mission data partial loss", "Reduced data", 2, 3, 2, "Capacity monitoring + auto-purge oldest"),
    ("Downlink data RF/SATCOM", "Interruzione bandwidth", "RF interference / weather", "Delay delivery", "Latency degradata", 2, 3, 4, "Buffer + retry + alt downlink + ACM"),
    ("Downlink data", "Authentication fail", "Crypto key error", "Loss telemetry/data link", "Manual override", 2, 2, 2, "Key rotation + redundant auth + ground backup"),
    ("LiDAR payload (Y3+)", "Range error", "Laser drift", "GSD inaccurato", "Mapping degradato", 2, 2, 3, "Calibration plate + crosscheck"),
    ("Multispectral MicaSense (Y2+)", "Filter mismatch", "Filter wheel jam", "Banda persa", "Index errato", 2, 2, 2, "Pre-flight test + redundant filter"),
    ("Power conversion DC-DC payload", "Voltage drop", "Capacitor failure", "Power loss sensor", "Sensor reset cascade", 3, 2, 2, "Caps redundant + monitor V real-time"),
    ("Payload interface bus", "Data bus corruption", "EMI / connector", "Wrong telemetry", "Data integrity loss", 2, 3, 3, "Shielding + EMI filter + checksum CRC"),
]

FMECA_AVIONICA = [
    ("FCS autopilot DAL-C", "Reboot unexpected", "SEU / SW exception", "Loss FCS momentaneo", "Risk loss vehicle", 4, 2, 3, "2oo3 voting + watchdog + ECC memory"),
    ("FCS", "Software defect runaway", "SW bug undetected", "Loss FCS control", "Loss of vehicle", 5, 1, 3, "DAL-C process + testing + formal verification"),
    ("IMU primary triplex", "Drift gyro out of spec", "Aging / temperature", "Attitude error", "Reduced nav accuracy", 3, 3, 2, "Triplex IMU + crosscheck + Kalman filter"),
    ("IMU", "Single IMU failure", "Hardware fault", "Reduced redundancy", "Continue with 2oo3", 2, 3, 2, "Triplex + auto-detection + isolation"),
    ("GNSS dual-frequency", "Spoofing detected", "Adversary signal", "Position error", "Risk navigation", 3, 3, 3, "RAIM + IMU dead-reckoning + multi-constellation"),
    ("GNSS", "Jamming sustained", "Adversary signal", "Position loss", "Risk navigation", 3, 3, 2, "Multi-frequency + IMU fallback + Lost-Link"),
    ("ADS-B IN transponder", "Failure no signal", "HW fault", "Loss traffic awareness", "Increased risk collision", 3, 2, 1, "Vendor MTBF + spare + ATC notification"),
    ("Communication C2 RF primary", "Lost-Link", "RF interference / range", "C2 loss", "Trigger Lost-Link Procedure", 3, 3, 2, "FHSS + SATCOM backup + RtB"),
    ("SATCOM Iridium backup", "Latency excessive", "Network congestion", "Delayed C2 response", "Increased pilot reaction time", 2, 2, 2, "SLA Iridium + buffer + critical command priority"),
    ("Power distribution avionics", "Voltage drop", "PSU failure", "Avionics reset", "FCS reboot", 3, 3, 2, "Redundant PSU + battery backup + monitor"),
    ("Lost-Link procedure logic", "Wrong RtB waypoint", "Configuration error", "Vehicle goes wrong direction", "Risk recovery", 2, 4, 2, "Pre-flight check + dual config + simulation"),
    ("Flight termination system", "False trigger", "Sensor anomaly", "Unintended FTS", "Loss vehicle but safe ground", 1, 5, 2, "Multi-sensor confirmation + manual confirm option"),
    ("FTS", "Failure to trigger", "HW or logic fault", "FTS does not deploy on command", "Risk loss vehicle uncontrolled", 1, 5, 3, "Periodic test + dual-channel arming + simulation"),
    ("Parachute deployment", "Failure to deploy", "Mechanism jam", "Hard landing", "Loss vehicle + risk to ground", 2, 5, 3, "Dual pyrotechnic + maintenance + test deployment"),
    ("Air Data System (Pitot)", "Pitot icing / clog", "Ice / debris", "Wrong airspeed", "Aerodynamic stall risk", 3, 3, 2, "Pitot heater + redundant Pitot + alpha-beta sensor"),
]

FMECA_PROP = [
    ("Motor electric brushless 6A", "Bearing failure", "Wear / contamination", "Power loss", "Loss propulsion", 3, 2, 3, "Vibration monitor + service interval + spare"),
    ("Motor electric brushless", "Winding short", "Insulation degradation", "Loss motor", "Loss propulsion", 4, 2, 3, "Insulation monitor + thermal cutoff + redundancy"),
    ("Propeller composite", "Blade strike", "FOD / bird", "Vibration + power loss", "Emergency landing", 3, 3, 3, "Inspection pre-flight + replacement schedule + spare"),
    ("Hybrid engine gasoline 6A", "Carburator clog", "Fuel quality / debris", "Power loss", "Emergency landing", 3, 3, 4, "Fuel filter + quality control + maintenance"),
    ("Hybrid engine", "Ignition failure", "Spark plug fault", "Engine stop", "Glide to recovery + battery", 3, 2, 3, "Dual ignition + spare plugs + battery override"),
    ("Battery LiPo VTOL 6A", "Cell unbalance", "Cell aging", "Reduced capacity", "Reduced autonomy", 4, 2, 2, "BMS active balancing + cell monitoring"),
    ("Battery LiPo", "Thermal runaway", "Cell defect / overcharge", "Fire / explosion", "Loss vehicle + ground risk", 2, 5, 4, "Cell-level fuse + intumescent + BMS + ATEX storage"),
    ("Battery LiS HALE 6B", "Capacity fade", "Cycle aging", "Reduced storage", "Reduced endurance", 4, 3, 3, "Cycle monitoring + DoD limit + replacement at 80% SoH"),
    ("Battery LiS", "Thermal runaway HALE", "Cell defect stratospheric thermal", "Fire / explosion HALE", "Loss vehicle", 2, 5, 4, "Cell-level fuse + thermal monitor + emergency vent"),
    ("Solar cells GaAs HALE", "Degradation UV/thermal", "Stratospheric environment", "Reduced output", "Reduced endurance", 3, 3, 3, "Margin design + redundant strings + bypass diodes"),
    ("Solar cells", "Hailstone / impact damage", "Atmospheric event ascent/descent", "String loss", "Reduced output", 2, 3, 2, "Bypass diodes + redundant strings + reinforced surface"),
    ("MPPT charge controller", "MPPT loss tracking", "SW defect / sensor", "Power output reduced", "Reduced storage", 3, 2, 2, "Redundant MPPT + monitoring + fallback fixed point"),
    ("Thermal management batteries", "Heater failure HALE", "Heater HW fault", "Battery freeze", "Loss storage capacity", 3, 4, 2, "Redundant heater + thermal monitor + insulation"),
    ("Thermal management", "Coolant loss payload", "Leak", "Payload overheat", "Payload shutdown", 2, 3, 3, "Closed loop + monitor + payload thermal cutoff"),
    ("Fuel system 6A (hybrid)", "Fuel leak", "Connector / hose fail", "Fuel loss + fire risk", "Emergency landing", 2, 4, 3, "Double-seal + leak detector + emergency cutoff"),
]

# FTA — top events
FTA_LOV = """
TOP EVENT: Loss of Vehicle in BVLOS (Percorso 6A)
Target: P < 1E-5 / flight hour (compliant SAIL III SORA 2.5)

Loss of Vehicle [TOP, OR-gate]
|
+--OR--> Lost-Link permanente + RtB fail
|  |
|  +--AND--> Lost-Link C2 (RF + SATCOM)        P ~= 1E-4/h
|  |    |
|  |    +--RF link lost (FHSS jam/range)       P ~= 1E-3/h
|  |    +--SATCOM Iridium backup lost          P ~= 1E-3/h
|  |
|  +--RtB procedure fail given Lost-Link       P ~= 1E-3 cond.
|  Cut set probability: 1E-4 * 1E-3 = 1E-7/h
|
+--OR--> Avaria FCS critica
|  |
|  +--Failure autopilot DAL-C primary          P ~= 1E-4/h
|  +--AND--> IMU1 + IMU2 + IMU3 failure        P ~= 1E-9/h
|  +--GNSS spoofing sustained + no IMU         P ~= 1E-5/h
|  +--Software runaway DAL-C undetected        P ~= 1E-6/h
|  Cut set probability: ~1E-5/h
|
+--OR--> Avaria propulsione + landing fail
|  |
|  +--Engine failure 6A (single)               P ~= 1E-4/h
|  +--AND--> Engine fail + parachute fail      P ~= 1E-4 * 1E-3 = 1E-7/h
|  +--Battery thermal runaway                  P ~= 1E-6/h
|  +--Fuel exhaustion (planning fail)          P ~= 1E-5/h
|  Cut set probability: ~1E-5/h
|
+--OR--> Cyber hijack
|  |
|  +--Crypto key compromise                    P ~= 1E-6/h
|  +--Authentication bypass FCS                P ~= 1E-6/h
|  Cut set probability: ~1E-6/h
|
+--OR--> Mid-air collision aviazione manned
|  |
|  +--ADS-B IN fail + ATC loss + no DAA        P ~= 1E-6/h
|  Cut set probability: ~1E-6/h
|
+--OR--> Severe weather encounter
|  |
|  +--Forecast fail + storm intercept          P ~= 1E-5/h
|  Cut set probability: ~1E-5/h

Total Loss of Vehicle P (top event) ~= 2-3E-5/h
Compliance check vs SAIL III target 1E-5/h:
  MARGINAL — necessita ulteriori mitigation per portare a < 1E-5/h:
  - Reduce FCS DAL-C failure rate (HW redundancy + formal verification)
  - Improve weather forecast integration (NOWCAST + abort)
  - Improve GNSS robustness (Galileo PRS opzionale Phase B)

Single points of failure identificati:
  SPOF-1: Single autopilot DAL-C primary (mitigation: 2oo3 voting)
  SPOF-2: Single parachute system (mitigation: dual pyrotechnic + ballistic backup)
  SPOF-3: Single SATCOM provider Iridium (mitigation: Inmarsat dual-provider Phase B)

Falsifying observation: se FTA dettagliato post-DOA mostra SPOF non mitigato
con P > 1E-4/h, design FCS richiede revisione prima del SAIL III approval.
"""

FTA_LOM = """
TOP EVENT: Loss of Mission EO (Percorso 6A pilota Pentema)
Target: P < 5% per missione (revenue impact)

Loss of Mission EO [TOP, OR-gate]
|
+--OR--> Payload EO failure
|  |
|  +--Camera RGB failure                       P ~= 0.5% / mission
|  +--AND--> RGB fail + IR fail                P ~= 0.05% / mission
|  +--Gimbal stuck off-target                  P ~= 1% / mission
|  +--IR calibration lost critical             P ~= 2% / mission
|  Cut set probability: ~3% / mission
|
+--OR--> Data downlink/storage failure
|  |
|  +--On-board storage corruption              P ~= 0.5% / mission
|  +--Downlink bandwidth congestion            P ~= 1% / mission
|  +--Cloud processing pipeline fail           P ~= 1% / mission
|  Cut set probability: ~2% / mission
|
+--OR--> Operational mission abort
|  |
|  +--Weather degradation in flight            P ~= 5% / mission (high)
|  +--Lost-Link triggering RtB                 P ~= 1% / mission
|  +--ATC restriction in-flight                P ~= 1% / mission
|  Cut set probability: ~7% / mission
|
+--OR--> Quality below SLA
|  |
|  +--GSD insufficient                         P ~= 0.5% / mission
|  +--Cloud cover > 30% area                   P ~= 5% / mission
|  +--Image blur (vibration / motion)          P ~= 1% / mission
|  Cut set probability: ~6% / mission
|
+--OR--> Permitting / regulatory abort
|  |
|  +--NOTAM not issued                         P ~= 0.5% / mission
|  +--Geographical zone restriction last-min   P ~= 0.5% / mission
|  Cut set probability: ~1% / mission

Total Loss of Mission P (top event) ~= 15-20% / mission
Vs target 5%/mission: NON-CONFORME, needs:
  - Weather planning + nowcasting integration (riduzione abort)
  - Buffer mission re-scheduling
  - Cloud cover prediction model
  - SLA realistico con cliente PA (revisione target a 10-15%/mission)

SLA realistico: 80-85% mission success rate (15-20% abort/quality issue)
incluso buffer climatico Liguria + operational margins.

Falsifying observation: se actuals Y1 mostrano < 70% mission success
rate sostenuto, revisione modello operativo + cost structure.
"""

# Early Warning Indicators per top-25
EWI_DATA = [
    ("RSK-TEC-001", "Sim energy balance update mensile", "Monthly + quarterly review", "Sim allegato A.7 deficit > 50% giorni", "propulsion-energy-engineer"),
    ("RSK-TEC-002", "Aeroelastic FEA subscale", "Trimestrale", "Flutter speed < 1.3x Vd", "aero-structures-engineer"),
    ("RSK-TEC-003", "EASA RMT HAPS calendar", "Trimestrale", "RMT non aperto entro Y3", "aviation-regulatory"),
    ("RSK-TEC-004", "Test bench HIL", "Mensile", "Mismatch ICD > 2 critical", "systems-engineer"),
    ("RSK-TEC-005", "GPS interference EASA bulletin", "Settimanale", "Eventi jamming Mar Ligure > 3/mese", "avionics-gnc-engineer"),
    ("RSK-TEC-008", "Battery thermal events sector", "Mensile", "Recall cella o vendor incident", "propulsion-energy-engineer"),
    ("RSK-REG-001", "EASA RMT HAPS published", "Trimestrale", "RMT non in workplan 2026-28", "aviation-regulatory"),
    ("RSK-REG-002", "ENAC SAIL pre-app feedback", "Mensile", "SAIL > III determination", "aviation-regulatory"),
    ("RSK-REG-018", "EUROCONTROL HAPS workplan", "Semestrale", "HAPS procedure absent", "avionics-gnc-engineer"),
    ("RSK-REG-019", "CISO hire + ISMS gap analysis", "Mensile", "CISO non hired entro M+6", "aviation-regulatory + CISO"),
    ("RSK-REG-020", "Settore UAS BVLOS incidents", "Mensile", "Incidente grave settore IT", "ops + safety"),
    ("RSK-REG-021", "AgID PSN qualifica list", "Trimestrale", "Cloud Aruba non qualificato", "data-privacy + IT"),
    ("RSK-REG-025", "Procedura affidamento Regione", "Mensile", "Avvio gara competitor incluso", "snai-funding + legal"),
    ("RSK-REG-027", "ACN classification notice", "Mensile", "Firmamento classificata essenziale", "CISO + legal"),
    ("RSK-REG-030", "ENAV procedure FL400+ pubbl.", "Trimestrale", "Procedure absent al Y3", "avionics + sovereign"),
    ("RSK-FIN-001", "Pipeline Phase B funding", "Mensile", "Commitment < 30% al gate G5", "financial-cfo"),
    ("RSK-FIN-004", "FTE hire pipeline regulatory", "Settimanale", "3 ruoli senior non riempiti M+9", "HR + CFO"),
    ("RSK-MKT-001", "Contratti Regione Liguria pipeline", "Mensile", "Tempo cycle > 18 mesi", "snai-funding"),
    ("RSK-MKT-002", "AALTO-Leonardo JV announcements", "Mensile", "Annuncio JV o pilota multi-regione", "competitive-intelligence"),
    ("RSK-MKT-005", "Cambio amministrazione Regione", "Annuale + eventi", "Elezioni 2025+", "sovereign-strategist + business"),
    ("RSK-OPS-001", "Giorni op. inverno Pentema", "Mensile", "< 60% giorni op. piano inverno", "vtol-uas-specialist"),
    ("RSK-OPS-002", "Near-miss + incident reports", "Settimanale", "Near-miss vs aviazione manned", "ops + safety"),
    ("RSK-SUP-001", "USA BIS notices + tariffe USA-CN", "Settimanale", "Restrizioni HAPS-related", "supply-chain + sovereign"),
    ("RSK-SEC-001", "SIEM alerts + ACN threat intel", "Settimanale", "Intrusion detected o ACN sector alert", "CISO"),
    ("RSK-HR-002", "Hire pipeline CISO+DPO+RegAff", "Settimanale", "Tempo hire > 6 mesi", "HR + CEO"),
    ("RSK-REP-001", "Settore drone incidents IT/EU", "Mensile", "Incidente coverage stampa nazionale", "CEO + comms"),
]

CHANGE_LOG = [
    ("v0.1", "2026-03-15", "Stesura iniziale top-10 Cap. 6.4", "risk-register-builder skill", "Top-10 + FMECA Payload preliminari"),
    ("v0.2", "2026-04-10", "Aggiunta 5 showstopper Cap. 10 §10.2", "risk-register-builder skill", "RSK-FIN-001 + ridondanza RSK-TEC-003 vs RSK-REG-001"),
    ("v0.3", "2026-05-01", "Aggiunta 5 RSK-GEO da RESERVED (reference, no detail)", "sovereign-strategist", "Geopolitical risks formalmente trackati"),
    ("v0.4", "2026-05-10", "Aggiunta 15 RSK-REG-016..030 da Cap. 5 §5.16 audit Regulatory Adversary", "regulatory-adversary audit", "Showstopper regolatori non coperti pre-audit"),
    ("v1.0", "2026-05-17", "Consolidamento Allegato A.2 + FMECA + FTA + EWI + Mitigation Plan", "senior risk manager + safety engineer", "Risk Register completo v1.0 per gate M+3"),
]

# ---------------------------------------------------------------------------
# WRITERS — 22 fogli Excel
# ---------------------------------------------------------------------------
RISK_COLS = [
    "ID", "Categoria", "Descrizione", "Trigger", "P (1-5)", "I (1-5)",
    "Score (PxI)", "Color", "Status", "Owner", "Response",
    "Mitigation actions", "Residual P", "Residual I", "Residual Score",
    "Fase critica", "Confidence", "Last review", "EWI",
]
RISK_WIDTHS = [13, 14, 50, 38, 8, 8, 9, 9, 16, 28, 14, 55, 9, 9, 11, 18, 12, 13, 38]


def color_label(score: int) -> str:
    if score >= 15:
        return "RED"
    if score >= 8:
        return "YELLOW"
    return "GREEN"


def risk_row_values(r: dict) -> list:
    return [
        r["id"], r["cat"], r["desc"], r["trigger"], r["P"], r["I"],
        r["score"], color_label(r["score"]), r["status"], r["owner"],
        r["response"], r["mitigation"], r["rP"], r["rI"], r["rscore"],
        r["phase"], r["confidence"], TODAY, r["ewi"],
    ]


def sheet_cover(wb):
    ws = wb.create_sheet("Cover", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90

    def block(row, title):
        c = ws.cell(row=row, column=1, value=title)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    def kv(row, k, v):
        a = ws.cell(row=row, column=1, value=k)
        b = ws.cell(row=row, column=2, value=v)
        a.font = FONT_BODY_BOLD
        b.font = FONT_BODY
        a.alignment = WRAP
        b.alignment = WRAP

    block(1, "Allegato A.2 - Risk Register v1.0 - Firmamento Technologies HALE/VTOL")
    kv(2, "Documento", "Allegato A.2 - Risk Register completo v1.0")
    kv(3, "Volume", "Volume 2 - Allegati tecnici dello Studio di Fattibilita")
    kv(4, "Studio di riferimento", "Studio di Fattibilita Piattaforma Aerea HALE/VTOL per Aree Interne")
    kv(5, "Bando", "Cooding Prototypes - Coopfond/Legacoop")
    kv(6, "Soggetto proponente", "Firmamento Technologies srl coop")
    kv(7, "Versione", VERSION)
    kv(8, "Data", TODAY)
    kv(9, "Stato", "Bozza M+3 consolidata")

    block(11, "Metodologia")
    kv(12, "Framework primario", "NASA NPR 8000.4A - Continuous Risk Management (CRM)")
    kv(13, "Failure analysis", "FMECA - MIL-STD-1629A / IEC 60812")
    kv(14, "Safety assessment", "ARP4761 - Aerospace Recommended Practice")
    kv(15, "Fault Tree Analysis", "ARP4761 + NUREG-0492 + IEC 61025")
    kv(16, "ISO standard", "ISO 31000:2018 - Risk Management Principles and Guidelines")
    kv(17, "Conformita italiana", "D.Lgs. 36/2023 art. 41 (analisi di rischio ingegneristico)")
    kv(18, "Conformita aviation", "EASA SORA 2.5 (ED Decision 2025/018/R) + Part-IS (Reg.UE 2023/203)")
    kv(19, "Conformita cyber", "NIS2 (D.Lgs. 138/2024) + ISO/IEC 27001 + Part-IS")

    block(21, "Sistema di scoring")
    kv(22, "Probabilita P (1-5)", "1=VeryLow (<5%) | 2=Low (5-20%) | 3=Medium (20-50%) | 4=High (50-80%) | 5=VeryHigh (>80%)")
    kv(23, "Impatto I (1-5)", "1=Negligible | 2=Minor | 3=Moderate | 4=Major | 5=Severe (showstopper/catastrofe)")
    kv(24, "Risk Score", "P x I (range 1-25)")
    kv(25, "Color coding", "GREEN 1-7 (accettabile, monitor) | YELLOW 8-14 (mitigation richiesta) | RED 15-25 (showstopper, response immediata)")
    kv(26, "Response options", "Avoid | Mitigate | Transfer | Accept (NASA + ISO 31000)")
    kv(27, "Residual risk", "P x I post-mitigation; richiesto <= GREEN o YELLOW per closure")

    block(29, "Struttura del Risk Register")
    kv(30, "Totale rischi tracciati", str(len(RISKS)))
    kv(31, "Categorie", "TEC (tecnici) | REG (regolatori) | FIN (finanziari) | MKT (mercato) | OPS (operativi) | SUP (supply chain) | PRV (privacy/legale) | SEC (cybersecurity) | HR (risorse umane) | REP (reputazionali) | GEO (geopolitici - RESERVED)")
    kv(32, "Fogli del workbook", "22 sheet (Cover, Top-25, 11 categorie, 3 FMECA, 2 FTA, Mitigation Plan, Residual Matrix, EWI, Audit Trail)")
    kv(33, "Boundary conditions", "B1: service-only + cooperative Legacoop | B2: EU sovereign stratospheric layer / complementare IRIS2")

    block(35, "Documenti correlati (cross-reference)")
    kv(36, "Cap. 6.4", "Top-10 rischi tecnici + FMECA Payload + FTA preliminare")
    kv(37, "Cap. 5.16", "15 showstopper regolatori RSK-REG-016..030")
    kv(38, "Cap. 10.2", "Risk residuo aggregato + 5 showstopper formali")
    kv(39, "RESERVED-rischi-geopolitici.md", "5 RSK-GEO-001..005 (accesso ristretto - NO contenuto sensibile in questo file pubblico)")
    kv(40, "Skill .claude/skills/risk-register-builder/SKILL.md", "Metodologia operativa")
    kv(41, "AUDIT-REDTEAM-VOLUME-1.md", "Audit Red Team M+3 + falsifying observations")
    kv(42, "AUDIT-REGULATORY-VOLUME-1.md", "Audit Regulatory Adversary M+3 (15 showstopper REG)")
    kv(43, "AUDIT-COMPETITOR-VOLUME-1.md", "Audit Competitor Intelligence M+3")

    block(45, "Disciplina epistemica")
    kv(46, "Confidence levels", "high / medium-high / medium / medium-low / low per ogni claim P e I")
    kv(47, "Falsifying observations", "Documentate per top-10 rischi - condizioni che falsificherebbero stima rischio")
    kv(48, "Source provenance", "Tutti i rischi tracciati a Cap. 5/6/10 dello Studio o audit M+3 dedicati")
    kv(49, "Re-assessment", "Trimestrale + dopo ogni gate review + trigger eventi esterni")

    block(51, "Owner principale")
    kv(52, "Risk Manager", "TBD (CISO + Head of Regulatory Affairs joint fino assunzione)")
    kv(53, "Approval gate review", "Steering Committee Firmamento + Coopfond observer + Legacoop")
    kv(54, "Esterno (gate G4 M+12)", "Auditor esterno terzo + advisory aerospace senior")


def sheet_top_25(wb):
    ws = wb.create_sheet("Top-25 Risks")
    write_header(ws, RISK_COLS, widths=RISK_WIDTHS)
    sorted_risks = sorted(RISKS, key=lambda r: (-r["score"], -r["rscore"], r["id"]))
    top25 = sorted_risks[:25]
    for i, r in enumerate(top25, 2):
        write_row(ws, i, risk_row_values(r), score_col=7)
        ws.cell(row=i, column=15).fill = color_for_score(r["rscore"])
        ws.cell(row=i, column=15).font = font_for_score(r["rscore"])
        ws.cell(row=i, column=9).font = Font(name="Calibri", size=10, bold=True, color=status_color(r["status"]))


def sheet_category(wb, name, cat_filter):
    ws = wb.create_sheet(name)
    write_header(ws, RISK_COLS, widths=RISK_WIDTHS)
    risks_cat = [r for r in RISKS if cat_filter(r)]
    risks_cat.sort(key=lambda r: (-r["score"], r["id"]))
    for i, r in enumerate(risks_cat, 2):
        write_row(ws, i, risk_row_values(r), score_col=7)
        ws.cell(row=i, column=15).fill = color_for_score(r["rscore"])
        ws.cell(row=i, column=15).font = font_for_score(r["rscore"])


def sheet_fmeca(wb, name, data, subsystem_title):
    ws = wb.create_sheet(name)
    headers = ["Item", "Failure Mode", "Cause", "Local Effect", "System Effect",
               "Severity (1-5)", "Frequency (1-5)", "Detection (1-5)", "RPN (SxFxD)", "Mitigation"]
    widths = [32, 26, 30, 26, 30, 9, 9, 9, 9, 50]
    # Title row
    ws.cell(row=1, column=1, value=f"FMECA - {subsystem_title}").font = Font(name="Calibri", size=13, bold=True, color="1F3864")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=2, column=1, value="Metodologia: MIL-STD-1629A + IEC 60812 + ARP4761").font = FONT_SUB
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    write_header(ws, headers, row=3, widths=widths)
    for i, row in enumerate(data, 4):
        s, f, d = row[5], row[6], row[7]
        rpn = s * f * d
        vals = list(row[:8]) + [rpn, row[8]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.alignment = WRAP
            cell.font = FONT_BODY
            cell.border = BORDER
        # color RPN
        rpn_cell = ws.cell(row=i, column=9)
        if rpn >= 40:
            rpn_cell.fill = FILL_RED
            rpn_cell.font = FONT_WHITE_BOLD
        elif rpn >= 20:
            rpn_cell.fill = FILL_YELLOW
        elif rpn >= 10:
            rpn_cell.fill = PatternFill("solid", fgColor="FFF2CC")
        else:
            rpn_cell.fill = FILL_GREEN
        rpn_cell.alignment = CENTER
    # legend
    leg_row = len(data) + 5
    ws.cell(row=leg_row, column=1, value="Legenda RPN").font = FONT_SUB
    ws.cell(row=leg_row + 1, column=1, value="RPN >= 40: mitigation obbligatoria (RED)")
    ws.cell(row=leg_row + 2, column=1, value="RPN 20-39: mitigation raccomandata (YELLOW)")
    ws.cell(row=leg_row + 3, column=1, value="RPN < 20: monitor (GREEN)")


def sheet_fta(wb, name, content, title):
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 110
    ws.cell(row=1, column=1, value=title).font = Font(name="Calibri", size=13, bold=True, color="1F3864")
    ws.cell(row=2, column=1, value="Metodologia: ARP4761 + NUREG-0492 + IEC 61025").font = FONT_SUB
    ws.cell(row=3, column=1, value="Notazione: AND-gate richiede tutti gli eventi; OR-gate richiede almeno uno; cut set = combinazione minima eventi che causa top event").font = FONT_BODY
    ws.cell(row=3, column=1).alignment = WRAP
    ws.row_dimensions[3].height = 40
    for i, line in enumerate(content.strip().split("\n"), 5):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(name="Consolas", size=10) if line.startswith("|") or line.startswith("+--") else FONT_BODY
        c.alignment = Alignment(wrap_text=False, vertical="top", horizontal="left")


def sheet_mitigation_plan(wb):
    ws = wb.create_sheet("Mitigation_Plan")
    headers = ["ID", "Status", "Score current", "Score residual", "Owner",
               "Action ID", "Action description", "Deadline", "Phase", "Acceptance criterion"]
    widths = [13, 16, 12, 12, 28, 11, 60, 16, 18, 50]
    write_header(ws, headers, widths=widths)
    row = 2
    # Per top-25 risks: una azione composta (estratta dal campo mitigation)
    sorted_risks = sorted(RISKS, key=lambda r: (-r["score"], r["id"]))
    for r in sorted_risks:
        actions = [a.strip() for a in r["mitigation"].replace(";", "|").split("|") if a.strip()]
        for j, a in enumerate(actions, 1):
            action_id = f"{r['id']}-A{j:02d}"
            deadline = {
                "Y0+": "M+0-3",
                "Y0+ immediato": "M+0-1",
                "Y0+ urgente": "M+0-3",
                "Y1": "M+3-12",
                "Y1+": "M+0-12 + ongoing",
                "Y1+ continuous": "ongoing",
                "Y1+ (M+0 → M+9)": "M+0-9",
                "Y1+ (M+0 → M+12) urgente": "M+0-12",
            }.get(r["phase"], "TBD per phase")
            criterion = "Mitigation deployed + verified by gate review next"
            vals = [r["id"], r["status"], r["score"], r["rscore"], r["owner"],
                    action_id, a, deadline, r["phase"], criterion]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.alignment = WRAP
                cell.font = FONT_BODY
                cell.border = BORDER
            ws.cell(row=row, column=3).fill = color_for_score(r["score"])
            ws.cell(row=row, column=4).fill = color_for_score(r["rscore"])
            ws.cell(row=row, column=3).alignment = CENTER
            ws.cell(row=row, column=4).alignment = CENTER
            row += 1


def sheet_residual_matrix(wb):
    ws = wb.create_sheet("Residual_Risk_Matrix")
    ws.cell(row=1, column=1, value="P x I Residual Risk Matrix (post-mitigation)").font = Font(name="Calibri", size=13, bold=True, color="1F3864")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=2, column=1, value="Rows = Impact (5 top - 1 bottom)  |  Cols = Probability (1 left - 5 right)").font = FONT_SUB
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    # Matrix 5x5
    header_row = 4
    ws.cell(row=header_row, column=1, value="I \\ P").fill = FILL_HEADER
    ws.cell(row=header_row, column=1).font = FONT_HEADER
    ws.cell(row=header_row, column=1).alignment = CENTER
    for p in range(1, 6):
        c = ws.cell(row=header_row, column=1 + p, value=f"P={p}")
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = CENTER

    # build matrix counts (residual)
    matrix = {(p, i): [] for p in range(1, 6) for i in range(1, 6)}
    for r in RISKS:
        matrix[(r["rP"], r["rI"])].append(r["id"])

    for i in range(5, 0, -1):
        row = header_row + (6 - i)
        c = ws.cell(row=row, column=1, value=f"I={i}")
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = CENTER
        for p in range(1, 6):
            score = p * i
            ids = matrix[(p, i)]
            val = f"{len(ids)}\n" + ("\n".join(ids[:6]) if ids else "-")
            cell = ws.cell(row=row, column=1 + p, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
            cell.font = FONT_BODY
            cell.border = BORDER
            cell.fill = color_for_score(score)
            if score >= 15:
                cell.font = FONT_WHITE_BOLD

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 22
    for r in range(header_row + 1, header_row + 6):
        ws.row_dimensions[r].height = 90

    # Summary block
    sumrow = header_row + 8
    ws.cell(row=sumrow, column=1, value="Statistiche residual risk").font = FONT_SUB
    red = sum(1 for r in RISKS if r["rscore"] >= 15)
    yellow = sum(1 for r in RISKS if 8 <= r["rscore"] < 15)
    green = sum(1 for r in RISKS if r["rscore"] < 8)
    ws.cell(row=sumrow + 1, column=1, value=f"RED (>=15) residual: {red}")
    ws.cell(row=sumrow + 2, column=1, value=f"YELLOW (8-14) residual: {yellow}")
    ws.cell(row=sumrow + 3, column=1, value=f"GREEN (<8) residual: {green}")
    ws.cell(row=sumrow + 4, column=1, value=f"Totale: {len(RISKS)}")
    # baseline pre
    red0 = sum(1 for r in RISKS if r["score"] >= 15)
    yellow0 = sum(1 for r in RISKS if 8 <= r["score"] < 15)
    green0 = sum(1 for r in RISKS if r["score"] < 8)
    ws.cell(row=sumrow + 6, column=1, value="Statistiche pre-mitigation (baseline)").font = FONT_SUB
    ws.cell(row=sumrow + 7, column=1, value=f"RED (>=15) baseline: {red0}")
    ws.cell(row=sumrow + 8, column=1, value=f"YELLOW (8-14) baseline: {yellow0}")
    ws.cell(row=sumrow + 9, column=1, value=f"GREEN (<8) baseline: {green0}")
    ws.cell(row=sumrow + 11, column=1, value=f"Reduzione RED: {red0} -> {red} ({(red0-red)/max(red0,1)*100:.0f}% reduction)")
    ws.cell(row=sumrow + 12, column=1, value=f"Mitigation effectiveness: {((red0-red)+(yellow0-yellow))/max(red0+yellow0,1)*100:.0f}% of RED+YELLOW reduced")


def sheet_ewi(wb):
    ws = wb.create_sheet("EWI")
    headers = ["RSK-ID", "EWI Indicatore", "Frequenza monitoring", "Threshold trigger", "Owner"]
    widths = [13, 50, 24, 50, 30]
    ws.cell(row=1, column=1, value="Early Warning Indicators - Top-25 Risks").font = Font(name="Calibri", size=13, bold=True, color="1F3864")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.cell(row=2, column=1, value="Monitoraggio quarterly (default); rischi RED monthly. Owner = funzione responsabile.").font = FONT_SUB
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    write_header(ws, headers, row=4, widths=widths)
    for i, (rid, ind, freq, thr, own) in enumerate(EWI_DATA, 5):
        vals = [rid, ind, freq, thr, own]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.alignment = WRAP
            cell.font = FONT_BODY
            cell.border = BORDER


def sheet_audit_trail(wb):
    ws = wb.create_sheet("Audit_Trail")
    headers = ["Versione", "Data", "Modifica", "Autore", "Note"]
    widths = [12, 14, 55, 35, 55]
    ws.cell(row=1, column=1, value="Audit Trail - Risk Register versioning").font = Font(name="Calibri", size=13, bold=True, color="1F3864")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    write_header(ws, headers, row=3, widths=widths)
    for i, (ver, date_, mod, auth, note) in enumerate(CHANGE_LOG, 4):
        vals = [ver, date_, mod, auth, note]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.alignment = WRAP
            cell.font = FONT_BODY
            cell.border = BORDER
    # next review
    last_row = 4 + len(CHANGE_LOG) + 2
    ws.cell(row=last_row, column=1, value="Prossimo aggiornamento previsto").font = FONT_SUB
    ws.cell(row=last_row + 1, column=1, value="v1.1")
    ws.cell(row=last_row + 1, column=2, value="2026-08-17 (M+6)")
    ws.cell(row=last_row + 1, column=3, value="Post gate G3 (M+10) review + chiusura RSK-FIN-001/REG-001 update")
    ws.cell(row=last_row + 1, column=4, value="risk-register-builder + steering")
    ws.cell(row=last_row + 1, column=5, value="Trigger: gate G3 outcomes + EWI quarterly review")


def sheet_geo(wb):
    """Foglio RSK-GEO con reference RESERVED (nessun contenuto sensibile)."""
    ws = wb.create_sheet("RSK-GEO")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 80
    c = ws.cell(row=1, column=1, value="RSK-GEO - Rischi Geopolitici (RESERVED - reference only)")
    c.font = Font(name="Calibri", size=13, bold=True, color="C00000")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    ws.cell(row=3, column=1, value="ATTENZIONE").font = FONT_HEADER
    ws.cell(row=3, column=1).fill = FILL_RED
    ws.cell(row=3, column=2, value="Questo foglio riassume i 5 RSK-GEO solo come reference. Il dettaglio operativo, mitigation e trigger osservabili sono CLASSIFICATI in riferimenti/RESERVED-rischi-geopolitici.md - accesso ristretto founder team + consulenti NDA + stakeholder istituzionali IT su richiesta motivata.")
    ws.cell(row=3, column=2).alignment = WRAP
    ws.cell(row=3, column=2).font = FONT_BODY
    ws.row_dimensions[3].height = 70

    headers = ["RSK-ID", "Categoria sintetica", "P (cond.)", "I (cond.)", "Score (cond.)", "Fase critica", "Owner"]
    widths = [13, 50, 12, 12, 14, 16, 35]
    write_header(ws, headers, row=5, widths=widths)
    rows = [
        ("RSK-GEO-001", "Frizione internazionale - posizione narrativa stratosferica", 3, 4, 12, "Y4+", "sovereign-strategist + CEO"),
        ("RSK-GEO-002", "Classificazione strategica nazionale - governance obblighi", 2, 4, 8, "Y3+", "sovereign-strategist + legal + CEO"),
        ("RSK-GEO-003", "Dipendenza supply chain non-EU", 3, 4, 12, "Y2+", "sovereign-strategist + supply-chain"),
        ("RSK-GEO-004", "Misalignment con architettura sovrana EU multi-orbita", 3, 4, 12, "Y2-Y4", "sovereign-strategist + CEO"),
        ("RSK-GEO-005", "Acquisizione difensiva da incumbent settore", 3, 4, 12, "Y3-Y4", "CEO + Board"),
    ]
    for i, row in enumerate(rows, 6):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.alignment = WRAP
            cell.font = FONT_BODY
            cell.border = BORDER
        ws.cell(row=i, column=5).fill = color_for_score(row[4])
        ws.cell(row=i, column=5).alignment = CENTER

    ws.cell(row=12, column=1, value="Documento dettaglio").font = FONT_SUB
    ws.cell(row=12, column=2, value="/home/user/HALE/riferimenti/RESERVED-rischi-geopolitici.md (NON pubblicabile)")
    ws.cell(row=13, column=1, value="Mitigation strategy").font = FONT_SUB
    ws.cell(row=13, column=2, value="Vedi documento RESERVED + Engagement Plan strategico §5 + tavoli istituzionali presidiati (MIMIT, DPE, ACN, ENAC, ESA/DG CNECT, DG DEFIS, NATO DIANA, ASD-Eurospace, ASI/CIRA)")
    ws.cell(row=14, column=1, value="Re-assessment").font = FONT_SUB
    ws.cell(row=14, column=2, value="Trimestrale - owner CEO + sovereign-infrastructure-strategist; review esterna annuale (advisor ex-MAE/MIMIT/Leonardo)")

# ---------------------------------------------------------------------------
# CSV writer (full register)
# ---------------------------------------------------------------------------
def write_csv():
    fields = [
        "id", "categoria", "descrizione", "trigger", "P", "I", "score", "color",
        "status", "owner", "response", "mitigation", "residual_P", "residual_I",
        "residual_score", "phase", "confidence", "last_review", "ewi",
        "falsifying_observation", "top10_flag",
    ]
    with open(CSV_PATH, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, quoting=csv.QUOTE_ALL)
        w.writeheader()
        for r in sorted(RISKS, key=lambda x: (-x["score"], x["id"])):
            w.writerow({
                "id": r["id"], "categoria": r["cat"], "descrizione": r["desc"],
                "trigger": r["trigger"], "P": r["P"], "I": r["I"],
                "score": r["score"], "color": color_label(r["score"]),
                "status": r["status"], "owner": r["owner"], "response": r["response"],
                "mitigation": r["mitigation"], "residual_P": r["rP"],
                "residual_I": r["rI"], "residual_score": r["rscore"],
                "phase": r["phase"], "confidence": r["confidence"],
                "last_review": TODAY, "ewi": r["ewi"],
                "falsifying_observation": r["fo"] or "",
                "top10_flag": "Y" if r.get("top") else "N",
            })


# ---------------------------------------------------------------------------
# Markdown report
# ---------------------------------------------------------------------------
def write_md_report():
    by_cat: dict[str, list] = {}
    for r in RISKS:
        by_cat.setdefault(r["cat"], []).append(r)

    red = [r for r in RISKS if r["score"] >= 15]
    yellow = [r for r in RISKS if 8 <= r["score"] < 15]
    green = [r for r in RISKS if r["score"] < 8]
    showstoppers = [r for r in RISKS if r["status"] == "Showstopper"]
    open_critical = [r for r in RISKS if r["status"] == "Open-Critical"]
    open_high = [r for r in RISKS if r["status"] == "Open-High"]
    monitor = [r for r in RISKS if r["status"] == "Monitor"]

    red_r = sum(1 for r in RISKS if r["rscore"] >= 15)
    yellow_r = sum(1 for r in RISKS if 8 <= r["rscore"] < 15)
    green_r = sum(1 for r in RISKS if r["rscore"] < 8)

    top25 = sorted(RISKS, key=lambda r: (-r["score"], -r["rscore"], r["id"]))[:25]

    md = []
    md.append(f"# Allegato A.2 - Risk Register Report v{VERSION[1:]}\n")
    md.append(f"> **Volume 2 - Allegati tecnici - Studio di Fattibilita Piattaforma Aerea HALE/VTOL**  ")
    md.append(f"> Firmamento Technologies srl coop - bando Cooding Prototypes (Coopfond / Legacoop)  ")
    md.append(f"> Versione: **{VERSION}** - Data: **{TODAY}** - Stato: **Bozza M+3 consolidata**  ")
    md.append(f"> Metodologia: NASA NPR 8000.4A + FMECA (MIL-STD-1629A) + FTA (ARP4761) + ISO 31000:2018  ")
    md.append(f"> Conformita: D.Lgs. 36/2023 art. 41 + EASA SORA 2.5 + Part-IS + NIS2\n")
    md.append("---\n")

    md.append("## 1. Metodologia\n")
    md.append("Il Risk Register Firmamento HALE/VTOL e' costruito secondo **NASA NPR 8000.4A - Continuous Risk Management (CRM)**, con integrazione di:\n")
    md.append("- **FMECA** (MIL-STD-1629A / IEC 60812) per analisi guasto sottosistema a livello item (Payload EO, Avionica, Propulsione)")
    md.append("- **FTA** (ARP4761 / NUREG-0492 / IEC 61025) per top events critici (Loss of Vehicle BVLOS, Loss of Mission EO)")
    md.append("- **ISO 31000:2018** per principi di risk management end-to-end (identificazione, analisi, valutazione, trattamento, monitoring, communication)")
    md.append("- Compliance specifica aviation: **EASA SORA 2.5** (ED Decision 2025/018/R) + **Part-IS** (Reg.UE 2023/203)")
    md.append("- Compliance cyber: **NIS2** (D.Lgs. 138/2024) + ISO/IEC 27001 + Part-IS\n")

    md.append("### 1.1 Sistema di scoring P x I\n")
    md.append("| P (Probabilita) | Descrizione | Range qualitativo |")
    md.append("|---|---|---|")
    md.append("| 1 Very Low | Improbabile | < 5% |")
    md.append("| 2 Low | Possibile ma raro | 5-20% |")
    md.append("| 3 Medium | Possibile | 20-50% |")
    md.append("| 4 High | Probabile | 50-80% |")
    md.append("| 5 Very High | Quasi certo | > 80% |\n")
    md.append("| I (Impatto) | Tecnico | Schedule | Costo | Safety | Reputational |")
    md.append("|---|---|---|---|---|---|")
    md.append("| 1 Negligible | Aggiornamento doc | < 1 sett. | < 5k EUR | Nessuno | Nessuno |")
    md.append("| 2 Minor | Modifica subsystem | 1-4 sett. | 5-50k EUR | Incidente minore | Locale |")
    md.append("| 3 Moderate | Re-design subsystem | 1-3 mesi | 50-200k EUR | Ferite leggere | Regionale |")
    md.append("| 4 Major | Re-design system | 3-12 mesi | 200k EUR - 1M EUR | Ferite gravi | Nazionale |")
    md.append("| 5 Severe | Showstopper / catastrofe | > 12 mesi | > 1M EUR | Decesso / danni terzi | Internazionale |\n")
    md.append("**Color coding** (NASA + ISO 31000):")
    md.append("- **GREEN (1-7)**: rischio accettabile, monitoring")
    md.append("- **YELLOW (8-14)**: mitigation richiesta")
    md.append("- **RED (15-25)**: showstopper, response immediata + Hold gate\n")

    md.append("### 1.2 Response options\n")
    md.append("| Response | Quando usarla |")
    md.append("|---|---|")
    md.append("| **Avoid** | Eliminare la causa (cambio architettura, eliminazione SPOF) |")
    md.append("| **Mitigate** | Ridurre P e/o I (design margin, ridondanza, test) |")
    md.append("| **Transfer** | Spostare il rischio (assicurazione, vendor contract, partnership) |")
    md.append("| **Accept** | Tollerare con monitoring (costo mitigation > exposure) |\n")
    md.append("---\n")

    md.append("## 2. Statistiche aggregate\n")
    md.append(f"**Totale rischi formalmente tracciati**: {len(RISKS)}\n")
    md.append("### 2.1 Per categoria")
    md.append("| Categoria | N rischi | Showstopper | RED | YELLOW | GREEN |")
    md.append("|---|---:|---:|---:|---:|---:|")
    for cat in sorted(by_cat.keys()):
        rl = by_cat[cat]
        n_ss = sum(1 for r in rl if r["status"] == "Showstopper")
        n_r = sum(1 for r in rl if r["score"] >= 15)
        n_y = sum(1 for r in rl if 8 <= r["score"] < 15)
        n_g = sum(1 for r in rl if r["score"] < 8)
        md.append(f"| {cat} | {len(rl)} | {n_ss} | {n_r} | {n_y} | {n_g} |")
    md.append(f"| **TOTALE** | **{len(RISKS)}** | **{len(showstoppers)}** | **{len(red)}** | **{len(yellow)}** | **{len(green)}** |\n")

    md.append("### 2.2 Per status\n")
    md.append("| Status | N rischi |")
    md.append("|---|---:|")
    md.append(f"| Showstopper | {len(showstoppers)} |")
    md.append(f"| Open-Critical | {len(open_critical)} |")
    md.append(f"| Open-High | {len(open_high)} |")
    md.append(f"| Monitor | {len(monitor)} |")
    md.append(f"| Open-Medium | {sum(1 for r in RISKS if r['status'] == 'Open-Medium')} |")
    md.append("")

    md.append("### 2.3 P x I matrix (baseline pre-mitigation)\n")
    matrix = {(p, i): 0 for p in range(1, 6) for i in range(1, 6)}
    for r in RISKS:
        matrix[(r["P"], r["I"])] += 1
    md.append("| I \\ P | P=1 | P=2 | P=3 | P=4 | P=5 |")
    md.append("|---|---|---|---|---|---|")
    for i in range(5, 0, -1):
        line = f"| I={i} |"
        for p in range(1, 6):
            score = p * i
            tag = "RED" if score >= 15 else "YEL" if score >= 8 else "GRN"
            line += f" {matrix[(p, i)]} ({tag}) |"
        md.append(line)
    md.append("")

    md.append("### 2.4 Mitigation effectiveness\n")
    md.append(f"- **Pre-mitigation**: RED={len(red)}, YELLOW={len(yellow)}, GREEN={len(green)}")
    md.append(f"- **Post-mitigation (residual)**: RED={red_r}, YELLOW={yellow_r}, GREEN={green_r}")
    md.append(f"- **RED reduction**: {len(red)} -> {red_r} ({(len(red)-red_r)/max(len(red),1)*100:.0f}%)")
    md.append(f"- **YELLOW reduction**: {len(yellow)} -> {yellow_r} ({(len(yellow)-yellow_r)/max(len(yellow),1)*100:.0f}%)\n")
    md.append("---\n")

    md.append("## 3. Top-25 rischi narrati\n")
    md.append("Ordinati per Score baseline (P x I) decrescente. Per ciascuno: descrizione, impatto sul progetto, mitigation status.\n")
    for i, r in enumerate(top25, 1):
        md.append(f"### {i}. {r['id']} - {r['cat']} - Score {r['score']} -> residual {r['rscore']} ({color_label(r['score'])})\n")
        md.append(f"**Descrizione**: {r['desc']}  ")
        md.append(f"**Trigger**: {r['trigger']}  ")
        md.append(f"**Owner**: {r['owner']}  ")
        md.append(f"**Status**: {r['status']}  ")
        md.append(f"**Response**: {r['response']}  ")
        md.append(f"**Mitigation**: {r['mitigation']}  ")
        md.append(f"**Residual P x I**: {r['rP']} x {r['rI']} = {r['rscore']}  ")
        md.append(f"**Fase critica**: {r['phase']}  ")
        md.append(f"**Confidence**: {r['confidence']}  ")
        md.append(f"**EWI**: {r['ewi']}  ")
        if r["fo"]:
            md.append(f"**Falsifying observation**: {r['fo']}  ")
        md.append("")

    md.append("---\n")
    md.append("## 4. Showstopper formali (5+5)\n")
    md.append("### 4.1 Showstopper originali Cap. 6.4 + Cap. 10.2 (5)\n")
    md.append("| ID | Rischio | Score | Percorso | Mitigation status |")
    md.append("|---|---|---:|---|---|")
    original_ss = ["RSK-TEC-001", "RSK-TEC-002", "RSK-TEC-003", "RSK-REG-001", "RSK-FIN-001"]
    for sid in original_ss:
        r = next(x for x in RISKS if x["id"] == sid)
        md.append(f"| **{r['id']}** | {r['desc'][:80]}... | {r['score']} | 6B | {r['response']}: {r['mitigation'][:80]}... |")
    md.append("")

    md.append("### 4.2 Showstopper critici aggiuntivi §5.16 (5)\n")
    md.append("Identificati dall'audit `regulatory-adversary` M+3, formalizzati in Cap. 5 §5.16. Score 15-20.\n")
    md.append("| ID | Rischio | Score | Owner | Deadline mitigation |")
    md.append("|---|---|---:|---|---|")
    additional_ss = ["RSK-REG-019", "RSK-REG-021", "RSK-REG-025", "RSK-REG-027", "RSK-REG-030", "RSK-REG-018"]
    for sid in additional_ss:
        r = next(x for x in RISKS if x["id"] == sid)
        md.append(f"| **{r['id']}** | {r['desc'][:80]}... | {r['score']} | {r['owner']} | {r['phase']} |")
    md.append("")

    md.append("### 4.3 Implicazione per il verdetto Cap. 10\n")
    md.append("Il verdetto Cap. 10 \"Go Condizionato 6A\" presuppone:")
    md.append("- Tutti i 5 RSK-REG critical aggiuntivi mitigated entro M+9-12 (Part-IS, AgID, NIS2 sono **urgenti** M+0-3)")
    md.append("- 3 FTE senior (CISO, DPO, Head of Regulatory Affairs) hired entro M+6-9 (RSK-HR-002)")
    md.append("- OpEx Y1 aggiornato con +450-800k EUR (RSK-FIN-004)")
    md.append("- RSK-FIN-001 (funding Phase B) tracciato come precondizione gate G5\n")
    md.append("Scenario realistico (post Red Team M+3): 60-80% percorsi sono **Hold con piano** vs **Go pieno** al M+10/M+11.\n")
    md.append("---\n")

    md.append("## 5. FMECA results - sintesi\n")
    md.append("Vedi fogli XLSX: `FMECA_Payload`, `FMECA_Avionica`, `FMECA_Propulsione`.\n")
    md.append("### 5.1 Payload EO\n")
    rpns_payload = sorted([(r[0], r[1], r[5] * r[6] * r[7], r[8]) for r in FMECA_PAYLOAD], key=lambda x: -x[2])
    md.append("**Top item RPN**:")
    for item, fm, rpn, mit in rpns_payload[:5]:
        md.append(f"- {item} - {fm} - RPN **{rpn}** - mitigation: {mit}")
    md.append("\n**Mitigation obbligatoria** (RPN >= 40): IR sensor calibrazione persa (RPN 48). NUC frequente + crosscheck con RGB + ground truth.\n")

    md.append("### 5.2 Avionica\n")
    rpns_avi = sorted([(r[0], r[1], r[5] * r[6] * r[7], r[8]) for r in FMECA_AVIONICA], key=lambda x: -x[2])
    md.append("**Top item RPN**:")
    for item, fm, rpn, mit in rpns_avi[:5]:
        md.append(f"- {item} - {fm} - RPN **{rpn}** - mitigation: {mit}")
    md.append("")

    md.append("### 5.3 Propulsione\n")
    rpns_prop = sorted([(r[0], r[1], r[5] * r[6] * r[7], r[8]) for r in FMECA_PROP], key=lambda x: -x[2])
    md.append("**Top item RPN**:")
    for item, fm, rpn, mit in rpns_prop[:5]:
        md.append(f"- {item} - {fm} - RPN **{rpn}** - mitigation: {mit}")
    md.append("")

    md.append("---\n")
    md.append("## 6. FTA results - sintesi\n")
    md.append("### 6.1 Top event: Loss of Vehicle in BVLOS (Percorso 6A)\n")
    md.append("**Target SAIL III SORA 2.5**: P < 1E-5 / flight hour  ")
    md.append("**Stima Firmamento (preliminare)**: P ~ 2-3E-5 / flight hour (**MARGINALE**)\n")
    md.append("**Cut sets dominanti**:")
    md.append("1. Avaria FCS critica (~1E-5/h) - SPOF mitigato da 2oo3 voting + watchdog + ECC")
    md.append("2. Avaria propulsione + landing fail (~1E-5/h) - mitigato da parachute dual + battery override")
    md.append("3. Severe weather encounter (~1E-5/h) - mitigato da NOWCAST integration + abort criteria")
    md.append("4. Cyber hijack (~1E-6/h) - mitigato da crypto + 2FA + air-gap FCS\n")
    md.append("**Single Points of Failure** (SPOF identificati):")
    md.append("- SPOF-1: autopilot DAL-C primary - **mitigato 2oo3 voting + formal verification**")
    md.append("- SPOF-2: parachute singolo - **mitigato dual pyrotechnic + ballistic backup**")
    md.append("- SPOF-3: SATCOM Iridium singolo - **mitigato Inmarsat dual-provider Phase B**\n")
    md.append("**Action items per SAIL III compliance**:")
    md.append("- Reduce FCS DAL-C failure rate (HW redundancy + formal verification)")
    md.append("- Improve weather forecast integration (NOWCAST + abort)")
    md.append("- Improve GNSS robustness (Galileo PRS opzionale Phase B)\n")

    md.append("### 6.2 Top event: Loss of Mission EO (Percorso 6A pilota Pentema)\n")
    md.append("**Target SLA cliente**: < 5% per missione  ")
    md.append("**Stima Firmamento (preliminare)**: 15-20% / missione (**NON-CONFORME al target 5%**)\n")
    md.append("**Cut sets dominanti**:")
    md.append("1. Operational mission abort - meteo + Lost-Link + ATC (~7%) - **driver primario**")
    md.append("2. Quality below SLA - cloud cover + blur (~6%)")
    md.append("3. Payload EO failure (~3%) - mitigato da ridondanza")
    md.append("4. Data downlink/storage failure (~2%) - mitigato da buffer + retry\n")
    md.append("**Action items**:")
    md.append("- SLA realistico con cliente PA = 80-85% mission success rate (revisione target a 10-15% abort)")
    md.append("- Integrazione NOWCAST meteo + cloud cover prediction")
    md.append("- Buffer mission re-scheduling automatico\n")

    md.append("---\n")
    md.append("## 7. Residual risk profile\n")
    md.append("Post-mitigation, il profilo rischio aggregato e':\n")
    md.append("| Profilo | RED | YELLOW | GREEN | Totale | Note |")
    md.append("|---|---:|---:|---:|---:|---|")
    md.append(f"| Baseline | {len(red)} | {len(yellow)} | {len(green)} | {len(RISKS)} | Pre-mitigation |")
    md.append(f"| Residual | {red_r} | {yellow_r} | {green_r} | {len(RISKS)} | Post-mitigation |")
    md.append("")
    md.append("### 7.1 Profilo per percorso\n")
    md.append("**Percorso 6A (VTOL pilota Pentema)**:")
    md.append("- Showstopper: 0 nessuno bloccante (RSK-REG-001 e RSK-TEC-001/002/003 sono 6B-specific)")
    md.append("- RED residuali: principalmente operativi/regolatori transizione (Part-IS, NIS2, AgID/PSN)")
    md.append("- Profilo: **medio-basso** - compatibile con verdetto Go Condizionato\n")
    md.append("**Percorso 6B (HALE stratosferico R&D)**:")
    md.append("- Showstopper: 5 (RSK-TEC-001/002/003 + RSK-REG-001 + RSK-FIN-001)")
    md.append("- Mitigation strategy esiste ma **non garantita**")
    md.append("- Profilo: **alto** - compatibile con verdetto Hold / Go Condizionato Estremo\n")

    md.append("### 7.2 Caveat epistemico\n")
    md.append("Tutti i residual score sono **stime expert judgment** del risk-register-builder + safety engineer, con confidence dichiarato per ogni rischio. La probabilita di mitigation effettiva al M+9-12 dipende da:")
    md.append("- Hiring 3 ruoli senior (RSK-HR-002)")
    md.append("- Pre-application ENAC outcomes (RSK-REG-002)")
    md.append("- Funding mix outcomes (RSK-FIN-001 + RSK-MKT-001)")
    md.append("- Audit Part-IS + AgID outcomes (RSK-REG-019 + RSK-REG-021)\n")
    md.append("Re-assessment quarterly con re-scoring trimestrale.\n")

    md.append("---\n")
    md.append("## 8. EWI quarterly monitoring plan\n")
    md.append(f"Top-{len(EWI_DATA)} rischi monitorati con Early Warning Indicators dedicati. Frequenza minimum quarterly; rischi RED monthly. Vedi foglio XLSX `EWI` per dettaglio.\n")
    md.append("### 8.1 EWI ad alta frequenza (settimanale/mensile)\n")
    weekly_monthly = [e for e in EWI_DATA if "Settimanal" in e[2] or "Mensile" in e[2]]
    md.append(f"**{len(weekly_monthly)} EWI** ad alta frequenza:")
    for rid, ind, freq, thr, own in weekly_monthly[:10]:
        md.append(f"- **{rid}** ({freq}): {ind} - trigger: {thr} - owner: {own}")
    md.append("")
    md.append("### 8.2 Quarterly review meeting\n")
    md.append("**Cadence**: Q+1, Q+2, Q+3, Q+4 (ogni 3 mesi)  ")
    md.append("**Partecipanti**: Risk Manager (=CISO joint Head of Regulatory Affairs fino assunzione), CEO, owner ogni RED risk, observer Coopfond/Legacoop  ")
    md.append("**Output**: aggiornamento P/I/Score, residual update, new risks identification, escalation Steering Committee  ")
    md.append("**Documenti generati**: Risk Register vN+1 (versioning) + EWI dashboard + escalation log\n")

    md.append("---\n")
    md.append("## 9. Versioning roadmap\n")
    md.append("| Versione | Data target | Trigger | Owner | Note |")
    md.append("|---|---|---|---|---|")
    md.append("| v1.0 | 2026-05-17 | Consolidamento M+3 | senior risk manager | **Attuale** |")
    md.append("| v1.1 | 2026-08-17 (M+6) | Gate G2 review + 3 FTE senior hired status | risk-register + steering | Post-CISO + DPO hire")
    md.append("| v1.2 | 2026-11-17 (M+9) | Pre-gate G3 (M+10/M+11) | risk-register + steering | Hard conditions C1-C5 status update")
    md.append("| v2.0 | 2027-02-17 (M+12) | Gate G3 outcome + Y1 close | risk-register + auditor esterno | Major re-baseline pre-Y2 operations")
    md.append("| v2.1 | 2027-08-17 (M+18) | Mid-Y2 update | risk-register | Gate G4 preparation")
    md.append("| v3.0 | 2028-02-17 (M+24) | Gate G5 outcome - Phase B decision | risk-register + senior advisor | Major re-baseline pre-Phase B")
    md.append("")

    md.append("### 9.1 Re-assessment triggers (oltre a versioning schedule)\n")
    md.append("- Cambio scope o requisiti (RTM update)")
    md.append("- Nuovo trade study completato")
    md.append("- Gate review imminente (G2/G3/G4/G5)")
    md.append("- Evento esterno (cambio regolatorio EASA/ENAC/AgID/AGCOM, market shock, geopolitical event)")
    md.append("- Incidente o near-miss interno o settore")
    md.append("- Hire 3 ruoli senior (CISO, DPO, Head Reg.Aff.) - re-balance owners")
    md.append("- EWI threshold breach (anche singolo)\n")

    md.append("---\n")
    md.append("## 10. Riferimenti\n")
    md.append("**Fonti normative e metodologiche**:")
    md.append("- NASA NPR 8000.4A - Agency Risk Management Procedural Requirements (vedi `fonti/NASA04. SysEng Handbook (NASA_SP-2016-6105 Rev 2).md` Annex N)")
    md.append("- MIL-STD-1629A - Procedures for Performing a Failure Mode, Effects and Criticality Analysis")
    md.append("- IEC 60812:2018 - Analysis techniques for system reliability - FMECA")
    md.append("- ARP4761 - Guidelines and Methods for Conducting the Safety Assessment Process on Civil Airborne Systems")
    md.append("- NUREG-0492 - Fault Tree Handbook")
    md.append("- IEC 61025:2006 - Fault tree analysis (FTA)")
    md.append("- ISO 31000:2018 - Risk Management - Principles and Guidelines")
    md.append("- D.Lgs. 36/2023 art. 41 (Codice dei Contratti - PFTE)")
    md.append("- Reg.UE 2019/947 + EASA SORA 2.5 (ED Decision 2025/018/R)")
    md.append("- Reg.UE 2023/203 - Part-IS Information Security")
    md.append("- D.Lgs. 138/2024 - recepimento NIS2")
    md.append("- Reg.UE 2024/1689 - AI Act")
    md.append("- D.Lgs. 81/2008 - Sicurezza sul lavoro\n")
    md.append("**Documenti di progetto** (cross-reference):")
    md.append("- `studio-di-fattibilita/cap-05-quadro-normativo.md` §5.16 - 15 showstopper regolatori")
    md.append("- `studio-di-fattibilita/cap-06-analisi-tecnica.md` §6.4 - Top-10 + FMECA Payload + FTA preliminari")
    md.append("- `studio-di-fattibilita/cap-10-raccomandazione-di-gate.md` §10.2 - Risk residuo aggregato")
    md.append("- `riferimenti/RESERVED-rischi-geopolitici.md` - 5 RSK-GEO (accesso ristretto)")
    md.append("- `studio-di-fattibilita/AUDIT-REDTEAM-VOLUME-1.md` - Red Team M+3")
    md.append("- `studio-di-fattibilita/AUDIT-COMPETITOR-VOLUME-1.md` - Competitor Intelligence M+3")
    md.append("- `studio-di-fattibilita/AUDIT-REGULATORY-VOLUME-1.md` - Regulatory Adversary M+3")
    md.append("- `.claude/skills/risk-register-builder/SKILL.md` - Metodologia operativa\n")

    md.append("---\n")
    md.append(f"*Fine documento - Allegato A.2 Risk Register Report v{VERSION[1:]} - {TODAY} - Firmamento Technologies*\n")

    MD_PATH.write_text("\n".join(md), encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    wb = Workbook()
    # remove default sheet (we replace with Cover)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 1 Cover
    sheet_cover(wb)
    # 2 Top-25
    sheet_top_25(wb)
    # 3-13 Categories
    sheet_category(wb, "RSK-TEC", lambda r: r["cat"] in ("Tecnico", "Tecnico/Regolatorio", "Tecnico/Privacy"))
    sheet_category(wb, "RSK-REG", lambda r: r["cat"] == "Regolatorio")
    sheet_category(wb, "RSK-FIN", lambda r: r["cat"] == "Finanziario")
    sheet_category(wb, "RSK-MKT", lambda r: r["cat"] == "Mercato")
    sheet_category(wb, "RSK-OPS", lambda r: r["cat"] == "Operativo")
    sheet_category(wb, "RSK-SUP", lambda r: r["cat"] == "Supply Chain")
    sheet_category(wb, "RSK-PRV", lambda r: r["cat"] == "Privacy/Legale")
    sheet_category(wb, "RSK-SEC", lambda r: r["cat"] == "Cybersecurity")
    sheet_category(wb, "RSK-HR", lambda r: r["cat"] == "Risorse Umane")
    sheet_category(wb, "RSK-REP", lambda r: r["cat"] == "Reputazionale")
    sheet_geo(wb)
    # 14-16 FMECA
    sheet_fmeca(wb, "FMECA_Payload", FMECA_PAYLOAD, "Sottosistema Payload EO (RGB + IR LWIR + storage + downlink)")
    sheet_fmeca(wb, "FMECA_Avionica", FMECA_AVIONICA, "Sottosistema Avionica (FCS DAL-C + IMU triplex + GNSS + C2 + FTS)")
    sheet_fmeca(wb, "FMECA_Propulsione", FMECA_PROP, "Sottosistema Propulsione/Energia (motori + batterie + solare + thermal)")
    # 17-18 FTA
    sheet_fta(wb, "FTA_LossOfVehicle", FTA_LOV, "FTA Top event: Loss of Vehicle in BVLOS")
    sheet_fta(wb, "FTA_LossOfMission", FTA_LOM, "FTA Top event: Loss of Mission EO")
    # 19 Mitigation plan
    sheet_mitigation_plan(wb)
    # 20 Residual matrix
    sheet_residual_matrix(wb)
    # 21 EWI
    sheet_ewi(wb)
    # 22 Audit trail
    sheet_audit_trail(wb)

    wb.save(XLSX_PATH)
    write_csv()
    write_md_report()

    print(f"OK  XLSX:    {XLSX_PATH}  ({len(wb.sheetnames)} sheets)")
    print(f"OK  CSV:     {CSV_PATH}  ({len(RISKS)} rischi)")
    print(f"OK  REPORT:  {MD_PATH}")
    print()
    print(f"Riepilogo: RED={sum(1 for r in RISKS if r['score']>=15)}, "
          f"YELLOW={sum(1 for r in RISKS if 8<=r['score']<15)}, "
          f"GREEN={sum(1 for r in RISKS if r['score']<8)}")


if __name__ == "__main__":
    main()
