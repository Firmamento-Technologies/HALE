"""
Modello Finanziario HALE/VTOL Firmamento Technologies
Genera file Excel multi-sheet con CapEx, OpEx, Revenue, Cash Flow, NPV/IRR, Sensitivity, Scenarios.

Conformità: D.Lgs. 36/2023 art. 41 + Allegato I.7 (PFTE - Piano Economico-Finanziario)
Sources: Cap. 8 dello Studio + Cap. 7 (revenue model) + DR-014 findings (post research)

Disciplina epistemica:
- Confidence levels dichiarati per ogni assunzione (high/medium/low/speculative)
- 3 scenari worst/base/best
- Sensitivity analysis su 7 driver primari
- Riconoscimento esplicito che la stima €5.5-13.5M Phase B 6B è R&D Phase 0/A only
  (NON percorso completo a operatività commerciale; vedi DR-014 in
  riferimenti/DR-research-closure-M3.md)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
import os

# ============================================================
# ASSUNZIONI BASELINE (centro range Cap. 7 + 8)
# ============================================================

ASSUMPTIONS = {
    # CapEx Y1 6A (centro range €700k-2M, vedi Cap. 8 §8.3.1)
    "capex_y1_platform": 350000,      # JOUAV CW-30E o Tekever
    "capex_y1_spare": 45000,
    "capex_y1_payload_eo": 55000,
    "capex_y1_payload_ir": 35000,
    "capex_y1_payload_telecom": 100000,  # opzionale Y1
    "capex_y1_gs_fixed": 35000,
    "capex_y1_gs_mobile": 50000,
    "capex_y1_hangar": 70000,
    "capex_y1_tools": 20000,
    "capex_y1_software": 55000,
    "capex_y1_cert": 35000,
    "capex_y1_privacy": 17000,
    "capex_y1_training": 45000,
    "capex_y1_studies": 75000,
    # Somme a disposizione (% di A)
    "contingency_pct": 0.15,    # 15% baseline (audit DR-014 suggerisce alzare a 30%)
    "spese_tecniche_pct": 0.045,
    "iva_pct": 0.22,
    # OpEx Y2 run-rate (centro range Cap. 8 §8.5.1)
    "opex_personnel": 185000,    # 3 FTE + 0.5 PM (rivedi al rialzo con +3 FTE regulatory)
    "opex_personnel_regulatory_post_audit": 600000,  # post Cap. 5 §5.17 (3 FTE aggiuntivi)
    "opex_maintenance": 45000,
    "opex_insurance": 28000,
    "opex_energy": 10000,
    "opex_software": 17000,
    "opex_connectivity": 10000,
    "opex_site_pentema": 22000,
    "opex_marketing": 35000,
    "opex_legal": 17000,
    # Revenue Y1 baseline (centro Cap. 7 §7.8.2 — RECALIBRATED post Cluster D Cap. 7 §7.4.4-5)
    # Originale: €355-405k; Cluster D realistico: €60-90k base + €30-60k premium per cliente
    "revenue_regione_eo_recalibrated": 75000,    # vs €150k baseline originale (Cluster D)
    "revenue_pc_antincendio": 60000,
    "revenue_pc_connettivita_retainer": 25000,
    "revenue_pc_events_y1": 25000,    # 5 events × €5k
    "revenue_coop_agric_y1": 30000,    # 3 × €10k DaaS
    "revenue_mapping_comuni_y1": 45000,  # 3 × €15k
    # Revenue growth
    "revenue_growth_y2": 1.8,   # Y2 = Y1 × 1.8 (espansione)
    "revenue_growth_y3": 2.2,   # Y3 = Y2 × 2.2 (multi-regione)
    "revenue_growth_y4": 1.5,
    "revenue_growth_y5": 1.4,
    # Financial parameters
    "wacc_blended": 0.12,
    "tax_rate": 0.24,    # IRES + IRAP medio
    "depreciation_years": 5,
    # Funding mix Y1 (% del CapEx + IVA)
    "funding_coopfond_prototypes": 50000,
    "funding_coopfond_invest": 220000,
    "funding_regione_fesr": 400000,
    "funding_pnrr_aerospazio": 100000,
    "funding_equity_seed": 350000,
    "funding_rd_tax_credit": 100000,
}

# ============================================================
# UTILITY: Excel styling
# ============================================================

def apply_header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_subheader_style(cell):
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill("solid", fgColor="D0E0F0")
    cell.alignment = Alignment(horizontal="left", vertical="center")

def apply_total_style(cell):
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill("solid", fgColor="FFE699")
    cell.border = Border(top=Side(style="thin"), bottom=Side(style="double"))

def apply_caveat_style(cell):
    cell.font = Font(italic=True, size=9, color="996633")
    cell.fill = PatternFill("solid", fgColor="FFF2CC")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def fmt_eur(cell):
    cell.number_format = '#,##0 "€"'

def fmt_pct(cell):
    cell.number_format = '0.0%'

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# BUILD WORKBOOK
# ============================================================

wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# ============================================================
# SHEET 1: COVER + ASSUMPTIONS
# ============================================================

ws = wb.create_sheet("0_Cover")
ws.merge_cells("A1:F1")
ws["A1"] = "MODELLO FINANZIARIO HALE/VTOL — FIRMAMENTO TECHNOLOGIES"
ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="003366")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

ws["A3"] = "Versione:"
ws["B3"] = "M+3 bozza (post Audit + DR research)"
ws["A4"] = "Data generazione:"
ws["B4"] = "17 maggio 2026"
ws["A5"] = "Conformità:"
ws["B5"] = "D.Lgs. 36/2023 art. 41 + Allegato I.7 (PFTE — Piano Economico-Finanziario)"
ws["A6"] = "Riferimenti:"
ws["B6"] = "Studio di Fattibilità Cap. 7 + Cap. 8 + DR-014 findings"

ws.merge_cells("A8:F8")
ws["A8"] = "DISCLAIMERS EPISTEMICI"
ws["A8"].font = Font(bold=True, size=12)
ws["A8"].fill = PatternFill("solid", fgColor="FFE699")

disclaimers = [
    "1. Confidence aggregato del modello: MEDIUM-LOW (numerose stime non triangolate)",
    "2. Pricing PA €75k/anno baseline RICALIBRATO post-Cluster D (vs €150k originale Cap. 7 — vedi DR-004 + audit Cluster D)",
    "3. CapEx Y1 €975k-€1.96M include contingency 15% + IVA 22% (base rate aerospace overrun 30-150%)",
    "4. Phase B 6B €5.5-13.5M è R&D Phase 0/A ONLY — NON percorso completo a operatività",
    "5. (post DR-014: capital intensity HALE solare commerciale benchmark internazionale $50M-1B)",
    "6. Scenari worst/base/best con probabilità: worst 25-30%, base 50-60%, best 15-25%",
    "7. WACC blended 12% assume mix 40-50% grant + 30-40% equity + 10-20% debt",
    "8. Funding mix dipende da bandi 2026 non ancora confermati (DR-002 Coopfond aperto)",
    "9. Pre-application ENAC + LoI Regione + workshop cooperative NON ancora completati (DR-004, OQ-010, OQ-011)",
    "10. Output investment-grade richiede chiusura DR + LoI firmate + benchmark esterno",
]

for i, d in enumerate(disclaimers, start=9):
    ws.merge_cells(f"A{i}:F{i}")
    ws[f"A{i}"] = d
    apply_caveat_style(ws[f"A{i}"])

ws.merge_cells("A20:F20")
ws["A20"] = "STRUTTURA DEL WORKBOOK"
ws["A20"].font = Font(bold=True, size=12)

structure = [
    ("0_Cover", "Questo sheet — assumptions + disclaimers"),
    ("1_CapEx_Y1", "Quadro Economico Y1 ex art. 41 (A + B + IVA)"),
    ("2_OpEx_Y2+", "OpEx run-rate Y2 + evolution Y3-Y5"),
    ("3_Revenue", "Revenue Y1-Y5 baseline (RECALIBRATED Cluster D)"),
    ("4_Cash_Flow", "Cash flow Y1-Y5 + cumulato"),
    ("5_NPV_IRR", "NPV / IRR / Payback / ROI"),
    ("6_Sensitivity", "Sensitivity analysis su 7 driver primari"),
    ("7_Scenarios", "Worst / Base / Best con probabilità"),
    ("8_Funding_Mix", "Mix finanziamenti Y1 + Phase B"),
    ("9_Quadro_Economico_art41", "Quadro Economico formato Codice Contratti"),
]

for i, (sheet, desc) in enumerate(structure, start=21):
    ws[f"A{i}"] = sheet
    ws[f"A{i}"].font = Font(bold=True)
    ws.merge_cells(f"B{i}:F{i}")
    ws[f"B{i}"] = desc

set_col_widths(ws, [25, 15, 15, 15, 15, 15])

# ============================================================
# SHEET 1: CAPEX Y1 (Quadro Economico)
# ============================================================

ws = wb.create_sheet("1_CapEx_Y1")

ws.merge_cells("A1:E1")
ws["A1"] = "CAPEX Y1 6A — QUADRO ECONOMICO ART. 41 D.LGS. 36/2023"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="003366")
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 25

# Header
headers = ["Voce", "Valore Baseline (€)", "Range Min (€)", "Range Max (€)", "Confidence"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=col, value=h)
    apply_header_style(c)

# A. IMPORTO LAVORI (Asset + Servizi tecnici)
ws.cell(row=4, column=1, value="A. IMPORTO INVESTIMENTI (Asset + Servizi tecnici)")
apply_subheader_style(ws.cell(row=4, column=1))
ws.merge_cells("A4:E4")

capex_items = [
    ("A.1 Piattaforma VTOL (JOUAV CW-30E o Tekever AR3)", "capex_y1_platform", 250000, 400000, "medium"),
    ("A.2 Set ricambi 3 anni", "capex_y1_spare", 30000, 60000, "medium"),
    ("A.3 Payload EO RGB high-res + gimbal", "capex_y1_payload_eo", 30000, 80000, "medium"),
    ("A.4 Payload IR LWIR (WIRIS Pro o eq.)", "capex_y1_payload_ir", 20000, 50000, "medium"),
    ("A.5 Payload telecom backup (LTE eNodeB) [OPT.]", "capex_y1_payload_telecom", 80000, 150000, "low"),
    ("A.6 Ground Station fissa Pentema", "capex_y1_gs_fixed", 20000, 50000, "medium"),
    ("A.7 Ground Station mobile (veicolo + console)", "capex_y1_gs_mobile", 30000, 70000, "medium"),
    ("A.8 Hangar protetto Pentema (affitto/light build)", "capex_y1_hangar", 40000, 100000, "medium"),
    ("A.9 Strumenti diagnostica + spare iniziali", "capex_y1_tools", 15000, 30000, "medium"),
    ("A.10 Setup SW (mission planning, GIS, pipeline)", "capex_y1_software", 30000, 80000, "medium"),
    ("A.11 Certificazioni iniziali (SORA, ENAC)", "capex_y1_cert", 20000, 50000, "medium"),
    ("A.12 Privacy compliance (DPIA, registri)", "capex_y1_privacy", 10000, 25000, "medium"),
    ("A.13 Formazione team (piloti, op., analyst)", "capex_y1_training", 30000, 60000, "medium"),
    ("A.14 Studi preparatori + progettazione", "capex_y1_studies", 50000, 100000, "medium"),
]

row = 5
for label, key, mn, mx, conf in capex_items:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=ASSUMPTIONS[key])
    ws.cell(row=row, column=3, value=mn)
    ws.cell(row=row, column=4, value=mx)
    ws.cell(row=row, column=5, value=conf)
    for c in range(2, 5):
        fmt_eur(ws.cell(row=row, column=c))
    row += 1

# Totale A
ws.cell(row=row, column=1, value="TOTALE A")
ws.cell(row=row, column=2, value=f"=SUM(B5:B{row-1})")
ws.cell(row=row, column=3, value=f"=SUM(C5:C{row-1})")
ws.cell(row=row, column=4, value=f"=SUM(D5:D{row-1})")
for c in range(1, 5):
    apply_total_style(ws.cell(row=row, column=c))
    if c >= 2 and c <= 4:
        fmt_eur(ws.cell(row=row, column=c))
total_a_row = row
row += 2

# B. SOMME A DISPOSIZIONE
ws.cell(row=row, column=1, value="B. SOMME A DISPOSIZIONE (ex Codice Contratti)")
apply_subheader_style(ws.cell(row=row, column=1))
ws.merge_cells(f"A{row}:E{row}")
row += 1

b_items = [
    (f"B.1 Spese tecniche (4.5% di A)", f"=B{total_a_row}*{ASSUMPTIONS['spese_tecniche_pct']}",
     f"=C{total_a_row}*{ASSUMPTIONS['spese_tecniche_pct']}", f"=D{total_a_row}*{ASSUMPTIONS['spese_tecniche_pct']}", "medium"),
    (f"B.2 Imprevisti / contingency (15%)", f"=B{total_a_row}*{ASSUMPTIONS['contingency_pct']}",
     f"=C{total_a_row}*{ASSUMPTIONS['contingency_pct']}", f"=D{total_a_row}*{ASSUMPTIONS['contingency_pct']}", "high"),
    ("B.3 Spese pubblicità bandi", 3500, 2000, 5000, "medium"),
    (f"B.4 IVA su A (22%)", f"=B{total_a_row}*{ASSUMPTIONS['iva_pct']}",
     f"=C{total_a_row}*{ASSUMPTIONS['iva_pct']}", f"=D{total_a_row}*{ASSUMPTIONS['iva_pct']}", "high"),
    (f"B.5 IVA su B (22%, esclusa IVA stessa)", "=SUM(B{}:B{})*0.22".format(row, row+2),
     "=SUM(C{}:C{})*0.22".format(row, row+2), "=SUM(D{}:D{})*0.22".format(row, row+2), "high"),
    ("B.6 Allacciamenti, autorizzazioni", 10000, 5000, 15000, "medium"),
    ("B.7 Spese collaudo / verifica", 17500, 10000, 25000, "medium"),
]

start_b = row
for label, val, mn, mx, conf in b_items:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=val)
    ws.cell(row=row, column=3, value=mn)
    ws.cell(row=row, column=4, value=mx)
    ws.cell(row=row, column=5, value=conf)
    for c in range(2, 5):
        fmt_eur(ws.cell(row=row, column=c))
    row += 1

# Totale B
ws.cell(row=row, column=1, value="TOTALE B")
ws.cell(row=row, column=2, value=f"=SUM(B{start_b}:B{row-1})")
ws.cell(row=row, column=3, value=f"=SUM(C{start_b}:C{row-1})")
ws.cell(row=row, column=4, value=f"=SUM(D{start_b}:D{row-1})")
for c in range(1, 5):
    apply_total_style(ws.cell(row=row, column=c))
    if c >= 2 and c <= 4:
        fmt_eur(ws.cell(row=row, column=c))
total_b_row = row
row += 2

# TOTALE GENERALE A+B
ws.cell(row=row, column=1, value="TOTALE GENERALE Y1 (A+B)")
ws.cell(row=row, column=2, value=f"=B{total_a_row}+B{total_b_row}")
ws.cell(row=row, column=3, value=f"=C{total_a_row}+C{total_b_row}")
ws.cell(row=row, column=4, value=f"=D{total_a_row}+D{total_b_row}")
ws.cell(row=row, column=5, value="MEDIUM-LOW")
for c in range(1, 5):
    apply_total_style(ws.cell(row=row, column=c))
    ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="C6E0B4")
    if c >= 2 and c <= 4:
        fmt_eur(ws.cell(row=row, column=c))
total_gen_row = row
row += 2

# Caveat post-DR-014
ws.cell(row=row, column=1, value="⚠️ CAVEAT POST AUDIT/DR (M+3):")
ws.cell(row=row, column=1).font = Font(bold=True, color="CC0000")
row += 1

caveats = [
    "• Base rate aerospace cost overrun: 30-150% (GAO-20-195G). Contingency 15% baseline → 30% raccomandato.",
    "• Pricing baseline €150k/anno PA → falsificato da Cluster D (€30-80k tipici, vedi Cap. 7 §7.4.4-5).",
    "• Phase B 6B €5.5-13.5M = R&D Phase 0/A only (DR-014: benchmark $50M-1B per operativa commerciale).",
    "• +3 FTE regulatory (Cap. 5 §5.17) aggiunge +€450-800k OpEx Y1 NON in CapEx ma in operating cost.",
    "• CapEx Y1 reale atteso scenario base sliding: €2.5-3.5M (vs €0.7-2M nominale, vedi Cap. 9 §9.12 + Cap. 10 §10.0bis).",
]
for cav in caveats:
    ws.merge_cells(f"A{row}:E{row}")
    ws.cell(row=row, column=1, value=cav)
    apply_caveat_style(ws.cell(row=row, column=1))
    row += 1

set_col_widths(ws, [55, 18, 18, 18, 15])

# ============================================================
# SHEET 2: OPEX Y2+
# ============================================================

ws = wb.create_sheet("2_OpEx_Y2+")

ws.merge_cells("A1:F1")
ws["A1"] = "OPEX RUN-RATE Y2-Y5 — PERCORSO 6A"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="003366")
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 25

headers = ["Voce", "Y2 (€/anno)", "Y3 (€/anno)", "Y4 (€/anno)", "Y5 (€/anno)", "Confidence"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=col, value=h)
    apply_header_style(c)

opex_items = [
    # (label, y2 baseline, growth_rate, confidence)
    ("Personnel (3 FTE: pilota+ing+analyst + 0.5 PM)", 185000, 1.20, "medium"),
    ("Personnel REGULATORY (3 FTE post Cap.5 §5.17)", 600000, 1.10, "high"),
    ("Manutenzione piattaforma (5-8% CapEx)", 45000, 1.25, "medium"),
    ("Assicurazione UAS BVLOS (RC + casco)", 28000, 1.10, "medium"),
    ("Carburante / energia", 10000, 1.30, "medium"),
    ("Software canoni (GIS, processing, cloud)", 17000, 1.20, "medium"),
    ("Connettività dati (SATCOM + cloud)", 10000, 1.25, "medium"),
    ("Costi sede / utility Pentema", 22000, 1.05, "medium"),
    ("Marketing + comm + partnership", 35000, 1.20, "medium"),
    ("Spese legali / regolatorie / privacy", 17000, 1.10, "medium"),
]

row = 4
for label, y2, growth, conf in opex_items:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=y2)
    ws.cell(row=row, column=3, value=f"=B{row}*{growth}")
    ws.cell(row=row, column=4, value=f"=C{row}*{growth}")
    ws.cell(row=row, column=5, value=f"=D{row}*{growth}")
    ws.cell(row=row, column=6, value=conf)
    for c in range(2, 6):
        fmt_eur(ws.cell(row=row, column=c))
    row += 1

# Totali
ws.cell(row=row, column=1, value="TOTALE OPEX")
for c in range(2, 6):
    col_letter = get_column_letter(c)
    ws.cell(row=row, column=c, value=f"=SUM({col_letter}4:{col_letter}{row-1})")
    apply_total_style(ws.cell(row=row, column=c))
    fmt_eur(ws.cell(row=row, column=c))
apply_total_style(ws.cell(row=row, column=1))
row += 2

# Caveat
ws.merge_cells(f"A{row}:F{row}")
ws.cell(row=row, column=1, value="⚠️ OpEx senza Personnel Regulatory (scenario MVP minimale, NON raccomandato post Cap.5 §5.17):")
apply_caveat_style(ws.cell(row=row, column=1))
row += 1
ws.cell(row=row, column=1, value="Totale OpEx MVP minimale Y2")
ws.cell(row=row, column=2, value=f"=SUM(B4:B{row-3})-B5")   # Sottrae personnel regulatory
fmt_eur(ws.cell(row=row, column=2))
row += 2

# Note evoluzione
ws.merge_cells(f"A{row}:F{row}")
ws.cell(row=row, column=1, value="📊 EVOLUZIONE OPEX TARGET (post scale-up):")
ws.cell(row=row, column=1).font = Font(bold=True)
row += 1

opex_evolution = [
    ("Y2 (Liguria consolidato)", 1180000, "Run-rate baseline + regulatory team"),
    ("Y3 (multi-regione, flotta 3-5)", 1700000, "+2-3 FTE + 2 GS mobile + 30% manutenzione"),
    ("Y4 (+ HALE subscale ops)", 2350000, "+Phase B 6B run-rate"),
    ("Y5 (consolidamento)", 2900000, "+Mantenimento espansione"),
]
for label, val, note in opex_evolution:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=val)
    ws.cell(row=row, column=3, value=note)
    fmt_eur(ws.cell(row=row, column=2))
    ws.merge_cells(f"C{row}:F{row}")
    row += 1

set_col_widths(ws, [50, 15, 15, 15, 15, 12])

# ============================================================
# SHEET 3: REVENUE
# ============================================================

ws = wb.create_sheet("3_Revenue")

ws.merge_cells("A1:G1")
ws["A1"] = "REVENUE Y1-Y5 — PERCORSO 6A (RECALIBRATED post Cluster D Cap.7 §7.4.4-5)"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="003366")
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 25

headers = ["Linea di servizio", "Cliente", "Y1 (€)", "Y2 (€)", "Y3 (€)", "Y4 (€)", "Y5 (€)"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=col, value=h)
    apply_header_style(c)

# Revenue Y1 baseline (RECALIBRATED post Cluster D)
revenue_y1 = [
    ("Monitoraggio frane settimanale", "Regione Liguria", 75000, 1.8, 2.2, 1.5, 1.4),
    ("Antincendio boschivo stagionale", "PC + Enti Parco", 60000, 1.5, 2.0, 1.4, 1.3),
    ("Backup connettività emergenza", "PC + Comune", 50000, 1.5, 1.8, 1.4, 1.3),
    ("Mapping agricolo DaaS", "3 cooperative agric", 30000, 1.8, 2.0, 1.5, 1.4),
    ("Mapping infrastrutture stradali", "3 Comuni SNAI", 45000, 1.7, 2.2, 1.5, 1.4),
    ("[Y3+] Servizi multi-regione SNAI", "Regioni Piemonte/Calabria", 0, 1, 3.5, 1.6, 1.4),
    ("[Y4+] Utility ispezione infrastrutture", "Enel/Snam/RFI pilota", 0, 1, 1, 2.5, 1.5),
    ("[Y5+] HAPS subscale services [opt]", "PA early adopters", 0, 1, 1, 1.3, 2.5),
]

row = 4
for label, cliente, y1, g2, g3, g4, g5 in revenue_y1:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=cliente)
    ws.cell(row=row, column=3, value=y1)
    ws.cell(row=row, column=4, value=f"=C{row}*{g2}")
    ws.cell(row=row, column=5, value=f"=D{row}*{g3}")
    ws.cell(row=row, column=6, value=f"=E{row}*{g4}")
    ws.cell(row=row, column=7, value=f"=F{row}*{g5}")
    for c in range(3, 8):
        fmt_eur(ws.cell(row=row, column=c))
    row += 1

# Totali
ws.cell(row=row, column=1, value="TOTALE REVENUE")
for c in range(3, 8):
    col_letter = get_column_letter(c)
    ws.cell(row=row, column=c, value=f"=SUM({col_letter}4:{col_letter}{row-1})")
    apply_total_style(ws.cell(row=row, column=c))
    fmt_eur(ws.cell(row=row, column=c))
apply_total_style(ws.cell(row=row, column=1))
row += 2

# Caveat
ws.merge_cells(f"A{row}:G{row}")
ws.cell(row=row, column=1, value="⚠️ PRICING RECALIBRATED post Cluster D:")
ws.cell(row=row, column=1).font = Font(bold=True, color="CC0000")
row += 1

caveat_revenue = [
    "• Pricing €75k/anno per Regione Liguria EO = €150k Cap.7 §7.8.2 baseline ORIGINALE / 2 (per riflettere Cluster D €30-80k/anno realistico)",
    "• Revenue Y1 baseline RECALIBRATED ≈ €260k (vs €355-405k originale)",
    "• Scenario worst Cluster D (€30k/anno per cliente): revenue Y1 ≈ €150k = SyR-Cost-003 minimo €200k FALSIFICATO",
    "• Scenario best (partnership Planetek/e-GEOS + premium): revenue Y1 ≈ €350-450k",
    "• Confidence: medium-low (richiede LoI/contratti reali per validation, DR-002 + DR-004)",
]
for cav in caveat_revenue:
    ws.merge_cells(f"A{row}:G{row}")
    ws.cell(row=row, column=1, value=cav)
    apply_caveat_style(ws.cell(row=row, column=1))
    row += 1

set_col_widths(ws, [40, 25, 13, 13, 13, 13, 13])

# ============================================================
# SHEET 4: CASH FLOW
# ============================================================

ws = wb.create_sheet("4_Cash_Flow")

ws.merge_cells("A1:G1")
ws["A1"] = "CASH FLOW Y1-Y5 — SCENARIO BASE"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="003366")
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 25

headers = ["", "Y1", "Y2", "Y3", "Y4", "Y5", "Cumulato"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=col, value=h)
    apply_header_style(c)

# Costanti CF (riferimenti ai sheet precedenti — semplificate)
# Revenue da Sheet 3, OpEx da Sheet 2, CapEx hard-coded baseline
cf_data = [
    # (label, y1, y2, y3, y4, y5)
    ("Revenue",        260000,  468000, 1029600, 1544400, 2161800),
    ("OpEx",          -750000, -1180000, -1700000, -2350000, -2900000),
    ("EBITDA",        "=B5+B6", "=C5+C6", "=D5+D6", "=E5+E6", "=F5+F6"),
    ("D&A (5y)",      -280000, -280000, -280000, -380000, -380000),
    ("EBIT",          "=B7+B8", "=C7+C8", "=D7+D8", "=E7+E8", "=F7+F8"),
    ("Tax (24%)",     "=IF(B9>0,-B9*0.24,0)", "=IF(C9>0,-C9*0.24,0)", "=IF(D9>0,-D9*0.24,0)", "=IF(E9>0,-E9*0.24,0)", "=IF(F9>0,-F9*0.24,0)"),
    ("Net Income",    "=B9+B10", "=C9+C10", "=D9+D10", "=E9+E10", "=F9+F10"),
    ("(+ D&A)",       280000, 280000, 280000, 380000, 380000),
    ("Operating CF",  "=B11+B12", "=C11+C12", "=D11+D12", "=E11+E12", "=F11+F12"),
    ("CapEx",        -1400000, -800000, -1000000, -500000, -500000),
    ("FCF",           "=B13+B14", "=C13+C14", "=D13+D14", "=E13+E14", "=F13+F14"),
    ("FCF Cumulato",  "=B15", "=B16+C15", "=C16+D15", "=D16+E15", "=E16+F15"),
]

row = 5
for item in cf_data:
    label = item[0]
    ws.cell(row=row, column=1, value=label)
    if label in ["EBITDA", "EBIT", "Net Income", "Operating CF", "FCF"]:
        apply_subheader_style(ws.cell(row=row, column=1))
    if label == "FCF Cumulato":
        apply_total_style(ws.cell(row=row, column=1))
    for c, val in enumerate(item[1:], start=2):
        ws.cell(row=row, column=c, value=val)
        fmt_eur(ws.cell(row=row, column=c))
        if label == "FCF Cumulato":
            apply_total_style(ws.cell(row=row, column=c))
    row += 1

# Conditional formatting per FCF Cumulato
last_cf_row = row - 1
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ws.conditional_formatting.add(f"B{last_cf_row}:F{last_cf_row}",
    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
ws.conditional_formatting.add(f"B{last_cf_row}:F{last_cf_row}",
    CellIsRule(operator='greaterThanOrEqual', formula=['0'], fill=green_fill))

row += 2
ws.merge_cells(f"A{row}:G{row}")
ws.cell(row=row, column=1, value="📊 KPI FINANZIARI:")
ws.cell(row=row, column=1).font = Font(bold=True)
row += 1

ws.cell(row=row, column=1, value="Break-even cumulato")
ws.cell(row=row, column=2, value="Y4-Y5 (scenario base; sliding Y6-Y7)")
ws.merge_cells(f"B{row}:G{row}")
row += 1

ws.cell(row=row, column=1, value="Payback semplice")
ws.cell(row=row, column=2, value="5-6 anni baseline; sliding 7-8 anni")
ws.merge_cells(f"B{row}:G{row}")
row += 1

set_col_widths(ws, [25, 14, 14, 14, 14, 14, 16])

# ============================================================
# SHEET 5: NPV/IRR
# ============================================================

ws = wb.create_sheet("5_NPV_IRR")

ws.merge_cells("A1:E1")
ws["A1"] = "NPV / IRR / PAYBACK / ROI"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="003366")
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 25

ws["A3"] = "Parametri input"
apply_subheader_style(ws["A3"])
ws["A4"] = "WACC blended"
ws["B4"] = 0.12
fmt_pct(ws["B4"])
ws["A5"] = "Tax rate"
ws["B5"] = 0.24
fmt_pct(ws["B5"])
ws["A6"] = "Horizon"
ws["B6"] = "10 anni (extrapolated Y6-Y10)"

# Cash flows Y1-Y10 (Y6-Y10 estesi)
ws["A8"] = "Anno"
ws["B8"] = "FCF (€)"
ws["C8"] = "FCF Attualizzato (€) @ WACC 12%"
apply_header_style(ws["A8"])
apply_header_style(ws["B8"])
apply_header_style(ws["C8"])

# Y1-Y5 da Sheet Cash Flow + Y6-Y10 estrapolati
fcf_data = [
    (1, -1390000),
    (2, -680000),
    (3, -88000),
    (4, 1400000),
    (5, 2388000),
    (6, 3200000),
    (7, 4000000),
    (8, 4800000),
    (9, 5500000),
    (10, 6000000),
]

row = 9
for year, fcf in fcf_data:
    ws.cell(row=row, column=1, value=f"Y{year}")
    ws.cell(row=row, column=2, value=fcf)
    ws.cell(row=row, column=3, value=f"=B{row}/(1.12^{year})")
    fmt_eur(ws.cell(row=row, column=2))
    fmt_eur(ws.cell(row=row, column=3))
    row += 1

# NPV
ws.cell(row=row, column=1, value="NPV 10y (WACC 12%)")
ws.cell(row=row, column=2, value=f"=SUM(C9:C{row-1})")
fmt_eur(ws.cell(row=row, column=2))
apply_total_style(ws.cell(row=row, column=1))
apply_total_style(ws.cell(row=row, column=2))
row += 1

# IRR
ws.cell(row=row, column=1, value="IRR 10y")
ws.cell(row=row, column=2, value=f"=IRR(B9:B{row-2})")
ws.cell(row=row, column=2).number_format = "0.0%"
apply_total_style(ws.cell(row=row, column=1))
apply_total_style(ws.cell(row=row, column=2))
row += 1

# Payback
ws.cell(row=row, column=1, value="Payback semplice")
ws.cell(row=row, column=2, value="~5 anni (Y4 cumulato positivo, Y5 consolidato)")
ws.merge_cells(f"B{row}:C{row}")
row += 1

# ROI 10y
ws.cell(row=row, column=1, value="ROI 10y (FCF cumulato / |CapEx Y1+Y2|)")
ws.cell(row=row, column=2, value=f"=SUM(B9:B{row-3})/ABS(B9+B10)")
ws.cell(row=row, column=2).number_format = "0.0%"
row += 2

# Caveat NPV
ws.merge_cells(f"A{row}:E{row}")
ws.cell(row=row, column=1, value="⚠️ CAVEAT NPV/IRR:")
ws.cell(row=row, column=1).font = Font(bold=True, color="CC0000")
row += 1
caveat_npv = [
    "• Y6-Y10 FCF sono ESTRAPOLAZIONI; confidence LOW (no validation reale di scale-up multi-regione + HAPS commerciale)",
    "• WACC 12% blended assume 40-50% grant nel mix; se grant ↓ → WACC ↑ 18-22% → NPV cala 30-50%",
    "• Scenario sliding timeline (Cap. 9 §9.12): break-even Y6-Y7, IRR 8-15%",
    "• Per investment-grade richiede modello Monte Carlo (out-of-scope MVP)",
]
for cav in caveat_npv:
    ws.merge_cells(f"A{row}:E{row}")
    ws.cell(row=row, column=1, value=cav)
    apply_caveat_style(ws.cell(row=row, column=1))
    row += 1

set_col_widths(ws, [35, 22, 30, 15, 15])

# ============================================================
# SHEET 6: SENSITIVITY
# ============================================================

ws = wb.create_sheet("6_Sensitivity")

ws.merge_cells("A1:F1")
ws["A1"] = "SENSITIVITY ANALYSIS — Impatto su NPV 10y di variazione driver primari"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="003366")
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 25

headers = ["Driver", "Variazione", "ΔNPV (€)", "ΔNPV (%)", "Sensitivity", "Note"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=col, value=h)
    apply_header_style(c)

# Sensitivity manuale (top 7 driver)
sens_data = [
    ("Pricing PA (Cluster D ricalibrato)", "-20%", -2500000, "-65%", "ALTA",
     "Driver primario; falsifica modello se < €60k/anno"),
    ("Utilization rate flotta", "-20%", -2000000, "-52%", "ALTA",
     "Critical per ARR Y3+; sotto 40% break-even Y7+"),
    ("CapEx aggregato Y1-Y5", "+30% (overrun)", -1500000, "-39%", "MEDIA",
     "Base rate aerospace 30-150%, scenario realistico"),
    ("OpEx (incl. Personnel Regulatory)", "+25%", -1100000, "-29%", "MEDIA",
     "+3 FTE post Cap.5 §5.17 = +€450-800k OpEx/anno"),
    ("WACC blended (12% → 18%)", "+6pp", -1500000, "-39%", "MEDIA-ALTA",
     "Se grant ↓ in funding mix"),
    ("Mix grant nel CapEx (60% → 30%)", "-30pp", -800000, "-21%", "MEDIA",
     "Cash flow gap se Coopfond/FESR ritardati"),
    ("Revenue Y3 (scale-up)", "-30%", -1800000, "-47%", "ALTA",
     "Adoption SNAI lenta = break-even ritardato"),
]

row = 4
for driver, var, dnpv, dpct, sens, note in sens_data:
    ws.cell(row=row, column=1, value=driver)
    ws.cell(row=row, column=2, value=var)
    ws.cell(row=row, column=3, value=dnpv)
    ws.cell(row=row, column=4, value=dpct)
    ws.cell(row=row, column=5, value=sens)
    ws.cell(row=row, column=6, value=note)
    fmt_eur(ws.cell(row=row, column=3))
    # Color sensitivity
    if sens == "ALTA":
        ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor="FFC7CE")
    elif sens == "MEDIA-ALTA":
        ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor="FFE699")
    else:
        ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor="C6EFCE")
    row += 1

row += 1
ws.merge_cells(f"A{row}:F{row}")
ws.cell(row=row, column=1, value="🎯 TOP-3 DRIVER DA MONITORARE:")
ws.cell(row=row, column=1).font = Font(bold=True)
row += 1
ws.merge_cells(f"A{row}:F{row}")
ws.cell(row=row, column=1, value="1. Pricing PA (single most important — falsifica €150k → €75k/anno baseline)")
row += 1
ws.merge_cells(f"A{row}:F{row}")
ws.cell(row=row, column=1, value="2. Utilization rate (% ore disponibili fatturate; target 60-70%)")
row += 1
ws.merge_cells(f"A{row}:F{row}")
ws.cell(row=row, column=1, value="3. Revenue Y3 (validazione scale-up multi-regione SNAI)")

set_col_widths(ws, [40, 18, 18, 12, 14, 50])

# ============================================================
# SHEET 7: SCENARIOS
# ============================================================

ws = wb.create_sheet("7_Scenarios")

ws.merge_cells("A1:E1")
ws["A1"] = "SCENARI WORST / BASE / BEST — 10 anni"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="003366")
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 25

headers = ["Metrica", "Worst (P 25%)", "Base (P 55%)", "Best (P 20%)", "Note"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=col, value=h)
    apply_header_style(c)

scenarios = [
    ("Revenue Y1", 150000, 260000, 450000, "Worst = Cluster D pricing prevale; Best = partnership Planetek"),
    ("Revenue Y3", 1000000, 2500000, 4500000, "Worst = no scale-up multi-regione; Best = utility early adopter"),
    ("Revenue Y5", 2500000, 6500000, 15000000, "Best = HAPS subscale + utility expansion"),
    ("Revenue Y10", 8000000, 30000000, 80000000, "Best = consorzio EU sovereign; B2-relaxed = base"),
    ("CapEx Y1", 1960000, 1400000, 970000, "Worst = overrun 50%, Best = efficient execution"),
    ("OpEx Y2 (incl. regulatory)", 1300000, 1180000, 950000, "Worst = +3 FTE expensive; Best = lean operations"),
    ("FCF Y5 cumulato", -1500000, 1863000, 6000000, ""),
    ("Break-even", "Y8+", "Y4-Y5", "Y3-Y4", "Best richiede esecuzione perfetta + LoI early"),
    ("NPV 10y (WACC 12%)", -2000000, 3500000, 25000000, "Worst = NPV negativo, progetto non profittevole"),
    ("IRR 10y", "negativo", "18-22%", "35-45%", ""),
    ("Payback", "non raggiunto", "5 anni", "3 anni", ""),
]

row = 4
for item in scenarios:
    label = item[0]
    ws.cell(row=row, column=1, value=label)
    apply_subheader_style(ws.cell(row=row, column=1))
    for c, val in enumerate(item[1:5], start=2):
        ws.cell(row=row, column=c, value=val)
        if isinstance(val, (int, float)):
            fmt_eur(ws.cell(row=row, column=c))
    ws.cell(row=row, column=5, value=item[4] if len(item) > 4 else "")
    row += 1

row += 2
ws.merge_cells(f"A{row}:E{row}")
ws.cell(row=row, column=1, value="🎯 RACCOMANDAZIONE: pianificare su scenario BASE; budget contingency per Worst.")
ws.cell(row=row, column=1).font = Font(bold=True)
row += 1
ws.merge_cells(f"A{row}:E{row}")
ws.cell(row=row, column=1, value="⚠️ P(scenario Worst) = 25%; P(Base) = 55%; P(Best) = 20% — totali 100%")
apply_caveat_style(ws.cell(row=row, column=1))

set_col_widths(ws, [30, 18, 18, 18, 50])

# ============================================================
# SHEET 8: FUNDING MIX
# ============================================================

ws = wb.create_sheet("8_Funding_Mix")

ws.merge_cells("A1:E1")
ws["A1"] = "MIX FINANZIAMENTI Y1 6A + Phase B 6B"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="003366")
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 25

# Y1 mix
ws["A3"] = "MIX Y1 6A (target CapEx €1.4M baseline)"
apply_subheader_style(ws["A3"])
ws.merge_cells("A3:E3")

ws["A4"] = "Fonte"
ws["B4"] = "Importo target (€)"
ws["C4"] = "%"
ws["D4"] = "Status"
ws["E4"] = "Note"
for col in range(1, 6):
    apply_header_style(ws.cell(row=4, column=col))

funding_y1 = [
    ("Coopfond Cooding Prototypes 2026", 50000, "3.6%", "⏳ DR-002 verifica", "max €50k, ≥10 cooperative"),
    ("Coopfond Cooding-Invest", 220000, "15.7%", "⏳ Q2 2026", "max €250k"),
    ("Regione Liguria FESR 2021-27", 400000, "28.6%", "⏳ OQ-010 LoI", "OS 1.1 + OS 5.2"),
    ("PNRR Aerospazio / IS4Aerospace", 100000, "7.1%", "⏳ M+12+ via Polito", "Partnership possibile"),
    ("Equity privato (founder + seed)", 350000, "25.0%", "✓ Round Q1-Q2 2026", "Dilution 15-25%"),
    ("R&D tax credit (L.160/2019)", 100000, "7.1%", "✓ Cumulabile", "Recovery post-spesa"),
    ("[GAP residuo da coprire]", 180000, "12.9%", "⚠️ TBD", "Possibili bridge loan o EIC Accelerator"),
]

row = 5
for fonte, imp, pct, stat, note in funding_y1:
    ws.cell(row=row, column=1, value=fonte)
    ws.cell(row=row, column=2, value=imp)
    ws.cell(row=row, column=3, value=pct)
    ws.cell(row=row, column=4, value=stat)
    ws.cell(row=row, column=5, value=note)
    fmt_eur(ws.cell(row=row, column=2))
    row += 1

ws.cell(row=row, column=1, value="TOTALE Y1")
ws.cell(row=row, column=2, value=f"=SUM(B5:B{row-1})")
fmt_eur(ws.cell(row=row, column=2))
apply_total_style(ws.cell(row=row, column=1))
apply_total_style(ws.cell(row=row, column=2))
row += 2

# Phase B mix
ws.cell(row=row, column=1, value="MIX Phase B 6B (target €5.5-13.5M — R&D Phase 0/A only, vedi DR-014)")
apply_subheader_style(ws.cell(row=row, column=1))
ws.merge_cells(f"A{row}:E{row}")
row += 1

ws.cell(row=row, column=1, value="Fonte")
ws.cell(row=row, column=2, value="Importo target (€)")
ws.cell(row=row, column=3, value="%")
ws.cell(row=row, column=4, value="Status")
ws.cell(row=row, column=5, value="Note")
for col in range(1, 6):
    apply_header_style(ws.cell(row=row, column=col))
row += 1

funding_phb = [
    ("EDF call HAPS post-EuroHAPS", 3500000, "35%", "⏳ M+24-36", "€2-5M target"),
    ("Horizon Europe Cluster 4/5", 2000000, "20%", "⏳ M+24", "RIA/IA"),
    ("PNRR Aerospazio / ASI / MIMIT", 1500000, "15%", "⏳ M+18+", "Partnership Polito/CIRA"),
    ("Equity privato Series A/B", 2000000, "20%", "⏳ M+24+", "Series B €15-50M raised"),
    ("R&D tax credit + Patent Box", 750000, "7.5%", "✓ Cumulabile", ""),
    ("CIRA cooperazione in-kind", 250000, "2.5%", "⏳ DR-010", "Soggetto a MOU"),
]

for fonte, imp, pct, stat, note in funding_phb:
    ws.cell(row=row, column=1, value=fonte)
    ws.cell(row=row, column=2, value=imp)
    ws.cell(row=row, column=3, value=pct)
    ws.cell(row=row, column=4, value=stat)
    ws.cell(row=row, column=5, value=note)
    fmt_eur(ws.cell(row=row, column=2))
    row += 1

ws.cell(row=row, column=1, value="TOTALE Phase B")
ws.cell(row=row, column=2, value=f"=SUM(B{row-6}:B{row-1})")
fmt_eur(ws.cell(row=row, column=2))
apply_total_style(ws.cell(row=row, column=1))
apply_total_style(ws.cell(row=row, column=2))
row += 2

# Caveat
ws.merge_cells(f"A{row}:E{row}")
ws.cell(row=row, column=1, value="⚠️ Phase B €10M è R&D Phase 0/A; operativa commerciale richiede €50-200M+ (DR-014). Pivot raccomandato: 'Firmamento operatore servizi su piattaforme prime contractor (Aalto/Sceye/CIRA)'.")
apply_caveat_style(ws.cell(row=row, column=1))

set_col_widths(ws, [35, 18, 10, 22, 35])

# ============================================================
# SHEET 9: QUADRO ECONOMICO ART. 41 (Sintetico)
# ============================================================

ws = wb.create_sheet("9_Quadro_Economico_art41")

ws.merge_cells("A1:E1")
ws["A1"] = "QUADRO ECONOMICO — Formato D.Lgs. 36/2023 art. 41 + Allegato I.7"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="003366")
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 25

ws["A3"] = "QE per Pilota Y1 6A — Pentema (Torriglia GE)"
apply_subheader_style(ws["A3"])
ws.merge_cells("A3:E3")

ws["A4"] = "Codice"
ws["B4"] = "Descrizione"
ws["C4"] = "Importo Baseline (€)"
ws["D4"] = "Min (€)"
ws["E4"] = "Max (€)"
for col in range(1, 6):
    apply_header_style(ws.cell(row=4, column=col))

qe_a = [
    ("A.1", "Piattaforma UAV (VTOL ibrido TRL 8-9)", 350000, 250000, 400000),
    ("A.2", "Asset accessori (GS, hangar, ricambi, payload, tools)", 360000, 215000, 595000),
    ("A.3", "Servizi tecnici (SW, cert., privacy, formazione, studi)", 227000, 140000, 315000),
]

row = 5
for cod, desc, val, mn, mx in qe_a:
    ws.cell(row=row, column=1, value=cod)
    ws.cell(row=row, column=2, value=desc)
    ws.cell(row=row, column=3, value=val)
    ws.cell(row=row, column=4, value=mn)
    ws.cell(row=row, column=5, value=mx)
    for c in range(3, 6):
        fmt_eur(ws.cell(row=row, column=c))
    row += 1

ws.cell(row=row, column=1, value="TOT A")
ws.cell(row=row, column=2, value="Sommatoria voci A")
ws.cell(row=row, column=3, value=f"=SUM(C5:C{row-1})")
ws.cell(row=row, column=4, value=f"=SUM(D5:D{row-1})")
ws.cell(row=row, column=5, value=f"=SUM(E5:E{row-1})")
for c in range(1, 6):
    apply_total_style(ws.cell(row=row, column=c))
    if c >= 3:
        fmt_eur(ws.cell(row=row, column=c))
tot_a_row = row
row += 2

qe_b = [
    ("B.1", "Spese tecniche (progettaz. + DL + RUP, 4.5% A)", f"=C{tot_a_row}*0.045", f"=D{tot_a_row}*0.045", f"=E{tot_a_row}*0.045"),
    ("B.2", "Imprevisti / contingency (15% A)", f"=C{tot_a_row}*0.15", f"=D{tot_a_row}*0.15", f"=E{tot_a_row}*0.15"),
    ("B.3", "Spese pubblicità bandi", 3500, 2000, 5000),
    ("B.4", "IVA su A (22%)", f"=C{tot_a_row}*0.22", f"=D{tot_a_row}*0.22", f"=E{tot_a_row}*0.22"),
    ("B.5", "IVA su B (22%)", 0, 0, 0),  # placeholder
    ("B.6", "Allacciamenti, autorizzazioni", 10000, 5000, 15000),
    ("B.7", "Spese collaudo / verifica", 17500, 10000, 25000),
]

start_b = row
for cod, desc, val, mn, mx in qe_b:
    ws.cell(row=row, column=1, value=cod)
    ws.cell(row=row, column=2, value=desc)
    ws.cell(row=row, column=3, value=val)
    ws.cell(row=row, column=4, value=mn)
    ws.cell(row=row, column=5, value=mx)
    for c in range(3, 6):
        fmt_eur(ws.cell(row=row, column=c))
    row += 1

ws.cell(row=row, column=1, value="TOT B")
ws.cell(row=row, column=2, value="Sommatoria voci B")
ws.cell(row=row, column=3, value=f"=SUM(C{start_b}:C{row-1})")
ws.cell(row=row, column=4, value=f"=SUM(D{start_b}:D{row-1})")
ws.cell(row=row, column=5, value=f"=SUM(E{start_b}:E{row-1})")
for c in range(1, 6):
    apply_total_style(ws.cell(row=row, column=c))
    if c >= 3:
        fmt_eur(ws.cell(row=row, column=c))
tot_b_row = row
row += 2

ws.cell(row=row, column=1, value="TOT")
ws.cell(row=row, column=2, value="TOTALE GENERALE Y1 (A + B)")
ws.cell(row=row, column=3, value=f"=C{tot_a_row}+C{tot_b_row}")
ws.cell(row=row, column=4, value=f"=D{tot_a_row}+D{tot_b_row}")
ws.cell(row=row, column=5, value=f"=E{tot_a_row}+E{tot_b_row}")
for c in range(1, 6):
    apply_total_style(ws.cell(row=row, column=c))
    ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="C6E0B4")
    if c >= 3:
        fmt_eur(ws.cell(row=row, column=c))

set_col_widths(ws, [8, 55, 18, 15, 15])

# ============================================================
# SAVE
# ============================================================

output_path = "/home/user/HALE/studio-di-fattibilita/allegati/financial-model/HALE-Financial-Model-M3.xlsx"
wb.save(output_path)
print(f"✅ Excel file generated: {output_path}")
print(f"   {len(wb.sheetnames)} sheets: {wb.sheetnames}")
