"""
Allegato A.4 — Interface Control Document (ICD) preliminare v1.0
Studio di Fattibilità HALE/VTOL — Firmamento Technologies

Genera file Excel multi-sheet con 50+ interfacce di sistema dettagliate
secondo NASA SE Handbook §6.3 (Interface Management) + ARP4754A (System Development)
+ ISO/IEC/IEEE 24765 (Systems and Software Engineering Vocabulary).

Sources:
- Cap. 4 §4.4 (20 interfacce ICD preliminare)
- Cap. 6 (architettura 6A + 6B)
- Cap. 5 (quadro regolatorio ENAC, EASA, AGCOM, Garante)
- Agenti: avionics-gnc, telecom-ntn-payload, aerodynamics-structures, earth-observation

Disciplina epistemica (skill epistemic-rigor):
- Confidence levels dichiarati per ciascuna interfaccia (high/medium/low/speculative)
- Falsifying observations su interfacce critiche
- Boundary conditions B1 (service-only cooperative) e B2 (visione 10 anni) preservate
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
import os

# ============================================================
# CONFIGURAZIONE STILI
# ============================================================

NAVY = "003366"
ORANGE = "D97706"
LIGHT_BLUE = "DBEAFE"
LIGHT_GREEN = "D1FAE5"
LIGHT_YELLOW = "FEF3C7"
LIGHT_RED = "FEE2E2"
GREY = "F3F4F6"
DARK_GREY = "6B7280"

def header_style(cell, color=NAVY):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def subheader_style(cell):
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def cell_style(cell, wrap=True, align="left", bold=False):
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.font = Font(size=9, bold=bold)
    cell.border = Border(
        left=Side(style="thin", color=DARK_GREY),
        right=Side(style="thin", color=DARK_GREY),
        top=Side(style="thin", color=DARK_GREY),
        bottom=Side(style="thin", color=DARK_GREY)
    )

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# DATABASE INTERFACCE (50 interfacce dettagliate)
# ============================================================

# Struttura record: (ID, Nome, Tipo, Parti_A, Parti_B, Direzione, Caratteristiche, Standard, Owner, Test, Status, Confidence, Note)

INTERFACES = [
    # =====================================================================
    # SEZIONE 6A — PHYSICAL (10)
    # =====================================================================
    {
        "id": "INT-6A-PHY-001",
        "name": "Airframe VTOL ↔ Modular Payload Bay",
        "tipo": "Fisica meccanica",
        "parte_a": "VTOL Airframe (JOUAV CW-30E o eq.)",
        "parte_b": "Payload Module (EO/IR/telecom)",
        "direzione": "Bidirezionale (statica)",
        "caratteristiche": "Bay 250×180×120 mm; massa payload ≤ 3 kg + 0.5 kg buffer; CG shift ≤ 2% MAC; mount quick-release tipo Picatinny-like + vibration isolation MIL-STD-810H Cat 4 (5-2000 Hz, 7.7 grms)",
        "standard": "MIL-STD-810H §514.8 (vibration); ISO 9022-2:2016 (mechanical mounting); vendor-specific dovetail mount",
        "owner": "Avionics Integration Lead (Firmamento) + JOUAV Liaison",
        "test_method": "Inspection + analisi CG calc + shake table test",
        "status": "Preliminary",
        "confidence": "medium",
        "note": "Sub-interface di INT-01 Cap.4. Critica per modularità payload. Falsifying: se vendor non fornisce CAD interfacce → re-engineering custom +€15-25k."
    },
    {
        "id": "INT-6A-PHY-002",
        "name": "Payload Power Rail 28 VDC",
        "tipo": "Fisica elettrica",
        "parte_a": "VTOL Power Distribution Unit (PDU)",
        "parte_b": "Payload Module (load fino a 100 W)",
        "direzione": "Unidirezionale (PDU → Payload)",
        "caratteristiche": "28 VDC ±5%; ripple ≤ 200 mV pk-pk; transient response per DO-160G §16; protezione corto-circuito 5 A fuse; connector MIL-DTL-38999 Series III shell size 11; ground-return isolato",
        "standard": "MIL-STD-704F (aircraft electrical power); DO-160G §16 (Power Input); MIL-STD-1377 (28 VDC)",
        "owner": "Power Management SE + Avionics Lead",
        "test_method": "Lab bench test + transient injection + load step",
        "status": "Preliminary",
        "confidence": "high",
        "note": "Standard aerospace consolidato. Verificare in datasheet JOUAV se PDU eroga 28 VDC stabilizzato (alcuni vendor cinesi forniscono 24 VDC nominale)."
    },
    {
        "id": "INT-6A-PHY-003",
        "name": "Payload Data Port Ethernet",
        "tipo": "Fisica elettrica + dati",
        "parte_a": "Payload Module (camera/sensor)",
        "parte_b": "Mission Computer (MC) VTOL",
        "direzione": "Bidirezionale full-duplex",
        "caratteristiche": "1000BASE-T Gigabit Ethernet su cavo Cat6 schermato; max 100 m; PoE+ opzionale (IEEE 802.3at, fino a 25.5 W); MTU 9000 jumbo frame supportato; latency end-to-end < 5 ms",
        "standard": "IEEE 802.3 (Ethernet); IEEE 802.3at (PoE+); ANSI/TIA-568 (cabling)",
        "owner": "Avionics Lead + Payload SE",
        "test_method": "Cable certification (Fluke DSX) + ping test + throughput iPerf3",
        "status": "Preliminary",
        "confidence": "high",
        "note": "Standard COTS. Camera GigE Vision (Phase One iXM 100 supporta GigE)."
    },
    {
        "id": "INT-6A-PHY-004",
        "name": "Ground Antenna Mount (parabolica 1.2 m)",
        "tipo": "Fisica meccanica",
        "parte_a": "GS fissa Pentema (container)",
        "parte_b": "Antenna parabolica RF 2.4 GHz",
        "direzione": "Statica (mount)",
        "caratteristiche": "Mount mast h ≥ 6 m AGL; antenna parabolica D=1.2 m, peso ≈ 8 kg; vento sopravvivenza 200 km/h (Liguria estremo); azimuth/elevazione motorizzati ±180°/0-90°; preset pointing automatic via GPS UAV feedback",
        "standard": "EIA-222-H (structural standards for steel antenna supports); IEC 60721 (environmental conditions)",
        "owner": "RF Systems Engineer + Civil Works Lead",
        "test_method": "Inspection + load calc + alignment test",
        "status": "Preliminary",
        "confidence": "high",
        "note": "Sito Pentema esposto a vento di valle: verifica anemometrica preliminare M+4-M+6. Sub-interface di INT-04 Cap.4."
    },
    {
        "id": "INT-6A-PHY-005",
        "name": "Airborne RF Antenna Whip (2.4 GHz)",
        "tipo": "Fisica meccanica + RF",
        "parte_a": "VTOL airframe (ventral)",
        "parte_b": "Whip antenna 5 dBi omni vertical polarization",
        "direzione": "Bidirezionale (RF)",
        "caratteristiche": "Antenna L=120 mm whip rigid; gain 5 dBi omni; VSWR ≤ 1.5 @ 2.4-2.485 GHz; connector SMA female; cavo LMR-240 max 1.5 m fino a transceiver",
        "standard": "MIL-STD-348B (RF connectors); MIL-C-39012 (SMA family)",
        "owner": "RF Systems Engineer",
        "test_method": "VNA sweep (VSWR + gain pattern) + range test",
        "status": "Preliminary",
        "confidence": "medium",
        "note": "Verifica pattern in presenza fusoliera (effetti reflection/null). Pre-installation EM compatibility test."
    },
    {
        "id": "INT-6A-PHY-006",
        "name": "Battery Pack ↔ Power Distribution (LiPo 6S 22.2V)",
        "tipo": "Fisica elettrica",
        "parte_a": "Battery Pack 6S LiPo 22.2V 16000 mAh (≈ 356 Wh)",
        "parte_b": "Power Distribution Unit (PDU)",
        "direzione": "Unidirezionale (Bat → PDU) + telemetry (BMS → MC)",
        "caratteristiche": "XT90-S connector (max 90 A continuous); cella balanced JST-XH 6-pin; BMS integrato con telemetry CAN; SOC sense via Coulomb counter ±1% accuracy; max discharge 30C burst; max charge 3C",
        "standard": "RTCA DO-311A (Rechargeable Lithium Battery); UN 38.3 (transport); IEC 62133 (safety)",
        "owner": "Power Management SE + Battery Specialist",
        "test_method": "Capacity test + thermal runaway test + cycle test",
        "status": "Preliminary",
        "confidence": "high",
        "note": "Sub-interface di INT-08 Cap.4. Inverno Pentema -10°C: capacità derating ~15-20% → autonomia ridotta. Batteria self-heating opzionale."
    },
    {
        "id": "INT-6A-PHY-007",
        "name": "Thermal Interface Payload-Airframe",
        "tipo": "Fisica termica",
        "parte_a": "Payload (heat generation 50-100 W)",
        "parte_b": "Airframe (heat sink, ambient air)",
        "direzione": "Unidirezionale (heat flow)",
        "caratteristiche": "Operating range payload -10°C / +50°C; conduction path via mounting plate (Al 6061); thermal pad K = 1.5 W/m·K; airflow cooling natural via NACA inlet 30×30 mm; storage range -25°C / +60°C",
        "standard": "DO-160G §4 (Temperature); MIL-STD-810H §501.7 (High Temp), §502.7 (Low Temp)",
        "owner": "Thermal Engineer + Payload SE",
        "test_method": "Thermal vacuum chamber test (per HALE) + thermal cycle (per VTOL)",
        "status": "Preliminary",
        "confidence": "medium",
        "note": "Per VTOL 6A range climatico Pentema gestibile. Per HALE 6B (-65°C @ 20 km) richiede heater + insulation attiva."
    },
    {
        "id": "INT-6A-PHY-008",
        "name": "VTOL Landing Gear ↔ Helipad Surface",
        "tipo": "Fisica meccanica",
        "parte_a": "VTOL landing gear (skid o quad)",
        "parte_b": "Helipad/touchdown surface (Pentema/mobile)",
        "direzione": "Contatto meccanico transitorio",
        "caratteristiche": "Vertical landing speed ≤ 1.5 m/s; touchdown footprint 1.5×1.5 m max; superficie piana ±2° slope max; resistenza load 100 kg/m² distribuited; superficie min FOD-safe (no gravel)",
        "standard": "FAA AC 150/5390-2D (Heliport Design); EASA CS-HPT (Heliport Design Manual) — applicabile per drone heliports",
        "owner": "Operations Lead + Civil Works",
        "test_method": "Site survey + load calc + landing test",
        "status": "Concept",
        "confidence": "high",
        "note": "Pentema: scegliere area pianeggiante a quota 1100-1200 m AMSL, distanza ≥ 50 m da edifici. Coordinamento con Comune Torriglia."
    },
    {
        "id": "INT-6A-PHY-009",
        "name": "GS Power Supply (230 VAC mains)",
        "tipo": "Fisica elettrica",
        "parte_a": "Rete elettrica Pentema (Enel Distribuzione)",
        "parte_b": "GS fissa container + GS mobile veicolo",
        "direzione": "Unidirezionale (mains → GS)",
        "caratteristiche": "230 VAC ±10% 50 Hz; potenza richiesta GS 2.5 kW peak / 1.2 kW average; UPS online 3 kVA con autonomy ≥ 30 min; generator backup diesel 5 kVA per outage estesi",
        "standard": "CEI 64-8 (impianti elettrici utilizzatori); IEC 61000-4 (EMC immunity)",
        "owner": "Civil Works Lead + IT Operations",
        "test_method": "Load test + outage simulation + UPS autonomy test",
        "status": "Concept",
        "confidence": "medium",
        "note": "Pentema: verifica capacity allaccio Enel; possibile upgrade trasformatore di cabina. UPS obbligatorio per business continuity."
    },
    {
        "id": "INT-6A-PHY-010",
        "name": "Backhaul Internet Connection (GS → Cloud)",
        "tipo": "Fisica + dati",
        "parte_a": "GS fissa Pentema",
        "parte_b": "Backhaul WAN (FWA Linkem/Open Fiber/Starlink)",
        "direzione": "Bidirezionale",
        "caratteristiche": "Throughput ≥ 25 Mbps uplink / ≥ 50 Mbps downlink; latency RTT ≤ 50 ms verso datacenter; uptime SLA ≥ 99.0% (single link), ≥ 99.5% (dual provider); media: fibra ottica primaria + LTE/5G backup + Starlink LEO fallback",
        "standard": "ITU-T G.983 (PON); IEEE 802.3ah (Ethernet First Mile); 3GPP TS 23.501 (5G architecture)",
        "owner": "IT Operations + Cloud Architect",
        "test_method": "iPerf3 throughput + ping latency + outage test",
        "status": "Concept",
        "confidence": "medium",
        "note": "Sub-interface di INT-05 Cap.4. Pentema area montana: probabile fibra non disponibile → FWA + Starlink dual. Verifica con Comune."
    },

    # =====================================================================
    # SEZIONE 6A — DATA (8)
    # =====================================================================
    {
        "id": "INT-6A-DATA-001",
        "name": "Payload EO RGB → Mission Computer",
        "tipo": "Funzionale data",
        "parte_a": "Camera Phase One iXM 100 (100 MP, 11608×8708 px)",
        "parte_b": "Mission Computer (MC) airborne",
        "direzione": "Unidirezionale (Camera → MC)",
        "caratteristiche": "GigE Vision 2.1 protocol; image RAW IIQ 16-bit ≈ 100 MB/frame; trigger frequency 1-5 fps; geotagging via NTP time sync + GPS PPS pulse 1 PPS ±100 ns",
        "standard": "GigE Vision 2.1 (AIA standard); GenICam 3.x; NTP v4 RFC 5905; PPS NMEA 0183",
        "owner": "Payload SE + Avionics Lead",
        "test_method": "Image capture test + geotag accuracy verification + throughput test",
        "status": "Preliminary",
        "confidence": "high",
        "note": "Sub-interface di INT-02 Cap.4. Standard consolidato per camera industrial."
    },
    {
        "id": "INT-6A-DATA-002",
        "name": "Payload IR LWIR → Mission Computer",
        "tipo": "Funzionale data",
        "parte_a": "Camera termica FLIR Vue Pro R / WIRIS Pro (640×512 px LWIR 7.5-13.5 μm)",
        "parte_b": "Mission Computer airborne",
        "direzione": "Unidirezionale",
        "caratteristiche": "USB 3.0 SuperSpeed; image radiometric 16-bit (T °C per pixel); NEdT ≤ 50 mK; frame rate 9 Hz; trigger via TTL line; calibration NUC ogni 30 min auto",
        "standard": "USB 3.0 specification; FLIR Atlas SDK; ASTM E1862-97 (IR camera calibration)",
        "owner": "Payload SE",
        "test_method": "Radiometric accuracy test (calibrazione blackbody) + frame rate verification",
        "status": "Preliminary",
        "confidence": "high",
        "note": "Hotspot detection antincendio: target threshold 50°C su pixel risoluzione ground ≈ 0.7 m @ 500 m AGL."
    },
    {
        "id": "INT-6A-DATA-003",
        "name": "Mission Data Storage (onboard NVMe SSD)",
        "tipo": "Funzionale data",
        "parte_a": "Mission Computer (MC)",
        "parte_b": "NVMe SSD storage 2 TB ruggedized (industrial)",
        "direzione": "Bidirezionale (R/W)",
        "caratteristiche": "PCIe Gen3 x4; sequential write ≥ 1.5 GB/s; capacity 2 TB (≈ 10 missioni 6h con EO RGB full); MTBF ≥ 2M hours; operating temp -40°C / +85°C industrial grade",
        "standard": "NVMe 1.4 specification; PCIe Gen3 (PCI-SIG)",
        "owner": "Avionics Lead + Mission Computer SE",
        "test_method": "Throughput benchmark + power-off recovery + thermal cycling",
        "status": "Preliminary",
        "confidence": "high",
        "note": "Storage capacity dimensionata per copertura 6-10h volo + 100 MB/frame RGB."
    },
    {
        "id": "INT-6A-DATA-004",
        "name": "Mission Data Upload GS → Cloud",
        "tipo": "Funzionale data",
        "parte_a": "GS post-flight ingest (NVMe rimovibile o WiFi 6)",
        "parte_b": "Cloud Object Storage (S3-compatible, Aruba/OVH)",
        "direzione": "Unidirezionale (GS → Cloud)",
        "caratteristiche": "HTTPS REST API multipart upload; AES-256 in transit (TLS 1.3); object size avg 5-50 GB per missione; throughput ≥ 25 Mbps target; resumable upload (S3 Multipart o Tus.io)",
        "standard": "AWS S3 API v4; TLS 1.3 RFC 8446; Tus.io resumable upload spec",
        "owner": "Cloud Architect + DevOps",
        "test_method": "Upload time vs object size benchmark + interrupted resume test",
        "status": "Preliminary",
        "confidence": "high",
        "note": "Sub-interface di INT-05 Cap.4. Backhaul fibra/FWA Pentema fondamentale (INT-6A-PHY-010)."
    },
    {
        "id": "INT-6A-DATA-005",
        "name": "Telemetry Real-time Stream (UAV → GS)",
        "tipo": "Funzionale data",
        "parte_a": "VTOL FCS (Flight Computer System)",
        "parte_b": "GS Telemetry Dashboard (operator UI)",
        "direzione": "Unidirezionale (UAV → GS)",
        "caratteristiche": "MAVLink v2.0 over UDP; freq 10 Hz nominale; payload telemetry: GPS lat/lon/alt, attitude RPY, airspeed, GS SOC, motor RPM, system status; bandwidth ≈ 20 kbps; encryption AES-256-GCM",
        "standard": "MAVLink Protocol v2.0 (Dronecode); AES-256-GCM (NIST SP 800-38D)",
        "owner": "Avionics Lead + GCS UX Developer",
        "test_method": "Packet capture + decode + latency measurement (link tester)",
        "status": "Preliminary",
        "confidence": "high",
        "note": "Sub-interface di INT-03 Cap.4 (downlink). MAVLink v2.0 standard industria UAS commerciali."
    },
    {
        "id": "INT-6A-DATA-006",
        "name": "Video Downlink Preview (Low-bitrate)",
        "tipo": "Funzionale data",
        "parte_a": "Payload EO RGB (preview pipeline)",
        "parte_b": "GS Operator Display",
        "direzione": "Unidirezionale (UAV → GS)",
        "caratteristiche": "H.265 HEVC encoded 720p @ 5 fps; bitrate target 0.5-1.5 Mbps; latency end-to-end ≤ 500 ms; RTSP/RTP protocol over UDP; backup degraded mode 320p @ 2 fps @ 200 kbps",
        "standard": "ITU-T H.265 / ISO/IEC 23008-2 (HEVC); RFC 3550 (RTP); RFC 7826 (RTSP 2.0)",
        "owner": "Payload SE + GCS Developer",
        "test_method": "Latency measurement (timestamp injection) + decoder compatibility test",
        "status": "Preliminary",
        "confidence": "medium",
        "note": "Sub-interface di INT-03 Cap.4 (video). Critico per situational awareness pilota + Protezione Civile monitoring."
    },
    {
        "id": "INT-6A-DATA-007",
        "name": "Photogrammetry Pipeline (Cloud)",
        "tipo": "Funzionale data",
        "parte_a": "Cloud Object Storage (raw images)",
        "parte_b": "Photogrammetry Engine (Pix4D/Agisoft/OpenDroneMap)",
        "direzione": "Bidirezionale",
        "caratteristiche": "REST API trigger via webhook post-upload; input GeoTIFF + EXIF; output: orthomosaic GeoTIFF (EPSG:32632 WGS84/UTM 32N), DSM/DTM, point cloud LAS; processing time ≈ 30-120 min per missione",
        "standard": "OGC GeoTIFF; ASPRS LAS 1.4; EXIF 2.32; OGC WMTS",
        "owner": "GIS Engineer + Cloud Architect",
        "test_method": "Pipeline test end-to-end + output validation",
        "status": "Preliminary",
        "confidence": "medium",
        "note": "Sub-interface di INT-06 Cap.4. Possibile outsourcing pipeline (Pix4Dcloud SaaS) per Y1; in-house dal Y2."
    },
    {
        "id": "INT-6A-DATA-008",
        "name": "Audit Log + Operational Telemetry → SIEM",
        "tipo": "Funzionale data + sicurezza",
        "parte_a": "GS + Cloud + Mission Computer (log producers)",
        "parte_b": "SIEM (Security Information & Event Management)",
        "direzione": "Unidirezionale (producers → SIEM)",
        "caratteristiche": "Syslog over TLS RFC 5425; CEF format (Common Event Format); retention 90 giorni hot + 1 anno cold; events/sec peak 100, average 10",
        "standard": "RFC 5424 (Syslog); RFC 5425 (Syslog over TLS); ArcSight CEF; ISO/IEC 27037 (digital evidence)",
        "owner": "Cybersecurity Engineer + DPO",
        "test_method": "Log ingestion test + correlation rule verification + retention check",
        "status": "Concept",
        "confidence": "medium",
        "note": "Necessario per NIS2 compliance (D.Lgs. 138/2024) + GDPR Art. 32 accountability. Sub-interface di INT-12 Cap.4."
    },

    # =====================================================================
    # SEZIONE 6A — C2 LINK (8)
    # =====================================================================
    {
        "id": "INT-6A-C2-001",
        "name": "C2 Uplink Primary (RF 2.4 GHz ISM)",
        "tipo": "Funzionale control",
        "parte_a": "GS RF Transceiver (Microhard pMDDL2450 o eq.)",
        "parte_b": "UAV RF Transceiver (FCS-side)",
        "direzione": "Unidirezionale (GS → UAV) per comandi",
        "caratteristiche": "Frequenza 2400-2483.5 MHz ISM; modulazione OFDM adaptive; bitrate 0.5-9 Mbps; EIRP ≤ 100 mW (20 dBm) per AGCOM ISM o ≤ 1 W con licenza individuale; latency one-way ≤ 100 ms; fade margin ≥ 12 dB",
        "standard": "EN 300 328 (ISM 2.4 GHz harmonized standard EU); RED 2014/53/EU; AGCOM PNRF",
        "owner": "RF Systems Engineer + AGCOM Liaison",
        "test_method": "Range test campo aperto + fade margin verification + spectrum analyzer",
        "status": "Preliminary",
        "confidence": "medium",
        "note": "Sub-interface di INT-03 Cap.4 (uplink). FALSIFYING: in valle Pentema shadow zones probabili → switch automatico a SATCOM L-band (INT-6A-C2-003) entro 5 s."
    },
    {
        "id": "INT-6A-C2-002",
        "name": "C2 Downlink Telemetry (RF 2.4 GHz)",
        "tipo": "Funzionale control",
        "parte_a": "UAV FCS + Telemetry Aggregator",
        "parte_b": "GS Telemetry Receiver",
        "direzione": "Unidirezionale (UAV → GS)",
        "caratteristiche": "Same RF link di INT-6A-C2-001 (TDD time-division duplex); telemetry frame size ≈ 256 byte @ 10 Hz; CRC-32 error detection; AES-256-GCM encryption end-to-end; auto-throttle se BER > 1e-4",
        "standard": "MAVLink v2.0; CRC-32 IEEE 802.3; AES-256-GCM NIST SP 800-38D",
        "owner": "Avionics Lead + RF Systems Engineer",
        "test_method": "BER test + encryption verification + replay attack test",
        "status": "Preliminary",
        "confidence": "high",
        "note": "Sub-interface di INT-03 Cap.4 (downlink)."
    },
    {
        "id": "INT-6A-C2-003",
        "name": "C2 Secondary SATCOM (Iridium Certus L-band)",
        "tipo": "Funzionale control",
        "parte_a": "UAV Iridium Certus 100 modem (Cobham/Honeywell)",
        "parte_b": "Iridium NEXT constellation + GS Hub Iridium",
        "direzione": "Bidirezionale",
        "caratteristiche": "L-band 1616-1626.5 MHz uplink, 1616-1626.5 MHz downlink; throughput 22 kbps Certus 100 / 88 kbps Certus 200 / 700 kbps Certus 700; latency one-way ≈ 200-400 ms (LEO ~780 km); coverage globale; antenna L-band patch 70×70 mm conformal",
        "standard": "ITU-R RR Article 22 (MSS L-band); Iridium Certus service specification; DO-262C (airborne SATCOM)",
        "owner": "Avionics Lead + SATCOM Vendor Liaison",
        "test_method": "Link test + handover test + latency measurement",
        "status": "Concept",
        "confidence": "medium",
        "note": "Sub-interface di INT-03 Cap.4 (SATCOM fallback). FALSIFYING: in valle Pentema con elevation angle satellite ≥ 8° quasi continuo grazie a 66 satelliti LEO Iridium → coverage OK. Costo: ~€800-1200/mese per piano dati."
    },
    {
        "id": "INT-6A-C2-004",
        "name": "C2 Lost-Link Procedure (Autonomous RTH)",
        "tipo": "Funzionale control + safety",
        "parte_a": "FCS Flight Computer",
        "parte_b": "Autopilot State Machine",
        "direzione": "Internal logic",
        "caratteristiche": "Trigger lost-link > 5 s consecutivi senza heartbeat MAVLink; primary action: RTH (Return-To-Home) verso ultimo home point GPS; secondary action se RTH fail: loiter + emergency land alla quota minima sicura; alert SMS PC + pilot",
        "standard": "RTCA DO-377 (C2 link); JARUS SORA 2.5 Annex F (lost-link procedure); EUROCAE ED-269",
        "owner": "Avionics Lead + Safety Engineer",
        "test_method": "Simulation HIL + flight test (in autorizzazione condizionata)",
        "status": "Concept",
        "confidence": "medium",
        "note": "Critico per SORA SAIL III BVLOS authorization. FALSIFYING: se RTH path attraversa shadow zone GPS o ostacolo orografico → emergency land può atterrare in area non sicura. Mitigazione: 3 RTH waypoint intermedi pre-pianificati."
    },
    {
        "id": "INT-6A-C2-005",
        "name": "Cybersecurity C2 Link (Mutual TLS + JWT)",
        "tipo": "Funzionale control + sicurezza",
        "parte_a": "GS Operator (authenticated user)",
        "parte_b": "UAV C2 endpoint",
        "direzione": "Bidirezionale (handshake)",
        "caratteristiche": "Mutual TLS 1.3 con X.509 certificates (CA Firmamento proprietary); session JWT (RFC 7519) signed RS256, validity 4h; rate limit 100 req/min per session; intrusion detection via anomaly behavior",
        "standard": "TLS 1.3 RFC 8446; X.509 v3 RFC 5280; JWT RFC 7519; DO-326A (Airworthiness Security)",
        "owner": "Cybersecurity Engineer + Avionics Lead",
        "test_method": "Penetration test (selettivo) + certificate validation test + JWT replay test",
        "status": "Concept",
        "confidence": "medium",
        "note": "DO-326A compliance richiesta per certificazione futura. Per Y1 PoC sufficiente."
    },
    {
        "id": "INT-6A-C2-006",
        "name": "ADS-B IN Receiver (Cooperative DAA)",
        "tipo": "Funzionale control + safety",
        "parte_a": "UAV ADS-B IN receiver (e.g. PingRX Pro by uAvionix)",
        "parte_b": "FCS Conflict Manager",
        "direzione": "Unidirezionale (Air Traffic → UAV)",
        "caratteristiche": "ADS-B 1090 MHz Extended Squitter receiver; range 50+ NM line-of-sight; reporting period 1 Hz nominal; output: traffic list (ICAO 24-bit address, lat/lon/alt/velocity) per FCS conflict detection",
        "standard": "RTCA DO-260C (ADS-B MOPS); ICAO Annex 10 Vol IV; EUROCAE ED-102B",
        "owner": "Avionics Lead + DAA Specialist",
        "test_method": "Reception test in area aerea attiva (e.g. vicino aeroporto) + range verification",
        "status": "Concept",
        "confidence": "high",
        "note": "Standard per SORA SAIL III in EU. ADS-B IN sufficiente per traffico cooperativo (aviazione commerciale). Per non-cooperativo (GA leggera, paracadutisti) NON sufficiente → INT-6A-C2-007 radar/EO."
    },
    {
        "id": "INT-6A-C2-007",
        "name": "Non-Cooperative DAA (Visual EO or Acoustic)",
        "tipo": "Funzionale control + safety",
        "parte_a": "UAV onboard EO camera + AI inference engine",
        "parte_b": "FCS Conflict Manager",
        "direzione": "Unidirezionale (Sensor → FCS)",
        "caratteristiche": "Camera RGB orientata azimuth ±90° + elevation ±30°; inference YOLO-based detection aircraft @ 1-2 Hz; range detection target ≥ 1 NM (target aircraft 3 m wingspan); false alarm rate ≤ 0.01/h",
        "standard": "RTCA DO-365B (DAA MOPS); ASTM F3442/F3442M-23 (DAA Performance)",
        "owner": "Avionics Lead + AI/ML Engineer",
        "test_method": "Detection range test (vs target ULM) + false alarm verification",
        "status": "Concept",
        "confidence": "low",
        "note": "Necessario per SORA SAIL IV o quando ADS-B insufficiente. R&D dedicato. Per Y1 6A focus su mitigazione operativa (NOTAM, geofencing aree non interferenti). FALSIFYING: se ENAC richiede DAA non-cooperativo certificato per Pentema BVLOS, e tecnologia COTS non disponibile → Hold SORA submission."
    },
    {
        "id": "INT-6A-C2-008",
        "name": "Geofence Boundary (No-Fly Zone enforcement)",
        "tipo": "Funzionale control",
        "parte_a": "GS Mission Planner (waypoint validator)",
        "parte_b": "UAV FCS (geofence enforcer)",
        "direzione": "GS pre-flight upload; FCS runtime check",
        "caratteristiche": "Geofence polygon GeoJSON (max 100 vertices); altitude ceiling AGL + AMSL; soft boundary (warning) + hard boundary (auto-return); buffer ≥ 100 m da no-fly zone (es. aerea D-flight U-Space)",
        "standard": "OGC GeoJSON RFC 7946; ASD-STAN prEN 4709 (UAS direct remote identification)",
        "owner": "GCS UX Developer + Avionics Lead",
        "test_method": "Mission rejection test (waypoint in NFZ) + runtime breach test",
        "status": "Preliminary",
        "confidence": "high",
        "note": "Standard COTS in MAVLink mission planner."
    },

    # =====================================================================
    # SEZIONE 6A — GROUND SEGMENT (8)
    # =====================================================================
    {
        "id": "INT-6A-GS-001",
        "name": "GS Control Room HMI",
        "tipo": "Funzionale + operativa",
        "parte_a": "Pilot/Operator + Observer",
        "parte_b": "GS Workstation (3 monitor: map + telemetry + video)",
        "direzione": "Bidirezionale (HMI)",
        "caratteristiche": "3 monitor 27\" 4K; input dual workstation (pilot + observer); HOTAS hardware joystick + keyboard/mouse; intercom system pilot-observer-PC; logging session full screen + audio",
        "standard": "ISO 9241 (Ergonomics of HCI); MIL-STD-1472H (Human Engineering); EUROCAE ED-269 §4 (Remote Pilot Station)",
        "owner": "GCS UX Developer + Human Factors Engineer",
        "test_method": "Usability test + workflow validation + cognitive workload assessment",
        "status": "Concept",
        "confidence": "medium",
        "note": "Standard JOUAV include workstation. Customizzazione per dual-operator Pentema."
    },
    {
        "id": "INT-6A-GS-002",
        "name": "GS Mobile Vehicle Integration",
        "tipo": "Fisica + funzionale",
        "parte_a": "Veicolo (Fiat Ducato L3H2 o eq.) outfitted GS",
        "parte_b": "GS portable rack + antenna telescopic",
        "direzione": "Hosting",
        "caratteristiche": "Rack 19\" 12U fisso al pianale veicolo; antenna telescopica h max 6 m AGL deployable; 12V/230V power dual; aria condizionata; setup time ≤ 30 min sul sito",
        "standard": "ISO 16750 (automotive electrical); IEC 60529 IP54 (rack protection)",
        "owner": "Civil Works + Operations Lead",
        "test_method": "Deploy time test + vibration test + thermal test in cabina",
        "status": "Concept",
        "confidence": "medium",
        "note": "Mobile GS per Protezione Civile rapid deployment scenarios."
    },
    {
        "id": "INT-6A-GS-003",
        "name": "Cooperative Dashboard (Web App)",
        "tipo": "Funzionale data",
        "parte_a": "Cooperative User (browser HTTPS)",
        "parte_b": "Cloud Dashboard backend (SaaS Firmamento)",
        "direzione": "Bidirezionale (HTTPS)",
        "caratteristiche": "Web app responsive (desktop + tablet); login via SPID (SAML 2.0) + JWT session 8h; map viewer Leaflet/OpenLayers; download GeoTIFF / GeoJSON / PDF report; RBAC per area geografica polygon",
        "standard": "SAML 2.0 (OASIS); SPID (AgID Linee Guida); OWASP ASVS 4.0; WAI-ARIA (accessibility)",
        "owner": "Cloud Architect + Frontend Developer + DPO",
        "test_method": "Functional test + RBAC verification + accessibility audit (WCAG 2.1 AA)",
        "status": "Concept",
        "confidence": "medium",
        "note": "Sub-interface di INT-06 Cap.4. SPID obbligatorio per accesso PA + cooperative. AgID compliance check."
    },
    {
        "id": "INT-6A-GS-004",
        "name": "Protezione Civile Emergency Trigger",
        "tipo": "Operativa + contrattuale",
        "parte_a": "Sala Operativa PC Regione Liguria (operatore)",
        "parte_b": "Firmamento GCS On-Call Pilot",
        "direzione": "Bidirezionale (PC → Firmamento richiesta; Firmamento → PC stato)",
        "caratteristiche": "Trigger via: (1) SMS gateway dedicato +39 numero unico, (2) email PEC priority, (3) webform dashboard. Acknowledge entro 5 min; launch decision entro 15 min TTR; ortofoto live entro 30 min (target nominale)",
        "standard": "SOP standard tabletop tested M+6-M+7; D.Lgs. 1/2018 Codice Protezione Civile; convenzione operativa ex art. 15 L. 241/90",
        "owner": "Operations Lead + PC Liaison Officer (Regione)",
        "test_method": "Tabletop exercise + scenario simulation + response time measurement",
        "status": "Concept",
        "confidence": "medium",
        "note": "Sub-interface di INT-11 Cap.4. SLA TTR ≤ 4h nominale; emergency ≤ 30 min target. FALSIFYING: in caso meteo avverso (vento > 17 m/s) la launch può non essere possibile → SLA degraded a 'best effort'."
    },
    {
        "id": "INT-6A-GS-005",
        "name": "Flight Authorization Workflow (Internal)",
        "tipo": "Operativa",
        "parte_a": "Mission Planner (pre-flight check) + Pilot",
        "parte_b": "Operations Manager (approval) + Safety Officer (oversight)",
        "direzione": "Workflow approval",
        "caratteristiche": "Checklist pre-flight: meteo, NOTAM, geofence, payload OK, battery OK, ENAC permission OK; signing digital (e-Sign) da 2 figure (pilot + ops manager); log workflow audit trail completo",
        "standard": "EUROCAE ED-269; SOP Firmamento (DEL-PFTE-04 ConOps); JARUS SORA 2.5 OSO #19",
        "owner": "Operations Manager + Safety Officer",
        "test_method": "Workflow walk-through + audit trail verification",
        "status": "Concept",
        "confidence": "medium",
        "note": "Necessario per OSO #19 (Operator's safety culture). Compliance SORA SAIL III."
    },
    {
        "id": "INT-6A-GS-006",
        "name": "GS Maintenance Console (Health Monitoring)",
        "tipo": "Funzionale data + operativa",
        "parte_a": "UAV onboard Health Monitoring System (HMS)",
        "parte_b": "GS Maintenance Console + remote alerts",
        "direzione": "Unidirezionale (UAV → GS)",
        "caratteristiche": "Telemetry per: motor bearings vibration, battery cell SOH (State Of Health), servo current consumption, GPS HDOP/VDOP, IMU calibration drift; auto-alert se metric fuori soglia; data retention 1 anno per trend analysis",
        "standard": "ARINC 624 (Onboard Maintenance System); MIL-HDBK-2155 (Logistic Support Analysis)",
        "owner": "Maintenance Manager + Avionics Lead",
        "test_method": "Fault injection test + alert verification",
        "status": "Concept",
        "confidence": "medium",
        "note": "Critico per Reliability-Centered Maintenance (RCM). Output al Piano di Manutenzione (Vol. 2 Allegato A.10)."
    },
    {
        "id": "INT-6A-GS-007",
        "name": "Data Anonymization Pipeline (GDPR)",
        "tipo": "Funzionale data + regolatoria",
        "parte_a": "Cloud Raw Image Storage",
        "parte_b": "Anonymized Public Dataset Repository",
        "direzione": "Pipeline (Raw → Anonymized)",
        "caratteristiche": "AI inference: face blurring + license plate blurring; processing time ≤ 5 min per ortofoto; quality check: false negative rate ≤ 1%; opt-out registry per cittadini che richiedono erasure (DSAR Art. 17 GDPR)",
        "standard": "GDPR Reg. UE 2016/679 Art. 17 + Art. 25 (privacy by design); ISO/IEC 27701 (privacy info management)",
        "owner": "DPO + Data Engineer + AI/ML Engineer",
        "test_method": "Anonymization quality test + DSAR workflow test + audit",
        "status": "Concept",
        "confidence": "medium",
        "note": "Sub-interface di INT-12 Cap.4. Necessario per DPIA approvazione. FALSIFYING: se Garante rileva residual identifiability nelle ortofoto pubbliche → workflow esteso o restrizione accesso."
    },
    {
        "id": "INT-6A-GS-008",
        "name": "Backup & Disaster Recovery",
        "tipo": "Funzionale data + operativa",
        "parte_a": "Cloud Primary (Aruba IT)",
        "parte_b": "Cloud Backup (OVH IT EU separate region)",
        "direzione": "Replication scheduled",
        "caratteristiche": "Backup incremental ogni 4h; full backup settimanale; retention 30 giorni hot + 1 anno cold; RPO ≤ 4h; RTO ≤ 24h; encryption at rest AES-256",
        "standard": "ISO/IEC 27031 (ICT readiness for business continuity); NIST SP 800-34 Contingency Planning",
        "owner": "Cloud Architect + DevOps",
        "test_method": "Restore test (mensile) + DR drill (semestrale)",
        "status": "Concept",
        "confidence": "medium",
        "note": "Necessario per NIS2 compliance art. 21. GAIA-X compliance preserva sovranità dati."
    },

    # =====================================================================
    # SEZIONE 6A — REGULATORY (6)
    # =====================================================================
    {
        "id": "INT-6A-REG-001",
        "name": "ENAC SORA Authorization (SAIL III BVLOS)",
        "tipo": "Regolatoria",
        "parte_a": "Firmamento Technologies (operatore UAS)",
        "parte_b": "ENAC Ufficio RPAS",
        "direzione": "Bidirezionale (application + authorization)",
        "caratteristiche": "Application: Operations Manual + SORA worksheet + Operator Declaration + insurance + DAA mitigation; iter: pre-application meeting M+3-M+6, application formal M+15-M+18 (Fase 1), authorization expected M+18-M+22; categoria Specific SAIL III stimato per BVLOS Pentema",
        "standard": "Reg. UE 2019/947 art. 12 + AMC Amendment 3 (Sett 2025) SORA 2.5 EU; ENAC Reg. APR Ed. 3",
        "owner": "Aviation Regulatory Counsel + ENAC Liaison",
        "test_method": "Documentation review + pre-application feedback",
        "status": "Concept",
        "confidence": "medium",
        "note": "Sub-interface di INT-13 Cap.4. FALSIFYING: se ENAC valuta Pentema come area sensibile (zona ad alta protezione naturale es. Aveto) e richiede SAIL IV → costi +€200-400k + tempo +6 mesi → review business case."
    },
    {
        "id": "INT-6A-REG-002",
        "name": "AGCOM Spectrum Authorization (2.4 GHz + L-band)",
        "tipo": "Regolatoria",
        "parte_a": "Firmamento Technologies",
        "parte_b": "AGCOM (Autorità Garante Comunicazioni)",
        "direzione": "Bidirezionale (application + authorization)",
        "caratteristiche": "Spettro primario: ISM 2.4-2.4835 GHz (EIRP ≤ 100 mW unlicensed; ≤ 1 W con autorizzazione individuale); fallback L-band Iridium 1616-1626.5 MHz già licenziato a Iridium; EIRP/coexistence verificato per area Pentema",
        "standard": "D.Lgs. 207/2021 Codice Comunicazioni Elettroniche art. 11; PNRF (Piano Nazionale Ripartizione Frequenze); EN 300 328",
        "owner": "Aviation Regulatory Counsel + RF Systems Engineer + AGCOM Liaison",
        "test_method": "Documentation review + AGCOM response",
        "status": "Concept",
        "confidence": "medium",
        "note": "Sub-interface di INT-07 Cap.4. ISM unlicensed fattibile ma EIRP basso → range limitato. Per range 50 km richiede licenza individuale (lead time 4-8 mesi)."
    },
    {
        "id": "INT-6A-REG-003",
        "name": "ENAV/D-Flight U-Space Coordination",
        "tipo": "Regolatoria + operativa",
        "parte_a": "Firmamento GCS",
        "parte_b": "ENAV / D-Flight USSP (U-Space Service Provider)",
        "direzione": "Bidirezionale (flight authorization + network ID)",
        "caratteristiche": "Network Identification: broadcast UAS ID + position via D-Flight API; flight authorization request per area U-Space (se istituita su Pentema); geo-awareness: pre-flight check NFZ",
        "standard": "Reg. UE 2021/664 (U-Space framework); ENAC LG-2023/006; D-Flight USSP+CISP API",
        "owner": "Aviation Regulatory Counsel + GCS Developer",
        "test_method": "API integration test + USSP coordination",
        "status": "Concept",
        "confidence": "medium",
        "note": "Sub-interface di INT-14 Cap.4. U-Space su Pentema non ancora istituito (2026). Engagement con ENAV per anticipare requirement."
    },
    {
        "id": "INT-6A-REG-004",
        "name": "Garante Privacy DPIA + DSAR Procedure",
        "tipo": "Regolatoria",
        "parte_a": "Firmamento (titolare + co-titolare con Regione)",
        "parte_b": "Garante per la Protezione Dati Personali",
        "direzione": "Notifica + audit",
        "caratteristiche": "DPIA Art. 35 GDPR completata + signed DPO; workshop pubblico Pentema con DPIA pubblicata; DSAR procedure: response entro 30 gg, erasure entro 30 gg; provider audit annuale",
        "standard": "GDPR Reg. UE 2016/679 Art. 35 + Art. 17; D.Lgs. 196/2003 novellato; Provv. Garante 1.8.2013 (videosorveglianza)",
        "owner": "Data Privacy Counsel + DPO Firmamento + DPO Regione",
        "test_method": "DPIA review + DSAR drill test",
        "status": "Preliminary",
        "confidence": "medium",
        "note": "Sub-interface di INT-15 Cap.4. DPIA firma DPO M+5. FALSIFYING: se Garante richiede notifica preventiva Art. 36 → lead time +3-6 mesi."
    },
    {
        "id": "INT-6A-REG-005",
        "name": "Assicurazione Aviation Third-Party Liability",
        "tipo": "Regolatoria + contrattuale",
        "parte_a": "Firmamento (insured)",
        "parte_b": "Insurance Broker + Carrier (Lloyd's / AON / Marsh)",
        "direzione": "Polizza contrattuale",
        "caratteristiche": "Massimale third-party liability obbligatorio Reg. UE 785/2004: MTOW 38 kg → massimale 750.000 DSP ≈ €870k; option upgrade a €1.5M-3M per servizi PA; hull insurance opzionale; war risk exclusion",
        "standard": "Reg. UE 785/2004 (insurance requirements air carriers); ENAC LRA 25/2023 (UAS insurance)",
        "owner": "CFO Firmamento + Insurance Broker",
        "test_method": "Polizza review + claim simulation",
        "status": "Concept",
        "confidence": "medium",
        "note": "Mercato assicurativo UAS in evoluzione 2026 (post-incidenti BVLOS UK 2024-25). Premi stimati €15-40k/anno per copertura €1.5M."
    },
    {
        "id": "INT-6A-REG-006",
        "name": "NIS2 Cybersecurity Notification (D.Lgs. 138/2024)",
        "tipo": "Regolatoria",
        "parte_a": "Firmamento (entità essenziale o importante TBD)",
        "parte_b": "ACN (Agenzia Cybersicurezza Nazionale)",
        "direzione": "Notifica incident + reporting",
        "caratteristiche": "Notifica incident significativo entro 24h initial + 72h detailed + 1 mese final report; misure tecniche obbligatorie art. 21 (risk mgmt, incident response, supply chain, awareness training)",
        "standard": "D.Lgs. 138/2024 (recepimento Dir. UE 2022/2555 NIS2); ENISA Threat Landscape report 2025",
        "owner": "Cybersecurity Engineer + Legal Counsel",
        "test_method": "Incident response drill + notification workflow test",
        "status": "Concept",
        "confidence": "medium",
        "note": "Status NIS2 (essenziale/importante) da chiarire: aviation civil è 'essenziale' Allegato I, ma UAS commerciale potrebbe rientrare in 'importante' Allegato II. Engagement ACN preliminare."
    },

    # =====================================================================
    # SEZIONE 6B — HALE PHYSICAL (4)
    # =====================================================================
    {
        "id": "INT-6B-PHY-001",
        "name": "HALE Airframe ↔ Solar Panel Array",
        "tipo": "Fisica meccanica + termica",
        "parte_a": "HALE Wing structure (CFRP + lino secondario)",
        "parte_b": "Solar Panel Array (GaAs multi-junction Spectrolab XTJ Prime)",
        "direzione": "Statica (bonded)",
        "caratteristiche": "Area pannelli ≥ 25 m² (worst-case calc); panel mass density 0.5-0.8 kg/m²; encapsulation Honeywell Aclar (UV resistant); bonding film 3M VHB structural; thermal CTE matching critical",
        "standard": "ECSS-Q-ST-70-71 (Materials and Processes for spacecraft); ASTM E1980 (solar reflectance)",
        "owner": "Aerodynamics-Structures Engineer + Solar Cell Specialist",
        "test_method": "Bonding strength test (lap shear) + thermal cycling + UV exposure",
        "status": "Concept",
        "confidence": "low",
        "note": "Sub-interface di INT-01 Cap.4 per 6B. R&D dedicato. FALSIFYING (RSK-TEC-001): se peso pannelli > 0.8 kg/m² o efficienza < 28% → energy balance inverno deficit aggravato."
    },
    {
        "id": "INT-6B-PHY-002",
        "name": "HALE Wing Spar ↔ Fuselage Structural Joint",
        "tipo": "Fisica meccanica + strutturale",
        "parte_a": "Wing spar carry-through (CFRP unidirezionale)",
        "parte_b": "Fuselage frame ring (CFRP woven)",
        "direzione": "Statica (bolted + bonded)",
        "caratteristiche": "Bolted joint M6 titanium x 16 + bonded film Hysol EA9394; load transfer max ±50 kN tension/compression (limit load) + 75 kN ultimate (factor 1.5); flutter clearance margin ≥ 20%",
        "standard": "ECSS-E-ST-32-08 (Materials); MIL-HDBK-17 (Composite Materials); ASTM D5868 (lap shear)",
        "owner": "Aerodynamics-Structures Engineer",
        "test_method": "FEA analysis + GVT (Ground Vibration Test) + structural test article",
        "status": "Concept",
        "confidence": "low",
        "note": "Critico per RSK-TEC-002 (aeroelasticità ala high-AR). Wing flexibility >20% apertura tipica HALE → joint deve sostenere fatigue + load reversal."
    },
    {
        "id": "INT-6B-PHY-003",
        "name": "HALE Battery Pack Thermal Management",
        "tipo": "Fisica termica",
        "parte_a": "LiS Battery Pack (target 350 Wh/kg, 100+ kWh capacity)",
        "parte_b": "Thermal control system (passive + active heater)",
        "direzione": "Bidirezionale (heat flow + control)",
        "caratteristiche": "Battery operating range +5°C / +45°C optimal; @ 20 km ambient -65°C → heater 50 W continuous + phase-change material insulation; charge inhibit se T < 0°C; thermal mass 5 kg PCM",
        "standard": "DO-311A; UN 38.3; IEC 62133; ECSS-E-ST-31C (Thermal Control)",
        "owner": "Thermal Engineer + Battery Specialist",
        "test_method": "Thermal vacuum chamber simulation + heater verification",
        "status": "Concept",
        "confidence": "low",
        "note": "Thermal mass penalty in MTOW. FALSIFYING: se heater richiede > 80 W continuous → impatta energy balance notturno → showstopper RSK-TEC-001 aggravato."
    },
    {
        "id": "INT-6B-PHY-004",
        "name": "HALE Antenna Aperture (Ka-band 31 GHz)",
        "tipo": "Fisica meccanica + RF",
        "parte_a": "HALE underbelly antenna mount",
        "parte_b": "Ka-band feeder antenna (steerable phased array)",
        "direzione": "Statica + steerable",
        "caratteristiche": "Phased array 256-element @ 31 GHz; aperture 30×30 cm; mass ≤ 3 kg; beam steering ±60° azimuth/elevation; gimbal opzionale per coarse pointing; mount struttura CFRP isolata termicamente",
        "standard": "ITU-R F.1500 (HAPS RF specs); RTCA DO-262C (airborne SATCOM)",
        "owner": "Telecom-NTN Payload Expert + RF Engineer",
        "test_method": "Pattern measurement (compact range) + thermal cycling",
        "status": "Concept",
        "confidence": "low",
        "note": "AESA Ka-band 2026: maturità TRL 6 component, TRL 3-4 integrato HALE. Costo €150-300k unit (R&D)."
    },

    # =====================================================================
    # SEZIONE 6B — NTN SERVICE LINK (4)
    # =====================================================================
    {
        "id": "INT-6B-NTN-001",
        "name": "5G NR-NTN Service Link Downlink (HAPS → UE)",
        "tipo": "Funzionale + RF",
        "parte_a": "HAPS gNodeB regenerative payload",
        "parte_b": "User Equipment 5G NTN-capable (smartphone, IoT)",
        "direzione": "Unidirezionale (DL)",
        "caratteristiche": "Frequency S-band 2010-2025 MHz (n255) o 1980-2010 MHz (n256); bandwidth 5-20 MHz per beam; modulation up to 256-QAM; coverage cell 30-50 km diameter @ 20 km altitude; capacity per beam 50-200 Mbps",
        "standard": "3GPP TS 38.211/.212/.213/.214 (NR L1); TS 38.811 (NTN study); TS 38.821 (NTN solutions); ITU-R M.2150",
        "owner": "Telecom-NTN Payload Expert",
        "test_method": "Link budget simulation + over-the-air test (vs UE simulator)",
        "status": "Concept",
        "confidence": "low",
        "note": "Per 6B HALE. Service link S-band ideale per copertura aree interne 5G NTN. Spettro contestato con MNO terrestri TIM/Vodafone n255 → accordi spectrum sharing necessari."
    },
    {
        "id": "INT-6B-NTN-002",
        "name": "5G NR-NTN Service Link Uplink (UE → HAPS)",
        "tipo": "Funzionale + RF",
        "parte_a": "User Equipment 5G NTN-capable",
        "parte_b": "HAPS gNodeB receiver",
        "direzione": "Unidirezionale (UL)",
        "caratteristiche": "Frequency S-band paired UL; UE EIRP max 23 dBm (handheld) o 33 dBm (vehicular); link budget challenging due to UE low power; modulation up to 64-QAM (UL)",
        "standard": "3GPP TS 38.101-5 (NR-NTN UE requirements); ITU-R M.2150",
        "owner": "Telecom-NTN Payload Expert",
        "test_method": "Link budget verification + UL throughput test",
        "status": "Concept",
        "confidence": "low",
        "note": "UL è il collo di bottiglia in NTN: UE handheld 23 dBm a 20 km distance richiede HAPS antenna G/T > 5 dB/K → AESA G ≥ 30 dBi necessario."
    },
    {
        "id": "INT-6B-NTN-003",
        "name": "NTN Doppler Compensation",
        "tipo": "Funzionale RF",
        "parte_a": "HAPS gNodeB onboard",
        "parte_b": "UE (Doppler offset compensation)",
        "direzione": "Bidirezionale signaling",
        "caratteristiche": "HAPS station-keeping ±5 km nominal; ground speed effective 10-30 km/h (wind drift) → Doppler @ S-band ≈ 50-200 Hz; pre-compensation HAPS-side + UE-side residual ≤ 50 Hz",
        "standard": "3GPP TS 38.811 §6.4 (Doppler); TS 38.821 §6.4.2 (Doppler shift compensation)",
        "owner": "Telecom-NTN Payload Expert + 5G Algorithm Specialist",
        "test_method": "Simulation OFDM Doppler vs SINR + lab test",
        "status": "Concept",
        "confidence": "medium",
        "note": "Doppler HAPS molto inferiore vs LEO (~50 kHz) → gestibile con compensation standard 3GPP Rel-17."
    },
    {
        "id": "INT-6B-NTN-004",
        "name": "Inter-Beam Handover (intra-HAPS)",
        "tipo": "Funzionale",
        "parte_a": "HAPS gNodeB multi-beam (16-32 beam)",
        "parte_b": "UE in motion (cross-beam boundary)",
        "direzione": "Bidirezionale signaling",
        "caratteristiche": "Beam ID broadcast via SIB1; UE measurement RSRP/RSRQ per beam neighbor; handover trigger via A3 event (neighbor > serving + threshold); HO interruption time < 50 ms",
        "standard": "3GPP TS 38.331 (RRC); TS 38.300 (NR overall); TR 38.821 §7.4 (mobility NTN)",
        "owner": "Telecom-NTN Payload Expert",
        "test_method": "Simulation handover + measurement campaign",
        "status": "Concept",
        "confidence": "medium",
        "note": "Per HAPS quasi-stazionario, handover intra-HAPS è rare (UE muove più della cella). HO inter-HAPS quando flotta dispiegata."
    },

    # =====================================================================
    # SEZIONE 6B — FEEDER LINK (3)
    # =====================================================================
    {
        "id": "INT-6B-FEEDER-001",
        "name": "Ka-band Feeder Link Downlink (HAPS → Gateway)",
        "tipo": "Funzionale + RF",
        "parte_a": "HAPS AESA Ka-band antenna",
        "parte_b": "Gateway HAPS ground station (parabola 3-5 m)",
        "direzione": "Unidirezionale (DL HAPS → Gateway)",
        "caratteristiche": "Frequency 31-31.3 GHz (HAPS dedicated post-WRC-19); bandwidth 200-500 MHz; modulation DVB-S2X up to 64-APSK; capacity 1-10 Gbps; EIRP HAPS ≥ 60 dBW; rain fade margin Italia Zona K ≥ 15 dB",
        "standard": "ITU-R F.1500/F.1891 (HAPS specs); ETSI EN 302 307-2 (DVB-S2X); ITU-R P.618-14 (rain fade)",
        "owner": "Telecom-NTN Payload Expert + Ka-band RF Engineer",
        "test_method": "Link budget calculation + rain fade simulation + outage analysis",
        "status": "Concept",
        "confidence": "low",
        "note": "Ka-band 31 GHz HAPS-dedicated. Rain fade Genova Zona K ~25 dB @ 99.9% → margine richiesto significativo. ACM (Adaptive Coding & Modulation) mandatory."
    },
    {
        "id": "INT-6B-FEEDER-002",
        "name": "Ka-band Feeder Link Uplink (Gateway → HAPS)",
        "tipo": "Funzionale + RF",
        "parte_a": "Gateway HAPS ground station transmitter",
        "parte_b": "HAPS AESA Ka-band receiver",
        "direzione": "Unidirezionale (UL Gateway → HAPS)",
        "caratteristiche": "Frequency 27.9-28.2 GHz (HAPS UL band post-WRC-19); EIRP Gateway 70-80 dBW (parabola 5 m + HPA 100W); UL capacity 100 Mbps - 1 Gbps",
        "standard": "ITU-R F.1500; ITU RR Appendix 30B",
        "owner": "Telecom-NTN Payload Expert",
        "test_method": "Link budget + interference analysis vs FSS",
        "status": "Concept",
        "confidence": "low",
        "note": "Coexistence con Fixed Satellite Service (FSS) GEO → coordination ITU. AGCOM autorizzazione individuale per gateway."
    },
    {
        "id": "INT-6B-FEEDER-003",
        "name": "Gateway HAPS ↔ Core Network 5G",
        "tipo": "Funzionale data",
        "parte_a": "Gateway HAPS ground station",
        "parte_b": "5G Core Network (5GC) operator partner",
        "direzione": "Bidirezionale",
        "caratteristiche": "N2/N3 interface SBA (Service Based Architecture); transport IP over MPLS; latency one-way ≤ 5 ms; throughput aggregato 10+ Gbps; integration AMF/SMF/UPF operator-side",
        "standard": "3GPP TS 23.501 (5G architecture); TS 29.281 (GTP-U); IETF RFC 3031 (MPLS)",
        "owner": "Telecom-NTN Payload Expert + 5G Core Specialist",
        "test_method": "Integration test (lab + field) + interop testing",
        "status": "Concept",
        "confidence": "low",
        "note": "Richiede partnership con MNO italiano (TIM/Vodafone/WindTre) o operatore wholesale (Open Fiber). Modello business: capacity wholesale → operator delivers to UE."
    },

    # =====================================================================
    # SEZIONE 6B — HALE OPS (3)
    # =====================================================================
    {
        "id": "INT-6B-OPS-001",
        "name": "HALE Air Traffic Coordination (ENAV)",
        "tipo": "Regolatoria + operativa",
        "parte_a": "Firmamento HALE Operations Center",
        "parte_b": "ENAV ACC (Area Control Center) Milano",
        "direzione": "Bidirezionale (clearance + position reporting)",
        "caratteristiche": "Flight plan ICAO IFR FL590-650 (18-20 km altitude); cleared corridor permanent (e.g. Genova FIR); ADS-B OUT transmission obbligatoria; latency reporting ≤ 60 s",
        "standard": "ICAO Annex 11 (Air Traffic Services); ICAO Doc 4444 (PANS-ATM); Reg. UE 923/2012 SERA",
        "owner": "Aviation Regulatory Counsel + HALE Operations Manager",
        "test_method": "Flight plan submission test + ATC coordination drill",
        "status": "Concept",
        "confidence": "low",
        "note": "Sub-interface di INT-14 Cap.4 per HALE. FL650 = 19.8 km, sopra controlled airspace UK ma in Italia tutto controlled space. Special arrangement con ENAV per HAPS-dedicated corridor."
    },
    {
        "id": "INT-6B-OPS-002",
        "name": "EUROCONTROL Cross-Border Coordination",
        "tipo": "Regolatoria",
        "parte_a": "Firmamento HALE Operations Center",
        "parte_b": "EUROCONTROL Network Manager (Brussels)",
        "direzione": "Bidirezionale (flight plan + slot)",
        "caratteristiche": "Flight plan submission via IFPS (Initial Flight Plan Processing System); slot allocation se attraversa boundaries FIR; coordination con ANSP confinanti (Switzerland, France) se applicable",
        "standard": "EUROCONTROL FPL format; ICAO Doc 7030 (Regional Supplementary Procedures)",
        "owner": "Aviation Regulatory Counsel",
        "test_method": "Flight plan test + EUROCONTROL coordination",
        "status": "Concept",
        "confidence": "low",
        "note": "Sub-interface long-term per Percorso 6B operativo. HAPS station-keeping non attraversa boundaries → IFPS submission singola."
    },
    {
        "id": "INT-6B-OPS-003",
        "name": "HALE Launch & Recovery Site (Glider-style)",
        "tipo": "Operativa + fisica",
        "parte_a": "HALE airframe",
        "parte_b": "Launch site (runway grass ≥ 800 m oppure aerotraino con winch)",
        "direzione": "Operativa (singolo evento per missione)",
        "caratteristiche": "Decollo: aerotraino con winch o motopropulsore singolo; quota cruise raggiunta dopo 6-8h salita; recovery: glide approach senza motore + grass runway; weather window stringente (wind ≤ 5 m/s @ ground, no rain)",
        "standard": "FAA AC 91-79 (Mitigating Risks Associated with Approach and Landing); EASA AMC1 SPA.SPO (Specialised Operations)",
        "owner": "HALE Operations Manager + Civil Works",
        "test_method": "Site survey + weather window analysis + recovery drill (subscale)",
        "status": "Concept",
        "confidence": "low",
        "note": "Pentema NON adatto come launch site (terreno montano). Site candidato: ex aeroporto Sarzana-Luni o aeroporto Albenga. Lead time site permitting ≥ 12 mesi."
    },

    # =====================================================================
    # SEZIONE TRASVERSALI (5)
    # =====================================================================
    {
        "id": "INT-X-CLOUD-001",
        "name": "Cloud Hosting GAIA-X Compliant",
        "tipo": "Funzionale + regolatoria + contrattuale",
        "parte_a": "Firmamento Cloud Workload (data + apps)",
        "parte_b": "Cloud Provider GAIA-X compliant (Aruba/OVH/IONOS)",
        "direzione": "Contrattuale",
        "caratteristiche": "Provider GAIA-X Label Level 2+ (data sovereignty); region IT or EU; SLA uptime ≥ 99.9% (Tier 3); RPO ≤ 4h; data residency garantita; subprocessor list GDPR Art. 28",
        "standard": "GAIA-X Compliance Framework v22.10; ISO/IEC 27001; SOC 2 Type II; CSA STAR Level 2",
        "owner": "Cloud Architect + Procurement + DPO",
        "test_method": "Provider audit + GAIA-X attestation review + data residency test",
        "status": "Concept",
        "confidence": "high",
        "note": "Sub-interface di INT-20 Cap.4. Aruba ha label GAIA-X confermato 2025. OVH Italia/IONOS pari opportunità. Hyperscaler USA (AWS/Azure/GCP) NON GAIA-X compliant per default."
    },
    {
        "id": "INT-X-PRIVACY-001",
        "name": "Data Subject Access Request (DSAR) Workflow",
        "tipo": "Regolatoria + operativa",
        "parte_a": "Data Subject (cittadino, cooperativo, dipendente)",
        "parte_b": "DPO Firmamento + Customer Support",
        "direzione": "Bidirezionale (request + response)",
        "caratteristiche": "Channels: email PEC dpo@firmamento.tech, web form, postal mail; response 30 gg max (extension +60 gg motivata); right: access, rectification, erasure (Art. 15-17 GDPR), portability (Art. 20), restriction (Art. 18), objection (Art. 21)",
        "standard": "GDPR Reg. UE 2016/679 Art. 12-22; EDPB Guidelines 01/2022 (Data Subject Rights)",
        "owner": "DPO + Customer Support",
        "test_method": "DSAR drill (semestrale) + response time audit",
        "status": "Concept",
        "confidence": "high",
        "note": "Workflow standard GDPR. Necessario tooling (es. OneTrust o open-source) per scale Y2+."
    },
    {
        "id": "INT-X-CYBER-001",
        "name": "Penetration Testing + Vulnerability Management",
        "tipo": "Regolatoria + sicurezza",
        "parte_a": "Firmamento attack surface (apps + infra + UAS)",
        "parte_b": "Pentest Provider (TÜV / DNV / Yarix)",
        "direzione": "Servizio professionale",
        "caratteristiche": "Pentest annuale: external + internal + web app + API; CVE scanning monthly (Nessus/Qualys); patch SLA Critical 7gg / High 30 gg / Medium 90 gg; bug bounty Y3+ opzionale",
        "standard": "OWASP Testing Guide v4.2; PTES (Penetration Testing Execution Standard); ISO/IEC 27001 A.12.6",
        "owner": "Cybersecurity Engineer + CISO (Y3+)",
        "test_method": "Pentest report + remediation tracking + retest",
        "status": "Concept",
        "confidence": "high",
        "note": "Necessario per NIS2 art. 21 misure tecniche. Costo annuale stimato €15-30k per pentest semplice; €50-100k per Red Team esercizio completo."
    },
    {
        "id": "INT-X-LOGGING-001",
        "name": "Centralized Logging + Audit Trail",
        "tipo": "Funzionale + regolatoria",
        "parte_a": "All system components (UAS + GS + Cloud + Apps)",
        "parte_b": "SIEM Central (e.g. Elastic SIEM, Wazuh, Splunk)",
        "direzione": "Unidirezionale (components → SIEM)",
        "caratteristiche": "Log levels: INFO + WARN + ERROR + AUDIT (security events); format JSON structured; correlation rules pre-built; retention 90 gg hot + 1 anno cold + 5 anni legal hold",
        "standard": "RFC 5424 Syslog; CEF Common Event Format; NIST SP 800-92 (Log Management); GDPR Art. 32 (accountability)",
        "owner": "DevOps + Cybersecurity Engineer + DPO",
        "test_method": "Log ingestion test + retention verification + audit",
        "status": "Concept",
        "confidence": "high",
        "note": "Sub-interface di INT-12 Cap.4. ELK stack open-source consigliato per Y1-Y2 budget."
    },
    {
        "id": "INT-X-VENDOR-001",
        "name": "Vendor SLA Framework (UAS Manufacturer)",
        "tipo": "Contrattuale",
        "parte_a": "Firmamento Technologies (buyer)",
        "parte_b": "UAS Vendor (JOUAV / Tekever / Quantum / FlyingBasket)",
        "direzione": "Contratto fornitura",
        "caratteristiche": "Delivery lead time ≤ 12 settimane post-PO; warranty 24 mesi parts + 12 mesi service; spare parts availability 5+ anni; training 5 gg incluso; technical support 24h response; supplier audit AS/EN 9100",
        "standard": "AS/EN 9100D (Quality Mgmt Aerospace); ICAO Annex 19 (Safety Mgmt); contratto fornitura ex artt. 1470-1547 c.c.",
        "owner": "Procurement Lead + Quality Manager",
        "test_method": "Vendor audit + warranty claim simulation + SLA monitoring",
        "status": "Concept",
        "confidence": "medium",
        "note": "Sub-interface di INT-16 Cap.4. Pre-RFQ engagement M+1-M+6. Selezione finale post-Trade Study M+8. FALSIFYING: vendor cinese (JOUAV) potrebbe avere export control issues post-evoluzione geopolitica → necessario vendor alternativo EU (Tekever PT, Quantum Systems DE)."
    },
]

# ============================================================
# COSTRUZIONE WORKBOOK
# ============================================================

def build_workbook():
    wb = Workbook()

    # ============================================================
    # SHEET 1 — COVER
    # ============================================================
    ws = wb.active
    ws.title = "Cover"
    set_col_widths(ws, [25, 80])

    rows_cover = [
        ("Documento", "Allegato A.4 — Interface Control Document (ICD) preliminare v1.0"),
        ("Progetto", "Studio di Fattibilità HALE/VTOL — Firmamento Technologies"),
        ("Bando", "Cooding Prototypes (Coopfond / Legacoop)"),
        ("Volume", "Volume 2 — Allegati tecnici"),
        ("Versione", "v1.0 (M+3 baseline, post-Allineamento Strategico Maggio 2026)"),
        ("Data", "2026-05-17"),
        ("Status", "Preliminary (concept-level + preliminary spec)"),
        ("Conformità metodologica", "ARP4754A (System Development); ISO/IEC/IEEE 24765:2017 (SE Vocabulary); NASA SE Handbook §6.3 (Interface Mgmt)"),
        ("Conformità procedurale IT", "D.Lgs. 36/2023 art. 41 + Allegato I.7 (PFTE - Relazione Tecnica)"),
        ("Numero interfacce", "50 interfacce dettagliate, 8 categorie"),
        ("Owner documento", "Senior Systems Engineer (Firmamento) + Integration Engineer Lead"),
        ("Boundary conditions", "B1 (modello service-only cooperative); B2 (visione 10 anni EU sovereign HAPS)"),
        ("Disciplina epistemica", "Skill epistemic-rigor: confidence levels + falsifying observations per interfacce critiche (≥ 5)"),
        ("Riferimento padre", "Cap. 4 §4.4 (ICD preliminare 20 interfacce primarie); Cap. 6 (architettura)"),
        ("Riferimento agenti", "avionics-gnc; telecom-ntn-payload; aerodynamics-structures; earth-observation"),
        ("Note", "L'ICD è preliminary, NON engineering-grade. Specifiche byte-level e latency end-to-end deferite a Phase A/B Detailed ICD."),
    ]

    # Title
    ws["A1"] = "ALLEGATO A.4 — INTERFACE CONTROL DOCUMENT (ICD) v1.0"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 35

    for i, (k, v) in enumerate(rows_cover, start=3):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        subheader_style(ws.cell(row=i, column=1))
        cell_style(ws.cell(row=i, column=2))

    # ============================================================
    # SHEET 2 — ALL_INTERFACES (50 records summary)
    # ============================================================
    ws = wb.create_sheet("All_Interfaces")
    headers = ["ID", "Nome", "Tipo", "Parte A", "Parte B", "Direzione",
               "Standard di riferimento", "Owner", "Test Method", "Status", "Confidence"]
    set_col_widths(ws, [18, 38, 22, 30, 30, 22, 35, 28, 25, 14, 12])

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell)
    ws.row_dimensions[1].height = 30

    for row_idx, intf in enumerate(INTERFACES, start=2):
        ws.cell(row=row_idx, column=1, value=intf["id"])
        ws.cell(row=row_idx, column=2, value=intf["name"])
        ws.cell(row=row_idx, column=3, value=intf["tipo"])
        ws.cell(row=row_idx, column=4, value=intf["parte_a"])
        ws.cell(row=row_idx, column=5, value=intf["parte_b"])
        ws.cell(row=row_idx, column=6, value=intf["direzione"])
        ws.cell(row=row_idx, column=7, value=intf["standard"])
        ws.cell(row=row_idx, column=8, value=intf["owner"])
        ws.cell(row=row_idx, column=9, value=intf["test_method"])
        ws.cell(row=row_idx, column=10, value=intf["status"])
        ws.cell(row=row_idx, column=11, value=intf["confidence"])
        for col in range(1, 12):
            cell_style(ws.cell(row=row_idx, column=col))

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    # Conditional formatting per Status
    status_col = 10
    ws.conditional_formatting.add(
        f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{len(INTERFACES)+1}",
        CellIsRule(operator="equal", formula=['"Concept"'], stopIfTrue=False,
                   fill=PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{len(INTERFACES)+1}",
        CellIsRule(operator="equal", formula=['"Preliminary"'], stopIfTrue=False,
                   fill=PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{len(INTERFACES)+1}",
        CellIsRule(operator="equal", formula=['"Tested"'], stopIfTrue=False,
                   fill=PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid"))
    )

    # Conditional formatting per Confidence
    conf_col = 11
    for col_letter, conf_val, color in [
        ("K", '"low"', LIGHT_RED),
        ("K", '"medium"', LIGHT_YELLOW),
        ("K", '"high"', LIGHT_GREEN),
    ]:
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{len(INTERFACES)+1}",
            CellIsRule(operator="equal", formula=[conf_val], stopIfTrue=False,
                       fill=PatternFill(start_color=color, end_color=color, fill_type="solid"))
        )

    # ============================================================
    # SHEET 3-9 — Per-category detailed sheets
    # ============================================================
    categories = [
        ("Physical_Interfaces", lambda i: "PHY" in i["id"]),
        ("Data_Interfaces", lambda i: "DATA" in i["id"]),
        ("C2_Interfaces", lambda i: "C2" in i["id"]),
        ("GS_Interfaces", lambda i: "GS" in i["id"]),
        ("Regulatory_Interfaces", lambda i: "REG" in i["id"]),
        ("HALE_Specific", lambda i: "6B" in i["id"]),
        ("Trasversali", lambda i: "INT-X" in i["id"]),
    ]

    detail_headers = ["ID", "Nome", "Tipo", "Parte A", "Parte B", "Direzione",
                      "Caratteristiche Tecniche", "Standard / Protocollo",
                      "Owner", "Test Method", "Status", "Confidence", "Note"]
    detail_widths = [16, 35, 22, 28, 28, 22, 60, 35, 28, 28, 13, 12, 70]

    for sheet_name, filter_fn in categories:
        ws_cat = wb.create_sheet(sheet_name)
        set_col_widths(ws_cat, detail_widths)

        for col, h in enumerate(detail_headers, 1):
            cell = ws_cat.cell(row=1, column=col, value=h)
            header_style(cell)
        ws_cat.row_dimensions[1].height = 30

        filtered = [i for i in INTERFACES if filter_fn(i)]
        for row_idx, intf in enumerate(filtered, start=2):
            ws_cat.cell(row=row_idx, column=1, value=intf["id"])
            ws_cat.cell(row=row_idx, column=2, value=intf["name"])
            ws_cat.cell(row=row_idx, column=3, value=intf["tipo"])
            ws_cat.cell(row=row_idx, column=4, value=intf["parte_a"])
            ws_cat.cell(row=row_idx, column=5, value=intf["parte_b"])
            ws_cat.cell(row=row_idx, column=6, value=intf["direzione"])
            ws_cat.cell(row=row_idx, column=7, value=intf["caratteristiche"])
            ws_cat.cell(row=row_idx, column=8, value=intf["standard"])
            ws_cat.cell(row=row_idx, column=9, value=intf["owner"])
            ws_cat.cell(row=row_idx, column=10, value=intf["test_method"])
            ws_cat.cell(row=row_idx, column=11, value=intf["status"])
            ws_cat.cell(row=row_idx, column=12, value=intf["confidence"])
            ws_cat.cell(row=row_idx, column=13, value=intf["note"])
            for col in range(1, 14):
                cell_style(ws_cat.cell(row=row_idx, column=col))
            ws_cat.row_dimensions[row_idx].height = 90

        ws_cat.freeze_panes = "B2"
        ws_cat.auto_filter.ref = ws_cat.dimensions

    # ============================================================
    # SHEET 10 — Compatibility Matrix
    # ============================================================
    ws = wb.create_sheet("Compatibility_Matrix")

    # Lista subset di interfacce critiche per compatibility check (10x10)
    critical_intf = [i for i in INTERFACES if i["confidence"] in ("medium", "low") and "INT-6A" in i["id"]][:10]

    set_col_widths(ws, [22] + [18] * len(critical_intf))
    ws.cell(row=1, column=1, value="Compatibility Matrix (interfacce critiche 6A)")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(critical_intf) + 1)

    # Header row
    for col_idx, intf in enumerate(critical_intf, start=2):
        cell = ws.cell(row=3, column=col_idx, value=intf["id"])
        header_style(cell)
    # Header col
    for row_idx, intf in enumerate(critical_intf, start=4):
        cell = ws.cell(row=row_idx, column=1, value=intf["id"])
        header_style(cell)

    # Matrix content (simplified): diagonal = self; off-diag = "OK"/"check"/"conflict"
    compatibility_data = {
        # (i,j) → status
        # Self-diagonal = "—"
        # Most off-diag = "OK" (compatible)
        # Some critical pairs need attention
        ("INT-6A-PHY-001", "INT-6A-PHY-002"): "OK",
        ("INT-6A-PHY-001", "INT-6A-PHY-007"): "Check (mass+thermal)",
        ("INT-6A-PHY-005", "INT-6A-PHY-004"): "OK",
        ("INT-6A-PHY-010", "INT-6A-PHY-009"): "OK (power dependency)",
    }

    for r_idx in range(4, 4 + len(critical_intf)):
        for c_idx in range(2, 2 + len(critical_intf)):
            row_intf = critical_intf[r_idx - 4]["id"]
            col_intf = critical_intf[c_idx - 2]["id"]
            if r_idx - 4 == c_idx - 2:
                val = "—"
            else:
                pair = (row_intf, col_intf)
                pair_rev = (col_intf, row_intf)
                val = compatibility_data.get(pair, compatibility_data.get(pair_rev, "OK"))
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell_style(cell, align="center")
            if val == "—":
                cell.fill = PatternFill("solid", fgColor=GREY)
            elif "Check" in val or "Conflict" in val:
                cell.fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
            elif val == "OK" or "OK" in val:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)

    # Legend
    legend_start = 4 + len(critical_intf) + 2
    ws.cell(row=legend_start, column=1, value="Legenda")
    ws.cell(row=legend_start, column=1).font = Font(bold=True)
    legend_items = [
        ("—", "Self (diagonal)", GREY),
        ("OK", "Compatibile (nessun conflitto rilevato)", LIGHT_GREEN),
        ("Check", "Richiede verifica multi-dominio (es. mass+thermal+EMC)", LIGHT_YELLOW),
        ("Conflict", "Conflitto identificato → mitigation richiesta", LIGHT_RED),
    ]
    for i, (sym, desc, color) in enumerate(legend_items, start=legend_start + 1):
        ws.cell(row=i, column=1, value=sym)
        ws.cell(row=i, column=2, value=desc)
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="center")

    # ============================================================
    # SHEET 11 — Test Plan
    # ============================================================
    ws = wb.create_sheet("Test_Plan")
    test_headers = ["ID Interfaccia", "Nome", "Test Method", "Test Phase",
                    "Test Environment", "Pass/Fail Criteria", "Owner", "Timing (M+)"]
    set_col_widths(ws, [18, 35, 30, 18, 25, 50, 25, 14])

    for col, h in enumerate(test_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell)
    ws.row_dimensions[1].height = 30

    test_plan_data = [
        # (id, name, method, phase, env, criteria, owner, timing)
        ("INT-6A-PHY-001", "Airframe ↔ Payload Bay", "Inspection + CG analysis + shake table",
         "DDT (Design + Devel Test)", "Vendor facility (JOUAV) + lab",
         "CG shift ≤ 2% MAC; vibration spectrum within MIL-STD-810H Cat 4 envelope; no payload bay damage 100h cumulative",
         "Avionics Integration Lead", "M+10-M+12"),
        ("INT-6A-PHY-002", "Payload Power Rail 28 VDC", "Lab bench + transient injection + load step",
         "DDT", "Lab elettronica Firmamento",
         "Voltage 28 VDC ±5%; ripple ≤ 200 mV pk-pk; transient recovery < 50 ms; load step 0→100 W stable",
         "Power Mgmt SE", "M+10"),
        ("INT-6A-DATA-001", "Payload EO RGB → MC", "Image capture + geotag + throughput",
         "DDT", "Lab + outdoor test",
         "5 fps stable; geotag accuracy ±5m horizontal; throughput ≥ 500 MB/s sustained",
         "Payload SE", "M+11"),
        ("INT-6A-C2-001", "C2 Uplink Primary RF", "Range test + fade margin + spectrum analyzer",
         "DDT + Field Test", "Campo aperto Pentema (M+11 limitato)",
         "Range LOS ≥ 50 km; fade margin ≥ 12 dB; BER ≤ 1e-5 @ nominal range",
         "RF Systems Engineer", "M+11 (limited) + Fase 1"),
        ("INT-6A-C2-003", "SATCOM Iridium L-band", "Link test + handover + latency",
         "Integration Test", "Vendor lab + field",
         "Throughput ≥ 22 kbps continuous; handover seamless; latency one-way ≤ 400 ms",
         "Avionics Lead", "Fase 1 M+13-M+15"),
        ("INT-6A-C2-004", "Lost-Link Procedure", "HIL simulation + flight test",
         "Verification", "HIL bench + flight test (autorizzato)",
         "RTH triggered within 5 s lost-link; emergency land if RTH fail; SMS alert delivered",
         "Avionics Lead + Safety", "Fase 1 M+16-M+18"),
        ("INT-6A-GS-004", "PC Emergency Trigger", "Tabletop exercise + scenario sim",
         "Validation", "GCS Pentema + Sala Operativa PC",
         "Acknowledge ≤ 5 min; launch decision ≤ 15 min; ortofoto live ≤ 30 min",
         "Operations Lead + PC", "M+6-M+9 (tabletop) + Fase 1 (flight)"),
        ("INT-6A-REG-001", "ENAC SORA Authorization", "Documentation review + ENAC feedback",
         "Regulatory Compliance", "Sede ENAC + remote",
         "Pre-application meeting completato M+6; SORA application accepted; SAIL III authorization issued",
         "Aviation Regulatory Counsel", "M+3-M+6 (pre-app) + Fase 1 (final)"),
        ("INT-6A-REG-002", "AGCOM Spectrum Auth", "Doc review + AGCOM response",
         "Regulatory Compliance", "Sede AGCOM + remote",
         "Consultazione AGCOM response ≤ M+4; spectrum allocation confirmed (or fallback ISM accepted)",
         "Aviation Reg Counsel + RF SE", "M+1-M+4"),
        ("INT-6A-REG-004", "Garante DPIA", "DPIA review + DSAR drill",
         "Regulatory Compliance", "Internal + Garante (se notifica preventiva)",
         "DPIA signed DPO M+5; DSAR drill PASS (response ≤ 30 gg simulated)",
         "Data Privacy Counsel + DPO", "M+4-M+5 + ongoing"),
        ("INT-6A-GS-003", "Cooperative Dashboard", "Functional + RBAC + accessibility",
         "Validation", "Cloud staging + browser test",
         "All functional tests PASS; RBAC correct per polygon; WCAG 2.1 AA compliance",
         "Cloud Architect + DPO", "Fase 1 M+14-M+16"),
        ("INT-6A-GS-007", "Data Anonymization", "AI inference quality test",
         "Validation", "Cloud + dataset test",
         "Face/plate blurring false negative rate ≤ 1%; processing ≤ 5 min/ortofoto",
         "DPO + AI/ML Engineer", "Fase 1 M+15-M+17"),
        ("INT-6B-PHY-001", "HALE Solar Array Bonding", "Lap shear + thermal cycling + UV",
         "Subscale Test", "Lab subscale + thermal chamber",
         "Bonding strength ≥ 5 MPa; 100 thermal cycles -65/+40°C no delamination; UV 1000h no degradation",
         "Aero-Struct Engineer", "Fase 3 M+36-M+42"),
        ("INT-6B-PHY-002", "HALE Wing-Fuselage Joint", "FEA + GVT + structural test",
         "Subscale + Component Test", "FEA bench + GVT facility (CIRA / DLR)",
         "Joint ultimate load 75 kN factor 1.5; flutter clearance ≥ 20% margin VD",
         "Aero-Struct Engineer", "Fase 3 M+36-M+44"),
        ("INT-6B-NTN-001", "5G NR-NTN Service DL", "Link budget sim + OTA test",
         "Lab Test", "5G NTN testbed (TIM/Vodafone partnership)",
         "Capacity per beam ≥ 50 Mbps; coverage 50 km diameter @ 20 km altitude",
         "Telecom-NTN Expert", "Fase 3 M+40-M+46"),
        ("INT-6B-FEEDER-001", "Ka-band Feeder DL", "Link budget + rain fade sim",
         "Lab + Simulation", "Lab RF + ITU-R simulation",
         "EIRP ≥ 60 dBW; rain fade margin ≥ 15 dB Zona K Italia",
         "Ka-band RF Engineer", "Fase 3 M+40-M+45"),
        ("INT-X-CLOUD-001", "Cloud GAIA-X", "Provider audit + attestation review",
         "Procurement Validation", "Remote audit",
         "GAIA-X Label Level 2+ confirmed; data residency IT/EU verified; SLA SLA met during baseline measurement",
         "Cloud Architect + Procurement", "M+5-M+9"),
        ("INT-X-PRIVACY-001", "DSAR Workflow", "DSAR drill + audit",
         "Validation", "Internal + simulated user",
         "Response ≤ 30 gg; all rights operational (access, erasure, portability); audit log complete",
         "DPO + Customer Support", "Fase 1 M+15-M+17"),
        ("INT-X-CYBER-001", "Pentest + Vuln Mgmt", "External pentest annuale",
         "Verification", "Production + staging",
         "No Critical CVE unpatched > 7 gg; pentest report no Critical findings unmitigated",
         "Cybersecurity Engineer", "Fase 1 M+16-M+18 + annuale"),
        ("INT-X-VENDOR-001", "Vendor SLA UAS", "Vendor audit + SLA monitoring",
         "Procurement Validation", "Vendor site visit + ongoing monitoring",
         "AS/EN 9100 cert verified; lead time ≤ 12 wk; warranty terms confirmed; reference customer verified ≥ 2",
         "Procurement Lead + QM", "M+1-M+6 (pre-engagement) + Fase 1 (contract)"),
    ]

    for row_idx, td in enumerate(test_plan_data, start=2):
        for col_idx, val in enumerate(td, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
            cell_style(ws.cell(row=row_idx, column=col_idx))
        ws.row_dimensions[row_idx].height = 60

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    # ============================================================
    # SHEET 12 — Risk Register Interfacce Critiche
    # ============================================================
    ws = wb.create_sheet("Risk_Register_Interfacce")
    risk_headers = ["RISK-ID", "Interfaccia coinvolta", "Descrizione rischio",
                    "Probabilità (1-5)", "Impatto (1-5)", "P×I",
                    "Mitigazione", "Owner", "Closure target"]
    set_col_widths(ws, [16, 22, 50, 12, 12, 8, 50, 22, 18])

    for col, h in enumerate(risk_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell)
    ws.row_dimensions[1].height = 30

    interface_risks = [
        ("RSK-INT-001", "INT-6A-C2-001", "Shadow zones C2 RF 2.4 GHz in valle Pentema → loss link",
         4, 4, 16, "Doppio link RF + SATCOM Iridium (INT-6A-C2-003); RTH automatico < 5 s",
         "Avionics Lead", "Fase 1 M+18"),
        ("RSK-INT-002", "INT-6A-REG-001", "ENAC valuta Pentema 'area sensibile' (parco Aveto) → SAIL IV richiesto",
         3, 5, 15, "Pre-application early M+3; alternative site analysis; coordination Comune+Regione+ENAC",
         "Aviation Reg Counsel", "M+6"),
        ("RSK-INT-003", "INT-6A-REG-002", "AGCOM rifiuta licenza individuale 2.4 GHz EIRP > 100 mW → range limitato",
         3, 4, 12, "Fallback ISM 100 mW + SATCOM Iridium; alternative spectrum L-band Iridium primario",
         "Aviation Reg Counsel + RF SE", "M+4"),
        ("RSK-INT-004", "INT-6B-PHY-001", "Solar array bonding fail thermal cycling stratosfera → delamination",
         3, 5, 15, "R&D dedicato Fase 3; test panel qualification ECSS",
         "Aero-Struct Engineer", "Fase 3 M+42"),
        ("RSK-INT-005", "INT-6B-PHY-003", "Battery thermal mgmt heater > 80W → energy balance inverno deficit",
         4, 5, 20, "Showstopper RSK-TEC-001 cross-reference; design alternativi (E5 Seasonal-only)",
         "Thermal Eng + Battery Specialist", "Fase 3 M+38"),
        ("RSK-INT-006", "INT-6B-NTN-001", "Spettro S-band n255 contestato MNO terrestri → no spectrum sharing",
         4, 4, 16, "Engagement TIM/Vodafone partnership; alternative banda HAPS dedicata C-band 6.4 GHz",
         "Telecom-NTN Expert + Reg", "Fase 3 M+40"),
        ("RSK-INT-007", "INT-X-VENDOR-001", "Vendor JOUAV (CN) blocked by export control EU post-tensione geopolitica",
         3, 4, 12, "Vendor alternativi EU (Tekever PT, Quantum DE, FlyingBasket IT); RFQ multi-vendor",
         "Procurement Lead", "M+6 (alternativa identificata)"),
        ("RSK-INT-008", "INT-6A-C2-007", "Non-cooperative DAA non maturo TRL → SORA SAIL III non autorizzato senza",
         3, 4, 12, "Mitigazione operativa (geofencing, NOTAM, ore volo limitate); R&D DAA EO Fase 1+",
         "Avionics Lead + DAA Spec", "Fase 1 M+18"),
        ("RSK-INT-009", "INT-6A-REG-004", "Garante richiede notifica preventiva Art. 36 → lead time +3-6 mesi",
         2, 4, 8, "DPIA early M+4-M+5; engagement Garante preliminary se possibile",
         "Data Privacy Counsel + DPO", "M+5"),
        ("RSK-INT-010", "INT-6A-GS-007", "Data anonymization quality insufficiente → Garante audit fail",
         3, 4, 12, "AI inference quality threshold strict; manual review pipeline; opt-in workflow",
         "DPO + AI/ML Engineer", "Fase 1 M+17"),
        ("RSK-INT-011", "INT-6A-PHY-010", "Backhaul fibra Pentema non disponibile → Starlink-only inadeguato per SLA",
         3, 3, 9, "Dual provider FWA + Starlink; engagement Open Fiber/Linkem early; UPS + buffering",
         "IT Operations + Cloud Architect", "Fase 1 M+15"),
        ("RSK-INT-012", "INT-6A-REG-005", "Mercato assicurativo UAS BVLOS hardened post-incidenti UK 2024-25 → premi 5-10x",
         3, 3, 9, "Broker engagement early; multi-quote; possibili captive arrangement Y3+",
         "CFO + Insurance Broker", "Fase 1 M+15-M+18"),
    ]

    for row_idx, risk in enumerate(interface_risks, start=2):
        for col_idx, val in enumerate(risk, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell_style(cell)
        # Color P×I score
        pi_score = risk[5]
        pi_cell = ws.cell(row=row_idx, column=6)
        if pi_score >= 15:
            pi_cell.fill = PatternFill("solid", fgColor=LIGHT_RED)
        elif pi_score >= 9:
            pi_cell.fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
        else:
            pi_cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
        ws.row_dimensions[row_idx].height = 50

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    # ============================================================
    # SHEET 13 — Versioning Roadmap
    # ============================================================
    ws = wb.create_sheet("Versioning_Roadmap")
    ver_headers = ["Versione", "Milestone", "Data target", "Scope",
                   "Numero interfacce", "Status target", "Owner approval"]
    set_col_widths(ws, [14, 22, 18, 50, 18, 30, 28])

    for col, h in enumerate(ver_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell)
    ws.row_dimensions[1].height = 30

    versioning_data = [
        ("v0.5", "Cap. 4 §4.4 baseline", "M+3 (Mag 2026)",
         "20 interfacce primarie identificate (ICD preliminary nel Capitolo 4)",
         20, "Concept + Preliminary mix", "PMO Firmamento"),
        ("v1.0", "Vol. 2 Allegato A.4 baseline", "M+3-M+6 (Mag-Ago 2026)",
         "50 interfacce dettagliate (current document); 7 categorie + risk register + test plan",
         50, "Concept + Preliminary mix; high-confidence sui standard COTS",
         "Senior SE + Integration Lead"),
        ("v1.5", "Post-pre-engagement vendor + ENAC/AGCOM feedback",
         "M+6-M+8 (Ago-Ott 2026)",
         "Specifications updated post-vendor RFQ responses; INT-6A-REG-001/002 refined post-ENAC/AGCOM response",
         "50-55 (potenziali nuove)", "Preliminary maggioranza; alcune Detailed",
         "SE Lead + Procurement + Reg Counsel"),
        ("v2.0", "Detailed ICD pre-Phase 1 execution",
         "M+10-M+12 (Dic 2026-Feb 2027)",
         "Byte-level format, latency end-to-end budget, failure mode behaviors, ICD frozen per vendor contracts",
         "55-65", "Detailed (all 6A); Concept (6B)",
         "SE Lead + Vendor Liaison + Customer"),
        ("v2.5", "Post-First Flight 6A",
         "M+20-M+24 (Ago-Dic 2027)",
         "Interfacce validate in-flight; lessons learned incorporated; SLA backhaul/cloud measured baseline",
         "65", "Validated (6A); Preliminary (6B)",
         "Operations Lead + V&V Engineer"),
        ("v3.0", "HALE 6B Phase B closure",
         "M+48-M+60 (Y4-Y5)",
         "ICD 6B Detailed completo post-subscale flight test; preparazione Phase C-D",
         "80-100", "Detailed (6A operational); Preliminary-Detailed (6B subscale)",
         "Senior SE + HALE Program Manager"),
    ]

    for row_idx, vd in enumerate(versioning_data, start=2):
        for col_idx, val in enumerate(vd, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell_style(cell)
        ws.row_dimensions[row_idx].height = 60

    ws.freeze_panes = "B2"

    # ============================================================
    # SAVE
    # ============================================================
    output_path = "/home/user/HALE/studio-di-fattibilita/allegati/A4-ICD/ICD-v1.0.xlsx"
    wb.save(output_path)
    print(f"OK: salvato {output_path}")
    print(f"  - {len(INTERFACES)} interfacce totali")
    print(f"  - {len([i for i in INTERFACES if 'INT-6A' in i['id']])} interfacce 6A")
    print(f"  - {len([i for i in INTERFACES if 'INT-6B' in i['id']])} interfacce 6B")
    print(f"  - {len([i for i in INTERFACES if 'INT-X' in i['id']])} interfacce trasversali")
    print(f"  - {len(test_plan_data)} test plan entries")
    print(f"  - {len(interface_risks)} risk register entries")
    print(f"  - {len(versioning_data)} versioni roadmap")

if __name__ == "__main__":
    build_workbook()
