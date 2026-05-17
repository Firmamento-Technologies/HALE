#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==============================================================================
A.5 -- Verification & Validation Plan  --  Excel workbook builder
Firmamento Technologies -- Studio di Fattibilita' HALE / VTOL
==============================================================================

Scope
-----
Genera VV-Plan-v1.0.xlsx (Volume 2 Allegato A.5) multi-sheet:

  Sheet 1: Cover
  Sheet 2: VV_Matrix_SyR        (60-70 SyR x metodo x fase x owner)
  Sheet 3: VV_Methods_Detail
  Sheet 4: Test_Facilities      (lab equipment, wind tunnel, GVT, range)
  Sheet 5: Test_Schedule        (Gantt)
  Sheet 6: Test_Costs           (CapEx + OpEx test)
  Sheet 7: Independent_Verification

Allineato con:
  - cap-03-requisiti-e-RTM.md  §3.7 V&V Plan preliminare
  - cap-06-analisi-tecnica.md  §6.6 V&V tecnica
  - cap-09-cronoprogramma-e-gate.md  Gate G3, G4, G5, G6 criteri

Methodology
-----------
NASA SE Handbook Rev 2 §5.3 (Verification) + §5.4 (Validation)
4 metodi standard: Inspection / Analysis / Demonstration / Test (I/A/D/T)
Allineamento per fase: Pre-Phase A / Phase A / Phase B / Phase C-D

Author: V&V Manager (synthetic)  --  Claude Code
Date  : 2026-05-17
==============================================================================
"""

from __future__ import annotations

import os
from typing import List, Dict, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# -----------------------------------------------------------------------------
OUT_DIR  = "/home/user/HALE/studio-di-fattibilita/allegati/A5-VV-Plan"
os.makedirs(OUT_DIR, exist_ok=True)
XLSX_PATH = os.path.join(OUT_DIR, "VV-Plan-v1.0.xlsx")

# -----------------------------------------------------------------------------
# STYLES
# -----------------------------------------------------------------------------
HEADER_FILL  = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
SUBHDR_FILL  = PatternFill("solid", fgColor="BDD7EE")
SUBHDR_FONT  = Font(bold=True, size=10)
ZONE_FILL    = {
    "I": PatternFill("solid", fgColor="E2EFDA"),
    "A": PatternFill("solid", fgColor="DDEBF7"),
    "D": PatternFill("solid", fgColor="FFF2CC"),
    "T": PatternFill("solid", fgColor="FCE4D6"),
}
OK_FILL      = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL    = PatternFill("solid", fgColor="FFC7CE")
THIN_BORDER  = Border(left=Side(style="thin", color="888888"),
                      right=Side(style="thin", color="888888"),
                      top=Side(style="thin", color="888888"),
                      bottom=Side(style="thin", color="888888"))


def style_header(ws, row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def autosize(ws, max_col: int, base: int = 14) -> None:
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        max_len = base
        for cell in ws[col_letter]:
            v = cell.value
            if v is None:
                continue
            length = len(str(v))
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


# =============================================================================
# 1. VV MATRIX -- 65 System Requirements x Verification method x phase x owner
# =============================================================================
# Columns: Req-ID, Description (short), Family, Method, Phase, Owner, Test facility,
#          Document evidence, Status, Risk-ID link, Trade Study link, Notes
# Family : F/P/O/S/E/C/Cost
# Method : I=Inspection / A=Analysis / D=Demonstration / T=Test (combinations allowed)
# Phase  : Pre-A / A / B / C
# =============================================================================

VV_MATRIX: List[Dict[str, str]] = [
    # ---- FUNCTIONAL (F) -----------------------------------------------------
    dict(req_id="SyR-F-001", desc="Sistema esegue missione EO sorvolando Pentema (waypoint-based)",
         family="F", method="D",   phase="B",  owner="vtol-uas-specialist",
         facility="Sito Pentema + GS mobile", evidence="Test report MIS-F-001",
         status="Planned", risk="–", ts="TS-PLATFORM-6A",
         note="Demo end-to-end mission ConOps"),
    dict(req_id="SyR-F-002", desc="Sistema esegue ricognizione area (raster) <=15 km^2",
         family="F", method="A+D", phase="A,B", owner="aerospace-SE",
         facility="Simulator + sito Pentema", evidence="Mission plan analysis + demo log",
         status="Planned", risk="–", ts="TS-PAYLOAD-EO",
         note="A: copertura geometrica, D: raster reale"),
    dict(req_id="SyR-F-003", desc="Rileva hotspot >=40 degC, alert + thumbnail in <=5 min",
         family="F", method="D+T", phase="B,C", owner="EO-expert",
         facility="Test bed fuoco controllato + GS", evidence="VR-F-003.1 + VR-F-003.2",
         status="Planned", risk="RSK-OPS-002", ts="–",
         note="Demo scenario fuoco + analysis FAR target <=5%"),
    dict(req_id="SyR-F-004", desc="Genera ortomosaico georef <=8 h dalla landing",
         family="F", method="A+D", phase="B",  owner="EO-expert",
         facility="Pipeline cloud processing", evidence="Analysis pipeline + demo timing",
         status="Planned", risk="–", ts="–",
         note="Catena Pix4D/OpenDroneMap on-premise"),
    dict(req_id="SyR-F-005", desc="Fornisce link voice/dati P2P emergenza durante evento",
         family="F", method="D+T", phase="B",  owner="telecom-payload",
         facility="Site Pentema + LTE emulator", evidence="Test report COMMS-F-005",
         status="Planned", risk="–", ts="TS-COMMS",
         note="Demo con apparato tactical LTE Athonet/Druid"),

    # ---- PERFORMANCE (P) ----------------------------------------------------
    dict(req_id="SyR-P-001", desc="Autonomia sortie >=4 h con payload 4 kg",
         family="P", method="A+T", phase="A,C", owner="propulsion-energy-engineer",
         facility="Cella batterie + flight test", evidence="Analysis energy + test report",
         status="Planned", risk="–", ts="TS-PLATFORM-6A",
         note="A: energy balance VTOL, T: flight log timestamp"),
    dict(req_id="SyR-P-002", desc="Velocita' cruise nominale 90-110 km/h",
         family="P", method="A+T", phase="A,C", owner="aerospace-SE",
         facility="XFLR5 + flight test", evidence="Polare + flight log",
         status="Planned", risk="–", ts="–",
         note="Polare drag preliminary + GNSS speed log"),
    dict(req_id="SyR-P-003", desc="Quota operativa nominale 500-1500 m AGL Appennino",
         family="P", method="A+T", phase="A,C", owner="vtol-uas-specialist",
         facility="Sito Pentema", evidence="Flight log altimeter",
         status="Planned", risk="–", ts="–",
         note="Verifica autorizzata SAIL III ENAC"),
    dict(req_id="SyR-P-004", desc="Resistenza vento sostenuto <=17 m/s",
         family="P", method="T",   phase="C",  owner="vtol-uas-specialist",
         facility="Sito di volo windy + meteo", evidence="Flight test wind envelope",
         status="Planned", risk="RSK-OPS-001", ts="–",
         note="Demonstrated handling 15-17 m/s gusts"),
    dict(req_id="SyR-P-005", desc="GSD camera EO <=5 cm @ 500 m AGL",
         family="P", method="A+T", phase="A,B", owner="EO-expert",
         facility="Camera test bench + cal range", evidence="Spec sheet + cal report",
         status="Planned", risk="–", ts="TS-PAYLOAD-EO",
         note="GSD = pixel_size * d / focal_length"),
    dict(req_id="SyR-P-006", desc="NEdT IR sensor <=50 mK",
         family="P", method="A+T", phase="A,B", owner="EO-expert",
         facility="Blackbody source lab", evidence="Calibration cert + lab report",
         status="Planned", risk="–", ts="TS-PAYLOAD-EO",
         note="Verifica con corpo nero a temperatura controllata"),
    dict(req_id="SyR-P-007", desc="C2 link RF range >=30 km LOS, fade margin >=12 dB",
         family="P", method="A+T", phase="A,C", owner="telecom-payload",
         facility="Range test ground + flight test", evidence="Link budget + RSSI log",
         status="Planned", risk="RSK-TEC-005", ts="TS-COMMS",
         note="Analisi: allegato A.7 link budget. Test: RSSI in volo."),

    # ---- OPERATIONAL (O) ----------------------------------------------------
    dict(req_id="SyR-O-001", desc="Lost-Link Return-to-Base profile <=120 s dall'evento",
         family="O", method="D",   phase="B",  owner="avionics-gnc-engineer",
         facility="HIL simulator + flight test", evidence="HIL run + flight test trigger",
         status="Planned", risk="RSK-TEC-007", ts="TS-AVI-6A",
         note="Demo deliberate link-loss test"),
    dict(req_id="SyR-O-002", desc="Set-up GS deployable in <=30 min (training operator)",
         family="O", method="D",   phase="B,C", owner="vtol-uas-specialist",
         facility="GS mobile in field", evidence="Time-motion demo",
         status="Planned", risk="–", ts="–",
         note="Cronometro setup field di operator trained"),
    dict(req_id="SyR-O-003", desc="Swap payload modulare ground-side <=90 min (target 30)",
         family="O", method="D",   phase="B",  owner="vtol-uas-specialist",
         facility="Hangar test", evidence="Time-motion demo + cal post-swap",
         status="Planned", risk="–", ts="–",
         note="Red Team review §6.8 Cap.6: 30 min e' stretch, 60-90 realistic"),

    # ---- SAFETY (S) ---------------------------------------------------------
    dict(req_id="SyR-S-001", desc="FMECA preliminary completo con RPN top 5 con mitigation",
         family="S", method="I+A", phase="A",  owner="systems-engineer",
         facility="Office (analysis)", evidence="Allegato A.2 Risk Register + FMECA",
         status="Done",    risk="–", ts="–",
         note="Inspection: existence + completeness"),
    dict(req_id="SyR-S-002", desc="FTA top event 'Loss of Vehicle in BVLOS' <= 10^-5/h",
         family="S", method="A",   phase="A",  owner="systems-engineer",
         facility="Office (analysis)", evidence="FTA report Vol.2 A.2",
         status="Planned", risk="RSK-TEC-007", ts="–",
         note="Target SAIL III ENAC = 10^-5/h"),
    dict(req_id="SyR-S-003", desc="Sistema parachute deploy in <=3 s da trigger",
         family="S", method="T",   phase="B",  owner="vtol-uas-specialist",
         facility="Drop test rig", evidence="HSC video drop test + log",
         status="Planned", risk="–", ts="–",
         note="Drop test indoor o esterno, alt 50 m, deployment time on log"),
    dict(req_id="SyR-S-004", desc="Geofence aree residenziali enforced (no overflight built-up)",
         family="S", method="D+I", phase="B,C", owner="avionics-gnc-engineer",
         facility="HIL + GS", evidence="HIL test report + ConOps doc inspection",
         status="Planned", risk="RSK-TEC-006", ts="–",
         note="Demo penetration attempt -> auto-redirect"),

    # ---- ENVIRONMENTAL (E) --------------------------------------------------
    dict(req_id="SyR-E-001", desc="Operazioni in T amb. -10 / +30 degC",
         family="E", method="T",   phase="C",  owner="vtol-uas-specialist",
         facility="Climatic chamber + sito Pentema", evidence="Cold-soak + hot-soak test",
         status="Planned", risk="RSK-OPS-001", ts="–",
         note="Lab climatic chamber + field winter test"),
    dict(req_id="SyR-E-002", desc="Operazioni in pioggia <=10 mm/24h",
         family="E", method="I+T", phase="C",  owner="vtol-uas-specialist",
         facility="Rain test rig + field", evidence="IP rating cert + field test",
         status="Planned", risk="–", ts="–",
         note="IP54 minimum vendor spec + field validation"),
    dict(req_id="SyR-E-003", desc="Storage shelter UAV +/- 5 degC ambient operation",
         family="E", method="I",   phase="C",  owner="vtol-uas-specialist",
         facility="Hangar Pentema", evidence="Spec sheet hangar + thermal log",
         status="Planned", risk="–", ts="–",
         note="Walkthrough hangar + datalogger 30 giorni"),

    # ---- COMPLIANCE (C) -----------------------------------------------------
    dict(req_id="SyR-C-001", desc="UAS classified Class C5 + tipo Specific Category",
         family="C", method="I",   phase="Pre-A,A", owner="aviation-regulatory-counsel",
         facility="Doc review", evidence="DoC + EASA UAS class label",
         status="Planned", risk="–", ts="–",
         note="Inspection: vendor DoC + EU Reg 2019/945"),
    dict(req_id="SyR-C-002", desc="ConOps SORA SAIL <=III Pentema BVLOS approved ENAC",
         family="C", method="I",   phase="Pre-A,A", owner="aviation-regulatory-counsel",
         facility="ENAC office (paper)", evidence="ENAC autorizzazione + GRC computation",
         status="Open",    risk="RSK-REG-002", ts="–",
         note="Pre-application meeting M+3-6, full sub M+6"),
    dict(req_id="SyR-C-003", desc="Spettro radio C2 conforme AGCOM/PNRF",
         family="C", method="I",   phase="Pre-A,A", owner="telecom-payload",
         facility="AGCOM (paper)", evidence="ETSI EN 300 328 DoC + license se >100mW",
         status="Planned", risk="–", ts="TS-COMMS",
         note="Inspection: EN 300 328 + EIRP measurement"),
    dict(req_id="SyR-C-004", desc="DPIA per missioni con immagini privacy compliant",
         family="C", method="I",   phase="A,B", owner="data-privacy-counsel",
         facility="Office", evidence="DPIA + audit Garante",
         status="Planned", risk="RSK-TEC-006", ts="–",
         note="Pubblica DPIA + workshop comunita' Pentema"),
    dict(req_id="SyR-C-005", desc="Insurance per BVLOS coverage TPL >=10 M EUR",
         family="C", method="I",   phase="B",  owner="program-manager",
         facility="Broker office", evidence="Insurance policy doc",
         status="Planned", risk="–", ts="–",
         note="Verifica polizza con specific endorsement BVLOS"),
    dict(req_id="SyR-C-006", desc="Cybersecurity compliance NIS2 Article 21",
         family="C", method="I+A", phase="A,B", owner="aerospace-SE",
         facility="Audit ACN-compliant", evidence="NIS2 self-assessment + 3rd-party audit",
         status="Planned", risk="–", ts="–",
         note="ACN reference framework + ENISA guidelines"),

    # ---- COST (Cost) --------------------------------------------------------
    dict(req_id="SyR-Cost-001", desc="CapEx Y1 <=1.2 M EUR (target 900k)",
         family="Cost", method="I+A", phase="A,B", owner="financial-cfo-analyst",
         facility="Office", evidence="Quadro Economico + Allegato Financial Model",
         status="Planned", risk="RSK-FIN-001", ts="–",
         note="Inspection budget + Analysis sensitivity"),
    dict(req_id="SyR-Cost-002", desc="OpEx Y2 run-rate <=600k EUR (incl. regulatory FTE)",
         family="Cost", method="I",   phase="B,C", owner="financial-cfo-analyst",
         facility="Office", evidence="Financial report Y1+Y2",
         status="Planned", risk="RSK-FIN-001", ts="–",
         note="Post audit DR-014 OpEx aggiornato +600k personnel reg."),
    dict(req_id="SyR-Cost-003", desc="Revenue Y1 cumulato >=200 k EUR (cf. Cap.7)",
         family="Cost", method="I",   phase="C",  owner="business-model-strategist",
         facility="Office", evidence="Revenue ledger Y1",
         status="Planned", risk="–", ts="–",
         note="Tracking commercial pipeline pilot"),

    # ---- AERO/STRUCT (SsR samples flagged for VV reporting) -----------------
    dict(req_id="SsR-AERO-001", desc="L/D cruise >=15 (HALE) o aerod. envelope VTOL",
         family="P", method="A+T", phase="A,B", owner="aero-structures-engineer",
         facility="XFLR5 + wind tunnel subscale", evidence="CFD report + wind tunnel test",
         status="Planned", risk="–", ts="TS-MATERIAL",
         note="Phase A CFD + Phase B wind tunnel 1:5"),
    dict(req_id="SsR-AERO-002", desc="Limit load factor n=+3.8/-1.5 (CS-23 equivalent)",
         family="S", method="A+T", phase="A,B", owner="aero-structures-engineer",
         facility="FEA + structural test", evidence="FEA report + test rig",
         status="Planned", risk="–", ts="TS-MATERIAL",
         note="Structural test ad ultimate load + 50%"),
    dict(req_id="SsR-AERO-003", desc="Flutter speed margin >=20% rispetto VD",
         family="S", method="A+T", phase="B",  owner="aero-structures-engineer",
         facility="GVT + aeroelastic analysis", evidence="GVT report + p-k analysis",
         status="Planned", risk="RSK-TEC-002", ts="–",
         note="Ground Vibration Test prima del flight test"),

    # ---- PROP/ENERGY (HALE & VTOL) ------------------------------------------
    dict(req_id="SsR-PROP-001", desc="Energy balance HALE inverno 44N margine >=20%",
         family="P", method="A",   phase="A",  owner="propulsion-energy-engineer",
         facility="Sim. solar + DLR data", evidence="Allegato A.7 Energy Balance Report",
         status="Done",    risk="RSK-TEC-001", ts="TS-PROP-6B",
         note="Sim. completa M+10 (gate G3)"),
    dict(req_id="SsR-PROP-002", desc="Pannello solare efficienza modulo >=24%",
         family="P", method="I+T", phase="B",  owner="propulsion-energy-engineer",
         facility="Solar simulator lab", evidence="IEC 60904 test cert",
         status="Planned", risk="–", ts="TS-PROP-6B",
         note="Verifica con flash solar simulator AM1.5"),
    dict(req_id="SsR-PROP-003", desc="Batteria pack-level energy density >=350 Wh/kg (LiS)",
         family="P", method="I+T", phase="B,C", owner="propulsion-energy-engineer",
         facility="Battery cycler lab", evidence="Discharge curve + spec",
         status="Planned", risk="RSK-TEC-001", ts="TS-PROP-6B",
         note="Test pack-level a C/3 discharge 0-100%"),
    dict(req_id="SsR-PROP-004", desc="Engine VTOL hybrid runtime >=4 h, fuel <=8 kg",
         family="P", method="T",   phase="C",  owner="propulsion-energy-engineer",
         facility="Engine test cell + flight test", evidence="Bench run + flight log",
         status="Planned", risk="–", ts="–",
         note="Cell test 4h continuous + correlate flight"),

    # ---- AVI/GNC ------------------------------------------------------------
    dict(req_id="SsR-AVI-001", desc="Autopilota FCS DAL-C (DO-178C objectives lvl C)",
         family="S", method="I+A", phase="B",  owner="avionics-gnc-engineer",
         facility="Test bed + audit DO-178C", evidence="DAL-C cert + verification trace",
         status="Planned", risk="–", ts="TS-AVI-6A",
         note="If commercial autopilot, check vendor DAL cert"),
    dict(req_id="SsR-AVI-002", desc="GNSS spoofing/jamming detection + alert",
         family="S", method="T",   phase="B",  owner="avionics-gnc-engineer",
         facility="GNSS sim. + RF chamber", evidence="Spoof test + alert log",
         status="Planned", risk="RSK-TEC-005", ts="–",
         note="GNSS simulator with replay/spoofing scenarios"),
    dict(req_id="SsR-AVI-003", desc="IMU redundancy 2 of 3 voting active",
         family="S", method="D",   phase="B",  owner="avionics-gnc-engineer",
         facility="HIL test", evidence="HIL fault injection report",
         status="Planned", risk="–", ts="TS-AVI-6A",
         note="Fault injection con IMU disabled"),
    dict(req_id="SsR-AVI-004", desc="Lost-Link auto-RTB profile loaded + tested",
         family="O", method="D",   phase="B,C", owner="avionics-gnc-engineer",
         facility="HIL + flight test", evidence="HIL + flight trigger test",
         status="Planned", risk="RSK-TEC-007", ts="TS-AVI-6A",
         note="OSO #9 SORA compliance"),

    # ---- PAYLOAD (EO/IR) ----------------------------------------------------
    dict(req_id="SsR-PAY-001", desc="Camera RGB 24 MP rolling/global shutter <=1/2000s",
         family="P", method="I+T", phase="B",  owner="EO-expert",
         facility="Cal range optical bench", evidence="Spec + MTF measurement",
         status="Planned", risk="–", ts="TS-PAYLOAD-EO",
         note="MTF and motion-blur lab measurement"),
    dict(req_id="SsR-PAY-002", desc="IR LWIR 640x512 NEdT <=50 mK GSD <=5 m @500 m AGL",
         family="P", method="I+T", phase="B",  owner="EO-expert",
         facility="Blackbody lab + cal range", evidence="Cal cert + lab report",
         status="Planned", risk="–", ts="TS-PAYLOAD-EO",
         note="NUC + crosscheck con RGB (cf. FMECA item RPN 48)"),
    dict(req_id="SsR-PAY-003", desc="Onboard storage redundant RAID, >=1 TB sortie capacity",
         family="P", method="I+T", phase="B",  owner="EO-expert",
         facility="Bench test rig", evidence="Spec + IOPS log",
         status="Planned", risk="–", ts="–",
         note="Storage cycle test + fault injection"),
    dict(req_id="SsR-PAY-004", desc="Edge AI hotspot detection FAR <=5% per 1k events",
         family="F", method="A+T", phase="B,C", owner="EO-expert",
         facility="Synthetic dataset + field trial", evidence="ROC curve + field test",
         status="Planned", risk="RSK-OPS-002", ts="–",
         note="Algorithm pre-training su dataset Italia"),

    # ---- COMMS --------------------------------------------------------------
    dict(req_id="SsR-COMMS-001", desc="C2 link RF + SATCOM fade margin >=12 dB",
         family="P", method="A+T", phase="A,C", owner="telecom-payload",
         facility="Range test + flight test", evidence="Allegato A.7 LB + RSSI log",
         status="Planned", risk="RSK-TEC-005", ts="TS-COMMS",
         note="Link budget A.7 + measured fade margin in flight"),
    dict(req_id="SsR-COMMS-002", desc="Data downlink throughput >=50 Mbps EO 20 km LOS",
         family="P", method="A+T", phase="A,B", owner="telecom-payload",
         facility="Range test + spectrum analyser", evidence="LB + measured throughput",
         status="Planned", risk="–", ts="TS-COMMS",
         note="5 GHz UHF licensed; verifica con iperf in volo"),
    dict(req_id="SsR-COMMS-003", desc="Encryption AES-256 + auth on C2 link",
         family="S", method="I+T", phase="B",  owner="telecom-payload",
         facility="Crypto bench + penetration test", evidence="Crypto cert + pen-test report",
         status="Planned", risk="RSK-TEC-005", ts="–",
         note="Verifica FIPS 140-2 module + pen-test"),
    dict(req_id="SsR-COMMS-004", desc="HAPS service link C/N0 >=80 dB-Hz @25 km nadir",
         family="P", method="A",   phase="A",  owner="telecom-payload",
         facility="Office (analysis)", evidence="Allegato A.7 LB ServiceLink",
         status="Done",    risk="–", ts="TS-COMMS",
         note="Riferimento allegato A.7 -- LB_ServiceLink_6B"),
    dict(req_id="SsR-COMMS-005", desc="HAPS feeder link Ka 31 GHz availability 99.9%",
         family="P", method="A",   phase="A,B", owner="telecom-payload",
         facility="Office + range test Phase B", evidence="LB Ka + site coordination doc",
         status="Planned", risk="–", ts="TS-COMMS",
         note="Site diversity gateway secondary >10 km"),

    # ---- GROUND SEGMENT -----------------------------------------------------
    dict(req_id="SsR-GS-001", desc="GS riceve alert + push interfaccia PC in <=60 s",
         family="P", method="D",   phase="B,C", owner="aerospace-SE",
         facility="GS test bed + PC interface", evidence="Latency test report",
         status="Planned", risk="–", ts="–",
         note="Demo end-to-end con stopwatch"),
    dict(req_id="SsR-GS-002", desc="GS 24/7 monitoring CPU/storage health auto-alert",
         family="O", method="I+D", phase="B,C", owner="aerospace-SE",
         facility="GS Pentema", evidence="Monitoring dashboard + drill",
         status="Planned", risk="–", ts="–",
         note="Walkthrough check NOC + simulated outage"),
    dict(req_id="SsR-GS-003", desc="Data export PA (Regione/PC) standard INSPIRE WMS/WFS",
         family="O", method="I",   phase="C",  owner="EO-expert",
         facility="Cloud server", evidence="Endpoint test + INSPIRE compliance",
         status="Planned", risk="–", ts="–",
         note="WMS/WFS endpoints accessible to clients"),

    # ---- HALE 6B ADDITIONAL --------------------------------------------------
    dict(req_id="SsR-HALE-001", desc="HALE subscale 1:3 prototype TRL 5 by M+36",
         family="P", method="D+T", phase="B",  owner="aero-structures-engineer",
         facility="Test site Sardegna / EuroHAPS analog", evidence="Subscale flight test",
         status="Open",    risk="RSK-TEC-002", ts="–",
         note="Phase B 6B exit criterion gate G6"),
    dict(req_id="SsR-HALE-002", desc="HALE stratospheric flight >=24 h continuous",
         family="P", method="T",   phase="C",  owner="aero-structures-engineer",
         facility="Test site + chaser", evidence="Flight test perennial >24h",
         status="Open",    risk="RSK-TEC-001", ts="–",
         note="Out-of-scope studio attuale (Phase C+ HALE)"),
    dict(req_id="SsR-HALE-003", desc="HALE controllo posizione +/- 10 km vento strato",
         family="P", method="A+T", phase="B,C", owner="avionics-gnc-engineer",
         facility="HIL strat + flight test", evidence="HIL strat sim + flight log",
         status="Open",    risk="–", ts="–",
         note="Modello vento stratosferico WMO climatology"),

    # ---- INTERFACE REQUIREMENTS (IR) ----------------------------------------
    dict(req_id="IR-PAY-001", desc="Payload power interface 28 VDC +/- 5%, 200 W max",
         family="C", method="I+T", phase="B",  owner="aerospace-SE",
         facility="Bench power supply test", evidence="ICD inspection + test report",
         status="Planned", risk="–", ts="–",
         note="ICD allegato A.4 + test bench measurement"),
    dict(req_id="IR-PAY-002", desc="Payload data interface Ethernet 1 GbE",
         family="C", method="I+T", phase="B",  owner="aerospace-SE",
         facility="Bench test", evidence="ICD + iperf test",
         status="Planned", risk="–", ts="–",
         note="Verify connector, pinout, throughput"),
    dict(req_id="IR-GS-001", desc="GS-Cloud upload bandwidth >=100 Mbps shared",
         family="P", method="I+T", phase="B",  owner="aerospace-SE",
         facility="Pentema connectivity test", evidence="Speedtest + ISP SLA",
         status="Planned", risk="–", ts="–",
         note="Verifica connettivita' Open Fiber/TIM Pentema"),

    # ---- ADDITIONAL OPERATIONAL ---------------------------------------------
    dict(req_id="SyR-O-004", desc="Pilota UAS certificato STS-01 SORA dedicated",
         family="O", method="I",   phase="A,B", owner="vtol-uas-specialist",
         facility="Doc review", evidence="ENAC pilot certificate",
         status="Planned", risk="–", ts="–",
         note="Verifica certificazione operatore + pilot"),
    dict(req_id="SyR-O-005", desc="Maintenance interval pre-flight check <=30 min",
         family="O", method="D",   phase="B,C", owner="vtol-uas-specialist",
         facility="Hangar", evidence="Checklist run + timing",
         status="Planned", risk="–", ts="–",
         note="Manuale MM allegato A.X"),
    dict(req_id="SyR-O-006", desc="Mission planner GS automatizza waypoint + no-fly check",
         family="O", method="D",   phase="B",  owner="avionics-gnc-engineer",
         facility="GS test bench", evidence="Demo planning + no-fly enforcement",
         status="Planned", risk="–", ts="–",
         note="Auto-check NOTAMs + restricted areas"),

    # ---- ADDITIONAL SAFETY/ENVIRONMENTAL ------------------------------------
    dict(req_id="SyR-S-005", desc="Lithium battery thermal runaway containment box",
         family="S", method="T",   phase="B",  owner="propulsion-energy-engineer",
         facility="Battery abuse test cell", evidence="Nail penetration + thermal test",
         status="Planned", risk="–", ts="–",
         note="UN 38.3 + JEDEC battery abuse tests"),
    dict(req_id="SyR-E-004", desc="Operazioni soggette a finestra meteo nominale (no temporali)",
         family="E", method="I",   phase="B,C", owner="vtol-uas-specialist",
         facility="Meteo forecast service", evidence="ConOps doc + go/no-go decision tree",
         status="Planned", risk="RSK-OPS-001", ts="–",
         note="ARPA Liguria forecast feed integration"),
    dict(req_id="SyR-E-005", desc="EMI/EMC compliance EN 55032 + DO-160G section 20",
         family="C", method="I+T", phase="B",  owner="avionics-gnc-engineer",
         facility="EMC chamber accredited", evidence="EMC test report",
         status="Planned", risk="–", ts="–",
         note="Conducted + radiated emissions"),

    # ---- ADDITIONAL FUNCTIONAL ----------------------------------------------
    dict(req_id="SyR-F-006", desc="Generazione DEM 3D area >=10 km^2 con precisione <=20 cm",
         family="F", method="A+T", phase="B,C", owner="EO-expert",
         facility="Cloud processing + GCP cal", evidence="DEM accuracy report",
         status="Planned", risk="–", ts="–",
         note="Confronto con GCPs noti misurati GPS RTK"),
    dict(req_id="SyR-F-007", desc="Trasmissione live video bassa risoluzione 720p<=3s latency",
         family="F", method="T",   phase="B,C", owner="telecom-payload",
         facility="Range test + GS", evidence="Latency test report",
         status="Planned", risk="–", ts="TS-COMMS",
         note="Critical per PC use case real-time"),
    dict(req_id="SyR-F-008", desc="Logging black-box flight data + payload data on landing",
         family="F", method="I+D", phase="B",  owner="avionics-gnc-engineer",
         facility="Bench + flight test", evidence="Black-box dump + replay",
         status="Planned", risk="–", ts="–",
         note="Crash-survivable enclosure spec"),

    # ---- HALE FIDELITY ANALYSIS ---------------------------------------------
    dict(req_id="SsR-HALE-004", desc="Aeroelastic analysis HALE wing nonlineare",
         family="S", method="A",   phase="A,B", owner="aero-structures-engineer",
         facility="CFD/CSD coupled solver", evidence="Aeroelastic analysis report",
         status="Open",    risk="RSK-TEC-002", ts="–",
         note="Phase B exit criterion gate G6"),
    dict(req_id="SsR-HALE-005", desc="Solar panel encapsulation UV-degradation <2% / year",
         family="E", method="T",   phase="B",  owner="propulsion-energy-engineer",
         facility="UV chamber lab", evidence="Accelerated UV test report",
         status="Planned", risk="–", ts="TS-PROP-6B",
         note="ASTM G154/G155 standard"),
]

# ===========================================================================
# 2. TEST FACILITIES catalog
# ===========================================================================

FACILITIES: List[Dict[str, str]] = [
    dict(name="GS test bed Firmamento (Pentema o Genova HQ)",
         category="Ground Segment", owner="In-house",
         cost_capex_k="35", cost_opex_k_year="20",
         availability="Q3 2026", purpose="GS integration + walkthrough + maintenance training",
         phase="A,B", notes="Container 20'' attrezzato"),
    dict(name="HIL simulator Avionics",
         category="Avionics/GNC", owner="In-house (build) o partner DiPSIM",
         cost_capex_k="80", cost_opex_k_year="30",
         availability="Q4 2026", purpose="Hardware-In-The-Loop FCS fault injection",
         phase="A,B", notes="PIL/HIL setup with ARINC 818 / MIL-STD-1553"),
    dict(name="Range test ground RF (Pentema o Albenga)",
         category="Comms", owner="Field deployment",
         cost_capex_k="20", cost_opex_k_year="10",
         availability="Q3 2026", purpose="C2 + payload link measurement",
         phase="A,C", notes="Antenna tracking + spectrum analyzer Rohde&Schwarz"),
    dict(name="Wind tunnel subscale (IIT Genova o UniGE DICCA)",
         category="Aero", owner="Partner academic",
         cost_capex_k="0 (rental)", cost_opex_k_year="50",
         availability="Q2 2027", purpose="Subscale wind tunnel test 1:5",
         phase="B", notes="Subscale 1:5 dimensions ~ 1.5 m; testing Re ~ 5e5"),
    dict(name="Wind tunnel POLITO (CIRA / POLITO partnership)",
         category="Aero", owner="Partner academic",
         cost_capex_k="0 (rental)", cost_opex_k_year="80",
         availability="Q4 2027", purpose="HALE subscale 1:3 test",
         phase="B (Phase B 6B)", notes="Solo Phase B 6B; budget Phase B"),
    dict(name="GVT (Ground Vibration Test) -- POLITO DIMEAS o CIRA",
         category="Structures", owner="Partner academic",
         cost_capex_k="0", cost_opex_k_year="60",
         availability="2028 (subscale)", purpose="Modal analysis full vehicle",
         phase="B (6A subscale), C (6B)", notes="Multi-axis shaker rig + accelerometers"),
    dict(name="Solar simulator lab",
         category="Energy", owner="Partner ENEA Casaccia o EURAC Bolzano",
         cost_capex_k="0", cost_opex_k_year="20",
         availability="Q1 2027", purpose="Panel efficiency cert AM1.5",
         phase="B", notes="IEC 60904 / ASTM E948 flash + steady-state"),
    dict(name="Battery cycler lab",
         category="Energy", owner="In-house o Partner CNR-ICMATE",
         cost_capex_k="40", cost_opex_k_year="15",
         availability="Q2 2027", purpose="LiS/SS-Li discharge curve + cycle life",
         phase="B,C", notes="Multi-channel cycler 0-5V, ±10A"),
    dict(name="Climatic chamber (env. test)",
         category="Environmental", owner="External lab (Pavia / Milano)",
         cost_capex_k="0", cost_opex_k_year="20",
         availability="Always", purpose="-40/+85 degC + humidity + thermal cycling",
         phase="B,C", notes="MIL-STD-810 / DO-160G qualified facility"),
    dict(name="EMC chamber accredited (Test House)",
         category="EMC", owner="External lab (IMQ, ICIM)",
         cost_capex_k="0", cost_opex_k_year="30",
         availability="Always", purpose="EN 55032 + DO-160G section 20 EMI",
         phase="B", notes="Anechoic chamber accreditata 17025"),
    dict(name="Blackbody source IR cal lab",
         category="Payload", owner="External (INRIM Torino)",
         cost_capex_k="0", cost_opex_k_year="10",
         availability="Always", purpose="NEdT cert IR sensor",
         phase="A,B", notes="Calibrazione INRIM con corpo nero standard"),
    dict(name="Drop test rig (parachute)",
         category="Safety", owner="In-house build",
         cost_capex_k="15", cost_opex_k_year="5",
         availability="Q3 2026", purpose="Parachute deploy time test",
         phase="B", notes="50 m drop rig + HSC video"),
    dict(name="GNSS simulator (spoofing / multipath)",
         category="GNC", owner="External (RegoLab Genova o vendor demo)",
         cost_capex_k="0 (rental)", cost_opex_k_year="15",
         availability="Q2 2027", purpose="GNSS spoofing detection test",
         phase="B", notes="Spirent / Skydel sim"),
    dict(name="Battery abuse test cell",
         category="Safety", owner="External (CNR-ITAE Messina)",
         cost_capex_k="0", cost_opex_k_year="20",
         availability="Q2 2027", purpose="Thermal runaway containment test",
         phase="B", notes="UN 38.3 + nail penetration"),
    dict(name="Flight test sites (Pentema + GATB Grottaglie)",
         category="Flight Test", owner="Partner DTA Puglia + Comune Torriglia",
         cost_capex_k="20 setup", cost_opex_k_year="30",
         availability="2026-2027", purpose="VLOS + BVLOS flight test",
         phase="C", notes="GATB Grottaglie certified for BVLOS test"),
    dict(name="Stratospheric test site (Sardegna o ESRANGE Kiruna)",
         category="HALE Flight Test", owner="Partner ESRANGE / ASI",
         cost_capex_k="0 (rental, plan B)", cost_opex_k_year="200",
         availability="2028+", purpose="Stratospheric subscale + perennial flight",
         phase="C (Phase B 6B HALE)", notes="EuroHAPS-adjacent partnership"),
]

# ===========================================================================
# 3. TEST SCHEDULE (Gantt)
# ===========================================================================
# Phases: Pre-Phase A (M0-M3), Phase A (M3-M12), Phase B (M12-M24), Phase C (M24-M36+)
# Activities aligned with gate G3, G4, G5, G6

SCHEDULE: List[Dict[str, str]] = [
    # Pre-Phase A -- documentation compliance
    dict(activity="Inspection compliance docs (C series)", start_m=0,  end_m=3,
         phase="Pre-A", gate="G1", owner="aviation-regulatory",
         output="Compliance matrix C-001..C-006 + RTM update"),
    dict(activity="Stakeholder workshop + Stakeholder Needs collection", start_m=0, end_m=3,
         phase="Pre-A", gate="G1", owner="snai-funding-expert",
         output="StNeeds-001..017 + provenance log"),

    # Phase A -- analysis
    dict(activity="Link budget analysis (Allegato A.7)", start_m=3, end_m=6,
         phase="A", gate="G2", owner="telecom-payload",
         output="A.7 Link Budget Report + Excel"),
    dict(activity="Energy balance analysis HALE 44N (Allegato A.6)", start_m=3, end_m=10,
         phase="A", gate="G3", owner="propulsion-energy-engineer",
         output="A.6 Energy Balance Report + Sim model"),
    dict(activity="FMECA preliminary (Allegato A.2)", start_m=3, end_m=9,
         phase="A", gate="G2", owner="systems-engineer",
         output="A.2 Risk Register + FMECA + FTA"),
    dict(activity="Trade Studies TS-PLATFORM-6A, TS-MATERIAL, TS-PROP-6B", start_m=3, end_m=9,
         phase="A", gate="G2", owner="aerospace-SE",
         output="TS reports A.3 DOCFAP"),
    dict(activity="ENAC pre-application + SORA stima", start_m=3, end_m=9,
         phase="A", gate="G3", owner="aviation-regulatory",
         output="SORA pre-app feedback + GRC computation"),
    dict(activity="AGCOM spectrum consultation", start_m=6, end_m=10,
         phase="A", gate="G3", owner="telecom-payload",
         output="Spectrum allocation plan + license roadmap"),
    dict(activity="DPIA preliminary + Garante engagement", start_m=3, end_m=9,
         phase="A", gate="G3", owner="data-privacy-counsel",
         output="DPIA preliminary doc"),

    # Phase B -- demonstration + subscale test
    dict(activity="GS test bed setup + walkthrough", start_m=12, end_m=15,
         phase="B", gate="G4", owner="aerospace-SE",
         output="GS Pentema operational + checklists"),
    dict(activity="HIL simulator FCS fault injection", start_m=13, end_m=20,
         phase="B", gate="G4", owner="avionics-gnc-engineer",
         output="HIL test report + AVI-002 IMU + AVI-003 IMU fault test"),
    dict(activity="Range test ground RF + payload swap test", start_m=14, end_m=18,
         phase="B", gate="G4", owner="telecom-payload",
         output="RF range test + payload swap timing"),
    dict(activity="Climatic chamber test (E series)", start_m=15, end_m=18,
         phase="B", gate="G4", owner="vtol-uas-specialist",
         output="-10/+30 degC operational env. + thermal cycle"),
    dict(activity="Solar panel + battery lab test", start_m=24, end_m=30,
         phase="B (6B)", gate="G6", owner="propulsion-energy-engineer",
         output="Panel cert IEC 60904 + battery cycle curve"),
    dict(activity="Wind tunnel subscale 1:5 + 1:3 (POLITO/CIRA)", start_m=24, end_m=30,
         phase="B (6B)", gate="G6", owner="aero-structures-engineer",
         output="WT report + aeroelastic correlation"),
    dict(activity="Subscale prototype 1:3 ground tests", start_m=30, end_m=36,
         phase="B (6B)", gate="G6", owner="aero-structures-engineer",
         output="Subscale prototype TRL 5"),

    # Phase C -- full-scale flight test + operations validation (only 6A)
    dict(activity="Flight test BVLOS Pentema + GATB Grottaglie", start_m=9, end_m=12,
         phase="C (6A)", gate="G4", owner="vtol-uas-specialist",
         output="Flight test log + ENAC operational auth"),
    dict(activity="Operations validation Y1 (50+ missioni)", start_m=10, end_m=24,
         phase="C (6A)", gate="G5", owner="program-manager",
         output="Operations Y1 report + customer feedback"),
    dict(activity="Subscale HALE stratospheric flight test", start_m=36, end_m=48,
         phase="C (6B)", gate="post-G6", owner="aero-structures-engineer",
         output="Stratospheric flight test report TRL 6"),

    # Documentation / Independent review
    dict(activity="Independent verification audit (RINA / DNV)", start_m=9, end_m=11,
         phase="A", gate="G3", owner="program-manager",
         output="Third-party feasibility audit report"),
    dict(activity="V&V documentation deliverables consolidation", start_m=9, end_m=11,
         phase="A", gate="G3", owner="aerospace-SE",
         output="V&V plan v1.0 + qualification matrix v1.0"),
]

# ===========================================================================
# 4. TEST COSTS  (CapEx + OpEx test)
# ===========================================================================

COSTS: List[Dict[str, str]] = [
    # CapEx test infrastructure (Y1-Y2 6A + Phase B 6B)
    dict(item="GS test bed setup (Pentema)", category="CapEx", phase="A,B",
         cost_eur_k=35.0, source="vendor quote container + radio + UPS"),
    dict(item="HIL simulator hardware + software", category="CapEx", phase="B",
         cost_eur_k=80.0, source="vendor quote NI/DiPSIM"),
    dict(item="Range test ground RF equipment", category="CapEx", phase="B",
         cost_eur_k=20.0, source="vendor quote antenna tracking + analyzer"),
    dict(item="Battery cycler lab in-house", category="CapEx", phase="B",
         cost_eur_k=40.0, source="vendor quote Arbin/Maccor 4-channel"),
    dict(item="Drop test rig (parachute)", category="CapEx", phase="B",
         cost_eur_k=15.0, source="in-house build"),
    dict(item="Flight test sites setup (Pentema + GATB hospital fee)", category="CapEx", phase="C",
         cost_eur_k=20.0, source="GATB Grottaglie fee + Pentema setup"),

    # External lab fees (OpEx test)
    dict(item="Wind tunnel rental (UniGE DICCA or POLITO)", category="OpEx", phase="B",
         cost_eur_k=50.0, source="POLITO/UniGE rental sheet"),
    dict(item="Wind tunnel HALE subscale 1:3 (CIRA)", category="OpEx Phase B 6B", phase="B",
         cost_eur_k=80.0, source="CIRA rate card"),
    dict(item="GVT external (POLITO DIMEAS / CIRA)", category="OpEx Phase B 6B", phase="B",
         cost_eur_k=60.0, source="POLITO partnership quote"),
    dict(item="Solar simulator (ENEA Casaccia / EURAC)", category="OpEx Phase B 6B", phase="B",
         cost_eur_k=20.0, source="ENEA rate card"),
    dict(item="Climatic chamber (external lab)", category="OpEx", phase="B,C",
         cost_eur_k=20.0, source="lab rate"),
    dict(item="EMC chamber accredited", category="OpEx", phase="B",
         cost_eur_k=30.0, source="IMQ ICIM rate"),
    dict(item="Blackbody IR cal (INRIM Torino)", category="OpEx", phase="B",
         cost_eur_k=10.0, source="INRIM accreditation fees"),
    dict(item="GNSS simulator rental", category="OpEx", phase="B",
         cost_eur_k=15.0, source="Spirent/Skydel rental"),
    dict(item="Battery abuse test cell (CNR-ITAE)", category="OpEx", phase="B",
         cost_eur_k=20.0, source="CNR partnership rate"),
    dict(item="Flight test operations Y1 (fuel + pilot + travel)", category="OpEx", phase="C",
         cost_eur_k=30.0, source="estimate 30 missioni @1k each"),
    dict(item="Stratospheric test site (Phase B 6B)", category="OpEx Phase B 6B", phase="C",
         cost_eur_k=200.0, source="ESRANGE / partner rate (estimate)"),

    # Independent verification
    dict(item="RINA / DNV feasibility audit", category="OpEx", phase="A",
         cost_eur_k=35.0, source="RINA quote feasibility audit aerospace"),
    dict(item="Cybersecurity audit (NIS2)", category="OpEx", phase="B",
         cost_eur_k=18.0, source="ACN-compliant cyber audit firm"),
    dict(item="DPIA + GDPR audit", category="OpEx", phase="A",
         cost_eur_k=12.0, source="data privacy counsel + 3rd party"),
]


# ===========================================================================
# 5. INDEPENDENT VERIFICATION
# ===========================================================================

INDEP_VERIF: List[Dict[str, str]] = [
    dict(scope="Feasibility audit overall (volumi 1+2+3)",
         body="RINA (preferred) o DNV / SGS", phase="A (M+9-11)",
         deliverable="Independent feasibility statement",
         cost_eur_k=35.0, mandatory="Recommended (not mandatory by bando)",
         note="Gate G3 'investment-grade' upgrade. Increase confidence dei finanziatori."),
    dict(scope="ConOps SORA + GRC validation",
         body="ENAC (pre-application)", phase="A (M+3-9)",
         deliverable="ENAC feedback letter + SAIL stima",
         cost_eur_k=0.0, mandatory="Necessary for SORA submission",
         note="Pre-application meeting non vincolante; full application formal."),
    dict(scope="Cybersecurity NIS2 audit",
         body="ACN-accredited audit firm", phase="B (M+12-18)",
         deliverable="NIS2 compliance audit report",
         cost_eur_k=18.0, mandatory="Necessary for operations (NIS2 enters force)",
         note="Verifica Article 21 NIS2 misure tecniche."),
    dict(scope="DPIA + GDPR audit",
         body="3rd-party data privacy counsel (external)", phase="A (M+3-9)",
         deliverable="DPIA report + Garante engagement",
         cost_eur_k=12.0, mandatory="Required for personal data processing",
         note="Pubblicazione DPIA + workshop comunita' Pentema."),
    dict(scope="Wind tunnel test report validation (subscale)",
         body="POLITO DIMEAS / UniGE DICCA", phase="B (M+24-30)",
         deliverable="Wind tunnel certified report",
         cost_eur_k=0.0, mandatory="Required for Phase B 6B aerodynamic baseline",
         note="Phase B 6B; budget Phase B."),
    dict(scope="GVT (Ground Vibration Test) certification",
         body="POLITO / CIRA", phase="B (M+30-36)",
         deliverable="GVT modal analysis report",
         cost_eur_k=0.0, mandatory="Required for flight test clearance",
         note="Phase B 6B; flutter margin verification."),
    dict(scope="Solar panel efficiency cert (IEC 60904)",
         body="ENEA Casaccia / EURAC Bolzano (TUV cert)", phase="B (M+24)",
         deliverable="IEC 60904 cert + flash test data",
         cost_eur_k=0.0, mandatory="Required for energy budget compliance",
         note="Phase B 6B."),
    dict(scope="Battery thermal runaway (UN 38.3)",
         body="CNR-ITAE / accredited lab", phase="B (M+24-30)",
         deliverable="UN 38.3 test cert",
         cost_eur_k=20.0, mandatory="Required for transport + flight clearance",
         note="Phase B 6B + 6A if Li-S pack."),
    dict(scope="EMC compliance (EN 55032 + DO-160G)",
         body="IMQ / ICIM / TUV", phase="B (M+15-18)",
         deliverable="EMC test report",
         cost_eur_k=30.0, mandatory="Required for ENAC operational authorization",
         note="Phase B 6A operational pre-requisite."),
]


# =============================================================================
# 6. SHEET BUILDERS
# =============================================================================

def add_cover(wb: Workbook) -> None:
    ws = wb.create_sheet("Cover", 0)
    ws["A1"] = "A.5 -- VERIFICATION & VALIDATION PLAN v1.0"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
    ws["A2"] = "Firmamento Technologies -- Studio di Fattibilita' HALE / VTOL"
    ws["A2"].font = Font(size=12, italic=True)
    ws["A3"] = "Volume 2 -- Allegato A.5 -- Generato 2026-05-17"
    ws["A3"].font = Font(size=10, italic=True, color="888888")

    row = 5
    ws.cell(row=row, column=1, value="Scope").font = Font(bold=True, size=12)
    row += 1
    for line in [
        "Plan di verifica e validazione di tutti i requisiti di sistema (SyR) e ",
        "subsystem (SsR) del progetto HALE/VTOL, suddiviso per i due percorsi:",
        "  -- Percorso 6A VTOL operativo (Phase A -> Phase B -> Phase C)",
        "  -- Percorso 6B HALE R&D (Phase B subscale ground + flight)",
        "",
        "Conformita': NASA SE Handbook Rev 2 §5.3 (Verification) + §5.4 (Validation)",
        "Allineamento con cap-03-requisiti §3.7, cap-06-tecnica §6.6, cap-09 Gate.",
    ]:
        ws.cell(row=row, column=1, value=line)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Metodi V&V standard (4)").font = Font(bold=True, size=12)
    row += 1
    methods = [
        ("I -- Inspection",   "Verifica documentale/visiva (datasheet, BoM, label, process audit)."),
        ("A -- Analysis",     "Calcoli, simulazioni, modeling (link budget, energy balance, FMECA)."),
        ("D -- Demonstration","Esercizio operativo del sistema (qualitative, no measurement)."),
        ("T -- Test",         "Misure quantitative su prototype/articolo di volo."),
    ]
    for k, v in methods:
        ws.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws.cell(row=row, column=2, value=v).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Fasi V&V (NASA SE V-model)").font = Font(bold=True, size=12)
    row += 1
    phases = [
        ("Pre-Phase A (M+0-3)", "Inspection compliance documentale (C series)"),
        ("Phase A (M+3-12)",    "Analysis (link budget, energy balance, FMECA, FTA)"),
        ("Phase B (M+12-24)",   "Demonstration test bed + Test subscale (6A operativo)"),
        ("Phase C (M+24+)",     "Test full-scale + operational validation"),
        ("Phase B 6B (M+24-48)","Lab + subscale 1:3 + stratospheric subscale flight"),
    ]
    for k, v in phases:
        ws.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws.cell(row=row, column=2, value=v).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Sheets in this workbook").font = Font(bold=True, size=12)
    row += 1
    sheets = [
        ("Cover",                    "Documento di copertura"),
        ("VV_Matrix_SyR",            f"V&V matrix: {len(VV_MATRIX)} requisiti SyR/SsR/IR"),
        ("VV_Methods_Detail",        "Dettaglio dei 4 metodi I/A/D/T con esempi"),
        ("Test_Facilities",          f"Test facility catalog ({len(FACILITIES)} facility)"),
        ("Test_Schedule",            f"Gantt test schedule ({len(SCHEDULE)} attivita')"),
        ("Test_Costs",               f"CapEx + OpEx test ({len(COSTS)} voci)"),
        ("Independent_Verification", f"Independent verification ({len(INDEP_VERIF)} review)"),
    ]
    for n, d in sheets:
        ws.cell(row=row, column=1, value=n).font = Font(bold=True)
        ws.cell(row=row, column=2, value=d)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    # ---- statistics ----
    row += 2
    ws.cell(row=row, column=1, value="VV Matrix statistics").font = Font(bold=True, size=12, color="C00000")
    row += 1
    # count methods
    method_count: Dict[str, int] = {"I": 0, "A": 0, "D": 0, "T": 0}
    for r in VV_MATRIX:
        for code in ["I", "A", "D", "T"]:
            if code in r["method"]:
                method_count[code] += 1
    ws.cell(row=row, column=1, value=f"Totale requisiti tracciati: {len(VV_MATRIX)}").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value=f"  Method I (Inspection): {method_count['I']}").fill = ZONE_FILL["I"]
    row += 1
    ws.cell(row=row, column=1, value=f"  Method A (Analysis):   {method_count['A']}").fill = ZONE_FILL["A"]
    row += 1
    ws.cell(row=row, column=1, value=f"  Method D (Demonstration): {method_count['D']}").fill = ZONE_FILL["D"]
    row += 1
    ws.cell(row=row, column=1, value=f"  Method T (Test):       {method_count['T']}").fill = ZONE_FILL["T"]
    row += 1

    # ---- phase distribution ----
    phase_count: Dict[str, int] = {"Pre-A": 0, "A": 0, "B": 0, "C": 0}
    for r in VV_MATRIX:
        for ph in phase_count.keys():
            if ph in r["phase"]:
                phase_count[ph] += 1
    row += 1
    ws.cell(row=row, column=1, value="Distribuzione per fase (req che includono la fase):").font = Font(bold=True)
    row += 1
    for ph, n in phase_count.items():
        ws.cell(row=row, column=1, value=f"  {ph}: {n}")
        row += 1

    autosize(ws, 6, base=20)


def add_vv_matrix(wb: Workbook) -> None:
    ws = wb.create_sheet("VV_Matrix_SyR")
    header = ["Req-ID", "Description", "Family", "Method", "Phase", "Owner",
              "Test Facility", "Document Evidence", "Status", "Risk-ID link",
              "Trade Study link", "Notes"]
    for c, h in enumerate(header, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(header))
    for i, r in enumerate(VV_MATRIX, start=2):
        ws.cell(row=i, column=1, value=r["req_id"])
        ws.cell(row=i, column=2, value=r["desc"])
        ws.cell(row=i, column=3, value=r["family"])
        ws.cell(row=i, column=4, value=r["method"])
        ws.cell(row=i, column=5, value=r["phase"])
        ws.cell(row=i, column=6, value=r["owner"])
        ws.cell(row=i, column=7, value=r["facility"])
        ws.cell(row=i, column=8, value=r["evidence"])
        ws.cell(row=i, column=9, value=r["status"])
        ws.cell(row=i, column=10, value=r["risk"])
        ws.cell(row=i, column=11, value=r["ts"])
        ws.cell(row=i, column=12, value=r["note"])
        for c in range(1, len(header)+1):
            ws.cell(row=i, column=c).border = THIN_BORDER
            ws.cell(row=i, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        # color by method-primary
        primary = r["method"].split("+")[0].strip()
        if primary in ZONE_FILL:
            ws.cell(row=i, column=4).fill = ZONE_FILL[primary]
        # color Status
        if r["status"] == "Done":
            ws.cell(row=i, column=9).fill = OK_FILL
        elif r["status"] == "Open":
            ws.cell(row=i, column=9).fill = WARN_FILL

    # freeze top + first col
    ws.freeze_panes = "C2"
    autosize(ws, len(header), base=10)
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["L"].width = 40


def add_vv_methods_detail(wb: Workbook) -> None:
    ws = wb.create_sheet("VV_Methods_Detail")
    ws["A1"] = "V&V Methods Detail"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A2"] = "Reference: NASA SE Handbook Rev 2 §5.3-5.4 + INCOSE SE Handbook v5"
    ws["A2"].font = Font(italic=True, size=9)

    header = ["Method", "Code", "Definition", "Esempio progetto HALE/VTOL",
              "Tipico per", "Limiti", "Documentazione output"]
    for c, h in enumerate(header, start=4):
        ws.cell(row=4, column=c, value=h)
    # use full row
    header2 = ["Method", "Code", "Definition", "Esempio progetto HALE/VTOL",
               "Tipico per", "Limiti", "Documentazione output"]
    for c, h in enumerate(header2, start=1):
        ws.cell(row=4, column=c, value=h)
    style_header(ws, 4, len(header2))

    rows = [
        ("Inspection", "I",
         "Verifica documentale / visiva: lettura di datasheet, BoM, label, processo, contract.",
         "SyR-C-001 Class UAS via vendor DoC, SyR-C-005 polizza assicurativa.",
         "Compliance requirements, contract terms, statici.",
         "Non valida prestazioni operative; sufficiente per requisiti documentali.",
         "Inspection record + signed-off checklist."),
        ("Analysis", "A",
         "Calcoli, simulazioni, modeling. Verifica per analytical proof.",
         "Allegato A.6 Energy Balance HALE 44N; Allegato A.7 Link Budget; FMECA; FTA.",
         "Requisiti di prestazione misurabili pre-test; design verification.",
         "Validita' dipende da assunzioni modello; richiede V&V del modello stesso.",
         "Analysis report + assumption list + uncertainty quantification."),
        ("Demonstration", "D",
         "Esercizio del sistema (qualitativo): functional demonstration end-to-end.",
         "SyR-F-001 mission EO Pentema; SyR-O-001 lost-link RTB; SsR-AVI-003 IMU fault injection.",
         "Functional + operational requirements; ConOps validation.",
         "Qualitativo: no measurement of margin/accuracy.",
         "Demo log + video record + sign-off da test director."),
        ("Test", "T",
         "Misure quantitative su prototype/articolo di volo.",
         "SsR-PROP-003 battery cycle test; SyR-P-004 wind envelope test; SsR-COMMS-001 RSSI flight.",
         "Performance + safety requirements con margin quantification.",
         "Costoso; richiede facility qualificata; test asset.",
         "Test plan + procedure + data sheet + test report + uncertainty budget."),
    ]
    for i, r in enumerate(rows, start=5):
        for c, v in enumerate(r, start=1):
            ws.cell(row=i, column=c, value=v).border = THIN_BORDER
            ws.cell(row=i, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=i, column=2).fill = ZONE_FILL[rows[i-5][1]]

    autosize(ws, len(header2), base=20)


def add_facilities(wb: Workbook) -> None:
    ws = wb.create_sheet("Test_Facilities")
    header = ["Facility", "Category", "Owner", "CapEx [k EUR]", "OpEx/year [k EUR]",
              "Availability", "Purpose", "Phase", "Notes"]
    for c, h in enumerate(header, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(header))
    for i, f in enumerate(FACILITIES, start=2):
        ws.cell(row=i, column=1, value=f["name"])
        ws.cell(row=i, column=2, value=f["category"])
        ws.cell(row=i, column=3, value=f["owner"])
        ws.cell(row=i, column=4, value=f["cost_capex_k"])
        ws.cell(row=i, column=5, value=f["cost_opex_k_year"])
        ws.cell(row=i, column=6, value=f["availability"])
        ws.cell(row=i, column=7, value=f["purpose"])
        ws.cell(row=i, column=8, value=f["phase"])
        ws.cell(row=i, column=9, value=f["notes"])
        for c in range(1, len(header)+1):
            ws.cell(row=i, column=c).border = THIN_BORDER
            ws.cell(row=i, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws, len(header), base=14)


def add_schedule(wb: Workbook) -> None:
    ws = wb.create_sheet("Test_Schedule")
    ws["A1"] = "V&V Test Schedule -- Gantt aligned to Gates G2/G3/G4/G5/G6"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")

    header = ["Activity", "Start (M+)", "End (M+)", "Duration [m]", "Phase", "Gate", "Owner",
              "Output / deliverable"]
    for c, h in enumerate(header, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(header))
    # Gantt bar columns (M+0 .. M+48 step 3)
    months = list(range(0, 49, 3))
    for j, m in enumerate(months, start=len(header)+1):
        ws.cell(row=3, column=j, value=f"M+{m}")
    # style timeline header
    for j in range(len(header)+1, len(header)+1+len(months)):
        cell = ws.cell(row=3, column=j)
        cell.fill = SUBHDR_FILL
        cell.font = SUBHDR_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for i, s in enumerate(SCHEDULE, start=4):
        ws.cell(row=i, column=1, value=s["activity"])
        ws.cell(row=i, column=2, value=s["start_m"])
        ws.cell(row=i, column=3, value=s["end_m"])
        ws.cell(row=i, column=4, value=s["end_m"] - s["start_m"])
        ws.cell(row=i, column=5, value=s["phase"])
        ws.cell(row=i, column=6, value=s["gate"])
        ws.cell(row=i, column=7, value=s["owner"])
        ws.cell(row=i, column=8, value=s["output"])
        for c in range(1, len(header)+1):
            ws.cell(row=i, column=c).border = THIN_BORDER
            ws.cell(row=i, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        # color gantt cells
        color_map = {"Pre-A": "BDD7EE", "A": "9BC2E6", "B": "FFE699", "C": "C6EFCE",
                     "B (6B)": "F4B084", "C (6A)": "B4C7E7", "C (6B)": "FF99CC"}
        ph = s["phase"]
        fill_color = color_map.get(ph, "DDDDDD")
        for j, m in enumerate(months, start=len(header)+1):
            cell = ws.cell(row=i, column=j)
            cell.border = THIN_BORDER
            if s["start_m"] <= m <= s["end_m"]:
                cell.fill = PatternFill("solid", fgColor=fill_color)

    ws.freeze_panes = "B4"
    autosize(ws, len(header), base=12)
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["H"].width = 35
    for j in range(len(header)+1, len(header)+1+len(months)):
        ws.column_dimensions[get_column_letter(j)].width = 6


def add_costs(wb: Workbook) -> None:
    ws = wb.create_sheet("Test_Costs")
    header = ["Item", "Category", "Phase", "Cost [k EUR]", "Source / vendor"]
    for c, h in enumerate(header, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(header))

    capex_total = 0.0
    opex_total = 0.0
    phase_b_6b_total = 0.0
    for i, x in enumerate(COSTS, start=2):
        ws.cell(row=i, column=1, value=x["item"])
        ws.cell(row=i, column=2, value=x["category"])
        ws.cell(row=i, column=3, value=x["phase"])
        ws.cell(row=i, column=4, value=x["cost_eur_k"])
        ws.cell(row=i, column=5, value=x["source"])
        for c in range(1, len(header)+1):
            ws.cell(row=i, column=c).border = THIN_BORDER
            ws.cell(row=i, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        if "CapEx" in x["category"]:
            capex_total += float(x["cost_eur_k"])
            ws.cell(row=i, column=4).fill = ZONE_FILL["I"]
        elif "Phase B 6B" in x["category"]:
            phase_b_6b_total += float(x["cost_eur_k"])
            ws.cell(row=i, column=4).fill = ZONE_FILL["D"]
        else:
            opex_total += float(x["cost_eur_k"])
            ws.cell(row=i, column=4).fill = ZONE_FILL["A"]

    # totals
    row = len(COSTS) + 3
    ws.cell(row=row, column=1, value="TOTAL CapEx Test (6A Phase A/B/C)").font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(capex_total, 1)).font = Font(bold=True)
    ws.cell(row=row, column=4).fill = ZONE_FILL["I"]
    row += 1
    ws.cell(row=row, column=1, value="TOTAL OpEx Test (6A Phase A/B/C) Y1+Y2").font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(opex_total, 1)).font = Font(bold=True)
    ws.cell(row=row, column=4).fill = ZONE_FILL["A"]
    row += 1
    ws.cell(row=row, column=1, value="TOTAL Phase B 6B (HALE R&D)").font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(phase_b_6b_total, 1)).font = Font(bold=True)
    ws.cell(row=row, column=4).fill = ZONE_FILL["D"]
    row += 1
    ws.cell(row=row, column=1, value="GRAND TOTAL (test V&V budget complete)").font = Font(bold=True, color="C00000")
    ws.cell(row=row, column=4, value=round(capex_total + opex_total + phase_b_6b_total, 1)).font = Font(bold=True, color="C00000")
    ws.cell(row=row, column=4).fill = WARN_FILL

    autosize(ws, len(header), base=14)


def add_independent_verif(wb: Workbook) -> None:
    ws = wb.create_sheet("Independent_Verification")
    header = ["Scope", "Body", "Phase / Timing", "Deliverable", "Cost [k EUR]", "Mandatory?", "Note"]
    for c, h in enumerate(header, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(header))
    for i, r in enumerate(INDEP_VERIF, start=2):
        ws.cell(row=i, column=1, value=r["scope"])
        ws.cell(row=i, column=2, value=r["body"])
        ws.cell(row=i, column=3, value=r["phase"])
        ws.cell(row=i, column=4, value=r["deliverable"])
        ws.cell(row=i, column=5, value=r["cost_eur_k"])
        ws.cell(row=i, column=6, value=r["mandatory"])
        ws.cell(row=i, column=7, value=r["note"])
        for c in range(1, len(header)+1):
            ws.cell(row=i, column=c).border = THIN_BORDER
            ws.cell(row=i, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        # color "Mandatory" cells
        if "Required" in r["mandatory"] or "Necessary" in r["mandatory"]:
            ws.cell(row=i, column=6).fill = WARN_FILL
        elif "Recommended" in r["mandatory"]:
            ws.cell(row=i, column=6).fill = OK_FILL

    # totals
    total = sum(float(r["cost_eur_k"]) for r in INDEP_VERIF)
    row = len(INDEP_VERIF) + 3
    ws.cell(row=row, column=1, value="TOTAL Independent Verification budget").font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(total, 1)).font = Font(bold=True)
    ws.cell(row=row, column=5).fill = WARN_FILL

    autosize(ws, len(header), base=14)


# =============================================================================
# 7. MAIN
# =============================================================================

def main() -> None:
    print("[1/3] Building V&V Plan workbook ...")
    wb = Workbook()
    wb.remove(wb.active)
    add_cover(wb)
    add_vv_matrix(wb)
    add_vv_methods_detail(wb)
    add_facilities(wb)
    add_schedule(wb)
    add_costs(wb)
    add_independent_verif(wb)

    print("[2/3] Saving ...")
    wb.save(XLSX_PATH)
    print(f"   -> {XLSX_PATH}")

    print("[3/3] Stats:")
    print(f"   - V&V matrix rows:    {len(VV_MATRIX)}")
    print(f"   - Test facilities:    {len(FACILITIES)}")
    print(f"   - Schedule activities:{len(SCHEDULE)}")
    print(f"   - Cost items:         {len(COSTS)}")
    print(f"   - Indep. verif. rows: {len(INDEP_VERIF)}")
    print("Done.")


if __name__ == "__main__":
    main()
